from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from .models import Framework, Policy, SubPolicy, FrameworkVersion, PolicyVersion, PolicyApproval, Users, Department, BusinessUnit, Entity, Location, RBAC, GRCLog
from .serializers import FrameworkSerializer, PolicySerializer, SubPolicySerializer, PolicyApprovalSerializer, UserSerializer   
from django.db import transaction
import traceback
import sys
import datetime
import re
from datetime import datetime, timedelta
from django.utils import timezone
from .routes.Consent import require_consent

from django.db.models import Count, Avg, Case, When, Value, FloatField, F
from django.db.models.functions import Coalesce, Cast
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from .models import Users
import json
import logging
from django.views.decorators.http import require_http_methods
from django.db import connection
from django.contrib.auth.hashers import make_password, check_password
from django.core.cache import cache
from django.core.mail import send_mail
from django.conf import settings
import time

logger = logging.getLogger(__name__)

# In-memory OTP storage as fallback for session issues
otp_storage = {}

# Helper function to get default framework for logging
def _get_default_framework():
    """Get a default framework for logging purposes"""
    try:
        # Try to get the first active framework
        framework = Framework.objects.filter(ActiveInactive='Active').first()
        if framework:
            return framework
        # If no active framework, get any framework
        framework = Framework.objects.first()
        if framework:
            return framework
    except Exception as e:
        logger.warning(f"Error getting default framework for logging: {str(e)}")
    return None

# Helper function to get client IP address
def _get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR', 'unknown')
    
    # Sanitize IP: remove port if present, truncate to 45 chars
    if ip and ip != 'unknown':
        # Remove port number if present (IPv4 only, not IPv6)
        if ':' in ip and not ip.startswith('['):
            parts = ip.split(':')
            if len(parts) == 2 and '.' in parts[0]:
                ip = parts[0]
        # Truncate to max 45 characters (database column limit)
        ip = ip[:45] if len(ip) > 45 else ip
    
    return ip

# Track verified emails for password reset
verified_emails = set()
# Map email to UserId to ensure correct user password reset
verified_users_mapping = {}

# Framework CRUD operations

"""
@api GET /api/frameworks/
Returns all frameworks with Status='Approved' and ActiveInactive='Active'.
Filtered by the serializer to include only policies with Status='Approved' and ActiveInactive='Active',
and subpolicies with Status='Approved'.

@api POST /api/frameworks/
Creates a new framework with associated policies and subpolicies.
New frameworks are created with Status='Under Review' and ActiveInactive='Inactive' by default.
CurrentVersion defaults to 1.0 if not provided.

Example payload:
{
  "FrameworkName": "ISO 27001",
  "FrameworkDescription": "Information Security Management System",
  "EffectiveDate": "2023-10-01",
  "CreatedByName": "John Doe",
  "CreatedByDate": "2023-09-15",
  "Category": "Information Security and Compliance",
  "DocURL": "https://example.com/iso27001",
  "Identifier": "ISO-27001",
  "StartDate": "2023-10-01",
  "EndDate": "2025-10-01",
  "policies": [
    {
      "PolicyName": "Access Control Policy",
      "PolicyDescription": "Guidelines for access control management",
      "StartDate": "2023-10-01",
      "Department": "IT",
      "Applicability": "All Employees",
      "Scope": "All IT systems",
      "Objective": "Ensure proper access control",
      "Identifier": "ACP-001",
      "subpolicies": [
        {
          "SubPolicyName": "Password Management",
          "Identifier": "PWD-001",
          "Description": "Password requirements and management",
          "PermanentTemporary": "Permanent",
          "Control": "Use strong passwords with at least 12 characters"
        }
      ]
    }
  ]
}
"""
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def framework_list(request):
    if request.method == 'GET':
        frameworks = Framework.objects.filter(Status='Approved', ActiveInactive='Active')
        serializer = FrameworkSerializer(frameworks, many=True)
        return Response(serializer.data)
 
    elif request.method == 'POST':
        try:
            with transaction.atomic():
                # Prepare incoming data
                data = request.data.copy()
 
                # Set default values if not provided
                data.setdefault('Status', 'Under Review')
                data.setdefault('ActiveInactive', 'Inactive')
               
                # Always set CreatedByDate to current date
                data['CreatedByDate'] = datetime.date.today()
 
                # Set version to 1.0 for all new frameworks
                new_version = 1.0
 
                # Create Framework
                framework_serializer = FrameworkSerializer(data=data)
                if not framework_serializer.is_valid():
                    return Response(framework_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
 
                framework = framework_serializer.save()
                framework.CurrentVersion = new_version
                framework.save()
 
                # Create FrameworkVersion
                framework_version = FrameworkVersion(
                    FrameworkId=framework,
                    Version=framework.CurrentVersion,
                    FrameworkName=framework.FrameworkName,
                    CreatedBy=framework.CreatedByName,
                    CreatedDate=datetime.date.today(),  # Always use current date
                    PreviousVersionId=None
                )
                framework_version.save()
 
                # Handle Policies if provided
                policies_data = request.data.get('policies', [])
                created_policies_count = 0
                created_subpolicies_count = 0
               
                for policy_data in policies_data:
                    policy_data = policy_data.copy()
                    policy_data['FrameworkId'] = framework.FrameworkId
                    policy_data['CurrentVersion'] = framework.CurrentVersion
                    policy_data.setdefault('Status', 'Under Review')
                    policy_data.setdefault('ActiveInactive', 'Inactive')
                    policy_data.setdefault('CreatedByName', framework.CreatedByName)
                    policy_data['CreatedByDate'] = datetime.date.today()  # Always use current date
                   
                    # Get reviewer's name if reviewer ID is provided
                    reviewer_id = policy_data.get('Reviewer')
                    if reviewer_id:
                        reviewer_obj = Users.objects.filter(UserId=reviewer_id).first()
                        if reviewer_obj:
                            # Store the reviewer's name in the policy
                            policy_data['Reviewer'] = reviewer_obj.UserName
 
                    policy_serializer = PolicySerializer(data=policy_data)
                    if not policy_serializer.is_valid():
                        return Response(policy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
 
                    policy = policy_serializer.save()
                    created_policies_count += 1
 
                    policy_version = PolicyVersion(
                        PolicyId=policy,
                        Version=policy.CurrentVersion,
                        PolicyName=policy.PolicyName,
                        CreatedBy=policy.CreatedByName,
                        CreatedDate=datetime.date.today(),  # Always use current date
                        PreviousVersionId=None
                    )
                    policy_version.save()
                   
                    # Handle SubPolicies if provided
                    subpolicies_data = policy_data.get('subpolicies', [])
                    for subpolicy_data in subpolicies_data:
                        subpolicy_data = subpolicy_data.copy()
                        subpolicy_data['PolicyId'] = policy.PolicyId
                        subpolicy_data.setdefault('Status', 'Under Review')
                        subpolicy_data.setdefault('CreatedByName', policy.CreatedByName)
                        subpolicy_data['CreatedByDate'] = datetime.date.today()  # Always use current date
 
                        subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
                        if not subpolicy_serializer.is_valid():
                            return Response(subpolicy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                        subpolicy_serializer.save()
                        created_subpolicies_count += 1
 
                # Create a detailed success message
                message = f'Framework "{framework.FrameworkName}" created successfully'
                if created_policies_count > 0:
                    message += f' with {created_policies_count} policies'
                    if created_subpolicies_count > 0:
                        message += f' and {created_subpolicies_count} subpolicies'
                message += '!'
               
                return Response({
                    'message': message,
                    'FrameworkId': framework.FrameworkId,
                    'Version': framework.CurrentVersion
                }, status=status.HTTP_201_CREATED)
 
        except Exception as e:
            return Response({
                'error': 'Error creating framework',
                'details': {
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
            }, status=status.HTTP_400_BAD_REQUEST)
 
 

"""
@api GET /api/frameworks/{pk}/
Returns a specific framework by ID if it has Status='Approved' and ActiveInactive='Active'.

@api PUT /api/frameworks/{pk}/
Updates an existing framework. Only frameworks with Status='Approved' and ActiveInactive='Active' can be updated.

Example payload:
{
  "FrameworkName": "ISO 27001:2022",
  "FrameworkDescription": "Updated Information Security Management System",
  "Category": "Information Security",
  "DocURL": "https://example.com/iso27001-2022",
  "EndDate": "2026-10-01"
}

@api DELETE /api/frameworks/{pk}/
Soft-deletes a framework by setting ActiveInactive='Inactive'.
Also marks all related policies as inactive and all related subpolicies with Status='Inactive'.
"""
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([AllowAny])
def framework_detail(request, pk):
    framework = get_object_or_404(Framework, FrameworkId=pk)
    
    if request.method == 'GET':
        # Only return details if framework is Approved and Active
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get all active and approved policies for this framework
        policies = Policy.objects.filter(
            FrameworkId=framework,
            Status='Approved',
            ActiveInactive='Active'
        )
        
        # Get all subpolicies for these policies
        policy_data = []
        for policy in policies:
            policy_dict = {
                'PolicyId': policy.PolicyId,
                'PolicyName': policy.PolicyName,
                'PolicyDescription': policy.PolicyDescription,
                'CurrentVersion': policy.CurrentVersion,
                'StartDate': policy.StartDate,
                'EndDate': policy.EndDate,
                'Department': policy.Department,
                'CreatedByName': policy.CreatedByName,
                'CreatedByDate': policy.CreatedByDate,
                'Applicability': policy.Applicability,
                'DocURL': policy.DocURL,
                'Scope': policy.Scope,
                'Objective': policy.Objective,
                'Identifier': policy.Identifier,
                'PermanentTemporary': policy.PermanentTemporary,
                'subpolicies': []
            }
            
            # Get all subpolicies for this policy
            subpolicies = SubPolicy.objects.filter(PolicyId=policy)
            for subpolicy in subpolicies:
                subpolicy_dict = {
                    'SubPolicyId': subpolicy.SubPolicyId,
                    'SubPolicyName': subpolicy.SubPolicyName,
                    'CreatedByName': subpolicy.CreatedByName,
                    'CreatedByDate': subpolicy.CreatedByDate,
                    'Identifier': subpolicy.Identifier,
                    'Description': subpolicy.Description,
                    'Status': subpolicy.Status,
                    'PermanentTemporary': subpolicy.PermanentTemporary,
                    'Control': subpolicy.Control
                }
                policy_dict['subpolicies'].append(subpolicy_dict)
            
            policy_data.append(policy_dict)
        
        # Create response data
        response_data = {
            'FrameworkId': framework.FrameworkId,
            'FrameworkName': framework.FrameworkName,
            'CurrentVersion': framework.CurrentVersion,
            'FrameworkDescription': framework.FrameworkDescription,
            'EffectiveDate': framework.EffectiveDate,
            'CreatedByName': framework.CreatedByName,
            'CreatedByDate': framework.CreatedByDate,
            'Category': framework.Category,
            'DocURL': framework.DocURL,
            'Identifier': framework.Identifier,
            'StartDate': framework.StartDate,
            'EndDate': framework.EndDate,
            'Status': framework.Status,
            'ActiveInactive': framework.ActiveInactive,
            'policies': policy_data
        }
        
        return Response(response_data)
    
    elif request.method == 'PUT':
        # Check if framework is approved and active before allowing update
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Only approved and active frameworks can be updated'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            with transaction.atomic():
                serializer = FrameworkSerializer(framework, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Framework updated successfully',
                        'FrameworkId': framework.FrameworkId,
                        'CurrentVersion': framework.CurrentVersion
                    })
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error updating framework', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        try:
            with transaction.atomic():
                # Instead of deleting, set ActiveInactive to 'Inactive'
                framework.ActiveInactive = 'Inactive'
                framework.save()
                
                # Set all related policies to inactive
                policies = Policy.objects.filter(FrameworkId=framework)
                for policy in policies:
                    policy.ActiveInactive = 'Inactive'
                    policy.save()
                    
                    # Update Status of subpolicies since they don't have ActiveInactive field
                    subpolicies = SubPolicy.objects.filter(PolicyId=policy)
                    for subpolicy in subpolicies:
                        subpolicy.Status = 'Inactive'
                        subpolicy.save()
                
                return Response({'message': 'Framework and related policies marked as inactive'}, status=status.HTTP_200_OK)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error marking framework as inactive', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

# Policy CRUD operations

"""
@api GET /api/policies/{pk}/
Returns a specific policy by ID if it has Status='Approved' and ActiveInactive='Active',
and its parent framework has Status='Approved' and ActiveInactive='Active'.

@api PUT /api/policies/{pk}/
Updates an existing policy. Only policies with Status='Approved' and ActiveInactive='Active'
whose parent framework is also Approved and Active can be updated.

Example payload:
{
  "PolicyName": "Updated Access Control Policy",
  "PolicyDescription": "Enhanced guidelines for access control management with additional security measures",
  "StartDate": "2023-12-01",
  "EndDate": "2025-12-01",
  "Department": "IT,Security",
  "Scope": "All IT systems and cloud services",
  "Objective": "Ensure proper access control with improved security"
}

@api DELETE /api/policies/{pk}/
Soft-deletes a policy by setting ActiveInactive='Inactive'.
Also marks all related subpolicies with Status='Inactive'.
"""
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([AllowAny])
def policy_detail(request, pk):
    policy = get_object_or_404(Policy, PolicyId=pk)
    
    if request.method == 'GET':
        # Only return details if policy is Approved and Active
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Policy is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get framework to check if it's approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = PolicySerializer(policy)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        # Check if policy is approved and active before allowing update
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Only approved and active policies can be updated'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if framework is approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            with transaction.atomic():
                # Add status and ActiveInactive to request data
                update_data = request.data.copy()
                update_data['Status'] = 'Under Review'
                update_data['ActiveInactive'] = 'Inactive'
                
                serializer = PolicySerializer(policy, data=update_data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Policy updated successfully and set to Under Review',
                        'PolicyId': policy.PolicyId,
                        'CurrentVersion': policy.CurrentVersion,
                        'Status': 'Under Review',
                        'ActiveInactive': 'Inactive'
                    })
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error updating policy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        try:
            with transaction.atomic():
                # Instead of deleting, set ActiveInactive to 'Inactive'
                policy.ActiveInactive = 'Inactive'
                policy.save()
                
                # Update Status of subpolicies since they don't have ActiveInactive field
                subpolicies = SubPolicy.objects.filter(PolicyId=policy)
                for subpolicy in subpolicies:
                    subpolicy.Status = 'Inactive'
                    subpolicy.save()
                
                return Response({'message': 'Policy and related subpolicies marked as inactive'}, status=status.HTTP_200_OK)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error marking policy as inactive', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/frameworks/{framework_id}/policies/
Adds a new policy to an existing framework.
New policies are created with Status='Under Review' and ActiveInactive='Inactive' by default.
CurrentVersion defaults to 1.0 if not provided.

Example payload:
{
  "PolicyName": "Data Classification Policy",
  "PolicyDescription": "Guidelines for data classification and handling",
  "StartDate": "2023-10-01",
  "Department": "IT,Legal",
  "Applicability": "All Employees",
  "Scope": "All company data",
  "Objective": "Ensure proper data classification and handling",
  "Identifier": "DCP-001",
  "subpolicies": [
    {
      "SubPolicyName": "Confidential Data Handling",
      "Identifier": "CDH-001",
      "Description": "Guidelines for handling confidential data",
      "PermanentTemporary": "Permanent",
      "Control": "Encrypt all confidential data at rest and in transit"
    }
  ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
@require_consent('create_policy')
def add_policy_to_framework(request, framework_id):
    framework = get_object_or_404(Framework, FrameworkId=framework_id)
   
    try:
        with transaction.atomic():
            # Set framework ID and default values in the request data
            policy_data = request.data.copy()
            policy_data['FrameworkId'] = framework.FrameworkId
            policy_data['CurrentVersion'] = framework.CurrentVersion  # Use framework's version
            
            # Set default values if not provided
            if 'Status' not in policy_data:
                policy_data['Status'] = 'Under Review'
            if 'ActiveInactive' not in policy_data:
                policy_data['ActiveInactive'] = 'Inactive'
            if 'CreatedByName' not in policy_data or not policy_data['CreatedByName']:
                policy_data['CreatedByName'] = framework.CreatedByName
            if 'CreatedByDate' not in policy_data:
                policy_data['CreatedByDate'] = datetime.date.today()
            if 'Reviewer' not in policy_data:
                policy_data['Reviewer'] = None
           
            print("DEBUG: Policy data before serialization:", policy_data)
            policy_serializer = PolicySerializer(data=policy_data)
            print("DEBUG: validating policy serializer")
            if not policy_serializer.is_valid():
                print("Policy serializer errors:", policy_serializer.errors)
                return Response({
                    'error': 'Policy validation failed',
                    'details': policy_serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            print("DEBUG: serializer is valid")
 
            policy = policy_serializer.save()
 
            # Get reviewer ID directly from the request data
            reviewer_id = policy_data.get('Reviewer')  # This should be a UserId (number)
           
            # Get reviewer's name for the Policy table
            reviewer_name = None
            if reviewer_id:
                try:
                    reviewer_id = int(reviewer_id)  # Ensure reviewer_id is an integer
                    reviewer_obj = Users.objects.filter(UserId=reviewer_id).first()
                    if reviewer_obj:
                        reviewer_name = reviewer_obj.UserName
                        # Store reviewer name in the policy object
                        policy.Reviewer = reviewer_name
                        policy.save()
                except (ValueError, TypeError):
                    print(f"Warning: Invalid reviewer ID format: {reviewer_id}")
 
            # Get user id from CreatedByName
            created_by_name = policy_data.get('CreatedByName')
            user_obj = Users.objects.filter(UserName=created_by_name).first()
            user_id = user_obj.UserId if user_obj else None
 
            if user_id is None:
                print(f"Warning: CreatedBy user not found for: {created_by_name}")
            if reviewer_id is None:
                print("Warning: Reviewer id missing in request data")

            try:
                print("Creating PolicyVersion with:", {
                    "PolicyId": policy.PolicyId,
                    "Version": policy.CurrentVersion,
                    "PolicyName": policy.PolicyName,
                    "CreatedBy": policy.CreatedByName,
                    "CreatedDate": policy.CreatedByDate,
                    "PreviousVersionId": None
                })
 
                policy_version = PolicyVersion(
                    PolicyId=policy,
                    Version=policy.CurrentVersion,
                    PolicyName=policy.PolicyName,
                    CreatedBy=policy.CreatedByName,
                    CreatedDate=policy.CreatedByDate,
                    PreviousVersionId=None
                )
                policy_version.save()
            except Exception as e:
                print("Error creating PolicyVersion:", str(e))
                raise
 
           
            # Create subpolicies if provided
            subpolicies_data = request.data.get('subpolicies', [])
            created_subpolicies_count = 0
           
            for subpolicy_data in subpolicies_data:
                # Set policy ID and default values
                subpolicy_data = subpolicy_data.copy() if isinstance(subpolicy_data, dict) else {}
                subpolicy_data['PolicyId'] = policy.PolicyId
                if 'CreatedByName' not in subpolicy_data or not subpolicy_data['CreatedByName']:
                    subpolicy_data['CreatedByName'] = policy.CreatedByName
                if 'CreatedByDate' not in subpolicy_data:
                    subpolicy_data['CreatedByDate'] = datetime.date.today()
                if 'Status' not in subpolicy_data:
                    subpolicy_data['Status'] = 'Under Review'
                if 'PermanentTemporary' not in subpolicy_data:
                    subpolicy_data['PermanentTemporary'] = 'Permanent'
               
                print("DEBUG: SubPolicy data before serialization:", subpolicy_data)
                subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
                if not subpolicy_serializer.is_valid():
                    print("SubPolicy serializer errors:", subpolicy_serializer.errors)
                    return Response({
                        'error': 'SubPolicy validation failed',
                        'details': subpolicy_serializer.errors
                    }, status=status.HTTP_400_BAD_REQUEST)
                subpolicy_serializer.save()
                created_subpolicies_count += 1
               
            # Create a detailed success message
            message = 'Policy added to framework successfully'
            if created_subpolicies_count > 0:
                message += f' with {created_subpolicies_count} subpolicies'
            message += '!'
           
            return Response({
                'message': message,
                'PolicyId': policy.PolicyId,
                'FrameworkId': framework.FrameworkId,
                'Version': policy.CurrentVersion
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print("DEBUG: Error details:", error_info)
        return Response({
            'error': 'Error adding policy to framework',
            'details': error_info
        }, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/policies/{policy_id}/subpolicies/
Adds a new subpolicy to an existing policy.
New subpolicies are created with Status='Under Review' by default.

Example payload:
{
  "SubPolicyName": "Multi-Factor Authentication",
  "Identifier": "MFA-001",
  "Description": "Requirements for multi-factor authentication",
  "PermanentTemporary": "Permanent",
  "Control": "Implement MFA for all admin access and sensitive operations"
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def add_policy_to_framework(request, framework_id):
    framework = get_object_or_404(Framework, FrameworkId=framework_id)
   
    try:
        with transaction.atomic():
            # Set framework ID and default values in the request data
            policy_data = request.data.copy()
            policy_data['FrameworkId'] = framework.FrameworkId
            policy_data['CurrentVersion'] = framework.CurrentVersion  # Use framework's version
            if 'Status' not in policy_data:
                policy_data['Status'] = 'Under Review'
            if 'ActiveInactive' not in policy_data:
                policy_data['ActiveInactive'] = 'Inactive'
            if 'CreatedByName' not in policy_data:
                policy_data['CreatedByName'] = framework.CreatedByName
            if 'CreatedByDate' not in policy_data:
                policy_data['CreatedByDate'] = datetime.date.today()
           
            policy_serializer = PolicySerializer(data=policy_data)
            print("DEBUG: validating policy serializer")
            if not policy_serializer.is_valid():
                print("Policy serializer errors:", policy_serializer.errors)
                return Response(policy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            print("DEBUG: serializer is valid")
 
            policy = policy_serializer.save()
 
            # Get reviewer ID directly from the request data
            reviewer_id = policy_data.get('Reviewer')  # Changed from request.data to policy_data
           
            # Get reviewer's name for the Policy table
            reviewer_name = None
            if reviewer_id:
                reviewer_obj = Users.objects.filter(UserId=reviewer_id).first()
                if reviewer_obj:
                    reviewer_name = reviewer_obj.UserName
                    # Store reviewer name in the policy object
                    policy.Reviewer = reviewer_name
                    policy.save()
 
            # Get user id from CreatedByName
            created_by_name = policy_data.get('CreatedByName')
            user_obj = Users.objects.filter(UserName=created_by_name).first()
            user_id = user_obj.UserId if user_obj else None
 
            if user_id is None:
                print(f"Warning: CreatedBy user not found for: {created_by_name}")
            if reviewer_id is None:
                print("Warning: Reviewer id missing in request data")
 
            # No policy approval logic here - removed completely
 
            try:
                print("Creating PolicyVersion with:", {
                    "PolicyId": policy.PolicyId,
                    "Version": policy.CurrentVersion,
                    "PolicyName": policy.PolicyName,
                    "CreatedBy": policy.CreatedByName,
                    "CreatedDate": policy.CreatedByDate,
                    "PreviousVersionId": None
                })
 
                policy_version = PolicyVersion(
                    PolicyId=policy,
                    Version=policy.CurrentVersion,
                    PolicyName=policy.PolicyName,
                    CreatedBy=policy.CreatedByName,
                    CreatedDate=policy.CreatedByDate,
                    PreviousVersionId=None
                )
                policy_version.save()
            except Exception as e:
                print("Error creating PolicyVersion:", str(e))
                raise
 
            # Create subpolicies if provided
            subpolicies_data = request.data.get('subpolicies', [])
            created_subpolicies_count = 0
           
            for subpolicy_data in subpolicies_data:
                # Set policy ID and default values
                subpolicy_data = subpolicy_data.copy() if isinstance(subpolicy_data, dict) else {}
                subpolicy_data['PolicyId'] = policy.PolicyId
                if 'CreatedByName' not in subpolicy_data:
                    subpolicy_data['CreatedByName'] = policy.CreatedByName
                if 'CreatedByDate' not in subpolicy_data:
                    subpolicy_data['CreatedByDate'] = policy.CreatedByDate
                if 'Status' not in subpolicy_data:
                    subpolicy_data['Status'] = 'Under Review'
               
                subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
                if not subpolicy_serializer.is_valid():
                    print("SubPolicy serializer errors:", subpolicy_serializer.errors)  # Add this debug
                    return Response(subpolicy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                subpolicy_serializer.save()
                created_subpolicies_count += 1
               
            # Create a detailed success message
            message = 'Policy added to framework successfully'
            if created_subpolicies_count > 0:
                message += f' with {created_subpolicies_count} subpolicies'
            message += '!'
           
            return Response({
                'message': message,
                'PolicyId': policy.PolicyId,
                'FrameworkId': framework.FrameworkId,
                'Version': policy.CurrentVersion
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error adding policy to framework', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([AllowAny])
def list_policy_approvals_for_reviewer(request):
    # For now, reviewer_id is hardcoded as 2
    reviewer_id = 2
   
    # Get the latest version of each policy by identifier
    unique_identifiers = PolicyApproval.objects.values('Identifier').distinct()
    latest_approvals = []
   
    for identifier_obj in unique_identifiers:
        identifier = identifier_obj['Identifier']
        # Find the latest approval record for this identifier
        latest = PolicyApproval.objects.filter(
            Identifier=identifier,
            ReviewerId=reviewer_id
        ).order_by('-ApprovalId').first()
       
        if latest:
            latest_approvals.append(latest)
   
    # Serialize the data
    data = [
        {
            "ApprovalId": a.ApprovalId,
            "Identifier": a.Identifier,
            "ExtractedData": a.ExtractedData,
            "UserId": a.UserId,
            "ReviewerId": a.ReviewerId,
            "ApprovedNot": a.ApprovedNot,
            "Version": a.Version
        }
        for a in latest_approvals
    ]
   
    return Response(data)
 
@api_view(['PUT'])
@permission_classes([AllowAny])
def update_policy_approval(request, approval_id):
    try:
        # Get the original approval
        approval = PolicyApproval.objects.get(ApprovalId=approval_id)
       
        # Create a new approval object instead of updating
        new_approval = PolicyApproval()
        new_approval.Identifier = approval.Identifier
        new_approval.ExtractedData = request.data.get('ExtractedData', approval.ExtractedData)
        new_approval.UserId = approval.UserId
        new_approval.ReviewerId = approval.ReviewerId
        new_approval.ApprovedNot = request.data.get('ApprovedNot', approval.ApprovedNot)
       
        # Determine version prefix based on who made the change
        # For reviewers (assuming ReviewerId is the one making changes in this endpoint)
        prefix = 'r'
       
        # Get the latest version with this prefix for this identifier
        latest_version = PolicyApproval.objects.filter(
            Identifier=approval.Identifier,
            Version__startswith=prefix
        ).order_by('-Version').first()
       
        if latest_version and latest_version.Version:
            # Extract number and increment
            try:
                version_num = int(latest_version.Version[1:])
                new_approval.Version = f"{prefix}{version_num + 1}"
            except ValueError:
                new_approval.Version = f"{prefix}1"
        else:
            new_approval.Version = f"{prefix}1"
       
        new_approval.save()
       
        return Response({
            'message': 'Policy approval updated successfully',
            'ApprovalId': new_approval.ApprovalId,
            'Version': new_approval.Version
        })
    except PolicyApproval.DoesNotExist:
        return Response({'error': 'Policy approval not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
@api_view(['PUT'])
@permission_classes([AllowAny])
def resubmit_policy_approval(request, approval_id):
    try:
        # Get the original approval
        approval = PolicyApproval.objects.get(ApprovalId=approval_id)
       
        # Validate data
        extracted_data = request.data.get('ExtractedData')
        if not extracted_data:
            return Response({'error': 'ExtractedData is required'}, status=status.HTTP_400_BAD_REQUEST)
       
        # Print debug info
        print(f"Resubmitting policy with ID: {approval_id}, Identifier: {approval.Identifier}")
       
        # Get all versions for this identifier with 'u' prefix
        all_versions = PolicyApproval.objects.filter(Identifier=approval.Identifier)
       
        # Find the highest 'u' version number
        highest_u_version = 0
        for pa in all_versions:
            if pa.Version and pa.Version.startswith('u') and len(pa.Version) > 1:
                try:
                    version_num = int(pa.Version[1:])
                    if version_num > highest_u_version:
                        highest_u_version = version_num
                except ValueError:
                    continue
       
        # Set the new version
        new_version = f"u{highest_u_version + 1}"
        print(f"Setting new version: {new_version}")

        # Ensure all required fields are present in ExtractedData
        required_fields = [
            'PolicyName', 'PolicyDescription', 'Objective', 'Scope',
            'Department', 'Applicability', 'PolicyType', 'PolicyCategory',
            'PolicySubCategory', 'Status', 'CreatedBy', 'CreatedByDate'
        ]
        
        for field in required_fields:
            if field not in extracted_data:
                extracted_data[field] = approval.ExtractedData.get(field, '')

        # Update metadata
        extracted_data['Status'] = 'Under Review'
        extracted_data['LastModifiedDate'] = datetime.date.today().isoformat()
        
        # Handle subpolicies
        if 'subpolicies' in extracted_data and isinstance(extracted_data['subpolicies'], list):
            for subpolicy in extracted_data['subpolicies']:
                # Ensure each subpolicy has required fields
                if not subpolicy.get('SubPolicyId'):
                    continue
                
                # Reset approval status for rejected subpolicies
                if subpolicy.get('Status') == 'Rejected':
                    subpolicy['Status'] = 'Under Review'
                    subpolicy['approval'] = {
                        'approved': None,
                        'remarks': ''
                    }
                    subpolicy['resubmitted'] = True
                
                # Preserve version and ID
                subpolicy['version'] = subpolicy.get('version', 'u1')
       
        # Create a new approval object
        new_approval = PolicyApproval(
            Identifier=approval.Identifier,
            ExtractedData=extracted_data,
            UserId=approval.UserId,
            ReviewerId=approval.ReviewerId,
            ApprovedNot=None,  # Reset approval status
            Version=new_version,
            PolicyId=approval.PolicyId  # Preserve PolicyId reference
        )
       
        # Save the new record
        new_approval.save()
        print(f"Saved new approval with ID: {new_approval.ApprovalId}, Version: {new_approval.Version}")
       
        return Response({
            'message': 'Policy resubmitted for review successfully',
            'ApprovalId': new_approval.ApprovalId,
            'Version': new_version
        })
       
    except PolicyApproval.DoesNotExist:
        return Response({'error': 'Policy approval not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print("Error in resubmit_policy_approval:", str(e))
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def list_rejected_policy_approvals_for_user(request, user_id):
    # Filter policies by ReviewerId (not UserId) since we want reviewer's view
    approvals = PolicyApproval.objects.filter(ReviewerId=user_id)
   
    # Group by Identifier to find the latest for each policy/compliance
    identifier_latest = {}
    for approval in approvals:
        identifier = approval.Identifier
        if identifier not in identifier_latest or approval.ApprovalId > identifier_latest[identifier].ApprovalId:
            identifier_latest[identifier] = approval
   
    result = []
    for identifier, approval in identifier_latest.items():
        extracted = approval.ExtractedData
       
        # Check if this is a compliance item
        is_compliance = extracted.get('type') == 'compliance'
       
        # Check if main policy/compliance is rejected
        main_rejected = approval.ApprovedNot is False
       
        if is_compliance:
            # For compliance items
            item_rejected = extracted.get('compliance_approval', {}).get('approved') is False
           
            if main_rejected or item_rejected:
                result.append({
                    "ApprovalId": approval.ApprovalId,
                    "Identifier": approval.Identifier,
                    "ExtractedData": approval.ExtractedData,
                    "UserId": approval.UserId,
                    "ReviewerId": approval.ReviewerId,
                    "ApprovedNot": approval.ApprovedNot,
                    "main_item_rejected": main_rejected,
                    "is_compliance": True,
                    "rejection_reason": extracted.get('compliance_approval', {}).get('remarks', "")
                })
        else:
            # For policy items (existing logic)
            main_policy_rejected = main_rejected or (
                extracted.get('policy_approval', {}).get('approved') is False
            )
            # Check for rejected subpolicies
            rejected_subpolicies = []
            for sub in extracted.get('subpolicies', []):
                if sub.get('approval', {}).get('approved') is False:
                    rejected_subpolicies.append({
                        "Identifier": sub.get("Identifier"),
                        "SubPolicyName": sub.get("SubPolicyName"),
                        "Description": sub.get("Description"),
                        "Control": sub.get("Control"),
                        "remarks": sub.get("approval", {}).get("remarks", "")
                    })
            # Only add if main policy or any subpolicy is rejected
            if main_policy_rejected or rejected_subpolicies:
                result.append({
                    "ApprovalId": approval.ApprovalId,
                    "Identifier": approval.Identifier,
                    "ExtractedData": approval.ExtractedData,
                    "UserId": approval.UserId,
                    "ReviewerId": approval.ReviewerId,
                    "ApprovedNot": approval.ApprovedNot,
                    "main_policy_rejected": main_policy_rejected,
                    "rejected_subpolicies": rejected_subpolicies,
                    "is_compliance": False
                })
    return Response(result)
 
@api_view(['PUT'])
@permission_classes([AllowAny])
def submit_policy_review(request, approval_id):
    try:
        # Get the original approval
        approval = PolicyApproval.objects.get(ApprovalId=approval_id)
        
        # Get the latest user version approval to ensure we use the same IDs
        latest_user_approval = PolicyApproval.objects.filter(
            PolicyId=approval.PolicyId,
            Version__startswith='u'
        ).order_by('-ApprovalId').first()
        
        if not latest_user_approval:
            return Response({'error': 'No user version found for this policy'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate and prepare data
        extracted_data = request.data.get('ExtractedData')
        if not extracted_data:
            return Response({'error': 'ExtractedData is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        approved_not = request.data.get('ApprovedNot')
        
        # Get the latest version with "r" prefix for this identifier
        try:
            r_versions = []
            for pa in PolicyApproval.objects.filter(Identifier=approval.Identifier):
                if pa.Version and pa.Version.startswith('r') and pa.Version[1:].isdigit():
                    r_versions.append(int(pa.Version[1:]))
            
            if r_versions:
                new_version = f"r{max(r_versions) + 1}"
            else:
                new_version = "r1"  # Default version for reviewer
        except Exception as version_err:
            print(f"Error determining version (using default): {str(version_err)}")
            new_version = "r1"  # Default fallback
        
        # Set approved date if policy is approved
        approved_date = None
        if approved_not == True or approved_not == 1:
            approved_date = datetime.date.today()
        
        # Create a new record using Django ORM with same user/reviewer IDs
        new_approval = PolicyApproval(
            PolicyId=approval.PolicyId,
            Identifier=approval.Identifier,
            ExtractedData=extracted_data,
            UserId=latest_user_approval.UserId,  # Use same user ID as user version
            ReviewerId=latest_user_approval.ReviewerId,  # Use same reviewer ID as user version
            ApprovedNot=approved_not,
            ApprovedDate=approved_date,
            Version=new_version
        )
        new_approval.save()
        
        # If policy is approved, update the policy status
        if approved_not == True or approved_not == 1:
            try:
                policy = Policy.objects.get(PolicyId=approval.PolicyId.PolicyId)
                policy.Status = 'Approved'
                policy.ActiveInactive = 'Active'
                policy.save()
                
                # Update all subpolicies for this policy
                subpolicies = SubPolicy.objects.filter(PolicyId=policy.PolicyId)
                for subpolicy in subpolicies:
                    if subpolicy.Status == 'Under Review':
                        subpolicy.Status = 'Approved'
                        subpolicy.save()
            except Exception as e:
                print(f"Error updating policy status: {str(e)}")
        
        return Response({
            'message': 'Policy review submitted successfully',
            'ApprovalId': new_approval.ApprovalId,
            'Version': new_approval.Version
        })
    except Exception as e:
        print(f"Error in submit_policy_review: {str(e)}")
        return Response({
            'error': 'Error submitting policy review',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
@permission_classes([AllowAny])
def add_subpolicy_to_policy(request, policy_id):
    policy = get_object_or_404(Policy, PolicyId=policy_id)
   
    try:
        with transaction.atomic():
            # Set policy ID and default values in the request data
            subpolicy_data = request.data.copy()
            subpolicy_data['PolicyId'] = policy.PolicyId
            if 'CreatedByName' not in subpolicy_data:
                subpolicy_data['CreatedByName'] = policy.CreatedByName
            if 'CreatedByDate' not in subpolicy_data:
                subpolicy_data['CreatedByDate'] = datetime.date.today()
            if 'Status' not in subpolicy_data:
                subpolicy_data['Status'] = 'Under Review'
           
            # Check if subpolicy name is unique within the policy
            subpolicy_name = subpolicy_data.get('SubPolicyName')
            if SubPolicy.objects.filter(PolicyId=policy.PolicyId, SubPolicyName=subpolicy_name).exists():
                return Response({
                    'error': f'A subpolicy with the name "{subpolicy_name}" already exists in this policy. Subpolicy names must be unique within a policy.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            subpolicy_serializer = SubPolicySerializer(data=subpolicy_data, context={'policy_id': policy.PolicyId})
            if not subpolicy_serializer.is_valid():
                return Response(subpolicy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
           
            subpolicy = subpolicy_serializer.save()
           
            return Response({
                'message': 'Subpolicy added to policy successfully',
                'SubPolicyId': subpolicy.SubPolicyId,
                'PolicyId': policy.PolicyId
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error adding subpolicy to policy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
 

"""
@api GET /api/subpolicies/{pk}/
Returns a specific subpolicy by ID if it has Status='Approved',
its parent policy has Status='Approved' and ActiveInactive='Active',
and its parent framework has Status='Approved' and ActiveInactive='Active'.

@api PUT /api/subpolicies/{pk}/
Updates an existing subpolicy. Only subpolicies with Status='Approved'
whose parent policy and framework are also Approved and Active can be updated.

Example payload:
{
  "SubPolicyName": "Enhanced Password Management",
  "Description": "Updated password requirements and management",
  "Control": "Use strong passwords with at least 16 characters, including special characters",
  "Identifier": "PWD-002",
}

@api DELETE /api/subpolicies/{pk}/
Soft-deletes a subpolicy by setting Status='Inactive'.
"""
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([AllowAny])
def subpolicy_detail(request, pk):
    subpolicy = get_object_or_404(SubPolicy, SubPolicyId=pk)
    
    if request.method == 'GET':
        # Only return details if subpolicy is Approved
        if subpolicy.Status != 'Approved':
            return Response({'error': 'Subpolicy is not approved'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get policy to check if it's approved and active
        policy = subpolicy.PolicyId
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Policy is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get framework to check if it's approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = SubPolicySerializer(subpolicy)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        # Check if subpolicy is approved before allowing update
        if subpolicy.Status != 'Approved':
            return Response({'error': 'Only approved subpolicies can be updated'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if policy is approved and active
        policy = subpolicy.PolicyId
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Policy is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if framework is approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            with transaction.atomic():
                serializer = SubPolicySerializer(subpolicy, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Subpolicy updated successfully',
                        'SubPolicyId': subpolicy.SubPolicyId
                    })
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error updating subpolicy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        try:
            with transaction.atomic():
                # Instead of deleting, set Status to 'Inactive'
                subpolicy.Status = 'Inactive'
                subpolicy.save()
                
                return Response({'message': 'Subpolicy marked as inactive'}, status=status.HTTP_200_OK)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error marking subpolicy as inactive', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/frameworks/{pk}/copy/
Copies an existing framework to create a new one with modified details.
The FrameworkName must be unique - the request will be rejected if a framework with the same name already exists.
The copied framework will have Status='Under Review' and ActiveInactive='Inactive' by default.
All policies and subpolicies will be copied with the same structure but will also be set as Under Review/Inactive.
You can also modify specific policies by including a 'policies' array with PolicyId and updated fields.

Example payload:
{
  "FrameworkName": "ISO 27001:2023",
  "FrameworkDescription": "Updated Information Security Management System 2023 version",
  "EffectiveDate": "2023-11-01",
  "CreatedByName": "Jane Smith",
  "CreatedByDate": "2023-10-15",
  "Category": "Information Security and Compliance",
  "Identifier": "ISO-27001-2023",
  "policies": [
    {
      "original_policy_id": 1,
      "PolicyName": "Updated Access Control Policy 2023",
      "PolicyDescription": "Enhanced guidelines for access control with zero trust approach",
      "Department": "IT,Security",
      "Scope": "All IT systems and cloud environments",
      "Objective": "Implement zero trust security model"
    },
    {
      "original_policy_id": 2,
      "PolicyName": "Data Protection Policy 2023",
      "exclude": true
    }
  ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def copy_framework(request, pk):
    # Get original framework
    original_framework = get_object_or_404(Framework, FrameworkId=pk)
    print(f"Original Framework: ID={original_framework.FrameworkId}, Name={original_framework.FrameworkName}")

    try:
        with transaction.atomic():
            # Verify original framework status
            print(f"Original Framework Status: {original_framework.Status}, ActiveInactive: {original_framework.ActiveInactive}")
            if original_framework.Status != 'Approved' or original_framework.ActiveInactive != 'Active':
                print("Original framework not Approved or Active - aborting copy.")
                return Response({
                    'error': 'Only Approved and Active frameworks can be copied'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check framework name in request
            framework_name = request.data.get('FrameworkName')
            print(f"Requested new framework name: {framework_name}")
            if not framework_name:
                print("FrameworkName missing in request")
                return Response({'error': 'FrameworkName is required'}, status=status.HTTP_400_BAD_REQUEST)

            if Framework.objects.filter(FrameworkName=framework_name).exists():
                print("Framework with given name already exists.")
                return Response({'error': 'A framework with this name already exists'}, status=status.HTTP_400_BAD_REQUEST)

            framework_version = 1.0
            print(f"New framework version set to: {framework_version}")

            # Prepare new framework data
            new_framework_data = {
                'FrameworkName': framework_name,
                'CurrentVersion': framework_version,
                'FrameworkDescription': request.data.get('FrameworkDescription', original_framework.FrameworkDescription),
                'EffectiveDate': request.data.get('EffectiveDate', original_framework.EffectiveDate),
                'CreatedByName': request.data.get('CreatedByName', original_framework.CreatedByName),
                'CreatedByDate': datetime.date.today(),
                'Category': request.data.get('Category', original_framework.Category),
                'DocURL': request.data.get('DocURL', original_framework.DocURL),
                'Identifier': original_framework.Identifier,
                'StartDate': request.data.get('StartDate', original_framework.StartDate),
                'EndDate': request.data.get('EndDate', original_framework.EndDate),
                'Status': 'Under Review',
                'ActiveInactive': 'Inactive',
                'Reviewer': request.data.get('Reviewer', original_framework.Reviewer)
            }

            print(f"Creating new framework with data: {new_framework_data}")
            new_framework = Framework.objects.create(**new_framework_data)
            print(f"New Framework created: ID={new_framework.FrameworkId}, Name={new_framework.FrameworkName}")

            # Create framework version record
            framework_version_record = FrameworkVersion(
                FrameworkId=new_framework,
                Version=str(framework_version),
                FrameworkName=new_framework.FrameworkName,
                CreatedBy=new_framework.CreatedByName,
                CreatedDate=datetime.date.today(),
                PreviousVersionId=None
            )
            framework_version_record.save()
            print(f"FrameworkVersion record created for Framework ID {new_framework.FrameworkId} with Version {framework_version}")

            # Initialize policy tracking variables
            policy_customizations = {}
            policies_to_exclude = []
            created_policies = []

            # Handle policies from request
            if 'policies' in request.data:
                print(f"Received policies to process: {len(request.data.get('policies', []))}")
                for policy_data in request.data.get('policies', []):
                    if 'original_policy_id' in policy_data:
                        policy_id = policy_data.get('original_policy_id')
                        if policy_data.get('exclude', False):
                            print(f"Policy ID {policy_id} marked for exclusion")
                            policies_to_exclude.append(policy_id)
                        else:
                            print(f"Policy ID {policy_id} customization received")
                            policy_customizations[policy_id] = policy_data

            # Query original policies to copy
            original_policies = Policy.objects.filter(
                FrameworkId=original_framework,
                Status='Approved',
                ActiveInactive='Active'
            )
            print(f"Original policies count: {original_policies.count()}")

            for original_policy in original_policies:
                if original_policy.PolicyId in policies_to_exclude:
                    print(f"Skipping excluded policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")
                    continue
                print(f"Including policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")

                custom_data = policy_customizations.get(original_policy.PolicyId, {})

                created_by_user_id = custom_data.get('CreatedByUserId')
                if created_by_user_id:
                    try:
                        created_by_user = Users.objects.get(UserId=created_by_user_id)
                        created_by_name = created_by_user.UserName
                    except Users.DoesNotExist:
                        error_msg = f'User not found for CreatedByUserId: {created_by_user_id}'
                        print(error_msg)
                        return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    created_by_name = original_policy.CreatedByName

                reviewer_name = custom_data.get('Reviewer') or original_policy.Reviewer

                new_policy_data = {
                    'FrameworkId': new_framework,
                    'CurrentVersion': framework_version,
                    'Status': 'Under Review',
                    'PolicyDescription': custom_data.get('PolicyDescription', original_policy.PolicyDescription),
                    'PolicyName': custom_data.get('PolicyName', original_policy.PolicyName),
                    'StartDate': custom_data.get('StartDate', original_policy.StartDate),
                    'EndDate': custom_data.get('EndDate', original_policy.EndDate),
                    'Department': custom_data.get('Department', original_policy.Department),
                    'CreatedByName': custom_data.get('CreatedByName', created_by_name),
                    'CreatedByDate': new_framework.CreatedByDate,
                    'Applicability': custom_data.get('Applicability', original_policy.Applicability),
                    'DocURL': custom_data.get('DocURL', original_policy.DocURL),
                    'Scope': custom_data.get('Scope', original_policy.Scope),
                    'Objective': custom_data.get('Objective', original_policy.Objective),
                    'Identifier': custom_data.get('Identifier', original_policy.Identifier),
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_policy.PermanentTemporary),
                    'ActiveInactive': 'Inactive',
                    'Reviewer': reviewer_name,
                    'CoverageRate': custom_data.get('CoverageRate', original_policy.CoverageRate)
                }

                print(f"Creating new policy with data: {new_policy_data}")
                new_policy = Policy.objects.create(**new_policy_data)
                created_policies.append(new_policy)
                print(f"Created policy: {new_policy.PolicyName} (ID {new_policy.PolicyId})")

                # Subpolicy handling initialization
                subpolicy_customizations = {}
                subpolicies_to_exclude = []

                if 'subpolicies' in custom_data:
                    print(f"Policy {new_policy.PolicyName} has subpolicies to process: {len(custom_data.get('subpolicies', []))}")
                    for subpolicy_data in custom_data.get('subpolicies', []):
                        if 'original_subpolicy_id' in subpolicy_data:
                            subpolicy_id = subpolicy_data.get('original_subpolicy_id')
                            if subpolicy_data.get('exclude', False):
                                print(f"Subpolicy ID {subpolicy_id} marked for exclusion")
                                subpolicies_to_exclude.append(subpolicy_id)
                            else:
                                print(f"Customization for subpolicy ID {subpolicy_id} received")
                                subpolicy_customizations[subpolicy_id] = subpolicy_data

                original_subpolicies = SubPolicy.objects.filter(PolicyId=original_policy)
                print(f"Original subpolicies count for policy {original_policy.PolicyName}: {original_subpolicies.count()}")

                for subpolicy in original_subpolicies:
                    if subpolicy.SubPolicyId not in subpolicies_to_exclude:
                        sub_custom_data = subpolicy_customizations.get(subpolicy.SubPolicyId, {})
                        new_subpolicy_data = {
                            'PolicyId': new_policy,
                            'SubPolicyName': sub_custom_data.get('SubPolicyName', subpolicy.SubPolicyName),
                            'CreatedByName': new_policy.CreatedByName,
                            'CreatedByDate': datetime.date.today(),
                            'Identifier': sub_custom_data.get('Identifier', subpolicy.Identifier),
                            'Description': sub_custom_data.get('Description', subpolicy.Description),
                            'Status': 'Under Review',
                            'PermanentTemporary': sub_custom_data.get('PermanentTemporary', subpolicy.PermanentTemporary),
                            'Control': sub_custom_data.get('Control', subpolicy.Control)
                        }
                        print(f"DEBUG: Creating subpolicy with data_inventory: {new_subpolicy_data.get('data_inventory')}")
                        new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                        print(f"DEBUG: SubPolicy created with ID: {new_subpolicy.SubPolicyId}, data_inventory saved: {new_subpolicy.data_inventory}")
                        print(f"Created subpolicy: {new_subpolicy_data['SubPolicyName']} for policy {new_policy.PolicyName}")

            if 'new_policies' in request.data:
                print(f"Processing {len(request.data.get('new_policies', []))} new policies")
                for new_policy_data in request.data.get('new_policies', []):
                    required_fields = ['PolicyName', 'PolicyDescription', 'Identifier']
                    missing_fields = [f for f in required_fields if f not in new_policy_data]
                    if missing_fields:
                        error_msg = f"Missing required fields for new policy: {', '.join(missing_fields)}"
                        print(error_msg)
                        return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)

                    subpolicies_data = new_policy_data.pop('subpolicies', [])

                    policy_data = new_policy_data.copy()
                    policy_data['FrameworkId'] = new_framework
                    policy_data['CurrentVersion'] = framework_version
                    policy_data['Status'] = 'Under Review'
                    policy_data['ActiveInactive'] = 'Inactive'
                    policy_data.setdefault('CreatedByName', new_framework.CreatedByName)
                    policy_data['CreatedByDate'] = datetime.date.today()

                    new_policy = Policy.objects.create(**policy_data)
                    created_policies.append(new_policy)
                    print(f"Created new policy: {new_policy.PolicyName} (ID {new_policy.PolicyId})")

                    PolicyVersion.objects.create(
                        PolicyId=new_policy,
                        Version=str(framework_version),
                        PolicyName=new_policy.PolicyName,
                        CreatedBy=new_policy.CreatedByName,
                        CreatedDate=datetime.date.today(),
                        PreviousVersionId=None
                    )

                    for subpolicy_data in subpolicies_data:
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [f for f in required_fields if f not in subpolicy_data]
                        if missing_fields:
                            error_msg = f"Missing required fields for subpolicy in new policy {new_policy.PolicyName}: {', '.join(missing_fields)}"
                            print(error_msg)
                            return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)

                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = new_policy
                        subpolicy.setdefault('CreatedByName', new_policy.CreatedByName)
                        subpolicy['CreatedByDate'] = datetime.date.today()
                        subpolicy.setdefault('Status', 'Under Review')

                        SubPolicy.objects.create(**subpolicy)
                        print(f"Created subpolicy: {subpolicy.get('SubPolicyName')} for new policy {new_policy.PolicyName}")

            response_data = {
                'message': 'Framework copied successfully',
                'FrameworkId': new_framework.FrameworkId,
                'FrameworkName': new_framework.FrameworkName,
                'Version': new_framework.CurrentVersion,
            }

            if created_policies:
                response_data['policies'] = [{
                    'PolicyId': p.PolicyId,
                    'PolicyName': p.PolicyName,
                    'Identifier': p.Identifier,
                    'Version': p.CurrentVersion
                } for p in created_policies]

            print(f"Copy framework operation completed successfully for Framework ID {new_framework.FrameworkId}")
            return Response(response_data, status=status.HTTP_201_CREATED)

    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print("Error in copy_framework:", error_info)
        return Response({'error': 'Error copying framework', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)


"""
@api POST /api/policies/{pk}/copy/
Copies an existing policy to create a new one with modified details within the same framework.
The PolicyName must be unique within the framework - the request will be rejected if a policy with the same name already exists.
The copied policy will have Status='Under Review' and ActiveInactive='Inactive' by default.
All subpolicies will be copied with the same structure but will also be set as Under Review by default.
You can also modify, exclude, or add new subpolicies.

Example payload:
{
  "PolicyName": "Enhanced Access Control Policy 2023",
  "PolicyDescription": "Updated guidelines for access control with zero trust approach",
  "StartDate": "2023-11-01",
  "EndDate": "2025-11-01",
  "Department": "IT,Security",
  "CreatedByName": "Jane Smith",
  "CreatedByDate": "2023-10-15",
  "Scope": "All IT systems and cloud environments",
  "Objective": "Implement zero trust security model",
  "Identifier": "ACP-ZT-001",
  "subpolicies": [
    {
      "original_subpolicy_id": 5,
      "SubPolicyName": "Enhanced Password Rules",
      "Description": "Updated password requirements with MFA",
      "Control": "16-character passwords with MFA for all access"
    },
    {
      "original_subpolicy_id": 6,
      "exclude": true
    }
  ],
  "new_subpolicies": [
    {
      "SubPolicyName": "Device Authentication",
      "Description": "Requirements for device-based authentication",
      "Control": "Implement device certificates for all company devices",
      "Identifier": "DEV-AUTH-001",
      "Status": "Under Review"
    }
  ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def copy_policy(request, pk):
    # Get original policy
    original_policy = get_object_or_404(Policy, PolicyId=pk)
    
    try:
        with transaction.atomic():
            # Verify the original policy is Approved and Active
            if original_policy.Status != 'Approved' or original_policy.ActiveInactive != 'Active':
                return Response({
                    'error': 'Only Approved and Active policies can be copied'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if policy name is unique within the framework
            policy_name = request.data.get('PolicyName')
            if not policy_name:
                return Response({'error': 'PolicyName is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Get target framework ID from request
            target_framework_id = request.data.get('TargetFrameworkId')
            if not target_framework_id:
                return Response({'error': 'TargetFrameworkId is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Get target framework
            try:
                target_framework = Framework.objects.get(FrameworkId=target_framework_id)
            except Framework.DoesNotExist:
                return Response({'error': 'Target framework not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Check if a policy with this name already exists in the target framework
            if Policy.objects.filter(FrameworkId=target_framework, PolicyName=policy_name).exists():
                return Response({'error': 'A policy with this name already exists in the target framework'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Create new policy with data from original and overrides from request
            new_policy_data = {
                'FrameworkId': target_framework,  # Use target framework instead of original
                'Status': 'Under Review',
                'PolicyName': policy_name,
                'PolicyDescription': request.data.get('PolicyDescription', original_policy.PolicyDescription),
                'StartDate': request.data.get('StartDate', original_policy.StartDate),
                'EndDate': request.data.get('EndDate', original_policy.EndDate),
                'Department': request.data.get('Department', original_policy.Department),
                'CreatedByName': request.data.get('CreatedByName', original_policy.CreatedByName),
                'CreatedByDate': request.data.get('CreatedByDate', datetime.date.today()),
                'Applicability': request.data.get('Applicability', original_policy.Applicability),
                'DocURL': request.data.get('DocURL', original_policy.DocURL),
                'Scope': request.data.get('Scope', original_policy.Scope),
                'Objective': request.data.get('Objective', original_policy.Objective),
                'Identifier': request.data.get('Identifier', original_policy.Identifier),
                'PermanentTemporary': request.data.get('PermanentTemporary', original_policy.PermanentTemporary),
                'ActiveInactive': 'Inactive',
                'CurrentVersion': 1.0,  # Start with version 1.0 for new policy
                'Reviewer': request.data.get('Reviewer', original_policy.Reviewer),  # Store reviewer name
                'CoverageRate': request.data.get('CoverageRate', original_policy.CoverageRate)  # Add coverage rate
            }
            
            # Create new policy
            new_policy = Policy.objects.create(**new_policy_data)
            
            # Create policy version record (no previous version link) - ONLY ONCE
            policy_version = PolicyVersion(
                PolicyId=new_policy,
                Version='1.0',  # Start with version 1.0
                PolicyName=new_policy.PolicyName,
                CreatedBy=new_policy.CreatedByName,
                CreatedDate=new_policy.CreatedByDate,
                PreviousVersionId=None  # No version linking
            )
            policy_version.save()
            
            # Handle subpolicy customizations if provided
            subpolicy_customizations = {}
            subpolicies_to_exclude = []
            created_subpolicies = []  # Keep track of created subpolicies
            
            # Process subpolicy customizations if provided
            if 'subpolicies' in request.data:
                for subpolicy_data in request.data.get('subpolicies', []):
                    if 'original_subpolicy_id' in subpolicy_data:
                        subpolicy_id = subpolicy_data.get('original_subpolicy_id')
                        
                        # Check if this subpolicy should be excluded
                        if subpolicy_data.get('exclude', False):
                            subpolicies_to_exclude.append(subpolicy_id)
                        else:
                            # Store customizations for this subpolicy
                            subpolicy_customizations[subpolicy_id] = subpolicy_data
            
            # Copy only Approved and Active subpolicies from original policy - ONLY ONCE
            original_subpolicies = SubPolicy.objects.filter(
                PolicyId=original_policy,
                Status='Approved'
            )
            
            for original_subpolicy in original_subpolicies:
                # Skip if this subpolicy should be excluded
                if original_subpolicy.SubPolicyId in subpolicies_to_exclude:
                    continue
                
                # Get customizations for this subpolicy if any
                custom_data = subpolicy_customizations.get(original_subpolicy.SubPolicyId, {})
                
                # Create new subpolicy with data from original and any customizations
                new_subpolicy_data = {
                    'PolicyId': new_policy,
                    'SubPolicyName': custom_data.get('SubPolicyName', original_subpolicy.SubPolicyName),
                    'CreatedByName': new_policy.CreatedByName,
                    'CreatedByDate': new_policy.CreatedByDate,
                    'Identifier': custom_data.get('Identifier', original_subpolicy.Identifier),
                    'Description': custom_data.get('Description', original_subpolicy.Description),
                    'Status': 'Under Review',
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_subpolicy.PermanentTemporary),
                    'Control': custom_data.get('Control', original_subpolicy.Control)
                }
                
                new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                created_subpolicies.append(new_subpolicy)
            
            # --- Add this block to process new subpolicies (no original_subpolicy_id) ---
            if 'subpolicies' in request.data:
                for subpolicy_data in request.data.get('subpolicies', []):
                    if 'original_subpolicy_id' not in subpolicy_data:
                        # This is a new subpolicy, create it
                        new_subpolicy_data = {
                            'PolicyId': new_policy,
                            'SubPolicyName': subpolicy_data.get('SubPolicyName'),
                            'CreatedByName': new_policy.CreatedByName,
                            'CreatedByDate': new_policy.CreatedByDate,
                            'Identifier': subpolicy_data.get('Identifier'),
                            'Description': subpolicy_data.get('Description'),
                            'Status': 'Under Review',
                            'PermanentTemporary': subpolicy_data.get('PermanentTemporary', 'Permanent'),
                            'Control': subpolicy_data.get('Control')
                        }
                        print(f"DEBUG: Creating subpolicy with data_inventory: {new_subpolicy_data.get('data_inventory')}")
                    new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                    print(f"DEBUG: SubPolicy created with ID: {new_subpolicy.SubPolicyId}, data_inventory saved: {new_subpolicy.data_inventory}")
            # --- End new subpolicy block ---
            
            return Response({
                'message': 'Policy copied successfully to target framework',
                'PolicyId': new_policy.PolicyId,
                'PolicyName': new_policy.PolicyName,
                'SourceFrameworkId': original_policy.FrameworkId.FrameworkId,
                'TargetFrameworkId': target_framework.FrameworkId,
                'Version': new_policy.CurrentVersion
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print("Error in copy_policy:", error_info)  # Add this to see full error on server console/logs
        return Response({'error': 'Error copying policy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api PUT /api/frameworks/{pk}/toggle-status/
Toggles the ActiveInactive status of a framework between 'Active' and 'Inactive'.
If the framework is currently 'Active', it will be set to 'Inactive' and vice versa.
When a framework is set to 'Inactive', all its policies will also be set to 'Inactive'.

Example response:
{
    "message": "Framework status updated successfully",
    "FrameworkId": 1,
    "FrameworkName": "ISO 27001",
    "ActiveInactive": "Inactive"
}
"""
@api_view(['PUT'])
@permission_classes([AllowAny])
def toggle_framework_status(request, pk):
    framework = get_object_or_404(Framework, FrameworkId=pk)
    
    try:
        with transaction.atomic():
            # Toggle the status
            new_status = 'Inactive' if framework.ActiveInactive == 'Active' else 'Active'
            framework.ActiveInactive = new_status
            framework.save()
            
            # If setting to Inactive, also set all policies to Inactive
            if new_status == 'Inactive':
                policies = Policy.objects.filter(FrameworkId=framework)
                for policy in policies:
                    policy.ActiveInactive = 'Inactive'
                    policy.save()
            
            return Response({
                'message': 'Framework status updated successfully',
                'FrameworkId': framework.FrameworkId,
                'FrameworkName': framework.FrameworkName,
                'ActiveInactive': framework.ActiveInactive
            })
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error updating framework status', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api PUT /api/policies/{pk}/toggle-status/
Toggles the ActiveInactive status of a policy between 'Active' and 'Inactive'.
If the policy is currently 'Active', it will be set to 'Inactive' and vice versa.
Note: A policy can only be set to 'Active' if its parent framework is also 'Active'.

Example response:
{
    "message": "Policy status updated successfully",
    "PolicyId": 1,
    "PolicyName": "Access Control Policy",
    "ActiveInactive": "Active"
}
"""
@api_view(['PUT'])
@permission_classes([AllowAny])
def toggle_policy_status(request, pk):
    policy = get_object_or_404(Policy, PolicyId=pk)
    
    try:
        with transaction.atomic():
            # Check if trying to activate a policy while framework is inactive
            if policy.ActiveInactive == 'Inactive' and policy.FrameworkId.ActiveInactive == 'Inactive':
                return Response({
                    'error': 'Cannot activate policy while parent framework is inactive'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Toggle the status
            new_status = 'Inactive' if policy.ActiveInactive == 'Active' else 'Active'
            policy.ActiveInactive = new_status
            policy.save()
            
            return Response({
                'message': 'Policy status updated successfully',
                'PolicyId': policy.PolicyId,
                'PolicyName': policy.PolicyName,
                'ActiveInactive': policy.ActiveInactive
            })
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error updating policy status', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/frameworks/{pk}/create-version/
Creates a new version of an existing framework by cloning it with an incremented version number.
For example, if the original framework has version 1.0, the new version will be 1.1.
All policies and subpolicies will be cloned with their details.

Example payload:
{
  "FrameworkName": "ISO 27001 v3.3",
  "FrameworkDescription": "Updated Information Security Management System 2024",
  "EffectiveDate": "2024-01-01",
  "CreatedByName": "John Doe",
  "CreatedByDate": "2023-12-15",
  "policies": [
    {
      "original_policy_id": 1052,
      "PolicyName": "Access Control Policy",
      "PolicyDescription": "Original access control policy",
      "Identifier": "ACP-001",
      "subpolicies": [
        {
          "original_subpolicy_id": 100,
          "SubPolicyName": "Password Management",
          "Description": "Original password requirements",
          "Control": "Use strong passwords",
          "Identifier": "PWD-001"
        }
      ]
    },
    {
      "original_policy_id": 2,
      "exclude": true
    },
    {
      "original_policy_id": 3,
      "PolicyName": "Data Protection Policy",
      "PolicyDescription": "Original data protection policy",
      "Identifier": "DPP-001",
      "subpolicies": [
        {
          "original_subpolicy_id": 4,
          "exclude": true
        },
        {
          "original_subpolicy_id": 5,
          "exclude": true
        }
      ]
    }
  ],
  "new_policies": [
    {
      "PolicyName": "New Security Policy",
      "PolicyDescription": "A completely new policy",
      "Identifier": "NSP-001",
      "Department": "IT,Security",
      "Scope": "All systems",
      "Objective": "Implement new security measures",
      "subpolicies": [
        {
          "SubPolicyName": "New Security Control",
          "Description": "New security requirements",
          "Control": "Implement new security measures",
          "Identifier": "NSC-001"
        }
      ]
    }
  ]
}

Example response:
{
    "message": "New framework version created successfully",
    "FrameworkId": 35,
    "FrameworkName": "ISO 27001 v3.3",
    "PreviousVersion": 1.0,
    "NewVersion": 1.1,
    "Identifier": "ISO",
    "policies": [
        {
            "PolicyId": 1074,
            "PolicyName": "Access Control Policy",
            "Identifier": "ACP-001",
            "Version": 1.1
        },
        {
            "PolicyId": 1075,
            "PolicyName": "New Security Policy",
            "Identifier": "NSP-001",
            "Version": 1.1
        }
    ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def create_framework_version(request, pk):
    original_framework = get_object_or_404(Framework, FrameworkId=pk)
    
    try:
        with transaction.atomic():
            reviewer_id = request.data.get('Reviewer')
            reviewer_name = None
            if reviewer_id:
                user_obj = Users.objects.filter(UserId=reviewer_id).first()
                if user_obj:
                    reviewer_name = user_obj.UserName

            framework_name = request.data.get('FrameworkName')
            if not framework_name:
                return Response({'error': 'FrameworkName is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            latest_version = FrameworkVersion.objects.filter(
                FrameworkId__Identifier=original_framework.Identifier
            ).order_by('-Version').first()
            
            current_version = int(float(original_framework.CurrentVersion)) if not latest_version else int(float(latest_version.Version))
            new_version = current_version + 1
            new_version_str = f"{new_version}.0"
            
            # Get data_inventory from request.data (extract BEFORE any processing, like TT does)
            framework_data_inventory_raw = request.data.get('data_inventory')
            print(f"DEBUG: Framework data_inventory RAW from request: {framework_data_inventory_raw}")
            print(f"DEBUG: Framework data_inventory RAW type: {type(framework_data_inventory_raw)}")
            
            framework_data_inventory = None
            if framework_data_inventory_raw is not None:
                if isinstance(framework_data_inventory_raw, str):
                    try:
                        import json
                        framework_data_inventory = json.loads(framework_data_inventory_raw)
                        print(f"DEBUG: Parsed JSON string to dict: {framework_data_inventory}")
                    except json.JSONDecodeError:
                        print(f"Warning: Invalid JSON in framework data_inventory, setting to None")
                        framework_data_inventory = None
                elif isinstance(framework_data_inventory_raw, dict):
                    # If it's already a dict, use it as-is (even if empty)
                    framework_data_inventory = framework_data_inventory_raw
                    print(f"DEBUG: Using dict as-is: {framework_data_inventory}")
                else:
                    print(f"DEBUG: framework_data_inventory_raw is not str or dict, type: {type(framework_data_inventory_raw)}")
            else:
                print(f"DEBUG: framework_data_inventory_raw is None, falling back to original")
                # Fall back to original framework's data_inventory if not provided
                framework_data_inventory = original_framework.data_inventory if hasattr(original_framework, 'data_inventory') else None
            
            print(f"DEBUG: Framework data_inventory FINAL: {framework_data_inventory}")
            print(f"DEBUG: Framework data_inventory FINAL type: {type(framework_data_inventory)}")
            
            # Get InternalExternal from request or use original
            internal_external = request.data.get('InternalExternal', original_framework.InternalExternal if hasattr(original_framework, 'InternalExternal') else 'Internal')
            
            # Ensure data_inventory is explicitly set (even if None or empty dict)
            # This is important for JSONField to save correctly
            if framework_data_inventory is None:
                framework_data_inventory = original_framework.data_inventory if hasattr(original_framework, 'data_inventory') else None
            
            new_framework_data = {
                'FrameworkName': framework_name,
                'CurrentVersion': new_version_str,
                'FrameworkDescription': request.data.get('FrameworkDescription', original_framework.FrameworkDescription),
                'EffectiveDate': request.data.get('EffectiveDate', original_framework.EffectiveDate),
                'CreatedByName': request.data.get('CreatedByName', original_framework.CreatedByName),
                'CreatedByDate': request.data.get('CreatedByDate', datetime.date.today()),
                'Category': request.data.get('Category', original_framework.Category),
                'DocURL': request.data.get('DocURL', original_framework.DocURL),
                'Identifier': original_framework.Identifier,
                'StartDate': request.data.get('StartDate', original_framework.StartDate),
                'EndDate': request.data.get('EndDate', original_framework.EndDate),
                'Status': 'Under Review',
                'ActiveInactive': 'Inactive',
                'Reviewer': reviewer_name,
                'InternalExternal': internal_external,
                'data_inventory': framework_data_inventory,  # Explicitly include data_inventory
            }
            
            print(f"DEBUG: new_framework_data keys: {list(new_framework_data.keys())}")
            print(f"DEBUG: data_inventory in new_framework_data: {'data_inventory' in new_framework_data}")
            print(f"DEBUG: Creating framework with data_inventory: {new_framework_data.get('data_inventory')}")
            print(f"DEBUG: data_inventory value type before create: {type(new_framework_data.get('data_inventory'))}")
            
            new_framework = Framework.objects.create(**new_framework_data)
            
            # Refresh from database to get the actual saved value
            new_framework.refresh_from_db()
            print(f"DEBUG: Framework created with ID: {new_framework.FrameworkId}")
            print(f"DEBUG: Framework data_inventory after save: {new_framework.data_inventory}")
            print(f"DEBUG: Framework data_inventory type after save: {type(new_framework.data_inventory)}")

            original_framework_version = FrameworkVersion.objects.filter(
                FrameworkId=original_framework,
                Version=str(original_framework.CurrentVersion)
            ).first()
            
            if not original_framework_version:
                original_framework_version = FrameworkVersion.objects.create(
                    FrameworkId=original_framework,
                    Version=str(original_framework.CurrentVersion),
                    FrameworkName=original_framework.FrameworkName,
                    CreatedBy=original_framework.CreatedByName,
                    CreatedDate=original_framework.CreatedByDate,
                    PreviousVersionId=None
                )

            framework_version = FrameworkVersion(
                FrameworkId=new_framework,
                Version=str(new_version_str),
                FrameworkName=new_framework.FrameworkName,
                CreatedBy=new_framework.CreatedByName,
                CreatedDate=new_framework.CreatedByDate,
                PreviousVersionId=original_framework_version.VersionId
            )
            framework_version.save()

            # Process policies
            policy_customizations = {}
            policies_to_exclude = []
            created_policies = []

            if 'policies' in request.data:
                for policy_data in request.data.get('policies', []):
                    if 'original_policy_id' in policy_data:
                        policy_id = policy_data.get('original_policy_id')
                        if policy_data.get('exclude', False):
                            policies_to_exclude.append(int(policy_id))
                        else:
                            policy_customizations[int(policy_id)] = policy_data
            
            print("Policies to exclude:", policies_to_exclude)

            original_policies = Policy.objects.filter(
                FrameworkId=original_framework,
                Status='Approved',
                ActiveInactive='Active'
            )

            excluded_set = set(policies_to_exclude)

            for original_policy in original_policies:
                if original_policy.PolicyId in excluded_set:
                    print(f"Skipping excluded policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")
                    continue
                print(f"Including policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")

                custom_data = policy_customizations.get(original_policy.PolicyId, {})

                # Determine Reviewer username for existing policy
                reviewer_id_from_custom = custom_data.get('Reviewer')
                reviewer_username = reviewer_name  # fallback
                
                if reviewer_id_from_custom:
                    user_obj = Users.objects.filter(UserId=reviewer_id_from_custom).first()
                    if user_obj:
                        reviewer_username = user_obj.UserName

                # Get data_inventory for policy
                policy_data_inventory = custom_data.get('data_inventory')
                if policy_data_inventory is not None:
                    if isinstance(policy_data_inventory, str):
                        try:
                            import json
                            policy_data_inventory = json.loads(policy_data_inventory)
                        except json.JSONDecodeError:
                            policy_data_inventory = None
                    # If it's already a dict, use it as-is (even if empty)
                    elif not isinstance(policy_data_inventory, dict):
                        policy_data_inventory = None
                else:
                    # Fall back to original policy's data_inventory if not provided
                    policy_data_inventory = original_policy.data_inventory if hasattr(original_policy, 'data_inventory') else None
                
                print(f"DEBUG: Policy {original_policy.PolicyName} (ID: {original_policy.PolicyId}) data_inventory from custom_data: {custom_data.get('data_inventory')}, final: {policy_data_inventory}")
                
                new_policy_data = {
                    'FrameworkId': new_framework,
                    'CurrentVersion': new_version_str,
                    'Status': 'Under Review',
                    'PolicyDescription': custom_data.get('PolicyDescription', original_policy.PolicyDescription),
                    'PolicyName': custom_data.get('PolicyName', original_policy.PolicyName),
                    'StartDate': custom_data.get('StartDate', original_policy.StartDate),
                    'EndDate': custom_data.get('EndDate', original_policy.EndDate),
                    'Department': custom_data.get('Department', original_policy.Department),
                    'CreatedByName': custom_data.get('CreatedByName', original_policy.CreatedByName),
                    'CreatedByDate': new_framework.CreatedByDate,
                    'Applicability': custom_data.get('Applicability', original_policy.Applicability),
                    'DocURL': custom_data.get('DocURL', original_policy.DocURL),
                    'Scope': custom_data.get('Scope', original_policy.Scope),
                    'Objective': custom_data.get('Objective', original_policy.Objective),
                    'Identifier': custom_data.get('Identifier', original_policy.Identifier),
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_policy.PermanentTemporary),
                    'ActiveInactive': 'Inactive',
                    'Reviewer': reviewer_username,
                    'data_inventory': policy_data_inventory,
                }
                
                print(f"DEBUG: Creating policy with data_inventory: {new_policy_data.get('data_inventory')}")
                new_policy = Policy.objects.create(**new_policy_data)
                print(f"DEBUG: Policy created with ID: {new_policy.PolicyId}, data_inventory saved: {new_policy.data_inventory}")
                created_policies.append(new_policy)

                original_policy_version = PolicyVersion.objects.filter(
                    PolicyId=original_policy,
                    Version=str(original_policy.CurrentVersion)
                ).first()

                if not original_policy_version:
                    return Response({
                        'error': f'No PolicyVersion found for PolicyId={original_policy.PolicyId} and Version={original_policy.CurrentVersion}. Data integrity issue.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                policy_version = PolicyVersion(
                    PolicyId=new_policy,
                    Version=str(new_version_str),
                    PolicyName=new_policy.PolicyName,
                    CreatedBy=new_policy.CreatedByName,
                    CreatedDate=new_policy.CreatedByDate,
                    PreviousVersionId=original_policy_version.VersionId
                )
                policy_version.save()

                # Handle subpolicies
                subpolicy_customizations = {}
                subpolicies_to_exclude = []

                if 'subpolicies' in custom_data:
                    for subpolicy_data in custom_data.get('subpolicies', []):
                        if subpolicy_data.get('exclude', False):
                            if 'original_subpolicy_id' in subpolicy_data:
                                subpolicies_to_exclude.append(int(subpolicy_data.get('original_subpolicy_id')))
                        else:
                            if 'original_subpolicy_id' in subpolicy_data:
                                subpolicy_customizations[int(subpolicy_data.get('original_subpolicy_id'))] = subpolicy_data
                
                print(f"Subpolicies to exclude for policy {new_policy.PolicyName}:", subpolicies_to_exclude)

                original_subpolicies = SubPolicy.objects.filter(PolicyId=original_policy)

                excluded_subpolicy_set = set(subpolicies_to_exclude)
                for original_subpolicy in original_subpolicies:
                    if original_subpolicy.SubPolicyId in excluded_subpolicy_set:
                        print(f"Skipping excluded subpolicy: {original_subpolicy.SubPolicyName} (ID {original_subpolicy.SubPolicyId})")
                        continue
                    
                    custom_subpolicy_data = subpolicy_customizations.get(original_subpolicy.SubPolicyId, {})

                    # Get data_inventory for subpolicy
                    subpolicy_data_inventory = custom_subpolicy_data.get('data_inventory')
                    if subpolicy_data_inventory is not None:
                        if isinstance(subpolicy_data_inventory, str):
                            try:
                                import json
                                subpolicy_data_inventory = json.loads(subpolicy_data_inventory)
                            except json.JSONDecodeError:
                                subpolicy_data_inventory = None
                        # If it's already a dict, use it as-is (even if empty)
                        elif not isinstance(subpolicy_data_inventory, dict):
                            subpolicy_data_inventory = None
                    else:
                        # Fall back to original subpolicy's data_inventory if not provided
                        subpolicy_data_inventory = original_subpolicy.data_inventory if hasattr(original_subpolicy, 'data_inventory') else None
                    
                    print(f"DEBUG: SubPolicy {original_subpolicy.SubPolicyName} (ID: {original_subpolicy.SubPolicyId}) data_inventory from custom_subpolicy_data: {custom_subpolicy_data.get('data_inventory')}, final: {subpolicy_data_inventory}")
                    
                    new_subpolicy_data = {
                        'PolicyId': new_policy,
                        'SubPolicyName': custom_subpolicy_data.get('SubPolicyName', original_subpolicy.SubPolicyName),
                        'CreatedByName': new_policy.CreatedByName,
                        'CreatedByDate': new_policy.CreatedByDate,
                        'Identifier': custom_subpolicy_data.get('Identifier', original_subpolicy.Identifier),
                        'Description': custom_subpolicy_data.get('Description', original_subpolicy.Description),
                        'Status': 'Under Review',
                        'PermanentTemporary': custom_subpolicy_data.get('PermanentTemporary', original_subpolicy.PermanentTemporary),
                        'Control': custom_subpolicy_data.get('Control', original_subpolicy.Control),
                        'data_inventory': subpolicy_data_inventory
                    }

                    print(f"DEBUG: Creating subpolicy with data_inventory: {new_subpolicy_data.get('data_inventory')}")
                    new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                    print(f"DEBUG: SubPolicy created with ID: {new_subpolicy.SubPolicyId}, data_inventory saved: {new_subpolicy.data_inventory}")

                # Handle new subpolicies from subpolicies array (for backward compatibility)
                for sp_data in custom_data.get('subpolicies', []):
                    if sp_data.get('exclude', False):
                        continue
                    if 'original_subpolicy_id' in sp_data:
                        continue

                    required_fields = ['SubPolicyName', 'Description', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in sp_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new subpolicy in existing policy {new_policy.PolicyName}: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Extract data_inventory for new subpolicy
                    new_subpolicy_data_inventory = sp_data.get('data_inventory')
                    if new_subpolicy_data_inventory is not None:
                        if isinstance(new_subpolicy_data_inventory, str):
                            try:
                                import json
                                new_subpolicy_data_inventory = json.loads(new_subpolicy_data_inventory)
                            except json.JSONDecodeError:
                                new_subpolicy_data_inventory = None
                        # If it's already a dict, use it as-is (even if empty)
                        elif not isinstance(new_subpolicy_data_inventory, dict):
                            new_subpolicy_data_inventory = None

                    new_subpolicy_data = {
                        'PolicyId': new_policy,
                        'SubPolicyName': sp_data.get('SubPolicyName'),
                        'Description': sp_data.get('Description'),
                        'Identifier': sp_data.get('Identifier'),
                        'CreatedByName': new_policy.CreatedByName,
                        'CreatedByDate': new_policy.CreatedByDate,
                        'Status': 'Under Review',
                        'PermanentTemporary': sp_data.get('PermanentTemporary', 'Permanent'),
                        'Control': sp_data.get('Control', ''),
                        'data_inventory': new_subpolicy_data_inventory
                    }

                    print(f"DEBUG: Creating subpolicy with data_inventory: {new_subpolicy_data.get('data_inventory')}")
                    new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                    print(f"DEBUG: SubPolicy created with ID: {new_subpolicy.SubPolicyId}, data_inventory saved: {new_subpolicy.data_inventory}")
                
                # Handle new subpolicies from new_subpolicies array
                for sp_data in custom_data.get('new_subpolicies', []):
                    if sp_data.get('exclude', False):
                        continue

                    required_fields = ['SubPolicyName', 'Description', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in sp_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new subpolicy in existing policy {new_policy.PolicyName}: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Extract data_inventory for new subpolicy
                    new_subpolicy_data_inventory = sp_data.get('data_inventory')
                    if new_subpolicy_data_inventory is not None:
                        if isinstance(new_subpolicy_data_inventory, str):
                            try:
                                import json
                                new_subpolicy_data_inventory = json.loads(new_subpolicy_data_inventory)
                            except json.JSONDecodeError:
                                new_subpolicy_data_inventory = None
                        # If it's already a dict, use it as-is (even if empty)
                        elif not isinstance(new_subpolicy_data_inventory, dict):
                            new_subpolicy_data_inventory = None

                    new_subpolicy_data = {
                        'PolicyId': new_policy,
                        'SubPolicyName': sp_data.get('SubPolicyName'),
                        'Description': sp_data.get('Description'),
                        'Identifier': sp_data.get('Identifier'),
                        'CreatedByName': new_policy.CreatedByName,
                        'CreatedByDate': new_policy.CreatedByDate,
                        'Status': 'Under Review',
                        'PermanentTemporary': sp_data.get('PermanentTemporary', 'Permanent'),
                        'Control': sp_data.get('Control', ''),
                        'data_inventory': new_subpolicy_data_inventory
                    }

                    print(f"DEBUG: Creating subpolicy with data_inventory: {new_subpolicy_data.get('data_inventory')}")
                    new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                    print(f"DEBUG: SubPolicy created with ID: {new_subpolicy.SubPolicyId}, data_inventory saved: {new_subpolicy.data_inventory}")

            # Handle new policies
            if 'new_policies' in request.data:
                for new_policy_data in request.data.get('new_policies', []):
                    required_fields = ['PolicyName', 'PolicyDescription', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in new_policy_data]
                    
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new policy: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Extract data_inventory for new policy
                    new_policy_data_inventory = new_policy_data.get('data_inventory')
                    if new_policy_data_inventory is not None:
                        if isinstance(new_policy_data_inventory, str):
                            try:
                                import json
                                new_policy_data_inventory = json.loads(new_policy_data_inventory)
                            except json.JSONDecodeError:
                                new_policy_data_inventory = None
                        # If it's already a dict, use it as-is (even if empty)
                        elif not isinstance(new_policy_data_inventory, dict):
                            new_policy_data_inventory = None
                    
                    print(f"DEBUG: New Policy {new_policy_data.get('PolicyName')} data_inventory: {new_policy_data_inventory}")
                    
                    subpolicies_data = new_policy_data.pop('subpolicies', [])
                    new_subpolicies_data = new_policy_data.pop('new_subpolicies', [])
                    
                    policy_data = new_policy_data.copy()
                    policy_data['FrameworkId'] = new_framework
                    policy_data['CurrentVersion'] = new_version_str
                    policy_data['Status'] = 'Under Review'
                    policy_data['ActiveInactive'] = 'Inactive'
                    policy_data.setdefault('CreatedByName', new_framework.CreatedByName)
                    policy_data['CreatedByDate'] = datetime.date.today()

                    reviewer_id_new_policy = policy_data.get('Reviewer')
                    reviewer_username_new_policy = reviewer_name  # fallback
                    if reviewer_id_new_policy:
                        user_obj = Users.objects.filter(UserId=reviewer_id_new_policy).first()
                        if user_obj:
                            reviewer_username_new_policy = user_obj.UserName

                    policy_data['Reviewer'] = reviewer_username_new_policy
                    policy_data['data_inventory'] = new_policy_data_inventory
                    
                    created_policy = Policy.objects.create(**policy_data)
                    created_policies.append(created_policy)
                    
                    PolicyVersion.objects.create(
                        PolicyId=created_policy,
                        Version=new_version_str,
                        PolicyName=created_policy.PolicyName,
                        CreatedBy=created_policy.CreatedByName,
                        CreatedDate=created_policy.CreatedByDate,
                        PreviousVersionId=None
                    )
                    
                    for subpolicy_data in subpolicies_data:
                        if subpolicy_data.get('exclude', False):
                            continue
                        
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [field for field in required_fields if field not in subpolicy_data]
                        
                        if missing_fields:
                            return Response({
                                'error': f'Missing required fields for subpolicy in new policy {created_policy.PolicyName}: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)
                        
                        # Extract data_inventory for new subpolicy
                        new_subpolicy_data_inventory = subpolicy_data.get('data_inventory')
                        if new_subpolicy_data_inventory is not None:
                            if isinstance(new_subpolicy_data_inventory, str):
                                try:
                                    import json
                                    new_subpolicy_data_inventory = json.loads(new_subpolicy_data_inventory)
                                except json.JSONDecodeError:
                                    new_subpolicy_data_inventory = None
                            # If it's already a dict, use it as-is (even if empty)
                            elif not isinstance(new_subpolicy_data_inventory, dict):
                                new_subpolicy_data_inventory = None
                        
                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = created_policy
                        subpolicy.setdefault('CreatedByName', created_policy.CreatedByName)
                        subpolicy.setdefault('CreatedByDate', created_policy.CreatedByDate)
                        subpolicy.setdefault('Status', 'Under Review')
                        subpolicy['data_inventory'] = new_subpolicy_data_inventory
                        
                        SubPolicy.objects.create(**subpolicy)
                    
                    # Handle new_subpolicies array for new policies
                    for subpolicy_data in new_subpolicies_data:
                        if subpolicy_data.get('exclude', False):
                            continue
                        
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [field for field in required_fields if field not in subpolicy_data]
                        
                        if missing_fields:
                            return Response({
                                'error': f'Missing required fields for subpolicy in new policy {created_policy.PolicyName}: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)
                        
                        # Extract data_inventory for new subpolicy (extract BEFORE any processing, like TT does)
                        new_subpolicy_data_inventory_raw = subpolicy_data.get('data_inventory')
                        new_subpolicy_data_inventory = None
                        if new_subpolicy_data_inventory_raw:
                            if isinstance(new_subpolicy_data_inventory_raw, str):
                                try:
                                    import json
                                    new_subpolicy_data_inventory = json.loads(new_subpolicy_data_inventory_raw)
                                except json.JSONDecodeError:
                                    print(f"Warning: Invalid JSON in new subpolicy data_inventory, setting to None")
                                    new_subpolicy_data_inventory = None
                            elif isinstance(new_subpolicy_data_inventory_raw, dict):
                                # If it's already a dict, use it as-is (even if empty)
                                new_subpolicy_data_inventory = new_subpolicy_data_inventory_raw
                        
                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = created_policy
                        subpolicy.setdefault('CreatedByName', created_policy.CreatedByName)
                        subpolicy.setdefault('CreatedByDate', created_policy.CreatedByDate)
                        subpolicy.setdefault('Status', 'Under Review')
                        subpolicy['data_inventory'] = new_subpolicy_data_inventory
                        
                        print(f"DEBUG: Creating new subpolicy with data_inventory: {subpolicy.get('data_inventory')}")
                        SubPolicy.objects.create(**subpolicy)
            
            print(f"Created policies count: {len(created_policies)}")
            
            response_data = {
                'message': 'New framework version created successfully',
                'FrameworkId': new_framework.FrameworkId,
                'FrameworkName': new_framework.FrameworkName,
                'PreviousVersion': current_version,
                'NewVersion': new_version,
                'Identifier': new_framework.Identifier,
            }
            
            if created_policies:
                response_data['policies'] = [{
                    'PolicyId': policy.PolicyId,
                    'PolicyName': policy.PolicyName,
                    'Identifier': policy.Identifier,
                    'Version': policy.CurrentVersion
                } for policy in created_policies]
            
            return Response(response_data, status=status.HTTP_201_CREATED)

    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error creating new framework version', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)


"""
@api POST /api/policies/{pk}/create-version/
Creates a new version of an existing policy by cloning it with an incremented version number.
For example, if the original policy has version 1.0, the new version will be 1.1.
All subpolicies will be cloned with their details.
 
Example payload:
{
  "PolicyName": "Access Control Policy v1.1",
  "PolicyDescription": "Updated guidelines for access control",
  "StartDate": "2024-01-01",
  "EndDate": "2025-01-01",
  "Department": "IT,Security",
  "CreatedByName": "John Doe",
  "CreatedByDate": "2023-12-15",
  "Scope": "All IT systems and cloud environments",
  "Objective": "Implement enhanced access control measures"
}
 
Example response:
{
    "message": "New policy version created successfully",
    "PolicyId": 2,
    "PolicyName": "Access Control Policy v1.1",
    "PreviousVersion": 1.0,
    "NewVersion": 1.1,
    "FrameworkId": 1
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def create_policy_version(request, pk):
    original_policy = get_object_or_404(Policy, PolicyId=pk)

    try:
        with transaction.atomic():
            # Validate policy name
            policy_name = request.data.get('PolicyName')
            if not policy_name:
                return Response({'error': 'PolicyName is required'}, status=status.HTTP_400_BAD_REQUEST)

            current_version = str(original_policy.CurrentVersion).strip()

            # Find latest minor version under same major version
            latest_version = PolicyVersion.objects.filter(
                PolicyId__Identifier=original_policy.Identifier,
                Version__startswith=current_version.split('.')[0] + '.'
            ).order_by('-Version').first()

            if latest_version:
                parts = latest_version.Version.split('.')
                if len(parts) == 2:
                    major, minor = parts[0], int(parts[1])
                    new_version = f"{major}.{minor + 1}"
                else:
                    new_version = f"{current_version}.1"
            else:
                new_version = f"{current_version.split('.')[0]}.1"

            # Resolve Reviewer UserName from UserId if given in request, fallback to original
            reviewer_id = request.data.get('Reviewer')
            reviewer_name = None
            if reviewer_id:
                user_obj = Users.objects.filter(UserId=reviewer_id).first()
                if user_obj:
                    reviewer_name = user_obj.UserName
            if not reviewer_name:
                reviewer_name = original_policy.Reviewer  # fallback to existing username

            # Get data_inventory for policy (extract BEFORE any processing, like TT does)
            policy_data_inventory_raw = request.data.get('data_inventory')
            policy_data_inventory = None
            if policy_data_inventory_raw:
                if isinstance(policy_data_inventory_raw, str):
                    try:
                        import json
                        policy_data_inventory = json.loads(policy_data_inventory_raw)
                    except json.JSONDecodeError:
                        print(f"Warning: Invalid JSON in policy data_inventory, setting to None")
                        policy_data_inventory = None
                elif isinstance(policy_data_inventory_raw, dict):
                    # If it's already a dict, use it as-is (even if empty)
                    policy_data_inventory = policy_data_inventory_raw
            # If policy_data_inventory is still None, fall back to original policy's data_inventory
            if policy_data_inventory is None:
                policy_data_inventory = original_policy.data_inventory if hasattr(original_policy, 'data_inventory') else None
            
            print(f"DEBUG: Policy data_inventory received: {policy_data_inventory_raw}")
            print(f"DEBUG: Policy data_inventory processed: {policy_data_inventory}")
            
            # Prepare new policy data with Reviewer as UserName
            new_policy_data = {
                'FrameworkId': original_policy.FrameworkId,
                'CurrentVersion': new_version,
                'Status': 'Under Review',
                'PolicyName': policy_name,
                'PolicyDescription': request.data.get('PolicyDescription', original_policy.PolicyDescription),
                'StartDate': request.data.get('StartDate', original_policy.StartDate),
                'EndDate': request.data.get('EndDate', original_policy.EndDate),
                'Department': request.data.get('Department', original_policy.Department),
                'CreatedByName': request.data.get('CreatedByName', original_policy.CreatedByName),
                'CreatedByDate': datetime.date.today(),
                'Applicability': request.data.get('Applicability', original_policy.Applicability),
                'DocURL': request.data.get('DocURL', original_policy.DocURL),
                'Scope': request.data.get('Scope', original_policy.Scope),
                'Objective': request.data.get('Objective', original_policy.Objective),
                'Identifier': original_policy.Identifier,
                'PermanentTemporary': request.data.get('PermanentTemporary', original_policy.PermanentTemporary),
                'ActiveInactive': 'Inactive',
                'Reviewer': reviewer_name,  # Save UserName here
                'data_inventory': policy_data_inventory,
            }
            
            print(f"DEBUG: Creating policy with data_inventory: {new_policy_data.get('data_inventory')}")

            # Create new policy record
            new_policy = Policy.objects.create(**new_policy_data)

            # Find user ID for CreatedByName (for PolicyApproval)
            created_by_name = new_policy_data['CreatedByName']
            user_obj = Users.objects.filter(UserName=created_by_name).first()
            user_id = user_obj.UserId if user_obj else None

            # Prepare extracted data for PolicyApproval
            extracted_data = {
                'PolicyId': new_policy.PolicyId,
                'PolicyName': new_policy.PolicyName,
                'PolicyDescription': new_policy.PolicyDescription,
                'StartDate': new_policy.StartDate.isoformat() if isinstance(new_policy.StartDate, datetime.date) else new_policy.StartDate,
                'EndDate': new_policy.EndDate.isoformat() if isinstance(new_policy.EndDate, datetime.date) else new_policy.EndDate,
                'Department': new_policy.Department,
                'CreatedByName': new_policy.CreatedByName,
                'CreatedByDate': new_policy.CreatedByDate.isoformat() if isinstance(new_policy.CreatedByDate, datetime.date) else new_policy.CreatedByDate,
                'Applicability': new_policy.Applicability,
                'DocURL': new_policy.DocURL,
                'Scope': new_policy.Scope,
                'Objective': new_policy.Objective,
                'Identifier': new_policy.Identifier,
                'Status': new_policy.Status,
                'ActiveInactive': new_policy.ActiveInactive,
                'FrameworkId': new_policy.FrameworkId.FrameworkId,
                'policy_approval': {
                    'approved': None,
                    'remarks': ''
                },
                'subpolicies': []
            }

            # Check for existing policy approval to prevent duplicates
            existing_approval = PolicyApproval.objects.filter(
                Identifier=new_policy.Identifier,
                UserId=user_id,
                ReviewerId=reviewer_id,
                Version="u1",
                ApprovedNot=None  # Only check for pending approvals
            ).first()
            
            if existing_approval:
                print(f"DEBUG: ⚠️ Duplicate prevention: Policy approval already exists for Identifier {new_policy.Identifier} with ApprovalId: {existing_approval.ApprovalId}")
                print(f"  - Skipping duplicate creation")
            else:
                # Create PolicyApproval record with ReviewerId as UserId
                PolicyApproval.objects.create(
                    Identifier=new_policy.Identifier,
                    ExtractedData=extracted_data,
                    UserId=user_id,
                    ReviewerId=reviewer_id,
                    ApprovedNot=None,
                    Version="u1"
                )
                print(f"DEBUG: ✅ Created new PolicyApproval for Identifier {new_policy.Identifier}")

            # Get original PolicyVersion to link new version
            original_policy_version = PolicyVersion.objects.filter(
                PolicyId=original_policy,
                Version=str(original_policy.CurrentVersion)
            ).first()

            if not original_policy_version:
                return Response({
                    'error': f'No PolicyVersion found for PolicyId={original_policy.PolicyId} and Version={original_policy.CurrentVersion}. Data integrity issue.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Create new PolicyVersion linked to previous
            policy_version = PolicyVersion(
                PolicyId=new_policy,
                Version=new_version,
                PolicyName=new_policy.PolicyName,
                CreatedBy=new_policy.CreatedByName,
                CreatedDate=new_policy.CreatedByDate,
                PreviousVersionId=original_policy_version.VersionId
            )
            policy_version.save()

            # Handle subpolicy customizations and new subpolicies (same as your original logic)
            subpolicy_customizations = {}
            subpolicies_to_exclude = []

            if 'subpolicies' in request.data:
                for sp_data in request.data.get('subpolicies', []):
                    if 'original_subpolicy_id' in sp_data:
                        sp_id = sp_data.get('original_subpolicy_id')
                        if sp_data.get('exclude', False):
                            subpolicies_to_exclude.append(sp_id)
                        else:
                            if 'Identifier' not in sp_data:
                                return Response({
                                    'error': 'Identifier is required for modified subpolicies',
                                    'subpolicy_id': sp_id
                                }, status=status.HTTP_400_BAD_REQUEST)
                            subpolicy_customizations[sp_id] = sp_data

            original_subpolicies = SubPolicy.objects.filter(PolicyId=original_policy)
            for original_subpolicy in original_subpolicies:
                if original_subpolicy.SubPolicyId in subpolicies_to_exclude:
                    continue
                custom_data = subpolicy_customizations.get(original_subpolicy.SubPolicyId, {})

                # Get data_inventory for subpolicy (extract BEFORE any processing, like TT does)
                subpolicy_data_inventory_raw = custom_data.get('data_inventory')
                subpolicy_data_inventory = None
                if subpolicy_data_inventory_raw:
                    if isinstance(subpolicy_data_inventory_raw, str):
                        try:
                            import json
                            subpolicy_data_inventory = json.loads(subpolicy_data_inventory_raw)
                        except json.JSONDecodeError:
                            print(f"Warning: Invalid JSON in subpolicy data_inventory, setting to None")
                            subpolicy_data_inventory = None
                    elif isinstance(subpolicy_data_inventory_raw, dict):
                        # If it's already a dict, use it as-is (even if empty)
                        subpolicy_data_inventory = subpolicy_data_inventory_raw
                # If subpolicy_data_inventory is still None, fall back to original subpolicy's data_inventory
                if subpolicy_data_inventory is None:
                    subpolicy_data_inventory = original_subpolicy.data_inventory if hasattr(original_subpolicy, 'data_inventory') else None

                new_subpolicy_data = {
                    'PolicyId': new_policy,
                    'SubPolicyName': custom_data.get('SubPolicyName', original_subpolicy.SubPolicyName),
                    'CreatedByName': new_policy.CreatedByName,
                    'CreatedByDate': new_policy.CreatedByDate,
                    'Identifier': custom_data.get('Identifier', original_subpolicy.Identifier),
                    'Description': custom_data.get('Description', original_subpolicy.Description),
                    'Status': 'Under Review',
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_subpolicy.PermanentTemporary),
                    'Control': custom_data.get('Control', original_subpolicy.Control),
                    'data_inventory': subpolicy_data_inventory
                }

                print(f"DEBUG: Creating subpolicy with data_inventory: {new_subpolicy_data.get('data_inventory')}")
                SubPolicy.objects.create(**new_subpolicy_data)

            # Add new subpolicies if any
            if 'new_subpolicies' in request.data:
                for new_subpolicy_data in request.data.get('new_subpolicies', []):
                    required_fields = ['SubPolicyName', 'Description', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in new_subpolicy_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new subpolicy: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Extract data_inventory for new subpolicy (extract BEFORE any processing, like TT does)
                    new_subpolicy_data_inventory_raw = new_subpolicy_data.get('data_inventory')
                    new_subpolicy_data_inventory = None
                    if new_subpolicy_data_inventory_raw:
                        if isinstance(new_subpolicy_data_inventory_raw, str):
                            try:
                                import json
                                new_subpolicy_data_inventory = json.loads(new_subpolicy_data_inventory_raw)
                            except json.JSONDecodeError:
                                print(f"Warning: Invalid JSON in new subpolicy data_inventory, setting to None")
                                new_subpolicy_data_inventory = None
                        elif isinstance(new_subpolicy_data_inventory_raw, dict):
                            # If it's already a dict, use it as-is (even if empty)
                            new_subpolicy_data_inventory = new_subpolicy_data_inventory_raw

                    subpolicy = new_subpolicy_data.copy()
                    subpolicy['PolicyId'] = new_policy
                    if 'CreatedByName' not in subpolicy:
                        subpolicy['CreatedByName'] = new_policy.CreatedByName
                    if 'CreatedByDate' not in subpolicy:
                        subpolicy['CreatedByDate'] = new_policy.CreatedByDate
                    if 'Status' not in subpolicy:
                        subpolicy['Status'] = 'Under Review'
                    subpolicy['data_inventory'] = new_subpolicy_data_inventory

                    print(f"DEBUG: Creating new subpolicy with data_inventory: {subpolicy.get('data_inventory')}")
                    SubPolicy.objects.create(**subpolicy)

            # Handle any new policies if specified
            created_policies = []
            if 'new_policies' in request.data:
                for new_policy_data in request.data.get('new_policies', []):
                    required_fields = ['PolicyName', 'PolicyDescription', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in new_policy_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new policy: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Extract data_inventory for new policy (extract BEFORE any processing, like TT does)
                    new_policy_data_inventory_raw = new_policy_data.get('data_inventory')
                    new_policy_data_inventory = None
                    if new_policy_data_inventory_raw:
                        if isinstance(new_policy_data_inventory_raw, str):
                            try:
                                import json
                                new_policy_data_inventory = json.loads(new_policy_data_inventory_raw)
                            except json.JSONDecodeError:
                                print(f"Warning: Invalid JSON in new policy data_inventory, setting to None")
                                new_policy_data_inventory = None
                        elif isinstance(new_policy_data_inventory_raw, dict):
                            # If it's already a dict, use it as-is (even if empty)
                            new_policy_data_inventory = new_policy_data_inventory_raw

                    subpolicies_data = new_policy_data.pop('subpolicies', [])
                    new_subpolicies_data = new_policy_data.pop('new_subpolicies', [])
                    policy_data = new_policy_data.copy()
                    policy_data['FrameworkId'] = original_policy.FrameworkId
                    policy_data['CurrentVersion'] = new_version
                    policy_data['Status'] = 'Under Review'
                    policy_data['ActiveInactive'] = 'Inactive'
                    if 'CreatedByName' not in policy_data:
                        policy_data['CreatedByName'] = original_policy.CreatedByName
                    if 'CreatedByDate' not in policy_data:
                        policy_data['CreatedByDate'] = datetime.date.today()
                    policy_data['data_inventory'] = new_policy_data_inventory

                    print(f"DEBUG: Creating new policy with data_inventory: {policy_data.get('data_inventory')}")
                    created_policy = Policy.objects.create(**policy_data)
                    print(f"DEBUG: New policy created with ID: {created_policy.PolicyId}, data_inventory saved: {created_policy.data_inventory}")
                    created_policies.append(created_policy)

                    PolicyVersion.objects.create(
                        PolicyId=created_policy,
                        Version=new_version,
                        PolicyName=created_policy.PolicyName,
                        CreatedBy=created_policy.CreatedByName,
                        CreatedDate=created_policy.CreatedByDate,
                        PreviousVersionId=None
                    )

                    for subpolicy_data in subpolicies_data:
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [field for field in required_fields if field not in subpolicy_data]
                        if missing_fields:
                            return Response({
                                'error': f'Missing required fields for subpolicy in new policy {created_policy.PolicyName}: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)

                        # Extract data_inventory for new subpolicy (extract BEFORE any processing, like TT does)
                        new_subpolicy_data_inventory_raw = subpolicy_data.get('data_inventory')
                        new_subpolicy_data_inventory = None
                        if new_subpolicy_data_inventory_raw:
                            if isinstance(new_subpolicy_data_inventory_raw, str):
                                try:
                                    import json
                                    new_subpolicy_data_inventory = json.loads(new_subpolicy_data_inventory_raw)
                                except json.JSONDecodeError:
                                    print(f"Warning: Invalid JSON in new subpolicy data_inventory, setting to None")
                                    new_subpolicy_data_inventory = None
                            elif isinstance(new_subpolicy_data_inventory_raw, dict):
                                # If it's already a dict, use it as-is (even if empty)
                                new_subpolicy_data_inventory = new_subpolicy_data_inventory_raw

                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = created_policy
                        if 'CreatedByName' not in subpolicy:
                            subpolicy['CreatedByName'] = created_policy.CreatedByName
                        if 'CreatedByDate' not in subpolicy:
                            subpolicy['CreatedByDate'] = created_policy.CreatedByDate
                        if 'Status' not in subpolicy:
                            subpolicy['Status'] = 'Under Review'
                        subpolicy['data_inventory'] = new_subpolicy_data_inventory

                        print(f"DEBUG: Creating new subpolicy with data_inventory: {subpolicy.get('data_inventory')}")
                        SubPolicy.objects.create(**subpolicy)

            # Prepare response
            response_data = {
                'message': 'New policy version created successfully',
                'PolicyId': new_policy.PolicyId,
                'PolicyName': new_policy.PolicyName,
                'PreviousVersion': current_version,
                'NewVersion': new_version,
                'FrameworkId': new_policy.FrameworkId.FrameworkId,
                'Identifier': new_policy.Identifier,
            }

            if created_policies:
                response_data['policies'] = [{
                    'PolicyId': p.PolicyId,
                    'PolicyName': p.PolicyName,
                    'Identifier': p.Identifier,
                    'Version': p.CurrentVersion
                } for p in created_policies]

            return Response(response_data, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error creating new policy version', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api GET /api/frameworks/{framework_id}/export/
Exports all policies and their subpolicies for a specific framework to an Excel file in the following format:
Identifier, PolicyName (PolicyFamily), SubpolicyIdentifier, SubpolicyName, Control, Description

Example response:
Returns an Excel file as attachment
"""
@api_view(['GET'])
@permission_classes([AllowAny])
def export_policies_to_excel(request, framework_id):
    try:
        # Get the framework
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from datetime import datetime
        import re
        
        # Sanitize framework name for sheet title (remove invalid characters)
        def sanitize_sheet_name(name):
            # Remove or replace invalid characters for Excel sheet names
            # Excel sheet names cannot contain: \ / * ? : [ ]
            invalid_chars = r'[\\/*?:\[\]]'
            sanitized = re.sub(invalid_chars, '-', name)
            # Excel sheet names must be <= 31 characters
            return sanitized[:31].strip('-')  # Remove trailing dash if present
        
        def sanitize_filename(name):
            # Remove or replace invalid characters for file names
            invalid_chars = r'[<>:"/\\|?*]'
            return re.sub(invalid_chars, '-', name)
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Sanitize and set the sheet title
        sheet_title = sanitize_sheet_name(f"{framework.FrameworkName} Policies")
        if not sheet_title:  # If all characters were invalid
            sheet_title = "Framework Policies"
        ws.title = sheet_title
        
        # Define headers
        headers = ['Identifier', 'PolicyFamily', 'SubpolicyIdentifier', 'SubpolicyName', 'Control', 'Description']
        
        # Style for headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Style for policy rows
        policy_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Fetch all policies for this framework and their subpolicies
        policies = Policy.objects.filter(FrameworkId=framework).select_related('FrameworkId')
        
        row = 2  # Start from second row after headers
        for policy in policies:
            # Get subpolicies for this policy
            subpolicies = SubPolicy.objects.filter(PolicyId=policy)
            
            if not subpolicies:
                # If no subpolicies, write just the policy row
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col <= 2:  # Only fill first two columns
                        cell.fill = policy_fill
                
                ws.cell(row=row, column=1, value=policy.Identifier)
                ws.cell(row=row, column=2, value=policy.PolicyName)
                row += 1
            else:
                # Store the starting row for this policy
                policy_start_row = row
                
                # Write policy with each subpolicy
                for subpolicy in subpolicies:
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        if col <= 2:  # Only fill first two columns
                            cell.fill = policy_fill
                    
                    ws.cell(row=row, column=1, value=policy.Identifier)
                    ws.cell(row=row, column=2, value=policy.PolicyName)
                    ws.cell(row=row, column=3, value=subpolicy.Identifier)
                    ws.cell(row=row, column=4, value=subpolicy.SubPolicyName)
                    ws.cell(row=row, column=5, value=subpolicy.Control)
                    ws.cell(row=row, column=6, value=subpolicy.Description)
                    row += 1
                
                # Merge policy cells if there are multiple subpolicies
                if row - policy_start_row > 1:
                    ws.merge_cells(start_row=policy_start_row, start_column=1,
                                 end_row=row-1, end_column=1)
                    ws.merge_cells(start_row=policy_start_row, start_column=2,
                                 end_row=row-1, end_column=2)
                    
                    # Center the merged cells vertically
                    merged_cell1 = ws.cell(row=policy_start_row, column=1)
                    merged_cell2 = ws.cell(row=policy_start_row, column=2)
                    merged_cell1.alignment = Alignment(vertical='center')
                    merged_cell2.alignment = Alignment(vertical='center')
        
        # Add styling to the data cells and wrap text
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                if not cell.alignment:  # Don't override merged cell alignment
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Adjust column widths
        column_widths = [20, 30, 20, 30, 40, 50]  # Preset widths for each column
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Create response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Create a safe filename
        safe_framework_name = sanitize_filename(framework.FrameworkName)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'{safe_framework_name}_policies_{timestamp}.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save the workbook to the response
        wb.save(response)
        return response
        
    except Exception as e:
        return Response({
            'error': 'Error exporting policies to Excel',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def policy_list(request):
    try:
        # Get query parameters
        framework_id = request.GET.get('framework_id')
        status = request.GET.get('status')
        active_inactive = request.GET.get('active_inactive')

        # Base queryset
        policies_query = Policy.objects.all()

        # Apply filters if provided
        if framework_id:
            policies_query = policies_query.filter(FrameworkId=framework_id)
        if status:
            policies_query = policies_query.filter(Status=status)
        if active_inactive:
            policies_query = policies_query.filter(ActiveInactive=active_inactive)

        # Get all policies
        policies = policies_query.select_related('FrameworkId')

        # Calculate summary counts
        summary_counts = {
            'active': Policy.objects.filter(ActiveInactive='Active').count(),
            'inactive': Policy.objects.filter(ActiveInactive='Inactive').count(),
            'approved': Policy.objects.filter(Status='Approved').count(),
            'rejected': Policy.objects.filter(Status='Rejected').count(),
            'under_review': Policy.objects.filter(Status='Under Review').count()
        }

        # Serialize policies
        serializer = PolicySerializer(policies, many=True)

        return Response({
            'policies': serializer.data,
            'summary_counts': summary_counts
        })

    except Exception as e:
        return Response({
            'error': 'Error fetching policies',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def list_users(request):
    try:
        print("DEBUG: list_users endpoint called")
        print(f"DEBUG: Request method: {request.method}")
        print(f"DEBUG: Request user: {request.user}")
        print(f"DEBUG: Request headers: {dict(request.headers)}")
        
        users = Users.objects.all()
        print(f"DEBUG: Found {users.count()} users in database")
        
        from .serializers import UserSerializer
        serializer = UserSerializer(users, many=True)
        
        print(f"DEBUG: Serialized {len(serializer.data)} users")
        if serializer.data:
            print(f"DEBUG: First user data: {serializer.data[0]}")
            print(f"DEBUG: Sample user keys: {list(serializer.data[0].keys()) if serializer.data else 'No users'}")
            # Debug IsActive values
            for idx, user_data in enumerate(serializer.data[:5]):  # Check first 5 users
                print(f"DEBUG: User {idx+1} - UserId: {user_data.get('UserId')}, IsActive: {user_data.get('IsActive')}, DepartmentId: {user_data.get('DepartmentId')}, DepartmentName: {user_data.get('DepartmentName')}")
        
        response_data = {
            'success': True,
            'users': serializer.data
        }
        print(f"DEBUG: Returning response with {len(serializer.data)} users")
        return Response(response_data, status=status.HTTP_200_OK)
    except Exception as e:
        print(f"ERROR: Error in list_users: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': 'Error fetching users',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['PATCH', 'PUT'])
@permission_classes([AllowAny])
def update_user_status(request, user_id):
    """
    Update user IsActive status (activate/deactivate user)
    Only GRC Administrators can update user status
    """
    try:
        # Check if the requesting user is a GRC Administrator
        from .rbac.utils import RBACUtils
        admin_user_id = RBACUtils.get_user_id_from_request(request)
        
        if not admin_user_id:
            return Response({
                'success': False,
                'message': 'Authentication required to update user status'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if user is GRC Administrator
        if not RBACUtils.is_system_admin(admin_user_id):
            return Response({
                'success': False,
                'message': 'Only GRC Administrators can update user status'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the user to update
        try:
            user = Users.objects.get(UserId=user_id)
        except Users.DoesNotExist:
            return Response({
                'success': False,
                'message': 'User not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get the new status from request
        data = request.data
        new_status = data.get('isActive', data.get('IsActive'))
        
        if new_status not in ['Y', 'N', 'y', 'n']:
            return Response({
                'success': False,
                'message': 'Invalid status. Must be Y or N'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Normalize to uppercase
        new_status = new_status.upper()
        old_status = user.IsActive
        
        # Update user status
        user.IsActive = new_status
        user.save(update_fields=['IsActive'])
        
        logger.info(f"User {user.UserName} (ID: {user.UserId}) status updated from {old_status} to {new_status} by admin {admin_user_id}")
        print(f"[DEBUG] User {user.UserName} (ID: {user.UserId}) status updated from {old_status} to {new_status}")
        
        return Response({
            'success': True,
            'message': f'User status updated to {"Active" if new_status == "Y" else "Inactive"}',
            'user': {
                'UserId': user.UserId,
                'UserName': user.UserName,
                'IsActive': user.IsActive
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating user status: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'error': 'Error updating user status',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_user_role_simple(request):
    """
    Get user role and permissions from RBAC table
    Supports both session and JWT authentication
    """
    try:
        # Try to get user_id from JWT token first
        user_id = None
        auth_header = request.headers.get('Authorization')
        
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
            try:
                from .authentication import verify_jwt_token
                payload = verify_jwt_token(token)
                if payload and 'user_id' in payload:
                    user_id = payload['user_id']
            except Exception as e:
                logger.warning(f"JWT token verification failed: {str(e)}")
        
        # Fallback to session if JWT not available
        if not user_id:
            user_id = request.session.get('user_id')
            
        if not user_id:
            return Response({
                'success': False,
                'error': 'No user ID found in JWT token or session'
            }, status=status.HTTP_401_UNAUTHORIZED)

        # Get RBAC record
        from .models import RBAC
        rbac_record = RBAC.objects.filter(user_id=user_id, is_active='Y').first()
        
        if not rbac_record:
            return Response({
                'success': False,
                'error': 'No active RBAC record found'
            }, status=status.HTTP_404_NOT_FOUND)

        # Get user details
        user = Users.objects.filter(UserId=user_id).first()
        if not user:
            return Response({
                'success': False,
                'error': 'User not found'
            }, status=status.HTTP_404_NOT_FOUND)

        return Response({
            'success': True,
            'user_id': user_id,
            'username': user.UserName,
            'role': rbac_record.role,
            'permissions': {
                'view_all_policy': rbac_record.view_all_policy,
                'create_policy': rbac_record.create_policy,
                'edit_policy': rbac_record.edit_policy,
                'approve_policy': rbac_record.approve_policy,
                'policy_performance_analytics': rbac_record.policy_performance_analytics
            }
        })

    except Exception as e:
        logger.error(f"Error in get_user_role_simple: {str(e)}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_users_for_dropdown_simple(request):
    """Simple users dropdown endpoint that gets users from session data"""
    try:
        # Get all users from RBAC table
        from .models import RBAC
        rbac_users = RBAC.objects.filter(is_active='Y').order_by('username')
        
        user_data = []
        for rbac_user in rbac_users:
            user_data.append({
                'UserId': rbac_user.user_id,
                'UserName': rbac_user.username,
                'Role': rbac_user.role
            })
        
        return Response(user_data)
    except Exception as e:
        print(f"Error fetching users for dropdown: {e}")
        return Response({"error": str(e)}, status=500)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_framework_explorer_data(request):
    """
    API endpoint for the Framework Explorer page
    Returns frameworks with their status and counts of active/inactive policies
    """
    try:
        # Get all frameworks
        frameworks = Framework.objects.all()
       
        # Prepare response data with additional counts
        framework_data = []
        for fw in frameworks:
            # Count policies for this framework
            active_policies = Policy.objects.filter(
                FrameworkId=fw.FrameworkId,
                ActiveInactive='Active'
            ).count()
           
            inactive_policies = Policy.objects.filter(
                FrameworkId=fw.FrameworkId,
                ActiveInactive='Inactive'
            ).count()
           
            framework_data.append({
                'id': fw.FrameworkId,
                'name': fw.FrameworkName,
                'category': fw.Category or 'Uncategorized',
                'description': fw.FrameworkDescription,
                'status': fw.ActiveInactive,  # 'Active' or 'Inactive'
                'internalExternal': fw.InternalExternal or 'Internal',  # Add Internal/External field
                'active_policies_count': active_policies,
                'inactive_policies_count': inactive_policies
            })
       
        # Calculate summary counts
        active_frameworks = Framework.objects.filter(ActiveInactive='Active').count()
        inactive_frameworks = Framework.objects.filter(ActiveInactive='Inactive').count()
        active_policies = Policy.objects.filter(ActiveInactive='Active').count()
        inactive_policies = Policy.objects.filter(ActiveInactive='Inactive').count()
       
        return Response({
            'frameworks': framework_data,
            'summary': {
                'active_frameworks': active_frameworks,
                'inactive_frameworks': inactive_frameworks,
                'active_policies': active_policies,
                'inactive_policies': inactive_policies
            }
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def get_framework_policies(request, framework_id):
    """
    API endpoint for the Framework Policies page
    Returns policies for a specific framework
    """
    try:
        # Check if framework exists
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
       
        # Get policies for this framework
        policies = Policy.objects.filter(FrameworkId=framework_id)
       
        # Prepare response data
        policy_data = []
        for policy in policies:
            policy_data.append({
                'id': policy.PolicyId,
                'name': policy.PolicyName,
                'category': policy.Department or 'General',
                'description': policy.PolicyDescription,
                'status': policy.ActiveInactive  # 'Active' or 'Inactive'
            })
       
        # Framework details
        framework_data = {
            'id': framework.FrameworkId,
            'name': framework.FrameworkName,
            'category': framework.Category,
            'description': framework.FrameworkDescription
        }
       
        return Response({
            'framework': framework_data,
            'policies': policy_data
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['POST'])
@permission_classes([AllowAny])
def toggle_framework_status(request, framework_id):
    """
    Toggle the ActiveInactive status of a framework
    When cascadeToApproved=True, also update the status of all approved policies
    but leave their subpolicies status unchanged
    When activating a framework, all other related versions (connected through PreviousVersionId chain)
    are set to inactive
    """
    try:
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
       
        # Toggle status
        new_status = 'Inactive' if framework.ActiveInactive == 'Active' else 'Active'
        framework.ActiveInactive = new_status
        framework.save()
        
        # When activating a framework, set all other related versions to inactive
        other_versions_deactivated = 0
        if new_status == 'Active':
            # Find all versions connected to this framework through the PreviousVersionId chain
            related_version_ids = set()
            
            # First, get all framework versions associated with this framework
            try:
                framework_versions = FrameworkVersion.objects.filter(FrameworkId=framework)
                if framework_versions.exists():
                    # For each version, find all related versions through PreviousVersionId chain
                    for version in framework_versions:
                        # Start with this version's ID
                        to_process = [version.VersionId]
                        processed = set()
                        
                        # Follow chain forward and backward
                        while to_process:
                            current_id = to_process.pop()
                            if current_id in processed:
                                continue
                            
                            processed.add(current_id)
                            
                            # Find versions where this is PreviousVersionId (forward)
                            next_versions = FrameworkVersion.objects.filter(PreviousVersionId=current_id)
                            for next_ver in next_versions:
                                related_version_ids.add(next_ver.FrameworkId.FrameworkId)
                                if next_ver.VersionId not in processed:
                                    to_process.append(next_ver.VersionId)
                            
                            # Find version that this version points to (backward)
                            try:
                                current_version = FrameworkVersion.objects.get(VersionId=current_id)
                                if current_version.PreviousVersionId and current_version.PreviousVersionId not in processed:
                                    prev_version = FrameworkVersion.objects.get(VersionId=current_version.PreviousVersionId)
                                    related_version_ids.add(prev_version.FrameworkId.FrameworkId)
                                    to_process.append(current_version.PreviousVersionId)
                            except FrameworkVersion.DoesNotExist:
                                pass
                    
                    # Remove the current framework from the related list
                    if framework.FrameworkId in related_version_ids:
                        related_version_ids.remove(framework.FrameworkId)
                    
                    # Set all related frameworks to inactive
                    if related_version_ids:
                        related_frameworks = Framework.objects.filter(FrameworkId__in=related_version_ids)
                        for related_framework in related_frameworks:
                            if related_framework.ActiveInactive == 'Active':
                                related_framework.ActiveInactive = 'Inactive'
                                related_framework.save()
                                other_versions_deactivated += 1
            except Exception as e:
                print(f"Error finding related framework versions: {str(e)}")
        
        policies_affected = 0
        subpolicies_affected = 0
        
        # Check if we should cascade to approved policies
        cascade_to_approved = request.data.get('cascadeToApproved', False)
        if cascade_to_approved:
            # Get all approved policies for this framework
            approved_policies = Policy.objects.filter(
                FrameworkId=framework,
                Status='Approved'
            )
            
            # Update their status to match the framework
            for policy in approved_policies:
                policy.ActiveInactive = new_status
                policy.save()
                policies_affected += 1
                
                # Count subpolicies but don't change their status
                subpolicies_count = SubPolicy.objects.filter(PolicyId=policy).count()
                subpolicies_affected += subpolicies_count
       
        return Response({
            'id': framework.FrameworkId,
            'status': new_status,
            'message': f'Framework status updated to {new_status}',
            'policies_affected': policies_affected,
            'subpolicies_affected': subpolicies_affected,
            'other_versions_deactivated': other_versions_deactivated
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def toggle_policy_status(request, policy_id):
    """
    Toggle the ActiveInactive status of a policy
    When cascadeSubpolicies=True, we don't change subpolicy Status but still count them
    When activating a policy, all other related versions (connected through PreviousVersionId chain)
    are set to inactive
    """
    try:
        policy = get_object_or_404(Policy, PolicyId=policy_id)
       
        # Toggle status
        new_status = 'Inactive' if policy.ActiveInactive == 'Active' else 'Active'
        policy.ActiveInactive = new_status
        policy.save()
        
        # When activating a policy, set all other related versions to inactive
        other_versions_deactivated = 0
        if new_status == 'Active':
            # Find all versions connected to this policy through the PreviousVersionId chain
            related_policy_ids = set()
            
            # First, get all policy versions associated with this policy
            try:
                policy_versions = PolicyVersion.objects.filter(PolicyId=policy)
                if policy_versions.exists():
                    # For each version, find all related versions through PreviousVersionId chain
                    for version in policy_versions:
                        # Start with this version's ID
                        to_process = [version.VersionId]
                        processed = set()
                        
                        # Follow chain forward and backward
                        while to_process:
                            current_id = to_process.pop()
                            if current_id in processed:
                                continue
                            
                            processed.add(current_id)
                            
                            # Find versions where this is PreviousVersionId (forward)
                            next_versions = PolicyVersion.objects.filter(PreviousVersionId=current_id)
                            for next_ver in next_versions:
                                related_policy_ids.add(next_ver.PolicyId_id)
                                if next_ver.VersionId not in processed:
                                    to_process.append(next_ver.VersionId)
                            
                            # Find version that this version points to (backward)
                            try:
                                current_version = PolicyVersion.objects.get(VersionId=current_id)
                                if current_version.PreviousVersionId and current_version.PreviousVersionId not in processed:
                                    prev_version = PolicyVersion.objects.get(VersionId=current_version.PreviousVersionId)
                                    related_policy_ids.add(prev_version.PolicyId_id)
                                    to_process.append(current_version.PreviousVersionId)
                            except PolicyVersion.DoesNotExist:
                                pass
                    
                    # Remove the current policy from the related list
                    if policy.PolicyId in related_policy_ids:
                        related_policy_ids.remove(policy.PolicyId)
                    
                    # Set all related policies to inactive
                    if related_policy_ids:
                        related_policies = Policy.objects.filter(PolicyId__in=related_policy_ids)
                        for related_policy in related_policies:
                            if related_policy.ActiveInactive == 'Active':
                                related_policy.ActiveInactive = 'Inactive'
                                related_policy.save()
                                other_versions_deactivated += 1
            except Exception as e:
                print(f"Error finding related policy versions: {str(e)}")
        
        subpolicies_affected = 0
        
        # Count subpolicies but don't change their status
        cascade_to_subpolicies = request.data.get('cascadeSubpolicies', False)
        if cascade_to_subpolicies:
            # Get count of all subpolicies for this policy
            subpolicies_affected = SubPolicy.objects.filter(PolicyId=policy).count()
       
        return Response({
            'id': policy.PolicyId,
            'status': new_status,
            'message': f'Policy status updated to {new_status}',
            'subpolicies_affected': subpolicies_affected,
            'other_versions_deactivated': other_versions_deactivated
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def get_framework_details(request, framework_id):
    """
    API endpoint for detailed framework information
    Returns all details of a framework regardless of status
    """
    try:
        # Get framework by ID
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
       
        # Create response data
        response_data = {
            'FrameworkId': framework.FrameworkId,
            'FrameworkName': framework.FrameworkName,
            'CurrentVersion': framework.CurrentVersion,
            'FrameworkDescription': framework.FrameworkDescription,
            'EffectiveDate': framework.EffectiveDate,
            'CreatedByName': framework.CreatedByName,
            'CreatedByDate': framework.CreatedByDate,
            'Category': framework.Category,
            'DocURL': framework.DocURL,
            'Identifier': framework.Identifier,
            'StartDate': framework.StartDate,
            'EndDate': framework.EndDate,
            'Status': framework.Status,
            'ActiveInactive': framework.ActiveInactive
        }
       
        return Response(response_data)
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_details(request, policy_id):
    """
    API endpoint for detailed policy information
    Returns all details of a policy regardless of status
    """
    try:
        # Get policy by ID
        policy = get_object_or_404(Policy, PolicyId=policy_id)
       
        # Get all subpolicies for this policy
        subpolicies = SubPolicy.objects.filter(PolicyId=policy)
        subpolicy_data = []
       
        for subpolicy in subpolicies:
            subpolicy_data.append({
                'SubPolicyId': subpolicy.SubPolicyId,
                'SubPolicyName': subpolicy.SubPolicyName,
                'CreatedByName': subpolicy.CreatedByName,
                'CreatedByDate': subpolicy.CreatedByDate,
                'Identifier': subpolicy.Identifier,
                'Description': subpolicy.Description,
                'Status': subpolicy.Status,
                'PermanentTemporary': subpolicy.PermanentTemporary,
                'Control': subpolicy.Control
            })
       
        # Create response data
        response_data = {
            'PolicyId': policy.PolicyId,
            'PolicyName': policy.PolicyName,
            'PolicyDescription': policy.PolicyDescription,
            'CurrentVersion': policy.CurrentVersion,
            'StartDate': policy.StartDate,
            'EndDate': policy.EndDate,
            'Department': policy.Department,
            'CreatedByName': policy.CreatedByName,
            'CreatedByDate': policy.CreatedByDate,
            'Applicability': policy.Applicability,
            'DocURL': policy.DocURL,
            'Scope': policy.Scope,
            'Objective': policy.Objective,
            'Identifier': policy.Identifier,
            'PermanentTemporary': policy.PermanentTemporary,
            'Status': policy.Status,
            'ActiveInactive': policy.ActiveInactive,
            'subpolicies': subpolicy_data
        }
       
        return Response(response_data)
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#all policies code
@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_frameworks(request):
    """
    API endpoint to get all frameworks for AllPolicies.vue component.
    """
    try:
        frameworks = Framework.objects.all()
        
        frameworks_data = []
        for framework in frameworks:
            framework_data = {
                'id': framework.FrameworkId,
                'name': framework.FrameworkName,
                'category': framework.Category,
                'status': framework.ActiveInactive,
                'description': framework.FrameworkDescription,
                'versions': []
            }
            
            # Get versions for this framework
            versions = FrameworkVersion.objects.filter(FrameworkId=framework)
            version_data = []
            for version in versions:
                version_data.append({
                    'id': version.VersionId,
                    'name': f"v{version.Version}",
                    'version': version.Version
                })
            
            framework_data['versions'] = version_data
            frameworks_data.append(framework_data)
            
        return Response(frameworks_data)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_framework_version_policies(request, version_id):
    """
    API endpoint to get all policies for a specific framework version for AllPolicies.vue component.
    """
    try:
        # Get the framework version
        framework_version = get_object_or_404(FrameworkVersion, VersionId=version_id)
        framework = framework_version.FrameworkId
        
        # Get ALL policies for this framework (regardless of CurrentVersion)
        # This ensures we show all policies that belong to the framework
        policies = Policy.objects.filter(FrameworkId=framework)
        
        policies_data = []
        for policy in policies:
            policy_data = {
                'id': policy.PolicyId,
                'name': policy.PolicyName,
                'category': policy.Department,
                'status': policy.Status,
                'description': policy.PolicyDescription,
                'versions': []
            }
            
            # Get ALL versions for this policy
            # This allows showing all versions of a policy when viewing any framework version
            policy_versions = PolicyVersion.objects.filter(PolicyId=policy).order_by('Version')
            versions_data = []
            for version in policy_versions:
                versions_data.append({
                    'id': version.VersionId,
                    'name': f"v{version.Version}",
                    'version': version.Version
                })
            
            policy_data['versions'] = versions_data
            policies_data.append(policy_data)
            
        return Response(policies_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_policies(request):
    """
    API endpoint to get all policies for AllPolicies.vue component.
    """
    try:
        # Optional framework filter
        framework_id = request.GET.get('framework_id')
        
        # Start with all policies
        policies_query = Policy.objects.all()
        
        # Apply framework filter if provided
        if framework_id:
            policies_query = policies_query.filter(FrameworkId_id=framework_id)
        
        policies_data = []
        for policy in policies_query:
            policy_data = {
                'id': policy.PolicyId,
                'name': policy.PolicyName,
                'category': policy.Department,
                'status': policy.Status,
                'description': policy.PolicyDescription,
                'versions': []
            }
            
            # Get versions for this policy
            policy_versions = PolicyVersion.objects.filter(PolicyId=policy)
            versions_data = []
            for version in policy_versions:
                versions_data.append({
                    'id': version.VersionId,
                    'name': f"v{version.Version}",
                    'version': version.Version
                })
            
            policy_data['versions'] = versions_data
            policies_data.append(policy_data)
            
        return Response(policies_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_policy_versions(request, policy_id):
    """
    API endpoint to get all versions of a specific policy for AllPolicies.vue component.
    Implements a dedicated version that handles version chains through PreviousVersionId.
    """
    try:
        print(f"Request received for policy versions, policy_id: {policy_id}, type: {type(policy_id)}")
        
        # Ensure we have a valid integer ID
        try:
            policy_id = int(policy_id)
        except (ValueError, TypeError):
            return Response({'error': f'Invalid policy ID format: {policy_id}'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Get the base policy
        try:
            policy = Policy.objects.get(PolicyId=policy_id)
            print(f"Found policy: {policy.PolicyName} (ID: {policy.PolicyId})")
        except Policy.DoesNotExist:
            print(f"Policy with ID {policy_id} not found")
            return Response({'error': f'Policy with ID {policy_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get the direct policy version
        try:
            direct_version = PolicyVersion.objects.get(PolicyId=policy)
            print(f"Found direct policy version: {direct_version.VersionId}")
        except PolicyVersion.DoesNotExist:
            print(f"No policy version found for policy ID {policy_id}")
            return Response({'error': f'No version found for policy with ID {policy_id}'}, 
                           status=status.HTTP_404_NOT_FOUND)
        except PolicyVersion.MultipleObjectsReturned:
            # If there are multiple versions, get all of them
            direct_versions = list(PolicyVersion.objects.filter(PolicyId=policy))
            print(f"Found {len(direct_versions)} direct versions for policy {policy_id}")
            direct_version = direct_versions[0]  # Just use the first one for starting the chain
        
        # Start building version chain
        all_versions = {}
        visited = set()
        to_process = [direct_version.VersionId]
        
        # Find all versions in the chain
        while to_process:
            current_id = to_process.pop(0)
            
            if current_id in visited:
                continue
                
            visited.add(current_id)
            
            try:
                current_version = PolicyVersion.objects.get(VersionId=current_id)
                all_versions[current_id] = current_version
                
                # Follow PreviousVersionId chain backward
                if current_version.PreviousVersionId and current_version.PreviousVersionId not in visited:
                    to_process.append(current_version.PreviousVersionId)
                    
                # Find versions that reference this one as their previous version
                next_versions = PolicyVersion.objects.filter(PreviousVersionId=current_id)
                for next_ver in next_versions:
                    if next_ver.VersionId not in visited:
                        to_process.append(next_ver.VersionId)
            except PolicyVersion.DoesNotExist:
                print(f"Version with ID {current_id} not found")
                continue
        
        versions_data = []
        for version_id, version in all_versions.items():
            try:
                # Get the policy this version belongs to
                version_policy = version.PolicyId
                
                # Count subpolicies for this policy
                subpolicy_count = SubPolicy.objects.filter(PolicyId=version_policy).count()
                
                # Get previous version details if available
                previous_version = None
                if version.PreviousVersionId:
                    try:
                        previous_version = PolicyVersion.objects.get(VersionId=version.PreviousVersionId)
                    except PolicyVersion.DoesNotExist:
                        pass
                
                # Create a descriptive name
                formatted_name = f"{version.PolicyName} v{version.Version}" if version.PolicyName else f"{version_policy.PolicyName} v{version.Version}"
                
                version_data = {
                    'id': version.VersionId,
                    'policy_id': version_policy.PolicyId,
                    'name': formatted_name,
                    'version': version.Version,
                    'category': version_policy.Department or 'General',
                    'status': version_policy.Status or 'Unknown',
                    'description': version_policy.PolicyDescription or '',
                    'created_date': version.CreatedDate,
                    'created_by': version.CreatedBy,
                    'subpolicy_count': subpolicy_count,
                    'previous_version_id': version.PreviousVersionId,
                    'previous_version_name': previous_version.PolicyName + f" v{previous_version.Version}" if previous_version else None
                }
                versions_data.append(version_data)
                print(f"Added version: {version.VersionId} - {formatted_name}, Previous: {version.PreviousVersionId}")
            except Exception as e:
                print(f"Error processing version {version_id}: {str(e)}")
                # Continue to next version
        
        # Sort versions by version number (descending)
        versions_data.sort(key=lambda x: float(x['version']), reverse=True)

        
        
        print(f"Returning {len(versions_data)} policy versions")
        return Response(versions_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_policy_versions: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_subpolicies(request):
    """
    API endpoint to get all subpolicies for AllPolicies.vue component.
    """
    try:
        print("Request received for all subpolicies")
        
        # Optional framework filter
        framework_id = request.GET.get('framework_id')
        print(f"Framework filter: {framework_id}")
        
        # Start with all subpolicies
        subpolicies_query = SubPolicy.objects.all()
        
        # If framework filter is provided, filter through policies
        if framework_id:
            try:
                policy_ids = Policy.objects.filter(FrameworkId_id=framework_id).values_list('PolicyId', flat=True)
                print(f"Found {len(policy_ids)} policies for framework {framework_id}")
                subpolicies_query = subpolicies_query.filter(PolicyId_id__in=policy_ids)
            except Exception as e:
                print(f"Error filtering by framework: {str(e)}")
                # Continue with all subpolicies if framework filtering fails
        
        print(f"Found {subpolicies_query.count()} subpolicies")
        
        subpolicies_data = []
        for subpolicy in subpolicies_query:
            try:
                # Get the policy this subpolicy belongs to
                try:
                    policy = Policy.objects.get(PolicyId=subpolicy.PolicyId_id)
                    policy_name = policy.PolicyName
                    department = policy.Department
                except Policy.DoesNotExist:
                    print(f"Policy with ID {subpolicy.PolicyId_id} not found for subpolicy {subpolicy.SubPolicyId}")
                    policy_name = "Unknown Policy"
                    department = "Unknown"
                
                subpolicy_data = {
                    'id': subpolicy.SubPolicyId,
                    'name': subpolicy.SubPolicyName,
                    'category': department or 'General',
                    'status': subpolicy.Status or 'Unknown',
                    'description': subpolicy.Description or '',
                    'control': subpolicy.Control or '',
                    'identifier': subpolicy.Identifier,
                    'permanent_temporary': subpolicy.PermanentTemporary,
                    'policy_id': subpolicy.PolicyId_id,
                    'policy_name': policy_name,
                    'created_by': subpolicy.CreatedByName,
                    'created_date': subpolicy.CreatedByDate
                }
                subpolicies_data.append(subpolicy_data)
                # print(f"Added subpolicy: {subpolicy.SubPolicyId} - {subpolicy.SubPolicyName}")
            except Exception as e:
                print(f"Error processing subpolicy {subpolicy.SubPolicyId}: {str(e)}")
                # Continue to next subpolicy
        
        print(f"Returning {len(subpolicies_data)} subpolicies")
        return Response(subpolicies_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_subpolicies: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_subpolicy_details(request, subpolicy_id):
    """
    API endpoint to get details of a specific subpolicy for AllPolicies.vue component.
    """
    try:
        subpolicy = get_object_or_404(SubPolicy, SubPolicyId=subpolicy_id)
        policy = subpolicy.PolicyId
        
        subpolicy_data = {
            'id': subpolicy.SubPolicyId,
            'name': subpolicy.SubPolicyName,
            'category': policy.Department,
            'status': subpolicy.Status,
            'description': subpolicy.Description,
            'control': subpolicy.Control,
            'identifier': subpolicy.Identifier,
            'permanent_temporary': subpolicy.PermanentTemporary,
            'policy_id': policy.PolicyId,
            'policy_name': policy.PolicyName
        }
        
        return Response(subpolicy_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_framework_versions(request, framework_id):
    """
    API endpoint to get all versions of a specific framework for AllPolicies.vue component.
    Implements a dedicated version that handles version chains through PreviousVersionId.
    """
    try:
        print(f"Request received for framework versions, framework_id: {framework_id}, type: {type(framework_id)}")
        
        # Ensure we have a valid integer ID
        try:
            framework_id = int(framework_id)
        except (ValueError, TypeError):
            return Response({'error': f'Invalid framework ID format: {framework_id}'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Get the base framework
        try:
            framework = Framework.objects.get(FrameworkId=framework_id)
            print(f"Found framework: {framework.FrameworkName} (ID: {framework.FrameworkId})")
        except Framework.DoesNotExist:
            print(f"Framework with ID {framework_id} not found")
            return Response({'error': f'Framework with ID {framework_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get direct versions that belong to this framework
        direct_versions = list(FrameworkVersion.objects.filter(FrameworkId=framework).order_by('-Version'))
        print(f"Found {len(direct_versions)} direct versions for framework {framework_id}")
        
        # Create a dictionary to track all versions by VersionId
        all_versions = {v.VersionId: v for v in direct_versions}
        
        # Create a queue to process versions and follow PreviousVersionId links
        to_process = [v.VersionId for v in direct_versions]
        
        # Process the version chain by following PreviousVersionId links
        while to_process:
            current_id = to_process.pop(0)
            
            # Find versions that reference this one as their previous version
            linked_versions = FrameworkVersion.objects.filter(PreviousVersionId=current_id)
            print(f"Found {len(linked_versions)} linked versions for version ID {current_id}")
            
            for linked in linked_versions:
                if linked.VersionId not in all_versions:
                    # Add newly discovered version to our collection
                    all_versions[linked.VersionId] = linked
                    to_process.append(linked.VersionId)
        
        versions_data = []
        for version_id, version in all_versions.items():
            try:
                # Get the framework this version belongs to
                version_framework = version.FrameworkId
                
                # Count policies for this framework (without filtering by version)
                # This gets all policies associated with this framework regardless of version
                policy_count = Policy.objects.filter(
                    FrameworkId=version_framework
                ).count()
                
                print(f"Found {policy_count} policies for framework {version_framework.FrameworkId}")
                
                # Get previous version details if available
                previous_version = None
                if version.PreviousVersionId:
                    try:
                        previous_version = FrameworkVersion.objects.get(VersionId=version.PreviousVersionId)
                    except FrameworkVersion.DoesNotExist:
                        pass
                
                # Create a more descriptive name using the FrameworkName from the database
                # and appending the version number like v1.0, v2.0, etc.
                formatted_name = f"{version.FrameworkName} v{version.Version}"
                
                version_data = {
                    'id': version.VersionId,
                    'name': formatted_name,
                    'version': version.Version,
                    'category': version_framework.Category or 'General',
                    'status': version_framework.ActiveInactive or 'Unknown',
                    'description': version_framework.FrameworkDescription or '',
                    'created_date': version.CreatedDate,
                    'created_by': version.CreatedBy,
                    'policy_count': policy_count,
                    'previous_version_id': version.PreviousVersionId,
                    'previous_version_name': previous_version.FrameworkName + f" v{previous_version.Version}" if previous_version else None,
                    'framework_id': version_framework.FrameworkId
                }
                versions_data.append(version_data)
                print(f"Added version: {version.VersionId} - {formatted_name}, Previous: {version.PreviousVersionId}")
            except Exception as e:
                print(f"Error processing version {version_id}: {str(e)}")
                # Continue to next version
        
        # Sort versions by version number (descending)
        versions_data.sort(key=lambda x: float(x['version']), reverse=True)
        
        print(f"Returning {len(versions_data)} versions")
        return Response(versions_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_framework_versions: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_policy_version_subpolicies(request, version_id):
    """
    API endpoint to get all subpolicies for a specific policy version for AllPolicies.vue component.
    Implements a dedicated version instead of using the existing get_policy_version_subpolicies function.
    """
    try:
        print(f"Request received for policy version subpolicies, version_id: {version_id}, type: {type(version_id)}")
        
        # Ensure we have a valid integer ID
        try:
            version_id = int(version_id)
        except (ValueError, TypeError):
            print(f"Invalid version ID format: {version_id}")
            return Response({'error': f'Invalid version ID format: {version_id}'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Get the policy version
        try:
            policy_version = PolicyVersion.objects.get(VersionId=version_id)
            print(f"Found policy version: {policy_version.VersionId} for policy {policy_version.PolicyId_id}")
        except PolicyVersion.DoesNotExist:
            print(f"Policy version with ID {version_id} not found")
            return Response({'error': f'Policy version with ID {version_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get the policy this version belongs to
        try:
            policy = Policy.objects.get(PolicyId=policy_version.PolicyId_id)
            print(f"Found policy: {policy.PolicyName} (ID: {policy.PolicyId})")
        except Policy.DoesNotExist:
            print(f"Policy with ID {policy_version.PolicyId_id} not found")
            return Response({'error': f'Policy with ID {policy_version.PolicyId_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get subpolicies for this policy
        subpolicies = SubPolicy.objects.filter(PolicyId=policy)
        print(f"Found {len(subpolicies)} subpolicies for policy {policy.PolicyId}")
        
        subpolicies_data = []
        for subpolicy in subpolicies:
            try:
                subpolicy_data = {
                    'id': subpolicy.SubPolicyId,
                    'name': subpolicy.SubPolicyName,
                    'category': policy.Department or 'General',
                    'status': subpolicy.Status or 'Unknown',
                    'description': subpolicy.Description or '',
                    'control': subpolicy.Control or '',
                    'identifier': subpolicy.Identifier,
                    'permanent_temporary': subpolicy.PermanentTemporary,
                    'policy_id': policy.PolicyId,
                    'policy_name': policy.PolicyName,
                    'created_by': subpolicy.CreatedByName,
                    'created_date': subpolicy.CreatedByDate
                }
                subpolicies_data.append(subpolicy_data)
                print(f"Added subpolicy: {subpolicy.SubPolicyId} - {subpolicy.SubPolicyName}")
            except Exception as e:
                print(f"Error processing subpolicy {subpolicy.SubPolicyId}: {str(e)}")
                # Continue to next subpolicy
        
        print(f"Returning {len(subpolicies_data)} subpolicies")
        return Response(subpolicies_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_policy_version_subpolicies: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_dashboard_summary(request):
    total_policies = Policy.objects.count()
    total_subpolicies = SubPolicy.objects.count()
    active_policies = Policy.objects.filter(ActiveInactive='Active').count()
    inactive_policies = Policy.objects.filter(ActiveInactive='Inactive').count()
    approved_policies = PolicyApproval.objects.filter(ApprovedNot=1).count()
    total_approvals = PolicyApproval.objects.count()
    approval_rate = (approved_policies / total_approvals) * 100 if total_approvals else 0

    # Get all policies with their details
    policies = Policy.objects.all().values(
        'PolicyId', 'PolicyName', 'Department', 'Status', 
        'Applicability', 'CurrentVersion', 'ActiveInactive',
        'PermanentTemporary', 'CreatedByDate'
    )

    return Response({
        'total_policies': total_policies,
        'total_subpolicies': total_subpolicies,
        'active_policies': active_policies,
        'inactive_policies': inactive_policies,
        'approval_rate': round(approval_rate, 2),
        'policies': list(policies)
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_status_distribution(request):
    status_counts = Policy.objects.values('Status').annotate(count=Count('Status'))
    return Response(status_counts)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_reviewer_workload(request):
    reviewer_counts = Policy.objects.values('Reviewer').annotate(count=Count('Reviewer')).order_by('-count')
    return Response(reviewer_counts)

from django.utils import timezone
from datetime import timedelta

@api_view(['GET'])
@permission_classes([AllowAny])
def get_recent_policy_activity(request):
    one_week_ago = timezone.now().date() - timedelta(days=7)
    recent_policies = Policy.objects.filter(CreatedByDate__gte=one_week_ago).order_by('-CreatedByDate')[:10]
    return Response([
        {
            'PolicyName': p.PolicyName,
            'CreatedBy': p.CreatedByName,
            'CreatedDate': p.CreatedByDate
        } for p in recent_policies
    ])

from django.db.models import F, ExpressionWrapper, DurationField

@api_view(['GET'])
@permission_classes([AllowAny])
def get_avg_policy_approval_time(request):
    # Get all approved policies with approval dates
    approved_policies = PolicyApproval.objects.filter(
        ApprovedNot=1,
        ApprovedDate__isnull=False
    )
    
    if not approved_policies:
        return Response({'average_days': 0})

    # Get the first and last approval for each policy
    policy_approvals = {}
    for approval in approved_policies:
        if approval.Identifier not in policy_approvals:
            policy_approvals[approval.Identifier] = {
                'first': approval,
                'last': approval
            }
        else:
            if approval.ApprovalId < policy_approvals[approval.Identifier]['first'].ApprovalId:
                policy_approvals[approval.Identifier]['first'] = approval
            if approval.ApprovalId > policy_approvals[approval.Identifier]['last'].ApprovalId:
                policy_approvals[approval.Identifier]['last'] = approval

    # Calculate average days between first submission and approval
    total_days = 0
    count = 0
    for approvals in policy_approvals.values():
        if approvals['first'].ApprovedDate and approvals['last'].ApprovedDate:
            days = (approvals['last'].ApprovedDate - approvals['first'].ApprovedDate).days
            if days >= 0:  # Only count positive durations
                total_days += days
                count += 1

    avg_days = total_days / count if count > 0 else 0
    return Response({'average_days': round(avg_days, 2)})

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_analytics(request):
    try:
        x_axis = request.GET.get('x_axis', 'time')
        y_axis = request.GET.get('y_axis', 'count')
        framework_id = request.GET.get('framework_id')
        policy_id = request.GET.get('policy_id')
        
        # Initialize base queryset
        if x_axis == 'subpolicy':
            queryset = SubPolicy.objects.all()
            base_model = 'subpolicy'
            
            # Apply framework filter to subpolicies through policy relationship
            if framework_id and framework_id != 'all':
                queryset = queryset.filter(PolicyId__FrameworkId=framework_id)
            
            # Apply policy filter to subpolicies
            if policy_id and policy_id != 'all':
                queryset = queryset.filter(PolicyId=policy_id)
                
        elif x_axis == 'framework':
            queryset = Framework.objects.all()
            base_model = 'framework'
            
            # Apply framework filter
            if framework_id and framework_id != 'all':
                queryset = queryset.filter(FrameworkId=framework_id)
                
        else:
            queryset = Policy.objects.all()
            base_model = 'policy'
            
            # Apply framework filter to policies
            if framework_id and framework_id != 'all':
                queryset = queryset.filter(FrameworkId=framework_id)
            
            # Apply policy filter
            if policy_id and policy_id != 'all':
                queryset = queryset.filter(PolicyId=policy_id)

        # Select base queryset based on x-axis and y-axis combination
        if y_axis == 'framework_policies' and x_axis == 'time':
            # Count policies created on each date
            queryset = Policy.objects.values(
                'CreatedByDate'
            ).annotate(
                label=F('CreatedByDate'),
                value=Count('PolicyId')
            ).order_by('CreatedByDate')
            
            # Format the dates
            data = list(queryset)
            for item in data:
                item['label'] = item['label'].strftime('%Y-%m-%d') if item['label'] else None
            
            return Response(data)
        
        # Group by X-axis
        if x_axis == 'status':
            queryset = queryset.values(
                'Status'
            ).annotate(
                label=F('Status'),
            )
        elif x_axis == 'policy':
            queryset = queryset.values(
                'PolicyId', 'PolicyName'
            ).annotate(
                label=F('PolicyName'),
            )
        elif x_axis == 'subpolicy':
            queryset = queryset.values(
                'SubPolicyId', 'SubPolicyName'
            ).annotate(
                label=F('SubPolicyName'),
            )
        elif x_axis == 'time':
            date_field = {
                'framework': 'CreatedByDate',
                'policy': 'CreatedByDate',
                'subpolicy': 'CreatedByDate'
            }[base_model]
            queryset = queryset.values(
                date_field
            ).annotate(
                label=F(date_field),
            ).order_by(date_field)
        elif x_axis == 'framework':
            queryset = queryset.values(
                'FrameworkName'
            ).annotate(
                label=F('FrameworkName'),
            )
        
        # Apply Y-axis aggregation
        if y_axis == 'version':
            if base_model == 'framework':
                # Get all framework versions
                framework_versions = FrameworkVersion.objects.values(
                    'FrameworkId__Identifier'  # Group by framework identifier
                ).annotate(
                    version_count=Count('VersionId', distinct=True)  # Count distinct versions
                )
                
                # Create a mapping of framework identifier to version count
                version_counts = {fv['FrameworkId__Identifier']: fv['version_count'] for fv in framework_versions}
                
                # Get all frameworks first
                frameworks = Framework.objects.all()
                framework_id_to_identifier = {f.FrameworkId: f.Identifier for f in frameworks}
                
                # Add version count to each item in queryset
                data = list(queryset)
                for item in data:
                    if x_axis == 'framework':
                        framework_id = item['FrameworkId']
                    else:
                        # For other x_axis types (like department), we need to aggregate versions
                        # Get all frameworks matching the current group
                        if x_axis == 'department':
                            group_frameworks = frameworks.filter(Department=item['Department'])
                        elif x_axis == 'status':
                            group_frameworks = frameworks.filter(Status=item['Status'])
                        elif x_axis == 'applicability':
                            group_frameworks = frameworks.filter(Applicability=item['Applicability'])
                        else:
                            group_frameworks = frameworks
                            
                        # Sum up versions for all frameworks in this group
                        total_versions = 0
                        for framework in group_frameworks:
                            total_versions += version_counts.get(framework.Identifier, 1)
                        item['value'] = total_versions
                        continue
                    
                    # For framework x_axis, use the direct mapping
                    identifier = framework_id_to_identifier.get(framework_id)
                    item['value'] = version_counts.get(identifier, 1) if identifier else 1
                
                return Response(data)
                
            elif base_model == 'policy':
                # Count versions by following PreviousVersionId chain
                policy_versions = PolicyVersion.objects.values('PolicyId').annotate(
                    version_count=Count('VersionId', distinct=True)
                )
                version_counts = {pv['PolicyId']: pv['version_count'] for pv in policy_versions}
                
                # Add version count to each policy
                data = list(queryset)
                for item in data:
                    item['value'] = version_counts.get(item.get('PolicyId'), 0)
                return Response(data)
            else:
                # SubPolicies don't have versions
                queryset = queryset.annotate(value=Value(0))
        elif y_axis == 'activeInactive':
            if base_model == 'framework':
                # For frameworks, use ActiveInactive field
                queryset = queryset.values(
                    'ActiveInactive'
                ).annotate(
                    label=Coalesce('ActiveInactive', Value('Unknown')),
                    value=Count('FrameworkId')
                )
            elif base_model == 'policy':
                # For policies, use ActiveInactive field
                queryset = queryset.values(
                    'ActiveInactive'
                ).annotate(
                    label=Coalesce('ActiveInactive', Value('Unknown')),
                    value=Count('PolicyId')
                )
            else:
                # For subpolicies, use parent policy's ActiveInactive status
                queryset = queryset.values(
                    'PolicyId__ActiveInactive'  # Get ActiveInactive from parent policy
                ).annotate(
                    label=Coalesce('PolicyId__ActiveInactive', Value('Unknown')),
                    value=Count('SubPolicyId')
                )
        elif y_axis == 'framework_policies':
            if base_model == 'framework':
                queryset = queryset.annotate(
                    value=Count('policy')
                )
            else:
                queryset = queryset.annotate(value=Value(0))

        elif y_axis == 'createdByDate':
            # Handle CreatedByDate aggregation based on X-axis selection
            if x_axis == 'framework':
                queryset = queryset.values(
                    'CreatedByDate'
                ).annotate(
                    label=F('CreatedByDate'),
                    value=Count('FrameworkId')
                ).order_by('CreatedByDate')
            elif x_axis == 'policy':
                queryset = queryset.values(
                    'CreatedByDate'
                ).annotate(
                    label=F('CreatedByDate'),
                    value=Count('PolicyId')
                ).order_by('CreatedByDate')
            elif x_axis == 'subpolicy':
                queryset = queryset.values(
                    'CreatedByDate'
                ).annotate(
                    label=F('CreatedByDate'),
                    value=Count('SubPolicyId')
                ).order_by('CreatedByDate')

            # Format the dates for display
            data = list(queryset)
            for item in data:
                if item['label']:
                    # Convert date to string in YYYY-MM-DD format
                    item['label'] = item['label'].strftime('%Y-%m-%d')
            return Response(data)
        elif y_axis == 'department':
            # Handle Department aggregation based on X-axis selection
            if x_axis == 'framework':
                # For frameworks, get departments through policy relationship
                base_queryset = Framework.objects.values(
                    'FrameworkId', 'FrameworkName'
                ).annotate(
                    policy_count=Count('policy')
                ).filter(policy_count__gt=0)

                # Get all policies for these frameworks
                framework_policies = Policy.objects.filter(
                    FrameworkId__in=[f['FrameworkId'] for f in base_queryset]
                ).values('FrameworkId', 'Department')

                # Process departments and count frameworks
                department_counts = {}
                framework_departments = {}  # Track which departments each framework has been counted for

                for policy in framework_policies:
                    framework_id = policy['FrameworkId']
                    dept_str = policy['Department']
                    
                    if not dept_str:
                        continue

                    # Initialize set for this framework if not exists
                    if framework_id not in framework_departments:
                        framework_departments[framework_id] = set()

                    # Split departments and process each
                    departments = [d.strip().upper() for d in dept_str.split(',')]
                    for dept in departments:
                        if dept and dept not in framework_departments[framework_id]:
                            # Count framework for this department only once
                            department_counts[dept] = department_counts.get(dept, 0) + 1
                            framework_departments[framework_id].add(dept)

            elif x_axis == 'policy':
                # For policies, use Department field directly
                base_queryset = queryset.values(
                    'PolicyId',
                    'Department'
                )
            elif x_axis == 'subpolicy':
                # For subpolicies, get department through policy relationship
                base_queryset = queryset.values(
                    'SubPolicyId',
                    'PolicyId__Department'
                )

            if x_axis != 'framework':
                # Process departments and split comma-separated values
                department_counts = {}
                for item in base_queryset:
                    # Get the department field based on the model type
                    dept_field = (
                        item.get('Department') or 
                        item.get('policy__Department') or 
                        item.get('PolicyId__Department')
                    )
                    
                    if dept_field:
                        # Split departments by comma and strip whitespace
                        departments = [d.strip().upper() for d in dept_field.split(',')]
                        for dept in departments:
                            if dept:  # Only count non-empty departments
                                department_counts[dept] = department_counts.get(dept, 0) + 1

            # Convert to list format expected by frontend
            data = [
                {
                    'label': f"{dept.title()} ({count} frameworks)" if x_axis == 'framework' else f"{dept.title()} ({count} items)",
                    'value': count
                }
                for dept, count in department_counts.items()
            ]

            # Sort by count (descending) then by department name
            data.sort(key=lambda x: (-x['value'], x['label']))

            # Add unassigned if no departments found
            if not data:
                label = 'Unassigned (0 frameworks)' if x_axis == 'framework' else 'Unassigned (0 items)'
                data.append({
                    'label': label,
                    'value': 0
                })

            return Response(data)
        elif y_axis == 'createdByName':
            # Handle CreatedByName aggregation based on X-axis selection
            if x_axis == 'framework':
                queryset = queryset.values(
                    'CreatedByName'
                ).annotate(
                    label=F('CreatedByName'),  # Use the actual CreatedByName value
                    value=Count('FrameworkId', distinct=True)  # Count unique frameworks
                ).order_by('-value', 'CreatedByName')  # Order by count desc, then name
            elif x_axis == 'policy':
                queryset = queryset.values(
                    'CreatedByName'
                ).annotate(
                    label=F('CreatedByName'),  # Use the actual CreatedByName value
                    value=Count('PolicyId', distinct=True)  # Count unique policies
                ).order_by('-value', 'CreatedByName')  # Order by count desc, then name
            elif x_axis == 'subpolicy':
                queryset = queryset.values(
                    'CreatedByName'
                ).annotate(
                    label=F('CreatedByName'),  # Use the actual CreatedByName value
                    value=Count('SubPolicyId', distinct=True)  # Count unique subpolicies
                ).order_by('-value', 'CreatedByName')  # Order by count desc, then name

            # Add creator label for clarity
            data = list(queryset)
            for item in data:
                item['label'] = f"{item['label']} ({item['value']} items)"
            return Response(data)
        elif y_axis == 'status':
            if base_model == 'framework':
                # For frameworks, directly use the Status field
                queryset = queryset.values(
                    'Status'
                ).annotate(
                    label=Coalesce('Status', Value('Unknown')),
                    value=Count('FrameworkId')
                )
            elif base_model == 'policy':
                # For policies, use their Status field
                queryset = queryset.values(
                    'Status'
                ).annotate(
                    label=Coalesce('Status', Value('Unknown')),
                    value=Count('PolicyId')
                )
            else:
                # For subpolicies, use their Status field
                queryset = queryset.values(
                    'Status'
                ).annotate(
                    label=Coalesce('Status', Value('Unknown')),
                    value=Count('SubPolicyId')
                )
        elif y_axis == 'category':
            if base_model == 'framework':
                # For frameworks, directly use the Category field
                queryset = queryset.values(
                    'Category'
                ).annotate(
                    label=Coalesce('Category', Value('Uncategorized')),
                    value=Count('FrameworkId')
                )
            elif base_model == 'policy':
                # For policies, get category through framework relationship
                queryset = queryset.values(
                    'FrameworkId__Category'
                ).annotate(
                    label=Coalesce('FrameworkId__Category', Value('Uncategorized')),
                    value=Count('PolicyId')
                )
            else:
                # For subpolicies, get category through policy->framework relationship
                queryset = queryset.values(
                    'PolicyId__FrameworkId__Category'
                ).annotate(
                    label=Coalesce('PolicyId__FrameworkId__Category', Value('Uncategorized')),
                    value=Count('SubPolicyId')
                )
        
        data = list(queryset)
        
        # Format dates for time-based analysis
        if x_axis == 'time':
            for item in data:
                date_value = item.get(date_field)
                item['label'] = date_value.strftime('%Y-%m-%d') if date_value else None
        
        return Response(data)
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=500
        )

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_kpis(request):
    try:
        # Get total policies count
        total_policies = Policy.objects.count()
        
        # Get active policies count
        active_policies = Policy.objects.filter(
            ActiveInactive='Active'
        ).count()

        # Get total users count for acknowledgement rate calculation
        total_users = Users.objects.count()

        # Get top 5 policies by acknowledgement rate
        policies = Policy.objects.filter(ActiveInactive='Active').annotate(
            acknowledgement_rate=Case(
                When(AcknowledgementCount__gt=0, 
                     then=ExpressionWrapper(
                         F('AcknowledgementCount') * 100.0 / total_users,
                         output_field=FloatField()
                     )),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ).order_by('-acknowledgement_rate')[:5]

        top_acknowledged_policies = [
            {
                'policy_id': policy.PolicyId,
                'policy_name': policy.PolicyName,
                'acknowledged_count': policy.AcknowledgementCount,
                'total_users': total_users,
                'acknowledgement_rate': round(float(policy.acknowledgement_rate), 1)
            }
            for policy in policies
        ]

        # Get historical active policy counts for the last 12 months
        twelve_months_ago = datetime.date.today() - timedelta(days=365)
        monthly_counts = []
        
        # Get all policies with their creation dates
        policies = Policy.objects.filter(
            CreatedByDate__gte=twelve_months_ago
        ).values('CreatedByDate', 'ActiveInactive')
        
        # Group by month and count active policies
        month_data = {}
        current_date = datetime.date.today()
        
        # Initialize all months with 0
        for i in range(12):
            month_date = current_date - timedelta(days=30 * i)
            month_key = month_date.strftime('%Y-%m')
            month_data[month_key] = 0
        
        # Count active policies for each month
        for policy in policies:
            month_key = policy['CreatedByDate'].strftime('%Y-%m')
            if month_key in month_data and policy['ActiveInactive'] == 'Active':
                month_data[month_key] += 1
        
        # Convert to sorted list for last 12 months
        monthly_counts = [
            {
                'month': k,
                'count': v
            }
            for k, v in sorted(month_data.items(), reverse=True)
        ][:12]
        
        # Calculate revision metrics
        three_months_ago = datetime.date.today() - timedelta(days=90)
        
        # Get all policy versions with previous version links
        policy_versions = PolicyVersion.objects.filter(
            PreviousVersionId__isnull=False  # Has a previous version
        ).select_related('PolicyId')
        
        # Track revised policies and their revision counts
        revision_counts = {}  # Dictionary to track revisions per PreviousVersionId
        revised_policies_set = set()
        total_revisions = 0
        
        # Process each version that has a previous version
        for version in policy_versions:
            policy_id = version.PolicyId.PolicyId
            prev_version_id = version.PreviousVersionId
            
            # Count this as a revision
            revised_policies_set.add(policy_id)
            total_revisions += 1
            
            # Track revisions per previous version
            if prev_version_id not in revision_counts:
                revision_counts[prev_version_id] = 1
            else:
                revision_counts[prev_version_id] += 1
        
        # Calculate policies with multiple revisions
        multiple_revisions_count = sum(1 for count in revision_counts.values() if count > 1)
        
        # Calculate final metrics
        revised_policies = len(revised_policies_set)
        
        # Calculate revision rate using total policies as denominator
        revision_rate = 0
        if total_policies > 0:  # Avoid division by zero
            revision_rate = (revised_policies / total_policies) * 100
            revision_rate = min(revision_rate, 100)  # Cap at 100%

        # Calculate policy coverage by department
        departments = Policy.objects.values_list('Department', flat=True).distinct()
        department_coverage = []
        
        for dept in departments:
            if dept:  # Skip empty department values
                dept_policies = Policy.objects.filter(Department=dept)
                total_dept_policies = dept_policies.count()
                if total_dept_policies > 0:
                    avg_coverage = dept_policies.aggregate(
                        avg_coverage=Coalesce(Avg('CoverageRate'), 0.0)
                    )['avg_coverage']
                    
                    department_coverage.append({
                        'department': dept,
                        'coverage_rate': round(float(avg_coverage), 2),
                        'total_policies': total_dept_policies
                    })
        
        # Sort departments by coverage rate in descending order
        department_coverage.sort(key=lambda x: x['coverage_rate'], reverse=True)
        
        # Calculate overall average coverage rate
        overall_coverage = Policy.objects.aggregate(
            avg_coverage=Coalesce(Avg('CoverageRate'), 0.0)
        )['avg_coverage']
        
        return Response({
            'total_policies': total_policies,
            'active_policies': active_policies,
            'active_policies_trend': monthly_counts,  # Add historical trend data
            'revision_rate': round(revision_rate, 2),
            'revised_policies': revised_policies,
            'total_revisions': total_revisions,
            'policies_with_multiple_revisions': multiple_revisions_count,
            'measurement_period': '3 months',
            'coverage_metrics': {
                'overall_coverage_rate': round(float(overall_coverage), 2),
                'department_coverage': department_coverage
            },
            'top_acknowledged_policies': top_acknowledged_policies
        })
    except Exception as e:
        print(f"Error in get_policy_kpis: {str(e)}")
        return Response({
            'error': 'Error fetching policy KPIs',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
@permission_classes([AllowAny])
def acknowledge_policy(request, policy_id):
    try:
        # Extract user_id from JWT token or request.user
        user_id = None
        
        # Try to get user_id from JWT token first
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
            try:
                from .authentication import verify_jwt_token
                payload = verify_jwt_token(token)
                if payload and 'user_id' in payload:
                    user_id = payload['user_id']
                    print(f"DEBUG: acknowledge_policy (views.py) - Extracted user_id from JWT: {user_id}")
            except Exception as e:
                print(f"DEBUG: acknowledge_policy (views.py) - Error extracting user_id from JWT: {e}")
        
        # Fallback to request.user if JWT extraction failed
        if not user_id:
            user_id = getattr(request.user, 'UserId', None)
            if not user_id:
                user_id = getattr(request.user, 'id', None)
        
        # Fallback to session if still no user_id
        if not user_id:
            user_id = request.session.get('user_id')
        
        if not user_id:
            return Response({
                'error': 'User not authenticated. Please login again.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        policy = Policy.objects.get(PolicyId=policy_id)
        
        # Initialize acknowledged_users list if None
        # Handle case where AcknowledgedUserIds might be a string or None
        acknowledged_users = policy.AcknowledgedUserIds if policy.AcknowledgedUserIds is not None else []
        
        # Ensure acknowledged_users is a list (not a string)
        if isinstance(acknowledged_users, str):
            try:
                import json
                acknowledged_users = json.loads(acknowledged_users)
            except (json.JSONDecodeError, ValueError):
                acknowledged_users = []
        
        # Ensure it's a list
        if not isinstance(acknowledged_users, list):
            acknowledged_users = []
        
        # Ensure user_id is an integer for consistent comparison
        user_id = int(user_id)
        
        # Convert all items in acknowledged_users to integers for consistent comparison
        acknowledged_users = [int(uid) for uid in acknowledged_users if uid is not None]
        
        # Check if user already acknowledged
        if user_id not in acknowledged_users:
            # Add user to acknowledged list
            acknowledged_users.append(user_id)
            policy.AcknowledgedUserIds = acknowledged_users
            policy.AcknowledgementCount = len(acknowledged_users)
            policy.save()
            
            return Response({
                'message': 'Policy acknowledged successfully',
                'acknowledged_users': policy.AcknowledgedUserIds,
                'acknowledgement_count': policy.AcknowledgementCount
            })
        else:
            return Response({
                'message': 'Policy already acknowledged by this user',
                'acknowledged_users': policy.AcknowledgedUserIds,
                'acknowledgement_count': policy.AcknowledgementCount
            })

    except Policy.DoesNotExist:
        return Response({
            'error': 'Policy not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(f"Error in acknowledge_policy: {str(e)}")  # Add logging
        return Response({
            'error': 'Error acknowledging policy',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def login_user(request):
    """Session-based login endpoint with rate limiting and account lockout"""
    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        login_type = data.get('login_type', 'username')  # Default to username if not specified
        
        logger.debug(f"Login attempt for {login_type}: {username}")
        
        if not username or not password:
            return Response({
                'status': 'error',
                'message': 'Username and password are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ========================================
        # RATE LIMITING - PER IP
        # ========================================
        client_ip = _get_client_ip(request)  # Use sanitized IP function
        ip_cache_key = f"session_login_rate_limit_ip_{client_ip}"
        ip_attempts = cache.get(ip_cache_key, 0)
        
        if ip_attempts >= 10:  # Max 10 login attempts per minute per IP
            return Response({
                'status': 'error',
                'message': 'Too many login attempts from this IP. Please wait 1 minute and try again.'
            }, status=status.HTTP_429_TOO_MANY_REQUESTS)
        
        # Increment IP counter
        cache.set(ip_cache_key, ip_attempts + 1, 60)  # 60 seconds
        
        # ========================================
        # RATE LIMITING - PER USERNAME (LOCKOUT)
        # ========================================
        # Normalize username for cache key
        username_normalized = str(username).lower().strip()
        user_cache_key = f"session_login_failed_attempts_{username_normalized}"
        lockout_cache_key = f"session_login_locked_until_{username_normalized}"
        
        # Check if account is locked
        locked_until = cache.get(lockout_cache_key)
        if locked_until:
            remaining_seconds = int(locked_until - time.time())
            if remaining_seconds > 0:
                remaining_minutes = remaining_seconds // 60
                return Response({
                    'status': 'error',
                    'message': f'Account temporarily locked due to too many failed login attempts. Please try again in {remaining_minutes + 1} minute(s).',
                    'locked_until': remaining_seconds
                }, status=status.HTTP_403_FORBIDDEN)
            else:
                # Lock expired, clear it
                cache.delete(lockout_cache_key)
                cache.delete(user_cache_key)
        
        # Clear any existing session data first
        request.session.flush()
        
        # Check if user exists in database based on login type (with hashed password support)
        user = None
        try:
            if login_type == 'userid':
                # Login with User ID
                user_id = int(username)  # Convert to integer
                candidate = Users.objects.get(UserId=user_id)
                logger.debug(f"User found by ID: {candidate.UserId} - {candidate.UserName}")
            else:
                # Login with Username (default) - handles encrypted usernames
                candidate = Users.find_by_username(username)
                if not candidate:
                    raise Users.DoesNotExist(f"User with username '{username}' not found")
                logger.debug(f"User found by username: {candidate.UserId} - {candidate.UserName}")

            # First try hashed password verification
            if check_password(password, candidate.Password):
                user = candidate
            # Backward compatibility: if stored password is plain text, migrate it
            elif candidate.Password == password:
                candidate.Password = make_password(password)
                candidate.save(update_fields=['Password'])
                user = candidate
                logger.warning(f"Password for user {candidate.UserName} was stored in plain text and has been hashed.")
            
            if not user:
                # Password didn't match - treat as failed login
                raise Users.DoesNotExist
            
            # Check if user is active (handle both boolean and string values)
            # NOTE: New users are created with IsActive='N' and must reset password and login to activate
            is_active = user.IsActive
            if isinstance(is_active, str):
                is_active = is_active.upper() == 'Y'
            elif isinstance(is_active, bool):
                is_active = is_active
            else:
                is_active = False  # Default to inactive if unknown type
            
            # Store whether user was inactive before login (for activation after password verification)
            user_was_inactive = not is_active
            
            # Allow inactive users to proceed - they will be activated after successful password verification
            # This allows new users to login after resetting their password
            
            # ========================================
            # LICENSE KEY VALIDATION PROCESS
            # ========================================
            from django.conf import settings as _dj_settings
            if not getattr(_dj_settings, 'LICENSE_CHECK_ENABLED', True):
                logger.warning("🔕 LICENSE CHECK DISABLED via settings. Proceeding without external verification.")
            else:
                # Step 1: Check if user has a license key assigned
                license_verification_result = None
                if user.license_key:
                    logger.info(f"🔐 LICENSE VALIDATION: User {user.UserName} has license key: {user.license_key[:10]}...")
                    try:
                        # Step 2: Import and initialize the licensing system
                        from licensing_system import VardaanLicensingSystem
                        licensing_system = VardaanLicensingSystem()
                        logger.info(f"🔐 LICENSE VALIDATION: Licensing system initialized for user {user.UserName}")
                        # Step 3: Call external API to verify the license key
                        logger.info(f"🔐 LICENSE VALIDATION: Calling external API to verify license for user {user.UserName}")
                        license_verification_result = licensing_system.verify_license(user.license_key)
                        # Step 4: Check if license verification was successful
                        if not license_verification_result.get("success"):
                            logger.warning(f"❌ LICENSE VALIDATION FAILED: User {user.UserName} - {license_verification_result.get('error')}")
                            # Log failed login attempt due to license verification failure
                            try:
                                from .models import GRCLog
                                framework = _get_default_framework()
                                if framework:
                                    log_entry = GRCLog(
                                        Module='Authentication',
                                        ActionType='LOGIN_FAILED',
                                        Description=f'Failed login attempt for user {user.UserName} (ID: {user.UserId}). Reason: License verification failed',
                                        UserId=str(user.UserId),
                                        UserName=user.UserName,
                                        LogLevel='WARNING',
                                        IPAddress=client_ip,
                                        FrameworkId=framework,
                                        AdditionalInfo={
                                            'login_type': login_type,
                                            'reason': 'License verification failed',
                                            'license_error': license_verification_result.get('error', 'Unknown license error')
                                        }
                                    )
                                    log_entry.save()
                                    logger.debug(f"Logged failed login (license verification) for {user.UserName} to grc_logs")
                            except Exception as log_error:
                                logger.error(f"Error logging failed login attempt to grc_logs: {str(log_error)}")
                            return Response({
                                'status': 'error',
                                'message': 'License verification failed. Please contact your administrator.',
                                'license_error': license_verification_result.get('error', 'Unknown license error')
                            }, status=status.HTTP_403_FORBIDDEN)
                        else:
                            logger.info(f"✅ LICENSE VALIDATION SUCCESS: User {user.UserName} license verified successfully")
                    except Exception as license_error:
                        logger.error(f"❌ LICENSE VALIDATION ERROR: User {user.UserName} - {str(license_error)}")
                        # Log failed login attempt due to license verification error
                        try:
                            from .models import GRCLog
                            framework = _get_default_framework()
                            if framework:
                                log_entry = GRCLog(
                                    Module='Authentication',
                                    ActionType='LOGIN_FAILED',
                                    Description=f'Failed login attempt for user {user.UserName} (ID: {user.UserId}). Reason: License verification error',
                                    UserId=str(user.UserId),
                                    UserName=user.UserName,
                                    LogLevel='ERROR',
                                    IPAddress=client_ip,
                                    FrameworkId=framework,
                                    AdditionalInfo={
                                        'login_type': login_type,
                                        'reason': 'License verification error',
                                        'license_error': str(license_error)
                                    }
                                )
                                log_entry.save()
                                logger.debug(f"Logged failed login (license error) for {user.UserName} to grc_logs")
                        except Exception as log_error:
                            logger.error(f"Error logging failed login attempt to grc_logs: {str(log_error)}")
                        return Response({
                            'status': 'error',
                            'message': 'License verification error. Please contact your administrator.',
                            'license_error': str(license_error)
                        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                else:
                    # Step 5: Handle case where user has no license key
                    logger.warning(f"❌ LICENSE VALIDATION: User {user.UserName} has no license key assigned")
                    # Log failed login attempt due to missing license key
                    try:
                        from .models import GRCLog
                        framework = _get_default_framework()
                        if framework:
                            log_entry = GRCLog(
                                Module='Authentication',
                                ActionType='LOGIN_FAILED',
                                Description=f'Failed login attempt for user {user.UserName} (ID: {user.UserId}). Reason: No license key assigned',
                                UserId=str(user.UserId),
                                UserName=user.UserName,
                                LogLevel='WARNING',
                                IPAddress=client_ip,
                                FrameworkId=framework,
                                AdditionalInfo={
                                    'login_type': login_type,
                                    'reason': 'No license key assigned'
                                }
                            )
                            log_entry.save()
                            logger.debug(f"Logged failed login (no license key) for {user.UserName} to grc_logs")
                    except Exception as log_error:
                        logger.error(f"Error logging failed login attempt to grc_logs: {str(log_error)}")
                    return Response({
                        'status': 'error',
                        'message': 'No license key assigned to this user. Please contact your administrator.'
                    }, status=status.HTTP_403_FORBIDDEN)
            
            # ========================================
            # PASSWORD EXPIRY CHECK
            # ========================================
            from .routes.Global.password_expiry_utils import (
                is_password_expired,
                is_password_expiring_soon,
                send_password_expiry_email,
                log_password_action
            )
            
            password_expired, days_until_expiry, days_since_change = is_password_expired(user)
            password_expiring_soon, _ = is_password_expiring_soon(user)
            
            # If password is expired, block login and force password reset
            if password_expired:
                # Send email notification about expired password
                send_password_expiry_email(user, is_expired=True, days_until_expiry=days_until_expiry)
                
                return Response({
                    'status': 'error',
                    'message': 'Your password has expired. Please reset your password using the "Forgot Password" option on the login page.',
                    'password_expired': True,
                    'days_since_expiry': abs(days_until_expiry)
                }, status=status.HTTP_403_FORBIDDEN)
            
            # If password is expiring soon, send warning email (but allow login)
            if password_expiring_soon:
                # Send warning email (only once per day to avoid spam)
                warning_cache_key = f"password_expiry_warning_sent_{user.UserId}_{timezone.now().date()}"
                if not cache.get(warning_cache_key):
                    send_password_expiry_email(user, is_expired=False, days_until_expiry=days_until_expiry)
                    cache.set(warning_cache_key, True, 86400)  # Cache for 24 hours
            
            # Note: Password logs are only saved when password is changed, not on every login
            # Login activities are logged to grc_logs instead
            
            # ========================================
            # SUCCESSFUL LOGIN - CLEAR FAILED ATTEMPT COUNTERS
            # ========================================
            cache.delete(user_cache_key)
            cache.delete(lockout_cache_key)
            
            # Update last login time and activate user on successful login
            fields_to_update = ['last_login']
            user.last_login = timezone.now()
            
            # Activate user on successful login (set IsActive to 'Y')
            # This happens after password verification, so new users can login after resetting password
            if user.IsActive != 'Y':
                user.IsActive = 'Y'
                fields_to_update.append('IsActive')
                logger.info(f"✅ User {user.UserName} (ID: {user.UserId}) activated on successful login (was inactive)")
                print(f"[DEBUG] ✅ User {user.UserName} (ID: {user.UserId}) activated on successful login")
            
            user.save(update_fields=fields_to_update)
            logger.info(f"✅ User {user.UserName} (ID: {user.UserId}) last login updated: {user.last_login}")
            
            # Password logging is now handled in the password expiry check above
            
            # Set session data
            import time
            request.session['user_id'] = user.UserId
            request.session['grc_user_id'] = user.UserId  # Backup key
            request.session['grc_username'] = user.UserName
            request.session['session_created_at'] = time.time()  # Store session creation time for timeout check
            request.session.save()  # Explicitly save session
            
            # ========================================
            # LOGIN SUCCESS RESPONSE WITH LICENSE VERIFICATION
            # ========================================
            # Step 6: Create success response indicating license was validated
            response_data = {
                'status': 'success',
                'message': 'Login successful',
                'license_verified': True,  # This indicates license validation was successful
                'user': {
                    'id': user.UserId,
                    'UserId': user.UserId,
                    'username': user.UserName,
                    'Email': user.email_plain,  # Use decrypted email
                    'firstName': user.FirstName,
                    'lastName': user.LastName,
                    'license_key': user.license_key  # Include the validated license key
                }
            }
            
            logger.info(f"✅ LOGIN SUCCESS: User {user.UserName} (ID: {user.UserId}) logged in successfully with license verification")
            
            # Log successful login to grc_logs - DIRECT DATABASE SAVE
            log_saved = False
            try:
                logger.info(f"🔍 Attempting to log successful login for user {user.UserName} (ID: {user.UserId})")
                from .models import GRCLog  # Import locally to avoid scoping issues
                framework = _get_default_framework()
                if framework:
                    client_ip = _get_client_ip(request)
                    log_entry = GRCLog(
                        Module='Authentication',
                        ActionType='LOGIN_SUCCESS',
                        Description=f'User {user.UserName} (ID: {user.UserId}) logged in successfully',
                        UserId=str(user.UserId),
                        UserName=user.UserName,
                        LogLevel='INFO',
                        IPAddress=client_ip,
                        FrameworkId=framework,
                        AdditionalInfo={
                            'login_type': login_type,
                            'license_verified': True,
                            'license_key': user.license_key[:10] + '...' if user.license_key else None,
                            'auth_method': 'SESSION'
                        }
                    )
                    log_entry.save()
                    logger.info(f"✅ Successfully logged login to grc_logs with ID: {log_entry.LogId} for user {user.UserName} (ID: {user.UserId})")
                    log_saved = True
                else:
                    logger.warning("⚠️  Cannot log successful login: No framework available")
            except Exception as log_error:
                logger.error(f"❌ Error logging successful login to grc_logs: {str(log_error)}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
            
            # If direct save failed, try one more time with minimal data
            if not log_saved:
                try:
                    logger.info(f"🔄 Retrying login log save with minimal data")
                    from .models import GRCLog  # Import locally to avoid scoping issues
                    framework = _get_default_framework()
                    if framework:
                        client_ip = _get_client_ip(request)
                        log_entry = GRCLog(
                            Module='Authentication',
                            ActionType='LOGIN_SUCCESS',
                            Description=f'User {user.UserName} logged in',
                            UserId=str(user.UserId),
                            UserName=user.UserName,
                            LogLevel='INFO',
                            IPAddress=client_ip or 'unknown',
                            FrameworkId=framework
                        )
                        log_entry.save()
                        logger.info(f"✅ RETRY SUCCESS: Logged login with ID: {log_entry.LogId}")
                    else:
                        logger.error(f"❌ CRITICAL: Cannot save login log - no framework available")
                except Exception as retry_error:
                    logger.error(f"❌ CRITICAL: Retry also failed: {str(retry_error)}")
            
            return Response(response_data)
            
        except Users.DoesNotExist:
            # ========================================
            # FAILED LOGIN - INCREMENT COUNTER & CHECK LOCKOUT
            # ========================================
            failed_attempts = cache.get(user_cache_key, 0) + 1
            cache.set(user_cache_key, failed_attempts, 900)  # Keep counter for 15 minutes
            
            # Log failed login attempt to grc_logs
            try:
                from .models import GRCLog
                framework = _get_default_framework()
                if framework:
                    log_entry = GRCLog(
                        Module='Authentication',
                        ActionType='LOGIN_FAILED',
                        Description=f'Failed login attempt for {login_type}: {username}. Reason: Invalid credentials',
                        UserName=username,
                        LogLevel='WARNING',
                        IPAddress=client_ip,
                        FrameworkId=framework,
                        AdditionalInfo={
                            'login_type': login_type,
                            'failed_attempts': failed_attempts,
                            'attempt_number': failed_attempts
                        }
                    )
                    log_entry.save()
                    logger.debug(f"Logged failed login attempt for {username} to grc_logs")
            except Exception as log_error:
                logger.error(f"Error logging failed login attempt to grc_logs: {str(log_error)}")
                # Don't fail login if logging fails
            
            if failed_attempts >= 5:
                # Lock account for 15 minutes
                lockout_time = time.time() + 900  # 15 minutes from now
                cache.set(lockout_cache_key, lockout_time, 900)
                cache.delete(user_cache_key)  # Clear attempt counter
                
                logger.warning(f"Login failed - account locked for {username} after {failed_attempts} attempts")
                return Response({
                    'status': 'error',
                    'message': f'Too many failed login attempts. Account locked for 15 minutes. (Attempt {failed_attempts}/5)'
                }, status=status.HTTP_403_FORBIDDEN)
            
            logger.warning(f"Login failed - invalid credentials for {login_type}: {username} (Attempt {failed_attempts}/5)")
            return Response({
                'status': 'error',
                'message': f'Invalid {login_type} or password. ({failed_attempts}/5 attempts)'
            }, status=status.HTTP_401_UNAUTHORIZED)
        except ValueError:
            # Log failed login attempt due to invalid format
            try:
                from .models import GRCLog
                framework = _get_default_framework()
                if framework:
                    log_entry = GRCLog(
                        Module='Authentication',
                        ActionType='LOGIN_FAILED',
                        Description=f'Failed login attempt - invalid user ID format: {username}',
                        UserName=username if username else 'Unknown',
                        LogLevel='WARNING',
                        IPAddress=client_ip,
                        FrameworkId=framework,
                        AdditionalInfo={
                            'login_type': login_type,
                            'reason': 'Invalid user ID format'
                        }
                    )
                    log_entry.save()
                    logger.debug(f"Logged failed login attempt (invalid format) for {username} to grc_logs")
            except Exception as log_error:
                logger.error(f"Error logging failed login attempt to grc_logs: {str(log_error)}")
            
            logger.warning(f"Login failed - invalid user ID format: {username}")
            return Response({
                'status': 'error',
                'message': 'Invalid user ID format. Please enter a valid number.'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        return Response({
            'status': 'error',
            'message': 'Server error during login',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def logout_user(request):
    print("=" * 80)
    print("🚪🚪🚪 SESSION LOGOUT FUNCTION CALLED 🚪🚪🚪")
    print("=" * 80)
    try:
        print("[DEBUG] Session logout: Starting logout process...")
        # Get user info before clearing session - CHECK ALL POSSIBLE SESSION KEYS
        logger.info(f"Session keys available: {list(request.session.keys())}")
        print(f"[DEBUG] Session keys available: {list(request.session.keys())}")
        
        # Try multiple session keys to find user_id
        user_id = (request.session.get('user_id') or 
                  request.session.get('grc_user_id') or
                  None)
        
        # Try to get username from multiple session keys
        username = (request.session.get('grc_username') or 
                   request.session.get('username') or 
                   'Unknown')
        
        client_ip = request.META.get('REMOTE_ADDR', 'unknown')
        framework_id = None
        
        logger.info(f"Initial session data - user_id: {user_id}, username: {username}, IP: {client_ip}")
        
        # Try to get framework_id from user if available
        if user_id:
            try:
                from .models import Users
                # Handle case where user_id might be string
                if isinstance(user_id, str) and user_id.isdigit():
                    user_id = int(user_id)
                user = Users.objects.get(UserId=user_id)
                username = user.UserName
                framework_id = user.FrameworkId.FrameworkId if user.FrameworkId else None
                logger.info(f"✅ Got user from database: {username} (ID: {user_id}), framework_id: {framework_id}")
            except (Users.DoesNotExist, ValueError, TypeError) as e:
                logger.warning(f"Could not get user from database: {str(e)}")
                pass
        
        # Invalidate session token for multi-session management
        if user_id:
            try:
                from .authentication import _invalidate_user_session
                _invalidate_user_session(user_id)
                logger.info(f"🔐 Session token invalidated for user {user_id} on session logout")
            except Exception as session_error:
                logger.warning(f"Error invalidating session token: {str(session_error)}")
        
        # Log logout to grc_logs before clearing session - ALWAYS LOG, even if user_id is None
        logger.info("=" * 80)
        logger.info("🚪 SESSION LOGOUT CALLED")
        logger.info(f"User info - user_id: {user_id}, username: {username}, IP: {client_ip}")
        logger.info("=" * 80)
        print("=" * 80)
        print("🚪 SESSION LOGOUT CALLED")
        print(f"User info - user_id: {user_id}, username: {username}, IP: {client_ip}")
        print("=" * 80)
        
        log_saved = False
        print("[DEBUG] About to enter logging block...")
        # ALWAYS try to log, even if user_id is None
        if True:  # Changed from "if user_id:" to always log
            print("[DEBUG] Inside logging block - attempting send_log...")
            try:
                from .routes.Global.logging_service import send_log
                logger.info(f"🔍 Attempting to log logout for user {username} (ID: {user_id})")
                print(f"[DEBUG] 🔍 Attempting to log logout for user {username} (ID: {user_id})")
                log_id = send_log(
                    module='Authentication',
                    actionType='LOGOUT',
                    description=f'User {username} (ID: {user_id or "Unknown"}) logged out successfully',
                    userId=str(user_id) if user_id else None,
                    userName=username if username != 'Unknown' else None,
                    logLevel='INFO',
                    ipAddress=client_ip,
                    additionalInfo={'auth_method': 'session', 'user_id_found': user_id is not None},
                    frameworkId=framework_id
                )
                print(f"[DEBUG] send_log returned: {log_id}")
                if log_id:
                    logger.info(f"✅ Successfully logged logout to grc_logs with ID: {log_id}")
                    print(f"[DEBUG] ✅ Successfully logged logout to grc_logs with ID: {log_id}")
                    log_saved = True
                else:
                    logger.warning(f"⚠️  send_log returned None for logout - trying direct database save")
                    print(f"[DEBUG] ⚠️  send_log returned None for logout - trying direct database save")
            except Exception as log_error:
                logger.error(f"❌ Error in send_log for logout: {str(log_error)}")
                print(f"[DEBUG] ❌ Error in send_log for logout: {str(log_error)}")
                import traceback
                error_trace = traceback.format_exc()
                logger.error(f"Traceback: {error_trace}")
                print(f"[DEBUG] Traceback: {error_trace}")
            
            # FALLBACK: Direct database save if send_log failed
            print(f"[DEBUG] log_saved status: {log_saved}")
            if not log_saved:
                print("[DEBUG] Entering direct database save fallback...")
                try:
                    logger.info(f"🔄 Attempting direct database save for logout log")
                    print(f"[DEBUG] 🔄 Attempting direct database save for logout log")
                    framework = _get_default_framework()
                    print(f"[DEBUG] Framework retrieved: {framework}")
                    if framework:
                        print(f"[DEBUG] Framework found: ID={framework.FrameworkId}, Name={framework.FrameworkName}")
                        log_entry = GRCLog(
                            Module='Authentication',
                            ActionType='LOGOUT',
                            Description=f'User {username} (ID: {user_id or "Unknown"}) logged out successfully',
                            UserId=str(user_id) if user_id else None,
                            UserName=username if username != 'Unknown' else None,
                            LogLevel='INFO',
                            IPAddress=client_ip,
                            FrameworkId=framework,
                            AdditionalInfo={
                                'auth_method': 'session',
                                'logged_via': 'direct_database_save',
                                'user_id_found': user_id is not None
                            }
                        )
                        print(f"[DEBUG] Creating GRCLog entry with:")
                        print(f"  - Module: Authentication")
                        print(f"  - ActionType: LOGOUT")
                        print(f"  - UserId: {user_id} (type: {type(user_id)})")
                        print(f"  - UserName: {username}")
                        print(f"  - FrameworkId: {framework.FrameworkId}")
                        print(f"  - IPAddress: {client_ip}")
                        
                        log_entry.save()
                        print(f"[DEBUG] ✅ GRCLog.save() called successfully, LogId: {log_entry.LogId}")
                        
                        # Verify the log was saved with user_id
                        try:
                            saved_log = GRCLog.objects.get(LogId=log_entry.LogId)
                            logger.info(f"✅ DIRECT SAVE SUCCESS: Logged logout to grc_logs with ID: {log_entry.LogId}")
                            logger.info(f"✅ VERIFIED: Saved log has UserId={saved_log.UserId}, UserName={saved_log.UserName}")
                            print(f"[DEBUG] ✅ VERIFIED: Saved log has UserId={saved_log.UserId}, UserName={saved_log.UserName}")
                            log_saved = True
                            print(f"[LOGOUT LOG] ✅ Saved logout log with ID: {log_entry.LogId} for user {username} (ID: {user_id})")
                        except Exception as verify_error:
                            print(f"[DEBUG] ❌ Verification failed: {str(verify_error)}")
                            logger.error(f"❌ Verification failed: {str(verify_error)}")
                    else:
                        logger.error(f"❌ Cannot save logout log: No framework available")
                        print(f"[DEBUG] ❌ Cannot save logout log: No framework available")
                except Exception as direct_save_error:
                    logger.error(f"❌ CRITICAL: Direct database save for logout also failed: {str(direct_save_error)}")
                    print(f"[DEBUG] ❌ CRITICAL: Direct database save for logout also failed: {str(direct_save_error)}")
                    import traceback
                    error_trace = traceback.format_exc()
                    logger.error(f"Traceback: {error_trace}")
                    print(f"[DEBUG] Traceback: {error_trace}")
                    print(f"[LOGOUT LOG ERROR] {str(direct_save_error)}")
        
        print(f"[DEBUG] Final log_saved status: {log_saved}")
        if not log_saved:
            logger.error(f"❌❌❌ CRITICAL WARNING: Logout log was NOT saved to database!")
            print(f"[DEBUG] ❌❌❌ CRITICAL WARNING: Logout log was NOT saved to database!")
            print(f"[LOGOUT LOG ERROR] ❌ Failed to save logout log - check Django logs for details")
        else:
            logger.info("=" * 80)
            logger.info("✅ LOGOUT LOGGING COMPLETED SUCCESSFULLY")
            logger.info("=" * 80)
            print("=" * 80)
            print("✅ LOGOUT LOGGING COMPLETED SUCCESSFULLY")
            print("=" * 80)
        
        # Clear all session data
        request.session.flush()
        
        # Delete the session completely
        request.session.delete()
        
        return Response({
            'status': 'success',
            'message': 'Logged out successfully'
        })
    except Exception as e:
        logger.error(f"Logout error: {str(e)}")
        return Response({
            'status': 'error',
            'message': 'Error during logout',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])  # Keep AllowAny for now, but add RBAC check inside
def register_user(request):
    """
    Register a new user
    """
    try:
        # Check if the requesting user is a GRC Administrator
        from .rbac.utils import RBACUtils
        user_id = RBACUtils.get_user_id_from_request(request)
        
        if not user_id:
            return Response({
                'success': False,
                'message': 'Authentication required to create users'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if user is GRC Administrator
        if not RBACUtils.is_system_admin(user_id):
            return Response({
                'success': False,
                'message': 'Only GRC Administrators can create users'
            }, status=status.HTTP_403_FORBIDDEN)
        
        data = request.data
        logger.info(f"Registration request data: {data}")
        username = data.get('username')
        # Password is always auto-generated - ignore any password provided in request
        # Handle both 'email' and 'Email' field names
        Email = data.get('Email') or data.get('email')
        firstName = data.get('firstName', '')
        lastName = data.get('lastName', '')
        phoneNumber = data.get('phoneNumber') or data.get('phone') or ''
        address = data.get('address') or data.get('Address') or ''
        
        logger.info(f"Extracted fields - username: {username}, email: {Email}, firstName: {firstName}")
        
        if not username or not Email:
            logger.error(f"Missing required fields - username: {bool(username)}, email: {bool(Email)}")
            return Response({
                'success': False,
                'message': 'Username and Email are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Always auto-generate password in format: Riskavaire@<FirstName><number>
        # Get the user's first name (or username if first name is not available)
        name_part = firstName.strip() if firstName else username
        if not name_part:
            name_part = username
        
        # Find the next available number by counting existing users with the same first name
        # This ensures incrementing numbers for users with the same name
        if firstName:
            # Count existing users with the same first name
            similar_users_count = Users.objects.filter(FirstName=firstName).count()
            password_number = similar_users_count + 1
        else:
            # If no first name, use username-based counting
            similar_users_count = Users.objects.filter(UserName__startswith=username[:5]).count()
            password_number = similar_users_count + 1
        
        # Generate password in format: Riskavaire@<name><number>
        password = f"Riskavaire@{name_part}{password_number}"
        logger.info(f"Auto-generated password for user {username}: Riskavaire@{name_part}{password_number}")
        print(f"[DEBUG] Auto-generated password for user {username}: Riskavaire@{name_part}{password_number}")
        
        # Check if user already exists
        logger.info(f"Checking if username '{username}' already exists...")
        print(f"[DEBUG] Checking if username '{username}' already exists...")
        if Users.objects.filter(UserName=username).exists():
            logger.error(f"Username '{username}' already exists")
            print(f"[DEBUG] ❌ Username '{username}' already exists")
            return Response({
                'success': False,
                'message': 'Username already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"Checking if email '{Email}' already exists...")
        print(f"[DEBUG] Checking if email '{Email}' already exists...")
        if Users.objects.filter(Email=Email).exists():
            logger.error(f"Email '{Email}' already exists")
            print(f"[DEBUG] ❌ Email '{Email}' already exists")
            return Response({
                'success': False,
                'message': 'Email already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"User checks passed, proceeding to license generation...")
        print(f"[DEBUG] ✅ User checks passed, proceeding to license generation...")
        
        # Generate license key using the licensing system (skip external API if disabled)
        logger.info("Starting license key generation...")
        print(f"[DEBUG] Starting license key generation...")
        license_key = None
        license_message = "License key generation skipped"
        
        try:
            from django.conf import settings as _dj_settings
            if not getattr(_dj_settings, 'LICENSE_CHECK_ENABLED', True):
                logger.warning("🔕 LICENSE CHECK DISABLED during registration. Generating license key locally without API creation.")
                print(f"[DEBUG] LICENSE CHECK DISABLED - generating locally")
                try:
                    from licensing_system import VardaanLicensingSystem
                    licensing_system = VardaanLicensingSystem()
                    license_key = licensing_system.generate_secure_license_code()
                    license_message = "License key generated locally (API disabled)."
                    logger.info(f"License key generated locally: {license_key}")
                    print(f"[DEBUG] ✅ License key generated locally: {license_key}")
                except Exception as license_error:
                    logger.error(f"License generation error (disabled mode): {str(license_error)}")
                    print(f"[DEBUG] ❌ License generation error (disabled mode): {str(license_error)}")
                    license_key = None
                    license_message = f"License generation failed (disabled mode): {str(license_error)}"
            else:
                print(f"[DEBUG] LICENSE CHECK ENABLED - generating via API")
                try:
                    from licensing_system import VardaanLicensingSystem
                    licensing_system = VardaanLicensingSystem()
                    license_key = licensing_system.generate_secure_license_code()
                    # Create license via API
                    api_result = licensing_system.create_license(license_key)
                    if api_result.get("success"):
                        logger.info(f"License key generated and created via API: {license_key}")
                        print(f"[DEBUG] ✅ License key generated and created via API: {license_key}")
                        license_message = f"License key generated and created via API successfully"
                    else:
                        logger.warning(f"License key generated but API creation failed: {api_result.get('error')}")
                        print(f"[DEBUG] ⚠️ License key generated but API creation failed: {api_result.get('error')}")
                        license_message = f"License key generated but API creation failed: {api_result.get('error')}"
                except Exception as license_error:
                    logger.error(f"License generation error: {str(license_error)}")
                    print(f"[DEBUG] ❌ License generation error: {str(license_error)}")
                    license_key = None
                    license_message = f"License generation failed: {str(license_error)}"
        except Exception as license_import_error:
            logger.warning(f"Could not import licensing system, skipping license generation: {str(license_import_error)}")
            print(f"[DEBUG] ⚠️ Could not import licensing system, skipping: {str(license_import_error)}")
            license_key = None
            license_message = "License generation skipped (module not available)"
        
        logger.info(f"License generation completed. Key: {license_key}, Message: {license_message}")
        print(f"[DEBUG] License generation completed. Key: {license_key}, Message: {license_message}")
        
        # Create new user with all available fields including license key
        # IMPORTANT: Store password as a secure hash, not plain text
        # Store the plain password temporarily for email (before hashing)
        plain_password = password
        
        # Convert departmentId to string if it's an integer (model expects CharField)
        department_id = data.get('departmentId', '')
        if department_id and not isinstance(department_id, str):
            department_id = str(department_id)
        
        user_data = {
            'UserName': username,
            'Password': make_password(password),
            'Email': Email,
            'FirstName': data.get('firstName', ''),
            'LastName': data.get('lastName', ''),
            'PhoneNumber': phoneNumber if phoneNumber else None,
            'Address': address if address else None,
            'DepartmentId': department_id,
            'IsActive': 'N'  # ALWAYS set to 'N' for new users - they must reset password and login to activate
        }
        # Log that we're forcing IsActive to 'N' regardless of frontend input
        logger.info(f"Setting IsActive='N' for new user {username} (frontend sent: {data.get('isActive', 'not provided')})")
        print(f"[DEBUG] Setting IsActive='N' for new user {username} (frontend sent: {data.get('isActive', 'not provided')})")
        
        # Add license key if generated successfully (only if not None and not empty)
        if license_key and license_key.strip():
            user_data['license_key'] = license_key
            logger.info(f"Adding license key to user data: {license_key}")
            print(f"[DEBUG] Adding license key to user data: {license_key}")
        else:
            logger.info("No license key to add (will be None in database)")
            print(f"[DEBUG] No license key to add (will be None in database)")
        
        # Log user data before creation
        logger.info(f"Attempting to create user with data: {user_data}")
        logger.info(f"DepartmentId type: {type(user_data.get('DepartmentId'))}, value: {user_data.get('DepartmentId')}")
        print(f"[DEBUG] Attempting to create user with data: {user_data}")
        print(f"[DEBUG] DepartmentId type: {type(user_data.get('DepartmentId'))}, value: {user_data.get('DepartmentId')}")
        
        # Create user with error handling
        try:
            user = Users.objects.create(**user_data)
            logger.info(f"✅ User created successfully: {user.UserId} - {user.UserName}")
            print(f"[DEBUG] ✅ User created successfully: {user.UserId} - {user.UserName}")
            
            # Log password creation to password_logs
            try:
                from .models import PasswordLog
                PasswordLog.objects.create(
                    UserId=user.UserId,
                    UserName=user.UserName,
                    OldPassword=None,  # No old password for new user
                    NewPassword=user.Password,  # Hashed password
                    ActionType='created',
                    IPAddress=request.META.get('REMOTE_ADDR', ''),
                    UserAgent=request.META.get('HTTP_USER_AGENT', ''),
                    AdditionalInfo={'created_by': 'admin', 'email': user.Email}
                )
                logger.info(f"✅ Password log created for new user: {user.UserName}")
                print(f"[DEBUG] ✅ Password log created for new user: {user.UserName}")
            except Exception as log_error:
                logger.error(f"❌ Failed to create password log: {str(log_error)}")
                print(f"[DEBUG] ❌ Failed to create password log: {str(log_error)}")
                # Don't fail user creation if logging fails
            
            # Also log user registration and password creation to grc_logs
            try:
                from .routes.Global.logging_service import send_log
                client_ip = request.META.get('REMOTE_ADDR', 'unknown')
                send_log(
                    module='User Management',
                    actionType='USER_REGISTERED',
                    description=f'New user {user.UserName} (ID: {user.UserId}) registered with email {user.Email}',
                    userId=str(user.UserId),
                    userName=user.UserName,
                    logLevel='INFO',
                    ipAddress=client_ip,
                    additionalInfo={'email': user.Email, 'created_by': 'admin', 'password_created': True},
                    frameworkId=user.FrameworkId.FrameworkId if user.FrameworkId else None
                )
                logger.info(f"✅ User registration logged to grc_logs: {user.UserName}")
            except Exception as log_error:
                logger.error(f"❌ Failed to log user registration to grc_logs: {str(log_error)}")
                # Don't fail user creation if logging fails
        except Exception as create_error:
            import traceback
            error_traceback = traceback.format_exc()
            error_msg = f"❌ Error creating user in database: {str(create_error)}"
            logger.error(error_msg)
            logger.error(f"❌ Error type: {type(create_error).__name__}")
            logger.error(f"❌ User data that failed: {user_data}")
            logger.error(f"❌ Full error traceback:\n{error_traceback}")
            print(f"[DEBUG] {error_msg}")
            print(f"[DEBUG] Error type: {type(create_error).__name__}")
            print(f"[DEBUG] User data that failed: {user_data}")
            print(f"[DEBUG] Full error traceback:\n{error_traceback}")
            return Response({
                'success': False,
                'message': f'Failed to create user: {str(create_error)}',
                'error_details': str(create_error),
                'error_type': type(create_error).__name__
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Send email with credentials to the user using Azure email backend directly
        try:
            # Prepare email content with HTML formatting
            user_full_name = f"{user.FirstName} {user.LastName}".strip() or user.UserName
            email_subject = "Your GRC Account Credentials"
            
            # Get frontend URL from settings for verification link
            frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:8080')
            reset_password_url = f"{frontend_url}/login?resetPassword=true&email={user.Email}"
            
            # Plain text version
            email_body_text = f"""
Dear {user_full_name},

Your account has been created successfully by the administrator.

Here are your login credentials:

Username: {user.UserName}
Password: {plain_password}

IMPORTANT: For security reasons, please verify your email and reset your password immediately after your first login.

To verify your email, click on the following verification link:
{reset_password_url}

Or manually:
1. Go to the login page
2. Click on "Forgot Password"
3. Enter your email address: {user.Email}
4. Follow the instructions to verify your email and reset your password

If you have any questions, please contact the administrator.

Best regards,
GRC System Administrator
            """.strip()
            
            # HTML version for better formatting
            email_body_html = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #3b82f6 0%, #1e40af 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
        .content {{ background: #f9fafb; padding: 30px; border-radius: 0 0 8px 8px; }}
        .credentials {{ background: white; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #3b82f6; }}
        .credential-item {{ margin: 10px 0; }}
        .label {{ font-weight: 600; color: #1f2937; }}
        .value {{ color: #374151; font-family: monospace; background: #f3f4f6; padding: 4px 8px; border-radius: 4px; }}
        .warning {{ background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 4px; }}
        .footer {{ text-align: center; color: #6b7280; font-size: 12px; margin-top: 30px; }}
        .button {{ display: inline-block; background: #3b82f6; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; margin: 10px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>Welcome to GRC System</h2>
        </div>
        <div class="content">
            <p>Dear <strong>{user_full_name}</strong>,</p>
            
            <p>Your account has been created successfully by the administrator.</p>
            
            <div class="credentials">
                <h3 style="margin-top: 0; color: #1f2937;">Your Login Credentials:</h3>
                <div class="credential-item">
                    <span class="label">Username:</span>
                    <span class="value">{user.UserName}</span>
                </div>
                <div class="credential-item">
                    <span class="label">Password:</span>
                    <span class="value">{plain_password}</span>
                </div>
            </div>
            
            <div class="warning">
                <strong>⚠️ IMPORTANT:</strong> For security reasons, please verify your email and reset your password immediately after your first login.
            </div>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_password_url}" class="button" style="display: inline-block; background: linear-gradient(135deg, #3b82f6 0%, #1e40af 100%); color: white; padding: 14px 32px; text-decoration: none; border-radius: 8px; font-weight: 600; font-size: 16px; box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3); transition: transform 0.2s;">
                    🔐 Verification Link
                </a>
            </div>
            
            <p style="text-align: center; color: #6b7280; font-size: 14px; margin-top: 20px;">
                Click the verification link above to verify your email, or manually go to the login page and click "Forgot Password"
            </p>
            
            <p>If you have any questions, please contact the administrator.</p>
            
            <div class="footer">
                <p>Best regards,<br>GRC System Administrator</p>
            </div>
        </div>
    </div>
</body>
</html>
            """.strip()
            
            # Use Azure email backend directly to ensure it uses Azure (not SMTP fallback)
            from .routes.Global.azure_email_backend import AzureADEmailBackend
            from django.core.mail import EmailMessage
            
            # Create EmailMessage with HTML content
            email_message = EmailMessage(
                subject=email_subject,
                body=email_body_html,  # Use HTML as body
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@grc.com'),
                to=[user.Email],
            )
            email_message.content_subtype = "html"  # Set content type to HTML
            
            # Use Azure backend directly (this will use Azure Graph API, not SMTP)
            azure_backend = AzureADEmailBackend(fail_silently=False)
            result = azure_backend.send_messages([email_message])
            
            if result > 0:
                logger.info(f"✅ User creation email sent via Azure Graph API to {user.Email}")
                print(f"[DEBUG] ✅ User creation email sent via Azure Graph API to {user.Email}")
            else:
                logger.warning(f"⚠️ Email sending returned 0 sent messages for {user.Email}")
                print(f"[DEBUG] ⚠️ Email sending returned 0 sent messages for {user.Email}")
        except Exception as email_error:
            import traceback
            error_traceback = traceback.format_exc()
            logger.error(f"❌ Failed to send user creation email to {user.Email}: {str(email_error)}")
            logger.error(f"❌ Email error traceback:\n{error_traceback}")
            print(f"[DEBUG] ❌ Failed to send user creation email to {user.Email}: {str(email_error)}")
            print(f"[DEBUG] ❌ Email error traceback:\n{error_traceback}")
            # Don't fail user creation if email fails, just log the error
        
        # Create RBAC entry if role is provided
        if data.get('role'):
            try:
                from .rbac.utils import RBACUtils
                # Create RBAC entry with role and permissions
                rbac_data = {
                    'user': user,
                    'username': user.UserName,
                    'role': data.get('role'),
                    'is_active': 'Y'
                }
                
                # Add permissions if provided
                permissions = data.get('permissions', {})
                for permission, value in permissions.items():
                    if hasattr(RBAC, permission):
                        rbac_data[permission] = value
                
                RBAC.objects.create(**rbac_data)
                
            except Exception as rbac_error:
                logger.warning(f"Could not create RBAC entry for user {user.UserId}: {rbac_error}")
        
        return Response({
            'success': True,
            'message': 'User registered successfully',
            'license_message': license_message,
            'user': {
                'id': user.UserId,
                'username': user.UserName,
                'Email': user.Email,
                'firstName': user.FirstName,
                'lastName': user.LastName,
                'phoneNumber': user.PhoneNumber,
                'address': user.Address,
                'departmentId': user.DepartmentId,
                'isActive': user.IsActive,
                'license_key': user.license_key
            }
        }, status=status.HTTP_201_CREATED)
        
        # Log user creation with masked sensitive data
        try:
            from .routes.Global.data_masking import mask_log_data, get_masking_service
            from .routes.Global.logging_service import send_log
            
            # Get client IP helper function
            def get_client_ip(request):
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                return ip
            
            # Prepare log data with sensitive information
            log_additional_info = {
                'username': username,
                'email': Email,
                'firstName': firstName,
                'lastName': lastName,
                'phoneNumber': phoneNumber,
                'address': address,
                'departmentId': department_id,
                'action': 'user_created'
            }
            
            # Mask the log data before saving
            masked_additional_info = get_masking_service().mask_dict(log_additional_info)
            
            # Get client IP
            client_ip = get_client_ip(request)
            
            # Get framework for log entry (required field)
            framework = Framework.objects.filter(Status='Approved', ActiveInactive='Active').first()
            if not framework:
                framework = Framework.objects.first()
            
            # Create log entry with masked data
            if framework:
                # Mask the username for the log
                masked_username = get_masking_service().mask_name(username)
                masked_description = f'User account created: {masked_username}'
                
                log_entry = GRCLog(
                    UserId=str(user_id),
                    UserName=masked_username,  # Masked username
                    Module='User Management',
                    ActionType='CREATE_USER',
                    EntityType='User',
                    EntityId=str(user.UserId),
                    LogLevel='INFO',
                    Description=masked_description,
                    IPAddress=client_ip,
                    AdditionalInfo=masked_additional_info,
                    FrameworkId=framework
                )
                log_entry.save()
                logger.info(f"✅ User creation logged with masked data: {user.UserId}, LogId: {log_entry.LogId}")
                print(f"[DEBUG] ✅ User creation logged with masked data: {user.UserId}, LogId: {log_entry.LogId}")
            else:
                logger.warning("No framework found for logging user creation")
                print(f"[DEBUG] ⚠️ No framework found for logging user creation")
        except Exception as log_error:
            logger.error(f"❌ Failed to log user creation: {str(log_error)}")
            import traceback
            error_trace = traceback.format_exc()
            logger.error(error_trace)
            print(f"[DEBUG] ❌ Failed to log user creation: {str(log_error)}")
            print(f"[DEBUG] Error traceback:\n{error_trace}")
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        logger.error(f"❌❌❌ Registration error at top level: {str(e)}")
        logger.error(f"❌❌❌ Error type: {type(e).__name__}")
        logger.error(f"❌❌❌ Registration error traceback:\n{error_traceback}")
        return Response({
            'success': False,
            'message': f'An error occurred during registration: {str(e)}',
            'error_details': str(e),
            'error_type': type(e).__name__
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([AllowAny])
def test_connection(request):
    """
    Test API connection
    """
    return Response({
        'success': True,
        'message': 'API connection successful'
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def get_user_details_by_id(request, user_id):
    """
    Get user details from Users table by user ID
    """
    try:
        # Get user from Users table
        user = Users.objects.get(UserId=user_id)
        
        # Get RBAC info if available
        rbac_info = None
        try:
            from .rbac.utils import RBACUtils
            rbac_info = RBACUtils.get_user_rbac_info(user_id)
        except Exception as rbac_error:
            logger.warning(f"Could not get RBAC info for user {user_id}: {rbac_error}")
        
        user_details = {
            'user_id': user.UserId,
            'username': user.UserName,
            'Email': user.Email,
            'created_at': user.CreatedAt,
            'updated_at': user.UpdatedAt,
            'rbac_info': rbac_info
        }
        
        logger.info(f"✅ Retrieved user details for ID {user_id}: {user.UserName}")
        
        return Response({
            'success': True,
            'user': user_details,
            'message': f'User details retrieved for {user.UserName}'
        })
        
    except Users.DoesNotExist:
        logger.warning(f"❌ User with ID {user_id} not found")
        return Response({
            'success': False,
            'message': f'User with ID {user_id} not found'
        }, status=404)
        
    except Exception as e:
        logger.error(f"❌ Error retrieving user details for ID {user_id}: {str(e)}")
        return Response({
            'success': False,
            'message': 'Error retrieving user details',
            'error': str(e)
        }, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def save_user_session(request):
    """
    Save user session information manually for testing
    """
    try:
        from django.contrib.auth import authenticate, login
        from django.contrib.auth.models import User as DjangoUser
        
        data = request.data
        user_id = data.get('user_id')
        
        if not user_id:
            return Response({
                'success': False,
                'message': 'user_id is required'
            }, status=400)
        
        # Get user from Users table
        try:
            grc_user = Users.objects.get(UserId=user_id)
        except Users.DoesNotExist:
            return Response({
                'success': False,
                'message': f'GRC User with ID {user_id} not found'
            }, status=404)
        
        # Create or get Django User for session management
        django_user, created = DjangoUser.objects.get_or_create(
            username=grc_user.UserName,
            defaults={
                'Email': grc_user.Email,
                'first_name': grc_user.UserName,
                'is_active': True,
                'is_staff': False,
                'is_superuser': grc_user.UserName == 'admin.grc'
            }
        )
        
        # Set a default password if user was created
        if created:
            django_user.set_password('default_password')
            django_user.save()
        
        # Login the user (creates session)
        login(request, django_user)
        
        # Store GRC user ID in session for RBAC
        request.session['user_id'] = grc_user.UserId  # Changed from grc_user_id to user_id for RBACUtils
        request.session['grc_user_id'] = grc_user.UserId  # Keep both for backwards compatibility
        request.session['grc_username'] = grc_user.UserName
        request.session['grc_Email'] = grc_user.Email
        request.session.save()
        
        # Log user permissions using RBAC system
        try:
            from .rbac.utils import RBACUtils
            RBACUtils.log_user_login_permissions(grc_user.UserId)
        except Exception as rbac_error:
            logger.warning(f"Could not log RBAC permissions for user {grc_user.UserId}: {rbac_error}")
        
        logger.info(f"✅ Session created for user: {grc_user.UserName} (ID: {grc_user.UserId})")
        
        return Response({
            'success': True,
            'message': f'Session created for user {grc_user.UserName}',
            'user': {
                'id': grc_user.UserId,
                'username': grc_user.UserName,
                'Email': grc_user.Email,
                'django_user_id': django_user.id,
                'session_key': request.session.session_key
            },
            'session_data': dict(request.session)
        })
        
    except Exception as e:
        logger.error(f"❌ Error saving user session: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'success': False,
            'message': 'Error saving user session',
            'error': str(e)
        }, status=500)

@require_http_methods(["GET"])
def get_user_profile(request, user_id):
    try:
        logger.debug(f"Fetching user profile for user_id: {user_id}")
        user = Users.objects.get(UserId=user_id)
        
        # Import masking service
        from .routes.Global.data_masking import get_masking_service
        masking_service = get_masking_service()
        
        # Get decrypted values for encrypted fields
        email_plain = user.email_plain
        phone_plain = user.phone_plain
        address_plain = user.address_plain
        
        # Mask sensitive data for display (using decrypted values)
        masked_data = {
            'firstName': masking_service.mask_name(user.FirstName) if user.FirstName else None,
            'lastName': masking_service.mask_name(user.LastName) if user.LastName else None,
            'email': masking_service.mask_email(email_plain) if email_plain else None,
            'phoneNumber': masking_service.mask_phone(phone_plain) if phone_plain else None,
            'address': masking_service.mask_address(address_plain) if address_plain else None,
            'username': masking_service.mask_name(user.UserName) if user.UserName else None,
            'isActive': user.IsActive,
            'departmentId': user.DepartmentId,
            # Include original values for editing (decrypted values)
            'original': {
                'firstName': user.FirstName,
                'lastName': user.LastName,
                'email': email_plain,  # Use decrypted email
                'phoneNumber': phone_plain,  # Use decrypted phone
                'address': address_plain,  # Use decrypted address
                'username': user.UserName
            }
        }
        
        logger.debug(f"User data fetched with masking")
        return JsonResponse({
            'status': 'success',
            'data': masked_data
        })

    except Users.DoesNotExist:
        logger.error(f"User not found with id: {user_id}")
        return JsonResponse({
            'status': 'error',
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error fetching user profile: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@require_http_methods(["GET"])
def get_user_business_info(request, user_id):
    try:
        logger.debug(f"Fetching business info for user_id: {user_id}")
        
        # Get user's department ID
        user = Users.objects.get(UserId=user_id)
        department_id = user.DepartmentId
        
        logger.debug(f"Found department_id: {department_id}")

        # Get department info with related data using raw SQL for efficient joins
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    d.DepartmentId,
                    d.DepartmentName,
                    d.DepartmentHead,
                    bu.BusinessUnitId,
                    bu.Name as BusinessUnitName,
                    bu.Code as BusinessUnitCode,
                    e.Id as EntityId,
                    e.EntityName,
                    e.EntityType,
                    l.LocationID,
                    CONCAT(l.AddressLine, ', ', l.City, 
                        CASE WHEN l.State IS NOT NULL THEN CONCAT(', ', l.State) ELSE '' END,
                        ', ', l.Country,
                        CASE WHEN l.PostalCode IS NOT NULL THEN CONCAT(' - ', l.PostalCode) ELSE '' END
                    ) as Location
                FROM department d
                LEFT JOIN businessunits bu ON d.BusinessUnitId = bu.BusinessUnitId
                LEFT JOIN mainentities e ON d.EntityId = e.Id
                LEFT JOIN locations l ON e.LocationId = l.LocationID
                WHERE d.DepartmentId = %s
            """, [department_id])
            
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()
            
            if row:
                result = dict(zip(columns, row))
                logger.debug(f"Raw business info fetched: {result}")

                # Get department head name
                if result['DepartmentHead']:
                    dept_head = Users.objects.filter(UserId=result['DepartmentHead']).first()
                    if dept_head:
                        result['DepartmentHead'] = f"{dept_head.FirstName} {dept_head.LastName}"
                        logger.debug(f"Department head name: {result['DepartmentHead']}")

                return JsonResponse({
                    'status': 'success',
                    'data': result
                })
            else:
                logger.warning(f"No business info found for department_id: {department_id}")
                return JsonResponse({
                    'status': 'error',
                    'message': 'No business information found'
                }, status=404)

    except Users.DoesNotExist:
        logger.error(f"User not found with id: {user_id}")
        return JsonResponse({
            'status': 'error',
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error fetching business info: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@require_http_methods(["GET"])
def get_user_permissions(request, user_id):
    try:
        logger.debug(f"Fetching permissions for user_id: {user_id}")
        
        # Get user from Users table
        user = Users.objects.get(UserId=user_id)
        
        # Get RBAC permissions for the user
        from .models import RBAC
        rbac_entries = RBAC.objects.filter(user=user, is_active='Y')
        
        if not rbac_entries.exists():
            logger.warning(f"No RBAC permissions found for user_id: {user_id}")
            return JsonResponse({
                'status': 'success',
                'data': {
                    'role': '',
                    'modules': {}
                }
            })
        
        # Use the first role entry (in case user has multiple roles)
        rbac = rbac_entries.first()
        
        # Format permissions by module
        permissions = {
            'role': rbac.role,
            'modules': {
                'compliance': {
                    'create_compliance': rbac.create_compliance,
                    'edit_compliance': rbac.edit_compliance,
                    'approve_compliance': rbac.approve_compliance,
                    'view_all_compliance': rbac.view_all_compliance,
                    'compliance_performance_analytics': rbac.compliance_performance_analytics
                },
                'policy': {
                    'create_policy': rbac.create_policy,
                    'edit_policy': rbac.edit_policy,
                    'approve_policy': rbac.approve_policy,
                    'create_framework': rbac.create_framework,
                    'approve_framework': rbac.approve_framework,
                    'view_all_policy': rbac.view_all_policy,
                    'policy_performance_analytics': rbac.policy_performance_analytics
                },
                'audit': {
                    'assign_audit': rbac.assign_audit,
                    'conduct_audit': rbac.conduct_audit,
                    'review_audit': rbac.review_audit,
                    'view_audit_reports': rbac.view_audit_reports,
                    'audit_performance_analytics': rbac.audit_performance_analytics
                },
                'risk': {
                    'create_risk': rbac.create_risk,
                    'edit_risk': rbac.edit_risk,
                    'approve_risk': rbac.approve_risk,
                    'assign_risk': rbac.assign_risk,
                    'evaluate_assigned_risk': rbac.evaluate_assigned_risk,
                    'view_all_risk': rbac.view_all_risk,
                    'risk_performance_analytics': rbac.risk_performance_analytics
                },
                'incident': {
                    'create_incident': rbac.create_incident,
                    'edit_incident': rbac.edit_incident,
                    'assign_incident': rbac.assign_incident,
                    'evaluate_assigned_incident': rbac.evaluate_assigned_incident,
                    'escalate_to_risk': rbac.escalate_to_risk,
                    'view_all_incident': rbac.view_all_incident,
                    'incident_performance_analytics': rbac.incident_performance_analytics
                }
            }
        }
        
        logger.debug(f"Permissions data prepared for user_id {user_id}: {permissions}")
        
        return JsonResponse({
            'status': 'success',
            'data': permissions
        })
        
    except Users.DoesNotExist:
        logger.error(f"User not found with id: {user_id}")
        return JsonResponse({
            'status': 'error',
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error fetching user permissions: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@csrf_exempt
@api_view(['PUT'])
@authentication_classes([])
@permission_classes([AllowAny])
def update_user_permissions(request, user_id):
    """
    Update RBAC permissions for a user
    """
    try:
        logger.debug(f"Updating permissions for user_id: {user_id}")
        
        # Get user from Users table
        user = Users.objects.get(UserId=user_id)
        
        # Get permissions from request body
        permissions = request.data.get('permissions', {})
        
        logger.debug(f"Received permissions data: {permissions}")
        logger.debug(f"Request data: {request.data}")
        
        if not permissions:
            return JsonResponse({
                'status': 'error',
                'message': 'No permissions provided'
            }, status=400)
        
        # Get RBAC entry for the user
        from .models import RBAC, Framework
        from django.db import transaction
        
        # Get or create framework - use user's framework or first available
        framework = getattr(user, 'FrameworkId', None)
        if not framework:
            framework = Framework.objects.filter(Status='Approved', ActiveInactive='Active').first()
            if not framework:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No active framework found. Please create a framework first.'
                }, status=400)
        
        # Use transaction to ensure atomicity
        with transaction.atomic():
            # Get existing RBAC entry - check by user and role if role is provided
            role = request.data.get('role')
            if role:
                rbac_entry = RBAC.objects.filter(user=user, role=role, is_active='Y').first()
            else:
                rbac_entry = RBAC.objects.filter(user=user, is_active='Y').first()
            
            if not rbac_entry:
                # Create new RBAC entry if it doesn't exist
                rbac_entry = RBAC.objects.create(
                    user=user,
                    username=user.UserName,
                    role=role or 'End User',
                    is_active='Y',
                    FrameworkId=framework
                )
                logger.info(f"Created new RBAC entry for user_id: {user_id}, role: {role or 'End User'}")
            else:
                logger.info(f"Found existing RBAC entry (ID: {rbac_entry.rbac_id}) for user_id: {user_id}, role: {rbac_entry.role}")
            
            # Update permissions
            # Map frontend permission field names to model field names
            permission_mapping = {
                # Compliance
                'create_compliance': 'create_compliance',
                'edit_compliance': 'edit_compliance',
                'approve_compliance': 'approve_compliance',
                'view_all_compliance': 'view_all_compliance',
                'compliance_performance_analytics': 'compliance_performance_analytics',
                # Policy
                'create_policy': 'create_policy',
                'edit_policy': 'edit_policy',
                'approve_policy': 'approve_policy',
                'create_framework': 'create_framework',
                'approve_framework': 'approve_framework',
                'view_all_policy': 'view_all_policy',
                'policy_performance_analytics': 'policy_performance_analytics',
                # Audit
                'assign_audit': 'assign_audit',
                'conduct_audit': 'conduct_audit',
                'review_audit': 'review_audit',
                'view_audit_reports': 'view_audit_reports',
                'audit_performance_analytics': 'audit_performance_analytics',
                # Risk
                'create_risk': 'create_risk',
                'edit_risk': 'edit_risk',
                'approve_risk': 'approve_risk',
                'assign_risk': 'assign_risk',
                'evaluate_assigned_risk': 'evaluate_assigned_risk',
                'view_all_risk': 'view_all_risk',
                'risk_performance_analytics': 'risk_performance_analytics',
                # Incident
                'create_incident': 'create_incident',
                'edit_incident': 'edit_incident',
                'assign_incident': 'assign_incident',
                'evaluate_assigned_incident': 'evaluate_assigned_incident',
                'escalate_to_risk': 'escalate_to_risk',
                'view_all_incident': 'view_all_incident',
                'incident_performance_analytics': 'incident_performance_analytics'
            }
            
            # Track updated fields for logging
            updated_fields = []
            
            # Update each permission field
            for field_name, value in permissions.items():
                if field_name in permission_mapping:
                    model_field = permission_mapping[field_name]
                    if hasattr(rbac_entry, model_field):
                        old_value = getattr(rbac_entry, model_field)
                        new_value = bool(value)
                        setattr(rbac_entry, model_field, new_value)
                        if old_value != new_value:
                            updated_fields.append(f"{model_field}: {old_value} -> {new_value}")
                else:
                    logger.warning(f"Unknown permission field: {field_name}")
            
            # Update role if provided
            if role and rbac_entry.role != role:
                rbac_entry.role = role
                updated_fields.append(f"role: {rbac_entry.role} -> {role}")
            
            # Log what will be updated
            if updated_fields:
                logger.info(f"Updating fields for user_id {user_id}: {', '.join(updated_fields)}")
            else:
                logger.warning(f"No fields to update for user_id {user_id}")
            
            # Save the RBAC entry - use direct save without update_fields to ensure all changes are saved
            try:
                # Force save all fields
                rbac_entry.save()
                
                logger.info(f"Saved RBAC entry (ID: {rbac_entry.rbac_id}) for user_id: {user_id}")
                
                # Use raw SQL to ensure database is updated directly
                from django.db import connection
                with connection.cursor() as cursor:
                    # Build UPDATE query with all permission fields
                    set_clauses = []
                    params = []
                    
                    for field_name, value in permissions.items():
                        if field_name in permission_mapping:
                            model_field = permission_mapping[field_name]
                            try:
                                # Get database column name from model
                                field_obj = RBAC._meta.get_field(model_field)
                                db_column = field_obj.db_column or model_field
                                set_clauses.append(f"`{db_column}` = %s")
                                params.append(1 if bool(value) else 0)
                                logger.debug(f"Adding field {model_field} -> {db_column} = {1 if bool(value) else 0}")
                            except Exception as field_error:
                                logger.warning(f"Could not get db_column for {model_field}: {field_error}")
                    
                    if set_clauses:
                        # Add updated_at
                        set_clauses.append("`UpdatedAt` = NOW()")
                        
                        # Build full UPDATE query
                        update_query = f"""
                            UPDATE `rbac` 
                            SET {', '.join(set_clauses)}
                            WHERE `RBACId` = %s
                        """
                        params.append(rbac_entry.rbac_id)
                        
                        logger.info(f"Executing raw SQL update for RBAC ID {rbac_entry.rbac_id}")
                        logger.info(f"SQL: {update_query}")
                        logger.info(f"Params count: {len(params)}")
                        
                        cursor.execute(update_query, params)
                        rows_affected = cursor.rowcount
                        
                        logger.info(f"Raw SQL update affected {rows_affected} row(s)")
                        
                        if rows_affected == 0:
                            logger.error(f"WARNING: No rows were updated! RBAC ID {rbac_entry.rbac_id} may not exist in database.")
                        else:
                            logger.info(f"SUCCESS: {rows_affected} row(s) updated in rbac table")
                        
                        # Note: No need to commit manually - transaction.atomic() handles it automatically
                
                # Refresh from database to verify the save
                rbac_entry.refresh_from_db()
                
                # Log some key permissions to verify they were saved
                logger.info(f"Verification - RBAC ID {rbac_entry.rbac_id} permissions after save:")
                logger.info(f"  create_compliance: {rbac_entry.create_compliance}")
                logger.info(f"  create_policy: {rbac_entry.create_policy}")
                logger.info(f"  create_risk: {rbac_entry.create_risk}")
                logger.info(f"  create_incident: {rbac_entry.create_incident}")
                logger.info(f"  role: {rbac_entry.role}")
                
            except Exception as save_error:
                logger.error(f"Error saving RBAC entry: {str(save_error)}")
                import traceback
                traceback.print_exc()
                raise save_error
        
        logger.info(f"Successfully updated permissions for user_id: {user_id}")
        
        return JsonResponse({
            'status': 'success',
            'message': 'User permissions updated successfully',
            'data': {
                'user_id': user_id,
                'username': user.UserName
            }
        })
        
    except Users.DoesNotExist:
        logger.error(f"User not found with id: {user_id}")
        return JsonResponse({
            'status': 'error',
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error updating user permissions: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def send_otp(request):
    """
    Send OTP to user's Email for password reset
    """
    logger.info("=" * 80)
    logger.info("[SEND OTP] ========== SEND OTP ENDPOINT CALLED ==========")
    logger.info(f"[SEND OTP] Request method: {request.method}")
    logger.info(f"[SEND OTP] Request path: {request.path}")
    logger.info(f"[SEND OTP] Request data: {request.data}")
    try:
        data = request.data
        Email = data.get('Email')
        
        logger.info(f"[SEND OTP] Received request to send OTP to email: {Email}")
        
        if not Email:
            logger.error("[SEND OTP] No email provided in request")
            return Response({
                'success': False,
                'message': 'Email is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if email is masked (contains ***) - if so, we need to get full email from username/ID
        if '***' in Email:
            logger.warning(f"[SEND OTP] Received masked email: {Email}. Cannot use masked email to send OTP.")
            return Response({
                'success': False,
                'message': 'Please use the full email address. If you entered username/ID, the system should automatically fetch and use your email.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if user exists with this Email (handles encrypted email fields)
        try:
            user = Users.find_by_email(Email)
            if not user:
                logger.warning(f"[SEND OTP] No user found with email: {Email}")
                return Response({
                    'success': False,
                    'message': 'No user found with this Email address'
                }, status=status.HTTP_404_NOT_FOUND)
            logger.info(f"[SEND OTP] Found user: UserId={user.UserId}, UserName={getattr(user, 'UserName_plain', None) or user.UserName}")
        except Exception as e:
            logger.error(f"[SEND OTP] Error finding user by email: {str(e)}")
            import traceback
            logger.error(f"[SEND OTP] Traceback: {traceback.format_exc()}")
            return Response({
                'success': False,
                'message': 'No user found with this Email address'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Generate 6-digit OTP
        import random
        otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        # Store OTP in session with expiration (5 minutes)
        request.session['reset_otp'] = otp
        request.session['reset_Email'] = Email
        request.session['otp_expiry'] = datetime.now().timestamp() + 300  # 5 minutes
        request.session.save()
        
        # Also store in memory as fallback
        otp_storage[Email] = {
            'otp': otp,
            'expiry': datetime.now().timestamp() + 300  # 5 minutes
        }
        
        # Debug logging
        logger.info(f"Send OTP - Stored session data: reset_otp={otp}, reset_Email={Email}, otp_expiry={request.session['otp_expiry']}")
        logger.info(f"Send OTP - Session ID: {request.session.session_key}")
        logger.info(f"Send OTP - Session modified: {request.session.modified}")
        
        # Send OTP via Email using Notification Service (Azure AD with SMTP fallback)
        try:
            from .routes.Global.notification_service import NotificationService
            
            notification_service = NotificationService()
            
            # Use the passwordResetOTP template from notification service
            user_name = user.FirstName or user.UserName or user.Email.split('@')[0]
            platform_name = "GRC Platform"
            expiry_time = "5 minutes"
            
            notification_data = {
                'notification_type': 'passwordResetOTP',
                'email': Email,
                'email_type': 'gmail',  # Will use Azure if configured, SMTP if not
                'template_data': [
                    user_name,
                    otp,
                    expiry_time,
                    platform_name
                ],
            }
            
            logger.info(f"[OTP] Attempting to send OTP email to {Email} via Azure (with SMTP fallback)")
            email_result = notification_service.send_multi_channel_notification(notification_data)
            
            # Log result
            if email_result.get('success'):
                email_details = email_result.get('details', {}).get('email')
                if email_details:
                    method_used = email_details.get('method', 'unknown')
                    logger.info(f"[OTP] ✅ OTP email sent successfully to {Email} via {method_used}")
                    
                    # For Azure, log additional info to help debug delivery issues
                    if method_used == 'azure_graph_api':
                        logger.info(f"[OTP] Azure Graph API accepted the email. If not received, verify:")
                        logger.info(f"[OTP]    - Azure AD app has 'Mail.Send' Application permission with admin consent")
                        logger.info(f"[OTP]    - Sender email '{email_details.get('from', 'N/A')}' has mailbox enabled")
                        logger.info(f"[OTP]    - Check spam/junk folder")
                else:
                    logger.warning(f"[OTP] ⚠️ Success reported but no email details. Result: {email_result}")
            else:
                error_msg = email_result.get('error', 'Unknown error')
                error_details = email_result.get('details', {}).get('errors', [])
                logger.error(f"[OTP] ❌ Failed to send OTP email: {error_msg}")
                if error_details:
                    for err in error_details:
                        logger.error(f"[OTP]   - Channel: {err.get('channel')}, Error: {err.get('error')}")
                
                # Still return success to user for security (don't reveal email failures)
                logger.warning("[OTP] Returning success to user despite email failure for security reasons")
            response = Response({
                'success': True,
                'message': 'OTP sent to your Email address'
            })
            
            # Ensure session cookie is set in response
            if request.session.session_key:
                response.set_cookie(
                    'grc_sessionid',  # Use the custom session cookie name from settings
                    request.session.session_key,
                    max_age=3600,  # 1 hour
                    httponly=False,
                    samesite='Lax',
                    secure=False,
                    domain=None,  # Allow all domains
                    path='/'
                )
                logger.info(f"Send OTP - Set session cookie: grc_sessionid={request.session.session_key}")
            else:
                logger.warning("Send OTP - No session key available to set cookie")
            
            return response
                
        except Exception as e:
            logger.error(f"Error sending OTP Email: {str(e)}")
            return Response({
                'success': False,
                'message': 'Failed to send OTP. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        logger.error(f"Error in send_otp function: {str(e)}")
        return Response({
            'success': False,
            'message': 'An error occurred while processing your request'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@csrf_exempt
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def get_user_email_by_username(request):
    """
    Get user email by username or user ID for forgot password functionality
    Supports both encrypted and plain text username lookups
    If auto_send_otp=true, automatically sends OTP after fetching email
    """
    try:
        # Support both GET and POST
        if request.method == 'POST':
            username = request.data.get('username')
            auto_send_otp = request.data.get('auto_send_otp', False)
        else:
            username = request.GET.get('username')
            auto_send_otp = request.GET.get('auto_send_otp', 'false').lower() == 'true'
        
        if not username:
            return Response({
                'success': False,
                'message': 'Username or User ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"[FORGOT PASSWORD] Looking up user with username/ID: {username}, auto_send_otp: {auto_send_otp}")
        
        # Try to find user by User ID first (if username is numeric)
        user = None
        try:
            # Check if username is a number (User ID)
            user_id = int(username)
            try:
                user = Users.objects.get(UserId=user_id)
                logger.info(f"[FORGOT PASSWORD] Found user by User ID: {user_id}")
            except Users.DoesNotExist:
                logger.info(f"[FORGOT PASSWORD] No user found with User ID: {user_id}")
        except ValueError:
            # Not a number, continue to username lookup
            logger.info(f"[FORGOT PASSWORD] Username is not numeric, searching by username")
        
        # If not found by ID, try username lookup
        # Need to search all users and compare decrypted UserName (since it's encrypted)
        if not user:
            all_users = Users.objects.all()
            for u in all_users:
                # Get decrypted username using the _plain property
                decrypted_username = getattr(u, 'UserName_plain', None) or getattr(u, 'UserName', None)
                if decrypted_username == username:
                    user = u
                    logger.info(f"[FORGOT PASSWORD] Found user by username: {username}")
                    break
        
        if user:
            # Get decrypted email
            email = user.email_plain
            
            if not email:
                logger.error(f"[FORGOT PASSWORD] User {username} has no email address")
                return Response({
                    'success': False,
                    'message': 'No email address associated with this account'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Mask email for security (show first 3 chars and domain)
            # Example: user@example.com -> use***@example.com
            email_parts = email.split('@')
            if len(email_parts) == 2:
                masked_email = email_parts[0][:3] + '***@' + email_parts[1]
            else:
                masked_email = email[:3] + '***'
            
            response_data = {
                'success': True,
                'email': masked_email,  # Return masked email for security
                'full_email': email,  # Include full email for actual OTP sending
                'message': 'User found'
            }
            
            # If auto_send_otp is true, automatically send OTP
            if auto_send_otp:
                logger.info(f"[FORGOT PASSWORD] Auto-sending OTP to {email}")
                try:
                    # Generate 6-digit OTP
                    import random
                    otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
                    
                    # Store OTP in session with expiration (5 minutes)
                    request.session['reset_otp'] = otp
                    request.session['reset_Email'] = email
                    request.session['reset_UserId'] = user.UserId  # Store UserId to ensure correct user
                    request.session['otp_expiry'] = datetime.now().timestamp() + 300  # 5 minutes
                    request.session.save()
                    
                    # Also store in memory as fallback
                    otp_storage[email] = {
                        'otp': otp,
                        'user_id': user.UserId,  # Store UserId to ensure correct user
                        'expiry': datetime.now().timestamp() + 300  # 5 minutes
                    }
                    
                    # Send OTP via Email using Notification Service
                    try:
                        from .routes.Global.notification_service import NotificationService
                        notification_service = NotificationService()
                        
                        user_name = user.FirstName or getattr(user, 'UserName_plain', None) or user.UserName or email.split('@')[0]
                        platform_name = "GRC Platform"
                        expiry_time = "5 minutes"
                        
                        notification_data = {
                            'notification_type': 'passwordResetOTP',
                            'email': email,
                            'email_type': 'gmail',
                            'template_data': [
                                user_name,
                                otp,
                                expiry_time,
                                platform_name
                            ],
                        }
                        
                        logger.info(f"[FORGOT PASSWORD] Sending OTP email to {email}")
                        email_result = notification_service.send_multi_channel_notification(notification_data)
                        
                        if email_result.get('success'):
                            logger.info(f"[FORGOT PASSWORD] ✅ OTP sent successfully to {email}")
                            response_data['otp_sent'] = True
                            response_data['message'] = 'OTP has been sent to your email address'
                        else:
                            logger.error(f"[FORGOT PASSWORD] ❌ Failed to send OTP: {email_result.get('error', 'Unknown error')}")
                            response_data['otp_sent'] = False
                            response_data['message'] = 'User found, but failed to send OTP. Please try again.'
                    except Exception as email_error:
                        logger.error(f"[FORGOT PASSWORD] Error sending OTP email: {str(email_error)}")
                        response_data['otp_sent'] = False
                        response_data['message'] = 'User found, but failed to send OTP. Please try again.'
                except Exception as otp_error:
                    logger.error(f"[FORGOT PASSWORD] Error generating OTP: {str(otp_error)}")
                    response_data['otp_sent'] = False
            
            logger.info(f"[FORGOT PASSWORD] Returning response for user: {username}")
            return Response(response_data)
        else:
            logger.warning(f"[FORGOT PASSWORD] No user found with username/ID: {username}")
            return Response({
                'success': False,
                'message': 'No user found with this username or User ID'
            }, status=status.HTTP_404_NOT_FOUND)
            
    except Exception as e:
        logger.error(f"[FORGOT PASSWORD] Error fetching user email: {str(e)}")
        import traceback
        logger.error(f"[FORGOT PASSWORD] Traceback: {traceback.format_exc()}")
        return Response({
            'success': False,
            'message': 'An error occurred while fetching user information'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    """
    Verify OTP entered by user
    """
    try:
        data = request.data
        Email = data.get('Email')
        otp = data.get('otp')
        
        if not Email or not otp:
            return Response({
                'success': False,
                'message': 'Email and OTP are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if OTP exists and is valid (try session first, then memory)
        stored_otp = request.session.get('reset_otp')
        stored_Email = request.session.get('reset_Email')
        stored_UserId = request.session.get('reset_UserId')
        otp_expiry = request.session.get('otp_expiry')
        
        # Debug logging
        logger.info(f"Verify OTP - Session data: stored_otp={stored_otp}, stored_Email={stored_Email}, stored_UserId={stored_UserId}, otp_expiry={otp_expiry}")
        logger.info(f"Verify OTP - Request data: Email={Email}, otp={otp}")
        logger.info(f"Verify OTP - Session ID: {request.session.session_key}")
        logger.info(f"Verify OTP - Session modified: {request.session.modified}")
        logger.info(f"Verify OTP - All session keys: {list(request.session.keys())}")
        logger.info(f"Verify OTP - Request cookies: {request.COOKIES}")
        logger.info(f"Verify OTP - Request headers: {dict(request.headers)}")
        
        # If session data is missing, try memory storage
        if not stored_otp or not stored_Email or not otp_expiry:
            logger.warning(f"Verify OTP - Missing session data, trying memory storage")
            memory_data = otp_storage.get(Email)
            if memory_data:
                stored_otp = memory_data['otp']
                stored_Email = Email
                stored_UserId = memory_data.get('user_id')
                otp_expiry = memory_data['expiry']
                logger.info(f"Verify OTP - Found OTP in memory: stored_otp={stored_otp}, user_id={stored_UserId}, expiry={otp_expiry}")
            else:
                logger.warning(f"Verify OTP - Missing session data: stored_otp={stored_otp}, stored_Email={stored_Email}, otp_expiry={otp_expiry}")
                return Response({
                    'success': False,
                    'message': 'No OTP request found. Please request a new OTP.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if OTP has expired
        if datetime.now().timestamp() > otp_expiry:
            # Clear expired OTP from session (only if they exist)
            session_keys_to_clear = ['reset_otp', 'reset_Email', 'reset_UserId', 'otp_expiry']
            for key in session_keys_to_clear:
                if key in request.session:
                    del request.session[key]
            request.session.save()
            
            # Clear expired OTP from memory
            if Email in otp_storage:
                del otp_storage[Email]
                logger.info(f"Verify OTP - Cleaned up expired OTP from memory for {Email}")
            
            return Response({
                'success': False,
                'message': 'OTP has expired. Please request a new OTP.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if Email matches
        if Email != stored_Email:
            return Response({
                'success': False,
                'message': 'Email does not match the OTP request.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify OTP
        if otp == stored_otp:
            # Mark OTP as verified
            request.session['otp_verified'] = True
            request.session['reset_UserId'] = stored_UserId  # Ensure UserId is preserved
            request.session.save()
            
            # Add to verified emails set with UserId mapping
            verified_emails.add(Email)
            if stored_UserId:
                # Store email -> UserId mapping
                verified_users_mapping[Email] = stored_UserId
            logger.info(f"Verify OTP - Added {Email} to verified emails set with UserId={stored_UserId}")
            
            # Clean up memory storage
            if Email in otp_storage:
                del otp_storage[Email]
                logger.info(f"Verify OTP - Cleaned up memory storage for {Email}")
            
            return Response({
                'success': True,
                'message': 'OTP verified successfully'
            })
        else:
            return Response({
                'success': False,
                'message': 'Invalid OTP. Please try again.'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error in verify_otp: {str(e)}")
        return Response({
            'success': False,
            'message': 'An error occurred while verifying OTP'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password(request):
    """
    Reset password after OTP verification
    """
    try:
        data = request.data
        Email = data.get('Email')
        new_password = data.get('new_password')
        
        if not Email or not new_password:
            return Response({
                'success': False,
                'message': 'Email and new password are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if OTP was verified (try session first, then memory)
        otp_verified = request.session.get('otp_verified', False)
        stored_Email = request.session.get('reset_Email')
        stored_UserId = request.session.get('reset_UserId')
        
        # Debug logging
        logger.info(f"Reset Password - Session data: otp_verified={otp_verified}, stored_Email={stored_Email}, stored_UserId={stored_UserId}")
        logger.info(f"Reset Password - Request data: Email={Email}")
        logger.info(f"Reset Password - Session ID: {request.session.session_key}")
        logger.info(f"Reset Password - All session keys: {list(request.session.keys())}")
        
        # If session data is missing, check if we have verified OTP in memory
        if not otp_verified or not stored_Email:
            logger.warning(f"Reset Password - Missing session data, checking verified emails set")
            # Check if this email was verified via the verified_emails set
            if Email in verified_emails:
                logger.info(f"Reset Password - Found {Email} in verified emails set")
                otp_verified = True
                stored_Email = Email
                # Get UserId from mapping
                stored_UserId = verified_users_mapping.get(Email)
                logger.info(f"Reset Password - Retrieved UserId from mapping: {stored_UserId}")
            else:
                logger.warning(f"Reset Password - Email {Email} not found in verified emails set")
                return Response({
                    'success': False,
                    'message': 'OTP verification required before password reset'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if not otp_verified:
            return Response({
                'success': False,
                'message': 'OTP verification required before password reset'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if Email != stored_Email:
            return Response({
                'success': False,
                'message': 'Email does not match the verified OTP request.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Find and update user password using UserId for accuracy
        try:
            # Prefer UserId lookup to avoid issues with duplicate emails
            if stored_UserId:
                try:
                    user = Users.objects.get(UserId=stored_UserId)
                    logger.info(f"Reset Password - Found user by UserId: {stored_UserId} ({user.UserName})")
                except Users.DoesNotExist:
                    logger.error(f"Reset Password - User not found with UserId: {stored_UserId}")
                    return Response({
                        'success': False,
                        'message': 'User not found'
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # Fallback to email lookup (for backward compatibility)
                logger.warning(f"Reset Password - No UserId available, falling back to email lookup")
                user = Users.find_by_email(Email)
                if not user:
                    return Response({
                        'success': False,
                        'message': 'No user found with this Email address'
                    }, status=status.HTTP_404_NOT_FOUND)
            
            # Check password history to prevent reuse
            from .routes.Global.password_expiry_utils import check_password_history, get_password_history_count
            is_reused, checked_count = check_password_history(user, new_password)
            if is_reused:
                history_count = get_password_history_count()
                logger.warning(f"⚠️ Password reuse blocked for user {user.UserName}: new password matches one of the last {history_count} passwords")
                return Response({
                    'success': False,
                    'message': f'Password has been used recently. Please choose a different password that is not one of your last {history_count} passwords.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Store old password hash for logging
            old_password_hash = user.Password
            # Always store password as a secure hash
            user.Password = make_password(new_password)
            user.save(update_fields=['Password'])
            
            # ========================================
            # CRITICAL: Clear account lockout cache after successful password reset
            # This allows the user to login immediately with the new password
            # ========================================
            try:
                from django.core.cache import cache
                # Clear lockout cache for both JWT and session login systems
                username_normalized = str(user.UserName).lower().strip()
                userid_normalized = str(user.UserId).lower().strip()
                
                # JWT login cache keys
                jwt_user_cache_key = f"login_failed_attempts_{username_normalized}"
                jwt_lockout_cache_key = f"login_locked_until_{username_normalized}"
                jwt_userid_cache_key = f"login_failed_attempts_{userid_normalized}"
                jwt_userid_lockout_cache_key = f"login_locked_until_{userid_normalized}"
                
                # Session login cache keys
                session_user_cache_key = f"session_login_failed_attempts_{username_normalized}"
                session_lockout_cache_key = f"session_login_locked_until_{username_normalized}"
                session_userid_cache_key = f"session_login_failed_attempts_{userid_normalized}"
                session_userid_lockout_cache_key = f"session_login_locked_until_{userid_normalized}"
                
                # Clear all cache entries
                cache.delete(jwt_user_cache_key)
                cache.delete(jwt_lockout_cache_key)
                cache.delete(jwt_userid_cache_key)
                cache.delete(jwt_userid_lockout_cache_key)
                cache.delete(session_user_cache_key)
                cache.delete(session_lockout_cache_key)
                cache.delete(session_userid_cache_key)
                cache.delete(session_userid_lockout_cache_key)
                
                logger.info(f"✅ Cleared account lockout cache for user {user.UserName} (ID: {user.UserId}) after password reset")
            except Exception as cache_error:
                logger.warning(f"⚠️ Failed to clear lockout cache for user {user.UserName}: {str(cache_error)}")
                # Don't fail password reset if cache clearing fails
            
            # Log password reset using utility function (to password_logs)
            try:
                from .routes.Global.password_expiry_utils import log_password_action
                log_password_action(
                    user, 
                    'reset', 
                    old_password_hash=old_password_hash,
                    new_password_hash=user.Password,
                    request=request
                )
                logger.info(f"✅ Password log created for reset: {user.UserName}")
            except Exception as log_error:
                logger.error(f"❌ Failed to create password log on reset: {str(log_error)}")
                # Don't fail password reset if logging fails
            
            # Also log password reset to grc_logs
            try:
                from .routes.Global.logging_service import send_log
                client_ip = request.META.get('REMOTE_ADDR', 'unknown')
                # Handle FrameworkId safely - Users model may not have FrameworkId field
                framework_id = None
                if hasattr(user, 'FrameworkId') and user.FrameworkId:
                    # If FrameworkId is a ForeignKey relationship
                    if hasattr(user.FrameworkId, 'FrameworkId'):
                        framework_id = user.FrameworkId.FrameworkId
                    # If FrameworkId is already an integer
                    elif isinstance(user.FrameworkId, int):
                        framework_id = user.FrameworkId
                
                send_log(
                    module='Authentication',
                    actionType='PASSWORD_RESET',
                    description=f'User {user.UserName} (ID: {user.UserId}) reset their password via forgot password flow',
                    userId=str(user.UserId),
                    userName=user.UserName,
                    logLevel='INFO',
                    ipAddress=client_ip,
                    additionalInfo={'reset_method': 'forgot_password', 'email': Email},
                    frameworkId=framework_id
                )
                logger.info(f"✅ Password reset logged to grc_logs for user: {user.UserName}")
            except Exception as log_error:
                logger.error(f"❌ Failed to log password reset to grc_logs: {str(log_error)}")
                # Don't fail password reset if logging fails
            
            # Clear session data (only if they exist)
            session_keys_to_clear = ['reset_otp', 'reset_Email', 'reset_UserId', 'otp_expiry', 'otp_verified']
            for key in session_keys_to_clear:
                if key in request.session:
                    del request.session[key]
            request.session.save()
            
            # Remove from verified emails set and mapping
            if Email in verified_emails:
                verified_emails.remove(Email)
                logger.info(f"Reset Password - Removed {Email} from verified emails set")
            if Email in verified_users_mapping:
                del verified_users_mapping[Email]
                logger.info(f"Reset Password - Removed {Email} from verified users mapping")
            
            logger.info(f"Password reset successful for user: {user.UserName}")
            
            return Response({
                'success': True,
                'message': 'Password reset successfully'
            })
            
        except Users.DoesNotExist:
            return Response({
                'success': False,
                'message': 'User not found'
            }, status=status.HTTP_404_NOT_FOUND)
            
    except Exception as e:
        logger.error(f"Error in reset_password: {str(e)}")
        return Response({
            'success': False,
            'message': 'An error occurred while resetting password'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def test_session_auth(request):
    """
    Simple test endpoint to verify session authentication is working
    """
    try:
        # Get user_id from session
        user_id = request.session.get('user_id')
        
        debug_info = {
            'session_available': hasattr(request, 'session'),
            'session_keys': list(request.session.keys()) if hasattr(request, 'session') else [],
            'user_id_in_session': user_id,
            'session_id': request.session.session_key if hasattr(request, 'session') else None,
            'cookies': dict(request.COOKIES),
            'headers': dict(request.headers)
        }
        
        if user_id:
            # Get user details
            try:
                user = Users.objects.get(UserId=user_id)
                debug_info['user_found'] = True
                debug_info['username'] = user.UserName
                debug_info['Email'] = user.Email
            except Users.DoesNotExist:
                debug_info['user_found'] = False
                debug_info['error'] = f'User with ID {user_id} not found'
        else:
            debug_info['user_found'] = False
            debug_info['error'] = 'No user_id in session'
        
        return Response({
            'success': True,
            'message': 'Session test completed',
            'debug_info': debug_info
        })
        
    except Exception as e:
        logger.error(f"Error in test_session_auth: {str(e)}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_rbac_roles(request):
    """
    Get all available RBAC roles for dropdown
    """
    try:
        from .models import RBAC
        
        # Get unique roles from RBAC model
        roles = RBAC.objects.values_list('role', flat=True).distinct()
        
        # Convert to list and add any missing default roles
        role_list = list(roles)
        
        # Add default roles if not present
        default_roles = [
            'GRC Administrator',
            'Compliance Manager', 
            'Compliance Officer',
            'Compliance Approver',
            'Executive/Senior Management',
            'Policy Manager',
            'Policy Approver',
            'Audit Manager',
            'Internal Auditor',
            'External Auditor',
            'Audit Reviewer',
            'Risk Manager',
            'Risk Analyst',
            'Risk Reviewer',
            'Incident Response Manager',
            'Incident Analyst',
            'Department Manager',
            'End User'
        ]
        
        # Add any missing default roles
        for role in default_roles:
            if role not in role_list:
                role_list.append(role)
        
        return Response({
            'success': True,
            'roles': sorted(role_list)
        })
        
    except Exception as e:
        logger.error(f"Error getting RBAC roles: {str(e)}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Removed duplicate get_departments function - using the one from policy.py instead

@api_view(['GET'])
@permission_classes([AllowAny])
def test_jwt_auth(request):
    """
    Test endpoint to verify JWT authentication is working
    """
    try:
        # Get user from request (set by middleware)
        user = getattr(request, 'user', None)
        
        if user and hasattr(user, 'UserId'):
            return Response({
                'success': True,
                'message': 'JWT authentication working',
                'user': {
                    'id': user.UserId,
                    'username': user.UserName,
                    'Email': user.email_plain,  # Use decrypted email
                    'firstName': user.FirstName,
                    'lastName': user.LastName
                },
                'auth_method': 'JWT' if request.META.get('HTTP_AUTHORIZATION') else 'Session'
            })
        else:
            return Response({
                'success': False,
                'message': 'No authenticated user found',
                'auth_method': 'None'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
    except Exception as e:
        logger.error(f"JWT test error: {str(e)}")
        return Response({
            'success': False,
            'message': 'Server error during JWT test',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





@api_view(['GET'])
@permission_classes([AllowAny])
@csrf_exempt
def serve_document(request, doc_type):
    """Serve EULA or Privacy Policy documents"""
    try:
        from django.http import FileResponse, HttpResponse
        import os
        from django.conf import settings
       
        # Define document paths
        document_paths = {
            'eula': 'END USER LICENSE AGREEMENT RiskaVaire.pdf',
            'privacy': 'Privacy and Security Policy1.pdf',
            'faqs': 'RiskaVaire FAQs.pdf',
            'admin-manual': 'RiskaVaire Admin User Manual.pdf',
            'non-admin-manual': 'RiskaVaire Non-admin user manual.pdf',
            'open-source-attribution': 'RiskaVaire Open-Source Attribution.pdf'
        }
       
        if doc_type not in document_paths:
            return JsonResponse({'error': 'Document not found'}, status=404)
       
        # Get the document path (documents are in frontend/src/assets/help_docs/)
        doc_path = os.path.join(settings.BASE_DIR.parent, 'frontend', 'src', 'assets', 'help_docs', document_paths[doc_type])
       
        if not os.path.exists(doc_path):
            return JsonResponse({'error': 'Document file not found'}, status=404)
       
        # Check if download is requested
        download = request.GET.get('download', 'false').lower() == 'true'
       
        if download:
            # Serve file for download
            response = FileResponse(
                open(doc_path, 'rb'),
                as_attachment=True,
                filename=document_paths[doc_type]
            )
            return response
        else:
            # Serve for viewing (try to extract content for DOCX files, serve PDFs directly)
            file_extension = os.path.splitext(doc_path)[1].lower()
           
            if file_extension == '.pdf':
                # For PDF files, serve directly for viewing
                response = FileResponse(
                    open(doc_path, 'rb'),
                    content_type='application/pdf'
                )
                return response
            elif file_extension == '.docx':
                # For DOCX files, try to extract content
                try:
                    import docx
                    doc = docx.Document(doc_path)
                    content = []
                   
                    for paragraph in doc.paragraphs:
                        if paragraph.text.strip():
                            content.append(paragraph.text.strip())
                   
                    # Return HTML content for viewing
                    html_content = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <title>{doc_type.upper()} Document</title>
                        <style>
                            body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }}
                            h1 {{ color: #1e3a8a; }}
                            p {{ line-height: 1.6; margin-bottom: 10px; }}
                        </style>
                    </head>
                    <body>
                        <h1>{doc_type.upper()} Document</h1>
                        {"<p>" + "</p><p>".join(content) + "</p>"}
                    </body>
                    </html>
                    """
                   
                    return HttpResponse(html_content, content_type='text/html')
                   
                except ImportError:
                    # If python-docx is not available, just serve the file
                    response = FileResponse(
                        open(doc_path, 'rb'),
                        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                    )
                    return response
            else:
                # For other file types, serve directly
                response = FileResponse(
                    open(doc_path, 'rb')
                )
                return response
               
    except Exception as e:
        logger.error(f"Error serving document {doc_type}: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)
