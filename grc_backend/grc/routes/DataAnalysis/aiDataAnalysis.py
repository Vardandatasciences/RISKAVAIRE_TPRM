from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from django.conf import settings
from django.http import HttpResponse
from grc.models import (
    Policy, Compliance, Audit, Incident, Risk, RiskInstance, Event,
    ConsentConfiguration, ConsentAcceptance, ConsentWithdrawal
)
import json
import requests
import logging
from datetime import datetime, timedelta
from collections import defaultdict
import io
import csv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)


def _call_openai_api(prompt, temperature=0.3, max_tokens=4000):
    """Call OpenAI API for AI processing"""
    api_key = getattr(settings, 'OPENAI_API_KEY', '')
    if not api_key or api_key == 'your-openai-api-key-here':
        raise Exception("OpenAI API key not configured. Please set OPENAI_API_KEY environment variable.")
    
    model_raw = getattr(settings, 'OPENAI_MODEL', 'gpt-4o-mini')
    model = str(model_raw).strip().strip('"').strip("'")
    timeout = getattr(settings, 'OPENAI_TIMEOUT', 120)
    
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json'
    }
    
    payload = {
        'model': model,
        'messages': [
            {
                'role': 'system',
                'content': 'You are an expert privacy and data protection analyst specializing in GDPR, CCPA, and other privacy regulations. You excel at analyzing data privacy maturity, data minimization practices, consent management, and providing actionable recommendations for privacy compliance. Always provide detailed, accurate analysis in valid JSON format.'
            },
            {'role': 'user', 'content': prompt}
        ],
        'temperature': temperature,
        'max_tokens': max_tokens,
        'response_format': {'type': 'json_object'} if 'gpt-4' in model.lower() and 'mini' not in model.lower() else None
    }
    
    try:
        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers=headers,
            json=payload,
            timeout=timeout
        )
        
        if response.status_code != 200:
            error_data = response.json() if response.content else {}
            error_obj = error_data.get('error', {})
            error_msg = error_obj.get('message', 'Unknown error') if isinstance(error_obj, dict) else str(error_obj)
            raise Exception(f"OpenAI API error {response.status_code}: {error_msg}")
        
        result = response.json()
        content = result['choices'][0]['message']['content']
        
        # Try to parse as JSON
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            # If not JSON, return as text
            return {'analysis': content}
            
    except requests.exceptions.Timeout:
        raise Exception("OpenAI API request timed out")
    except requests.exceptions.RequestException as e:
        raise Exception(f"OpenAI API request failed: {str(e)}")


def collect_privacy_data_across_modules(framework_id=None):
    """
    Centralized aggregation: Collect privacy data across all modules
    Returns comprehensive privacy data inventory
    """
    filter_query = Q()
    if framework_id and framework_id != 'all' and framework_id != 'null':
        try:
            framework_id = int(framework_id)
            filter_query = Q(FrameworkId=framework_id)
        except (ValueError, TypeError):
            pass
    
    privacy_data = {
        'modules': {},
        'total_fields': 0,
        'total_records': 0,
        'personal_data_count': 0,
        'regular_data_count': 0,
        'confidential_data_count': 0,
        'modules_with_inventory': []
    }
    
    # Module configurations
    modules_config = [
        ('policy', Policy),
        ('compliance', Compliance),
        ('audit', Audit),
        ('incident', Incident),
        ('risk', Risk),
        ('risk_instance', RiskInstance),
        ('event', Event)
    ]
    
    for module_name, model_class in modules_config:
        queryset = model_class.objects.filter(filter_query)
        module_data = {
            'name': module_name,
            'total_records': queryset.count(),
            'total_fields': 0,
            'personal_count': 0,
            'regular_count': 0,
            'confidential_count': 0,
            'records_with_inventory': 0,
            'sample_fields': [],
            'all_fields': {}  # Store all fields with their classifications for miscategorization analysis
        }
        
        for record in queryset:
            data_inventory = getattr(record, 'data_inventory', None)
            if data_inventory:
                if isinstance(data_inventory, str):
                    try:
                        data_inventory = json.loads(data_inventory)
                    except json.JSONDecodeError:
                        continue
                
                if isinstance(data_inventory, dict):
                    module_data['records_with_inventory'] += 1
                    module_data['total_fields'] += len(data_inventory)
                    privacy_data['total_fields'] += len(data_inventory)
                    
                    # Collect ALL fields with their classifications
                    for field_name, field_type in data_inventory.items():
                        if isinstance(field_type, str):
                            field_type_lower = field_type.lower()
                            # Store field with its classification
                            if field_name not in module_data['all_fields']:
                                module_data['all_fields'][field_name] = {
                                    'classification': field_type_lower,
                                    'count': 0
                                }
                            module_data['all_fields'][field_name]['count'] += 1
                            
                            # Collect sample fields (first 20 unique)
                            if len(module_data['sample_fields']) < 20:
                                if not any(f['name'] == field_name for f in module_data['sample_fields']):
                                    module_data['sample_fields'].append({
                                        'name': field_name,
                                        'type': field_type_lower
                                    })
                            
                            # Count by type
                            if field_type_lower == 'personal':
                                module_data['personal_count'] += 1
                                privacy_data['personal_data_count'] += 1
                            elif field_type_lower == 'regular':
                                module_data['regular_count'] += 1
                                privacy_data['regular_data_count'] += 1
                            elif field_type_lower == 'confidential':
                                module_data['confidential_count'] += 1
                                privacy_data['confidential_data_count'] += 1
        
        privacy_data['total_records'] += module_data['total_records']
        
        if module_data['records_with_inventory'] > 0:
            privacy_data['modules_with_inventory'].append(module_name)
            # Calculate percentages
            total_module_fields = module_data['personal_count'] + module_data['regular_count'] + module_data['confidential_count']
            if total_module_fields > 0:
                module_data['personal_percentage'] = round((module_data['personal_count'] / total_module_fields) * 100, 2)
                module_data['regular_percentage'] = round((module_data['regular_count'] / total_module_fields) * 100, 2)
                module_data['confidential_percentage'] = round((module_data['confidential_count'] / total_module_fields) * 100, 2)
            else:
                module_data['personal_percentage'] = 0
                module_data['regular_percentage'] = 0
                module_data['confidential_percentage'] = 0
        
        privacy_data['modules'][module_name] = module_data
    
    # Calculate overall percentages
    total_all_fields = privacy_data['personal_data_count'] + privacy_data['regular_data_count'] + privacy_data['confidential_data_count']
    if total_all_fields > 0:
        privacy_data['overall_personal_percentage'] = round((privacy_data['personal_data_count'] / total_all_fields) * 100, 2)
        privacy_data['overall_regular_percentage'] = round((privacy_data['regular_data_count'] / total_all_fields) * 100, 2)
        privacy_data['overall_confidential_percentage'] = round((privacy_data['confidential_data_count'] / total_all_fields) * 100, 2)
    else:
        privacy_data['overall_personal_percentage'] = 0
        privacy_data['overall_regular_percentage'] = 0
        privacy_data['overall_confidential_percentage'] = 0
    
    return privacy_data


def calculate_privacy_metrics(privacy_data):
    """
    Calculate privacy metrics: Maturity score, Minimization score, Consent rates
    """
    metrics = {
        'maturity_score': 0,
        'minimization_score': 0,
        'consent_rate': 0,
        'data_inventory_coverage': 0,
        'module_scores': {}
    }
    
    total_records = privacy_data['total_records']
    total_fields = privacy_data['total_fields']
    modules_with_inventory = len(privacy_data['modules_with_inventory'])
    total_modules = len(privacy_data['modules'])
    
    # Data Inventory Coverage Score (0-100)
    # Based on how many modules have data inventory configured
    if total_modules > 0:
        metrics['data_inventory_coverage'] = round((modules_with_inventory / total_modules) * 100, 2)
    
    # Maturity Score Calculation (0-100)
    # Factors:
    # 1. Data inventory coverage (40%)
    # 2. Proper classification of personal/confidential data (30%)
    # 3. Module coverage (30%)
    maturity_factors = []
    
    # Factor 1: Data inventory coverage
    coverage_score = metrics['data_inventory_coverage']
    maturity_factors.append(('coverage', coverage_score, 0.4))
    
    # Factor 2: Proper classification (if personal/confidential data is identified)
    total_classified = privacy_data['personal_data_count'] + privacy_data['confidential_data_count']
    if total_fields > 0:
        classification_score = (total_classified / total_fields) * 100
    else:
        classification_score = 0
    maturity_factors.append(('classification', classification_score, 0.3))
    
    # Factor 3: Module coverage (how many modules have data inventory)
    module_coverage_score = (modules_with_inventory / total_modules * 100) if total_modules > 0 else 0
    maturity_factors.append(('module_coverage', module_coverage_score, 0.3))
    
    # Calculate weighted maturity score
    maturity_score = sum(score * weight for _, score, weight in maturity_factors)
    metrics['maturity_score'] = round(maturity_score, 2)
    
    # Minimization Score Calculation (0-100)
    # Higher score = better minimization (less personal/confidential data relative to regular data)
    total_sensitive = privacy_data['personal_data_count'] + privacy_data['confidential_data_count']
    if total_fields > 0:
        sensitive_percentage = (total_sensitive / total_fields) * 100
        # Invert: lower sensitive data percentage = higher minimization score
        # But we want to reward having less sensitive data, so:
        # If 0% sensitive = 100 score, if 100% sensitive = 0 score
        minimization_score = max(0, 100 - sensitive_percentage)
    else:
        minimization_score = 0
    metrics['minimization_score'] = round(minimization_score, 2)
    
    # Calculate per-module scores
    for module_name, module_data in privacy_data['modules'].items():
        if module_data['records_with_inventory'] > 0:
            module_total = module_data['personal_count'] + module_data['regular_count'] + module_data['confidential_count']
            if module_total > 0:
                module_sensitive = module_data['personal_count'] + module_data['confidential_count']
                module_minimization = max(0, 100 - ((module_sensitive / module_total) * 100))
                module_coverage = (module_data['records_with_inventory'] / module_data['total_records'] * 100) if module_data['total_records'] > 0 else 0
                module_maturity = (module_coverage * 0.5) + (module_minimization * 0.5)
            else:
                module_minimization = 0
                module_maturity = 0
                module_coverage = 0
        else:
            module_minimization = 0
            module_maturity = 0
            module_coverage = 0
        
        metrics['module_scores'][module_name] = {
            'maturity': round(module_maturity, 2),
            'minimization': round(module_minimization, 2),
            'coverage': round(module_coverage, 2)
        }
    
    return metrics


def get_consent_rates(framework_id=None):
    """
    Calculate consent rates from consent management data
    """
    filter_query = Q()
    if framework_id and framework_id != 'all' and framework_id != 'null':
        try:
            framework_id = int(framework_id)
            filter_query = Q(FrameworkId=framework_id)
        except (ValueError, TypeError):
            pass
    
    # Get consent configurations
    consent_configs = ConsentConfiguration.objects.filter(filter_query)
    total_configs = consent_configs.count()
    enabled_configs = consent_configs.filter(is_enabled=True).count()
    
    # Get consent acceptances
    consent_acceptances = ConsentAcceptance.objects.filter(filter_query)
    total_acceptances = consent_acceptances.count()
    
    # Get consent withdrawals
    consent_withdrawals = ConsentWithdrawal.objects.filter(filter_query)
    total_withdrawals = consent_withdrawals.count()
    
    # Calculate rates
    consent_rate = 0
    if total_configs > 0:
        consent_rate = round((enabled_configs / total_configs) * 100, 2)
    
    withdrawal_rate = 0
    if total_acceptances > 0:
        withdrawal_rate = round((total_withdrawals / total_acceptances) * 100, 2)
    
    return {
        'consent_configuration_rate': consent_rate,
        'total_configurations': total_configs,
        'enabled_configurations': enabled_configs,
        'total_acceptances': total_acceptances,
        'total_withdrawals': total_withdrawals,
        'withdrawal_rate': withdrawal_rate
    }


def generate_ai_insights(privacy_data, metrics, consent_data):
    """
    Use OpenAI to generate automated insights and recommendations
    Includes field-level miscategorization analysis
    """
    # Prepare data summary for AI
    data_summary = {
        'total_modules': len(privacy_data['modules']),
        'modules_with_inventory': len(privacy_data['modules_with_inventory']),
        'total_records': privacy_data['total_records'],
        'total_fields': privacy_data['total_fields'],
        'personal_data_percentage': privacy_data['overall_personal_percentage'],
        'confidential_data_percentage': privacy_data['overall_confidential_percentage'],
        'maturity_score': metrics['maturity_score'],
        'minimization_score': metrics['minimization_score'],
        'data_inventory_coverage': metrics['data_inventory_coverage'],
        'consent_rate': consent_data['consent_configuration_rate'],
        'module_breakdown': {},
        'field_classifications': {}  # Detailed field-level data for miscategorization analysis
    }
    
    # Add module breakdown and field-level data
    for module_name, module_data in privacy_data['modules'].items():
        if module_data['records_with_inventory'] > 0:
            data_summary['module_breakdown'][module_name] = {
                'records': module_data['total_records'],
                'fields': module_data['total_fields'],
                'personal_percentage': module_data.get('personal_percentage', 0),
                'confidential_percentage': module_data.get('confidential_percentage', 0),
                'maturity': metrics['module_scores'].get(module_name, {}).get('maturity', 0),
                'minimization': metrics['module_scores'].get(module_name, {}).get('minimization', 0)
            }
            
            # Add field-level classifications for miscategorization analysis
            if 'all_fields' in module_data and module_data['all_fields']:
                # Limit to top 50 fields per module to avoid token limits and improve performance
                sorted_fields = sorted(
                    module_data['all_fields'].items(),
                    key=lambda x: x[1]['count'],
                    reverse=True
                )[:50]
                
                data_summary['field_classifications'][module_name] = {
                    field_name: {
                        'classification': field_info['classification'],
                        'count': field_info['count']
                    }
                    for field_name, field_info in sorted_fields
                }
    
    prompt = f"""Analyze the following privacy and data protection metrics for a GRC (Governance, Risk & Compliance) system and provide comprehensive insights and recommendations.

IMPORTANT: You must analyze the field_classifications data to identify potential miscategorizations of data fields. Review each field name and its current classification (personal, regular, or confidential) and determine if the classification is appropriate based on:
- Field name patterns (e.g., fields containing "email", "phone", "ssn", "address", "name" are typically personal data)
- Industry standards and privacy regulations (GDPR, CCPA)
- Context of the module where the field appears

DATA SUMMARY:
{json.dumps(data_summary, indent=2)}

Please provide a JSON response with the following structure:
{{
    "executive_summary": "A brief 2-3 sentence overview of the privacy posture",
    "key_findings": [
        "Finding 1",
        "Finding 2",
        "Finding 3"
    ],
    "strengths": [
        "Strength 1",
        "Strength 2"
    ],
    "weaknesses": [
        "Weakness 1",
        "Weakness 2"
    ],
    "recommendations": [
        {{
            "priority": "high|medium|low",
            "category": "data_inventory|data_minimization|consent_management|governance",
            "title": "Recommendation title",
            "description": "Detailed recommendation description",
            "impact": "Expected impact of implementing this recommendation"
        }}
    ],
    "miscategorizations": [
        {{
            "module": "module_name",
            "field_name": "field_name",
            "current_classification": "personal|regular|confidential",
            "suggested_classification": "personal|regular|confidential",
            "confidence": "high|medium|low",
            "reason": "Detailed explanation of why this field may be miscategorized",
            "risk_level": "low|medium|high|critical",
            "recommendation": "Specific action to take"
        }}
    ],
    "risk_assessment": {{
        "overall_risk_level": "low|medium|high|critical",
        "risk_factors": [
            "Risk factor 1",
            "Risk factor 2"
        ],
        "compliance_gaps": [
            "Gap 1",
            "Gap 2"
        ]
    }},
    "maturity_level": "initial|developing|defined|managed|optimizing",
    "next_steps": [
        "Step 1",
        "Step 2",
        "Step 3"
    ]
}}

Focus on:
1. Data privacy maturity assessment
2. Data minimization practices
3. Consent management effectiveness
4. Regulatory compliance (GDPR, CCPA, etc.)
5. Field-level miscategorization analysis (CRITICAL - analyze all fields in field_classifications)
6. Actionable recommendations prioritized by impact

For miscategorizations, be thorough and identify:
- Fields that should be "personal" but are classified as "regular" or "confidential"
- Fields that should be "confidential" but are classified as "regular" or "personal"
- Fields that are over-classified (e.g., regular data marked as personal)
- Fields with ambiguous names that need review
"""
    
    try:
        ai_response = _call_openai_api(prompt, temperature=0.3, max_tokens=5000)
        # Ensure miscategorizations array exists
        if 'miscategorizations' not in ai_response:
            ai_response['miscategorizations'] = []
        return ai_response
    except Exception as e:
        logger.error(f"Error generating AI insights: {str(e)}")
        return {
            'error': f"Failed to generate AI insights: {str(e)}",
            'executive_summary': 'AI analysis temporarily unavailable',
            'recommendations': [],
            'miscategorizations': []
        }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_ai_privacy_analysis(request):
    """
    Get AI-powered privacy analysis across all modules.
    Returns:
    - Centralized privacy data aggregation
    - Privacy metrics (maturity score, minimization score, consent rates)
    - Cross-module analysis
    - Automated AI insights and recommendations
    """
    try:
        framework_id = request.query_params.get('framework_id', None)
        
        # Step 1: Collect privacy data across modules
        logger.info("Collecting privacy data across modules...")
        privacy_data = collect_privacy_data_across_modules(framework_id)
        
        # Step 2: Calculate privacy metrics
        logger.info("Calculating privacy metrics...")
        metrics = calculate_privacy_metrics(privacy_data)
        
        # Step 3: Get consent rates
        logger.info("Calculating consent rates...")
        consent_data = get_consent_rates(framework_id)
        
        # Step 4: Generate AI insights (with timeout handling)
        logger.info("Generating AI insights...")
        try:
            ai_insights = generate_ai_insights(privacy_data, metrics, consent_data)
        except Exception as ai_error:
            logger.error(f"AI insights generation failed: {str(ai_error)}")
            # Return partial data even if AI fails
            ai_insights = {
                'error': f"AI analysis failed: {str(ai_error)}",
                'executive_summary': 'AI analysis temporarily unavailable. Basic metrics are available below.',
                'recommendations': [],
                'miscategorizations': []
            }
        
        # Compile comprehensive response
        response_data = {
            'privacy_data': privacy_data,
            'metrics': metrics,
            'consent_data': consent_data,
            'ai_insights': ai_insights,
            'generated_at': datetime.now().isoformat(),
            'framework_id': framework_id
        }
        
        return Response({
            'status': 'success',
            'data': response_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in AI privacy analysis: {str(e)}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_privacy_dashboard_metrics(request):
    """
    Get privacy dashboard metrics for visualization
    Returns simplified metrics for dashboard display
    """
    try:
        framework_id = request.query_params.get('framework_id', None)
        
        privacy_data = collect_privacy_data_across_modules(framework_id)
        metrics = calculate_privacy_metrics(privacy_data)
        consent_data = get_consent_rates(framework_id)
        
        # Prepare dashboard-friendly response
        dashboard_data = {
            'overall_scores': {
                'maturity_score': metrics['maturity_score'],
                'minimization_score': metrics['minimization_score'],
                'data_inventory_coverage': metrics['data_inventory_coverage'],
                'consent_rate': consent_data['consent_configuration_rate']
            },
            'data_distribution': {
                'personal': {
                    'count': privacy_data['personal_data_count'],
                    'percentage': privacy_data['overall_personal_percentage']
                },
                'regular': {
                    'count': privacy_data['regular_data_count'],
                    'percentage': privacy_data['overall_regular_percentage']
                },
                'confidential': {
                    'count': privacy_data['confidential_data_count'],
                    'percentage': privacy_data['overall_confidential_percentage']
                }
            },
            'module_breakdown': {},
            'trends': {
                'modules_analyzed': len(privacy_data['modules_with_inventory']),
                'total_modules': len(privacy_data['modules']),
                'total_records': privacy_data['total_records'],
                'total_fields': privacy_data['total_fields']
            }
        }
        
        # Add module breakdown
        for module_name, module_data in privacy_data['modules'].items():
            if module_name in privacy_data['modules_with_inventory']:
                module_scores = metrics['module_scores'].get(module_name, {})
                dashboard_data['module_breakdown'][module_name] = {
                    'maturity': module_scores.get('maturity', 0),
                    'minimization': module_scores.get('minimization', 0),
                    'coverage': module_scores.get('coverage', 0),
                    'personal_percentage': module_data.get('personal_percentage', 0),
                    'confidential_percentage': module_data.get('confidential_percentage', 0),
                    'total_records': module_data['total_records'],
                    'total_fields': module_data['total_fields']
                }
        
        return Response({
            'status': 'success',
            'data': dashboard_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in privacy dashboard metrics: {str(e)}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_privacy_compliance_report(request):
    """
    Generate privacy compliance report with AI insights
    """
    try:
        framework_id = request.query_params.get('framework_id', None)
        include_ai = request.query_params.get('include_ai', 'true').lower() == 'true'
        
        privacy_data = collect_privacy_data_across_modules(framework_id)
        metrics = calculate_privacy_metrics(privacy_data)
        consent_data = get_consent_rates(framework_id)
        
        report = {
            'report_metadata': {
                'generated_at': datetime.now().isoformat(),
                'framework_id': framework_id,
                'report_type': 'privacy_compliance'
            },
            'executive_summary': {
                'maturity_score': metrics['maturity_score'],
                'minimization_score': metrics['minimization_score'],
                'overall_assessment': 'compliant' if metrics['maturity_score'] >= 70 else 'needs_improvement'
            },
            'privacy_metrics': metrics,
            'consent_management': consent_data,
            'module_analysis': {}
        }
        
        # Add module analysis
        for module_name, module_data in privacy_data['modules'].items():
            if module_name in privacy_data['modules_with_inventory']:
                module_scores = metrics['module_scores'].get(module_name, {})
                report['module_analysis'][module_name] = {
                    'metrics': module_scores,
                    'data_inventory': {
                        'total_fields': module_data['total_fields'],
                        'personal_count': module_data['personal_count'],
                        'confidential_count': module_data['confidential_count'],
                        'regular_count': module_data['regular_count']
                    },
                    'records': {
                        'total': module_data['total_records'],
                        'with_inventory': module_data['records_with_inventory']
                    }
                }
        
        # Add AI insights if requested
        if include_ai:
            try:
                ai_insights = generate_ai_insights(privacy_data, metrics, consent_data)
                report['ai_insights'] = ai_insights
            except Exception as e:
                logger.error(f"Error generating AI insights for report: {str(e)}")
                report['ai_insights'] = {'error': 'AI insights unavailable'}
        
        return Response({
            'status': 'success',
            'data': report
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error generating privacy compliance report: {str(e)}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def generate_excel_report(report_data):
    """Generate Excel report from privacy compliance data"""
    if not OPENPYXL_AVAILABLE:
        raise Exception("openpyxl library not installed. Install with: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Privacy Compliance Report"
    
    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Privacy Compliance Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Report metadata
    row = 3
    ws[f'A{row}'] = "Generated At:"
    ws[f'B{row}'] = report_data.get('report_metadata', {}).get('generated_at', datetime.now().isoformat())
    row += 1
    ws[f'A{row}'] = "Framework ID:"
    ws[f'B{row}'] = report_data.get('report_metadata', {}).get('framework_id', 'All')
    
    row += 2
    
    # Executive Summary
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Executive Summary"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    exec_summary = report_data.get('executive_summary', {})
    ws[f'A{row}'] = "Maturity Score:"
    ws[f'B{row}'] = exec_summary.get('maturity_score', 0)
    row += 1
    ws[f'A{row}'] = "Minimization Score:"
    ws[f'B{row}'] = exec_summary.get('minimization_score', 0)
    row += 1
    ws[f'A{row}'] = "Overall Assessment:"
    ws[f'B{row}'] = exec_summary.get('overall_assessment', 'N/A')
    
    row += 2
    
    # Privacy Metrics
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Privacy Metrics"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    metrics = report_data.get('privacy_metrics', {})
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    ws[f'A{row}'] = "Maturity Score"
    ws[f'B{row}'] = metrics.get('maturity_score', 0)
    row += 1
    ws[f'A{row}'] = "Minimization Score"
    ws[f'B{row}'] = metrics.get('minimization_score', 0)
    row += 1
    ws[f'A{row}'] = "Data Inventory Coverage"
    ws[f'B{row}'] = f"{metrics.get('data_inventory_coverage', 0)}%"
    
    row += 2
    
    # Module Analysis
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Module-Level Analysis"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    # Module headers
    headers = ['Module', 'Maturity', 'Minimization', 'Coverage', 'Personal Data %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    # Module data
    module_analysis = report_data.get('module_analysis', {})
    for module_name, module_data in module_analysis.items():
        metrics_data = module_data.get('metrics', {})
        ws[f'A{row}'] = module_name.replace('_', ' ').title()
        ws[f'B{row}'] = metrics_data.get('maturity', 0)
        ws[f'C{row}'] = metrics_data.get('minimization', 0)
        ws[f'D{row}'] = f"{metrics_data.get('coverage', 0)}%"
        data_inv = module_data.get('data_inventory', {})
        total = data_inv.get('personal_count', 0) + data_inv.get('regular_count', 0) + data_inv.get('confidential_count', 0)
        if total > 0:
            personal_pct = (data_inv.get('personal_count', 0) / total) * 100
        else:
            personal_pct = 0
        ws[f'E{row}'] = f"{personal_pct:.2f}%"
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(report_data):
    """Generate PDF report from privacy compliance data"""
    if not REPORTLAB_AVAILABLE:
        raise Exception("reportlab library not installed. Install with: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a202c'),
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("Privacy Compliance Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Report metadata
    metadata = report_data.get('report_metadata', {})
    story.append(Paragraph(f"<b>Generated At:</b> {metadata.get('generated_at', datetime.now().isoformat())}", styles['Normal']))
    story.append(Paragraph(f"<b>Framework ID:</b> {metadata.get('framework_id', 'All')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Executive Summary
    story.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
    exec_summary = report_data.get('executive_summary', {})
    story.append(Paragraph(f"<b>Maturity Score:</b> {exec_summary.get('maturity_score', 0)}", styles['Normal']))
    story.append(Paragraph(f"<b>Minimization Score:</b> {exec_summary.get('minimization_score', 0)}", styles['Normal']))
    story.append(Paragraph(f"<b>Overall Assessment:</b> {exec_summary.get('overall_assessment', 'N/A')}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Privacy Metrics
    story.append(Paragraph("<b>Privacy Metrics</b>", styles['Heading2']))
    metrics = report_data.get('privacy_metrics', {})
    metrics_data = [
        ['Metric', 'Value'],
        ['Maturity Score', str(metrics.get('maturity_score', 0))],
        ['Minimization Score', str(metrics.get('minimization_score', 0))],
        ['Data Inventory Coverage', f"{metrics.get('data_inventory_coverage', 0)}%"]
    ]
    metrics_table = Table(metrics_data)
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Module Analysis
    story.append(Paragraph("<b>Module-Level Analysis</b>", styles['Heading2']))
    module_analysis = report_data.get('module_analysis', {})
    
    module_data = [['Module', 'Maturity', 'Minimization', 'Coverage']]
    for module_name, module_info in module_analysis.items():
        metrics_info = module_info.get('metrics', {})
        module_data.append([
            module_name.replace('_', ' ').title(),
            str(metrics_info.get('maturity', 0)),
            str(metrics_info.get('minimization', 0)),
            f"{metrics_info.get('coverage', 0)}%"
        ])
    
    module_table = Table(module_data)
    module_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(module_table)
    
    # AI Insights summary
    ai_insights = report_data.get('ai_insights', {})
    if ai_insights and not ai_insights.get('error'):
        story.append(PageBreak())
        story.append(Paragraph("<b>AI-Generated Insights</b>", styles['Heading2']))
        
        if ai_insights.get('executive_summary'):
            story.append(Paragraph(f"<b>Summary:</b> {ai_insights['executive_summary']}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        if ai_insights.get('recommendations'):
            story.append(Paragraph("<b>Key Recommendations:</b>", styles['Heading3']))
            for rec in ai_insights['recommendations'][:5]:  # Top 5
                story.append(Paragraph(f"• <b>{rec.get('title', 'N/A')}</b> ({rec.get('priority', 'N/A')} priority)", styles['Normal']))
                story.append(Paragraph(f"  {rec.get('description', '')}", styles['Normal']))
                story.append(Spacer(1, 0.1*inch))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_module_ai_analysis(module_name, module_data, module_metrics, framework_id=None):
    """
    Generate AI-powered analysis for a specific module.
    Returns recommendations to improve scores and detects miscategorized columns.
    """
    # Prepare module-specific data for AI
    module_summary = {
        'module_name': module_name,
        'total_records': module_data.get('total_records', 0),
        'total_fields': module_data.get('total_fields', 0),
        'records_with_inventory': module_data.get('records_with_inventory', 0),
        'personal_count': module_data.get('personal_count', 0),
        'regular_count': module_data.get('regular_count', 0),
        'confidential_count': module_data.get('confidential_count', 0),
        'personal_percentage': module_data.get('personal_percentage', 0),
        'regular_percentage': module_data.get('regular_percentage', 0),
        'confidential_percentage': module_data.get('confidential_percentage', 0),
        'maturity_score': module_metrics.get('maturity', 0),
        'minimization_score': module_metrics.get('minimization', 0),
        'coverage_score': module_metrics.get('coverage', 0),
        'field_classifications': {}
    }
    
    # Add all field classifications for miscategorization analysis
    if 'all_fields' in module_data and module_data['all_fields']:
        # Include all fields, not just top 50, for thorough analysis
        module_summary['field_classifications'] = {
            field_name: {
                'classification': field_info['classification'],
                'count': field_info['count']
            }
            for field_name, field_info in module_data['all_fields'].items()
        }
    
    prompt = f"""Analyze the following module-level privacy and data protection metrics for a GRC (Governance, Risk & Compliance) system.

MODULE DATA:
{json.dumps(module_summary, indent=2)}

Your task is to:
1. Provide specific, actionable recommendations to improve the module's maturity score, minimization score, and coverage score
2. Identify miscategorized data fields by analyzing field names and their current classifications
3. Prioritize recommendations by impact and feasibility

For field miscategorization analysis, carefully review each field name in field_classifications and determine if its classification (personal, regular, or confidential) is appropriate based on:
- Field name patterns (e.g., fields containing "email", "phone", "ssn", "address", "name", "id", "dob", "birth", "salary", "credit", "card" are typically personal data)
- Fields containing "secret", "password", "key", "token", "api", "auth" are typically confidential
- Industry standards and privacy regulations (GDPR, CCPA, HIPAA)
- Context of the module where the field appears

Please provide a JSON response with the following structure:
{{
    "recommendations": [
        {{
            "priority": "high|medium|low",
            "category": "maturity|minimization|coverage|data_inventory|governance",
            "title": "Specific recommendation title",
            "description": "Detailed recommendation description with actionable steps",
            "impact": "Expected impact on score (e.g., 'Can improve maturity score by 10-15 points')",
            "feasibility": "high|medium|low",
            "estimated_score_improvement": {{
                "maturity": 0-20,
                "minimization": 0-20,
                "coverage": 0-20
            }}
        }}
    ],
    "miscategorizations": [
        {{
            "field_name": "field_name",
            "current_classification": "personal|regular|confidential",
            "suggested_classification": "personal|regular|confidential",
            "confidence": "high|medium|low",
            "reason": "Detailed explanation of why this field may be miscategorized based on field name patterns and privacy regulations",
            "risk_level": "low|medium|high|critical",
            "recommendation": "Specific action to take (e.g., 'Reclassify this field as personal data to comply with GDPR')"
        }}
    ],
    "score_analysis": {{
        "maturity_analysis": "Detailed analysis of why maturity score is at current level and how to improve it",
        "minimization_analysis": "Detailed analysis of minimization score and recommendations",
        "coverage_analysis": "Analysis of data inventory coverage and how to improve it"
    }},
    "priority_actions": [
        "Top 3 most important actions to take immediately"
    ]
}}

Focus on:
1. Specific, actionable recommendations that can directly improve scores
2. Field-level miscategorization detection (CRITICAL - analyze ALL fields in field_classifications)
3. Prioritization by impact and compliance risk
4. Clear explanations of why fields may be miscategorized
"""
    
    try:
        ai_response = _call_openai_api(prompt, temperature=0.3, max_tokens=4000)
        
        # Ensure required fields exist
        if 'recommendations' not in ai_response:
            ai_response['recommendations'] = []
        if 'miscategorizations' not in ai_response:
            ai_response['miscategorizations'] = []
        if 'score_analysis' not in ai_response:
            ai_response['score_analysis'] = {
                'maturity_analysis': 'Analysis unavailable',
                'minimization_analysis': 'Analysis unavailable',
                'coverage_analysis': 'Analysis unavailable'
            }
        if 'priority_actions' not in ai_response:
            ai_response['priority_actions'] = []
        
        return ai_response
    except Exception as e:
        logger.error(f"Error generating module AI analysis: {str(e)}")
        return {
            'error': f"Failed to generate AI analysis: {str(e)}",
            'recommendations': [],
            'miscategorizations': [],
            'score_analysis': {
                'maturity_analysis': 'AI analysis unavailable',
                'minimization_analysis': 'AI analysis unavailable',
                'coverage_analysis': 'AI analysis unavailable'
            },
            'priority_actions': []
        }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_module_ai_analysis(request):
    """
    Get AI-powered analysis for a specific module.
    Returns:
    - Recommendations to improve maturity, minimization, and coverage scores
    - Detected miscategorized columns with suggested reclassifications
    - Score analysis and priority actions
    """
    try:
        module_name = request.query_params.get('module_name', None)
        framework_id = request.query_params.get('framework_id', None)
        
        if not module_name:
            return Response({
                'status': 'error',
                'message': 'module_name parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Normalize module name
        module_name = module_name.lower().strip()
        
        # Map module names to model classes
        module_map = {
            'policy': Policy,
            'compliance': Compliance,
            'audit': Audit,
            'incident': Incident,
            'risk': Risk,
            'risk_instance': RiskInstance,
            'event': Event
        }
        
        if module_name not in module_map:
            return Response({
                'status': 'error',
                'message': f'Invalid module name. Valid modules: {", ".join(module_map.keys())}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Build filter query
        filter_query = Q()
        if framework_id and framework_id != 'all' and framework_id != 'null':
            try:
                framework_id = int(framework_id)
                filter_query = Q(FrameworkId=framework_id)
            except (ValueError, TypeError):
                pass
        
        # Collect module-specific data
        model_class = module_map[module_name]
        queryset = model_class.objects.filter(filter_query)
        
        module_data = {
            'name': module_name,
            'total_records': queryset.count(),
            'total_fields': 0,
            'personal_count': 0,
            'regular_count': 0,
            'confidential_count': 0,
            'records_with_inventory': 0,
            'all_fields': {}
        }
        
        # Process data inventory
        for record in queryset:
            data_inventory = getattr(record, 'data_inventory', None)
            if data_inventory:
                if isinstance(data_inventory, str):
                    try:
                        data_inventory = json.loads(data_inventory)
                    except json.JSONDecodeError:
                        continue
                
                if isinstance(data_inventory, dict):
                    module_data['records_with_inventory'] += 1
                    module_data['total_fields'] += len(data_inventory)
                    
                    for field_name, field_type in data_inventory.items():
                        if isinstance(field_type, str):
                            field_type_lower = field_type.lower()
                            
                            # Store field with its classification
                            if field_name not in module_data['all_fields']:
                                module_data['all_fields'][field_name] = {
                                    'classification': field_type_lower,
                                    'count': 0
                                }
                            module_data['all_fields'][field_name]['count'] += 1
                            
                            # Count by type
                            if field_type_lower == 'personal':
                                module_data['personal_count'] += 1
                            elif field_type_lower == 'regular':
                                module_data['regular_count'] += 1
                            elif field_type_lower == 'confidential':
                                module_data['confidential_count'] += 1
        
        # Calculate percentages
        total_module_fields = module_data['personal_count'] + module_data['regular_count'] + module_data['confidential_count']
        if total_module_fields > 0:
            module_data['personal_percentage'] = round((module_data['personal_count'] / total_module_fields) * 100, 2)
            module_data['regular_percentage'] = round((module_data['regular_count'] / total_module_fields) * 100, 2)
            module_data['confidential_percentage'] = round((module_data['confidential_count'] / total_module_fields) * 100, 2)
        else:
            module_data['personal_percentage'] = 0
            module_data['regular_percentage'] = 0
            module_data['confidential_percentage'] = 0
        
        # Calculate module metrics
        if module_data['records_with_inventory'] > 0:
            module_total = module_data['personal_count'] + module_data['regular_count'] + module_data['confidential_count']
            if module_total > 0:
                module_sensitive = module_data['personal_count'] + module_data['confidential_count']
                module_minimization = max(0, 100 - ((module_sensitive / module_total) * 100))
                module_coverage = (module_data['records_with_inventory'] / module_data['total_records'] * 100) if module_data['total_records'] > 0 else 0
                module_maturity = (module_coverage * 0.5) + (module_minimization * 0.5)
            else:
                module_minimization = 0
                module_maturity = 0
                module_coverage = 0
        else:
            module_minimization = 0
            module_maturity = 0
            module_coverage = 0
        
        module_metrics = {
            'maturity': round(module_maturity, 2),
            'minimization': round(module_minimization, 2),
            'coverage': round(module_coverage, 2)
        }
        
        # Generate AI analysis
        logger.info(f"Generating AI analysis for module: {module_name}")
        try:
            ai_analysis = generate_module_ai_analysis(module_name, module_data, module_metrics, framework_id)
        except Exception as ai_error:
            logger.error(f"AI analysis generation failed: {str(ai_error)}")
            ai_analysis = {
                'error': f"AI analysis failed: {str(ai_error)}",
                'recommendations': [],
                'miscategorizations': [],
                'score_analysis': {
                    'maturity_analysis': 'AI analysis unavailable',
                    'minimization_analysis': 'AI analysis unavailable',
                    'coverage_analysis': 'AI analysis unavailable'
                },
                'priority_actions': []
            }
        
        # Compile response
        response_data = {
            'module_name': module_name,
            'module_data': {
                'total_records': module_data['total_records'],
                'total_fields': module_data['total_fields'],
                'records_with_inventory': module_data['records_with_inventory'],
                'personal_count': module_data['personal_count'],
                'regular_count': module_data['regular_count'],
                'confidential_count': module_data['confidential_count'],
                'personal_percentage': module_data['personal_percentage'],
                'regular_percentage': module_data['regular_percentage'],
                'confidential_percentage': module_data['confidential_percentage']
            },
            'module_metrics': module_metrics,
            'ai_analysis': ai_analysis,
            'generated_at': datetime.now().isoformat(),
            'framework_id': framework_id
        }
        
        return Response({
            'status': 'success',
            'data': response_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error in module AI analysis: {str(e)}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_privacy_report(request):
    """
    Export privacy compliance report in PDF or Excel format
    Query params: format (pdf/excel), framework_id (optional)
    """
    try:
        export_format = request.query_params.get('format', 'pdf').lower()
        framework_id = request.query_params.get('framework_id', None)
        include_ai = request.query_params.get('include_ai', 'true').lower() == 'true'
        
        if export_format not in ['pdf', 'excel', 'xlsx']:
            return Response({
                'status': 'error',
                'message': 'Invalid format. Use "pdf" or "excel"'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get report data
        privacy_data = collect_privacy_data_across_modules(framework_id)
        metrics = calculate_privacy_metrics(privacy_data)
        consent_data = get_consent_rates(framework_id)
        
        report = {
            'report_metadata': {
                'generated_at': datetime.now().isoformat(),
                'framework_id': framework_id,
                'report_type': 'privacy_compliance'
            },
            'executive_summary': {
                'maturity_score': metrics['maturity_score'],
                'minimization_score': metrics['minimization_score'],
                'overall_assessment': 'compliant' if metrics['maturity_score'] >= 70 else 'needs_improvement'
            },
            'privacy_metrics': metrics,
            'consent_management': consent_data,
            'module_analysis': {}
        }
        
        # Add module analysis
        for module_name, module_data in privacy_data['modules'].items():
            if module_name in privacy_data['modules_with_inventory']:
                module_scores = metrics['module_scores'].get(module_name, {})
                report['module_analysis'][module_name] = {
                    'metrics': module_scores,
                    'data_inventory': {
                        'total_fields': module_data['total_fields'],
                        'personal_count': module_data['personal_count'],
                        'confidential_count': module_data['confidential_count'],
                        'regular_count': module_data['regular_count']
                    },
                    'records': {
                        'total': module_data['total_records'],
                        'with_inventory': module_data['records_with_inventory']
                    }
                }
        
        # Add AI insights if requested
        if include_ai:
            try:
                ai_insights = generate_ai_insights(privacy_data, metrics, consent_data)
                report['ai_insights'] = ai_insights
            except Exception as e:
                logger.error(f"Error generating AI insights for export: {str(e)}")
                report['ai_insights'] = {'error': 'AI insights unavailable'}
        
        # Generate file
        if export_format in ['excel', 'xlsx']:
            try:
                excel_file = generate_excel_report(report)
                response = HttpResponse(
                    excel_file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"privacy_compliance_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except Exception as e:
                return Response({
                    'status': 'error',
                    'message': f'Failed to generate Excel: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:  # PDF
            try:
                pdf_file = generate_pdf_report(report)
                response = HttpResponse(
                    pdf_file.read(),
                    content_type='application/pdf'
                )
                filename = f"privacy_compliance_report_{datetime.now().strftime('%Y%m%d')}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            except Exception as e:
                return Response({
                    'status': 'error',
                    'message': f'Failed to generate PDF: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    except Exception as e:
        logger.error(f"Error exporting privacy report: {str(e)}")
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

