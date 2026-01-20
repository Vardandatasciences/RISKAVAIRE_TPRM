"""
Vendor Dashboard Views
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from django.db.models import Count, F, Q
from django.db.models.functions import Cast
from django.db.models import FloatField
from django.db import connection
from django.db import connections
from django.http import HttpResponse
from tprm_backend.apps.vendor_core.models import ExternalScreeningResult, ScreeningMatch, TempVendor
from tprm_backend.apps.vendor_questionnaire.models import QuestionnaireAssignments
import logging
from datetime import datetime, timedelta
import io
import csv

# RBAC imports
from tprm_backend.apps.vendor_core.vendor_authentication import JWTAuthentication, SimpleAuthenticatedPermission, VendorPermission

# MULTI-TENANCY: Import tenant utilities for filtering
from tprm_backend.core.tenant_utils import (
    get_tenant_id_from_request,
    filter_queryset_by_tenant,
    get_tenant_aware_queryset,
    require_tenant,
    tenant_filter
)

# Database connection helper - Use tprm_integration database for all vendor operations
def get_db_connection():
    """
    Get the correct database connection for tprm_integration database.
    Returns 'tprm' connection if available, otherwise falls back to 'default'.
    """
    if 'tprm' in connections.databases:
        return connections['tprm']
    return connections['default']

# Try to import openpyxl, but don't fail if it's not available
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import reportlab, but don't fail if it's not available
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics import renderPDF
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Get logger for this module
logger = logging.getLogger(__name__)


class ScreeningMatchRateAPIView(APIView):
    """API view for calculating vendor screening match rate with RBAC protection
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get screening match rate statistics"""
        logger.info("ScreeningMatchRateAPIView.get called")
        
        try:
            # MULTI-TENANCY: Get tenant ID from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                return Response({'error': 'Tenant context not found'}, status=status.HTTP_403_FORBIDDEN)

            with get_db_connection().cursor() as cursor:
                # Step 1: Get total vendors count from temp_vendor table
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("SELECT COUNT(*) AS total_vendors FROM temp_vendor WHERE TenantId = %s", [tenant_id])
                total_vendors = cursor.fetchone()[0] or 0
                logger.info(f"Total vendors: {total_vendors}")
                
                # Step 2: Get vendors with valid matches (excluding false positives and resolved matches)
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT COUNT(DISTINCT v.id) AS matched_vendors
                    FROM temp_vendor v
                    JOIN external_screening_results esr ON v.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND v.TenantId = %s
                """, [tenant_id])
                matched_vendors = cursor.fetchone()[0] or 0
                logger.info(f"Matched vendors: {matched_vendors}")
                
                # Step 3: Calculate match rate percentage
                match_rate = (matched_vendors / total_vendors * 100) if total_vendors > 0 else 0
                match_rate = round(match_rate, 1)
                logger.info(f"Match rate: {match_rate}%")
                
                # Step 4: Get breakdown by match type (since screening_type doesn't exist in screening_matches)
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT 
                        sm.match_type,
                        COUNT(DISTINCT v.id) as vendor_count,
                        ROUND((COUNT(DISTINCT v.id) * 100.0 / %s), 1) as percentage
                    FROM temp_vendor v
                    JOIN external_screening_results esr ON v.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND v.TenantId = %s
                    GROUP BY sm.match_type
                    ORDER BY sm.match_type
                """, [total_vendors, tenant_id])
                
                screening_breakdown = cursor.fetchall()
                logger.info(f"Screening breakdown: {screening_breakdown}")
                
                # Format breakdown data for chart
                chart_data = {
                    'labels': [],
                    'values': []
                }
                
                for match_type, vendor_count, percentage in screening_breakdown:
                    chart_data['labels'].append(match_type)
                    chart_data['values'].append(float(percentage))
                
                # If no matches, provide default chart data
                if not chart_data['labels']:
                    chart_data = {
                        'labels': ['OFAC', 'PEP', 'SANCTIONS'],
                        'values': [0, 0, 0]
                    }
                
                # Step 5: Determine status based on match rate
                variant = 'success' if match_rate <= 5 else 'destructive'
                status_text = 'On Target' if match_rate <= 5 else 'Off Target'
                
                response_data = {
                    'value': f"{match_rate}%",
                    'target': "<= 5%",
                    'variant': variant,
                    'chartData': chart_data,
                    'total_vendors': total_vendors,
                    'matched_vendors': matched_vendors,
                    'match_rate': match_rate,
                    'status': status_text,
                    'debug_info': {
                        'total_vendors': total_vendors,
                        'matched_vendors': matched_vendors,
                        'match_rate': match_rate,
                        'screening_breakdown': screening_breakdown
                    }
                }
                
                logger.info(f"Returning response: {response_data}")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Provide more detailed error information
            logger.error(f"Error in ScreeningMatchRateAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to calculate screening match rate',
                'total_vendors': 'unknown',
                'matched_vendors': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuestionnaireOverdueRateAPIView(APIView):
    """API view for calculating questionnaire overdue rate with RBAC protection
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get questionnaire overdue rate statistics"""
        logger.info("QuestionnaireOverdueRateAPIView.get called")
        
        try:
            # MULTI-TENANCY: Get tenant ID from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                return Response({'error': 'Tenant context not found'}, status=status.HTTP_403_FORBIDDEN)

            with get_db_connection().cursor() as cursor:
                # Step 1: Get total questionnaires count from questionnaire_assignments table
                # MULTI-TENANCY: Filter by tenant through temp_vendor
                cursor.execute("""
                    SELECT COUNT(*) AS total_questionnaires 
                    FROM questionnaire_assignments qa
                    JOIN temp_vendor tv ON qa.temp_vendor_id = tv.id
                    WHERE tv.TenantId = %s
                """, [tenant_id])
                total_questionnaires = cursor.fetchone()[0] or 0
                logger.info(f"Total questionnaires: {total_questionnaires}")
                
                # Step 2: Get overdue questionnaires count
                # MULTI-TENANCY: Filter by tenant through temp_vendor
                cursor.execute("""
                    SELECT COUNT(*) AS overdue_questionnaires
                    FROM questionnaire_assignments qa
                    JOIN temp_vendor tv ON qa.temp_vendor_id = tv.id
                    WHERE due_date < NOW()
                      AND (submission_date IS NULL OR submission_date > due_date)
                      AND tv.TenantId = %s
                """, [tenant_id])
                overdue_questionnaires = cursor.fetchone()[0] or 0
                logger.info(f"Overdue questionnaires: {overdue_questionnaires}")
                
                # Step 3: Calculate overdue rate percentage
                overdue_rate = (overdue_questionnaires / total_questionnaires * 100) if total_questionnaires > 0 else 0
                overdue_rate = round(overdue_rate, 1)
                logger.info(f"Overdue rate: {overdue_rate}%")
                
                # Step 4: Get breakdown by overdue vs on-time questionnaires
                # Get on-time questionnaires count
                # MULTI-TENANCY: Filter by tenant through temp_vendor
                cursor.execute("""
                    SELECT COUNT(*) AS ontime_questionnaires
                    FROM questionnaire_assignments qa
                    JOIN temp_vendor tv ON qa.temp_vendor_id = tv.id
                    WHERE NOT (due_date < NOW() AND (submission_date IS NULL OR submission_date > due_date))
                      AND tv.TenantId = %s
                """, [tenant_id])
                ontime_questionnaires = cursor.fetchone()[0] or 0
                logger.info(f"On-time questionnaires: {ontime_questionnaires}")
                
                # Calculate percentages for chart
                overdue_percentage = round(overdue_rate, 1)
                ontime_percentage = round((ontime_questionnaires / total_questionnaires * 100) if total_questionnaires > 0 else 0, 1)
                
                # Create meaningful chart data showing overdue vs on-time
                chart_data = {
                    'labels': ['Overdue', 'On-Time'],
                    'values': [overdue_percentage, ontime_percentage]
                }
                logger.info(f"Chart data: {chart_data}")
                
                # Step 5: Determine status based on overdue rate
                variant = 'success' if overdue_rate <= 5 else 'destructive'
                status_text = 'On Target' if overdue_rate <= 5 else 'Off Target'
                
                response_data = {
                    'value': f"{overdue_rate}%",
                    'target': "<= 5%",
                    'variant': variant,
                    'chartData': chart_data,
                    'total_questionnaires': total_questionnaires,
                    'overdue_questionnaires': overdue_questionnaires,
                    'overdue_rate': overdue_rate,
                    'status': status_text,
                    'debug_info': {
                        'total_questionnaires': total_questionnaires,
                        'overdue_questionnaires': overdue_questionnaires,
                        'ontime_questionnaires': ontime_questionnaires,
                        'overdue_rate': overdue_rate,
                        'ontime_percentage': ontime_percentage,
                        'chart_data': chart_data
                    }
                }
                
                logger.info(f"Returning response: {response_data}")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Provide more detailed error information
            logger.error(f"Error in QuestionnaireOverdueRateAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to calculate questionnaire overdue rate',
                'total_questionnaires': 'unknown',
                'overdue_questionnaires': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorsFlaggedOFACPEPAPIView(APIView):
    """API view for calculating vendors flagged in OFAC/PEP lists with RBAC protection
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get vendors flagged in OFAC/PEP lists statistics"""
        logger.info("VendorsFlaggedOFACPEPAPIView.get called")
        
        try:
            # MULTI-TENANCY: Get tenant ID from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                return Response({'error': 'Tenant context not found'}, status=status.HTTP_403_FORBIDDEN)

            with get_db_connection().cursor() as cursor:
                # Step 1: Get total vendors count from temp_vendor table
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("SELECT COUNT(*) AS total_vendors FROM temp_vendor WHERE TenantId = %s", [tenant_id])
                total_vendors = cursor.fetchone()[0] or 0
                logger.info(f"Total vendors: {total_vendors}")
                
                # Step 2: Get vendors flagged in OFAC/PEP lists
                # Join temp_vendor with external_screening_results and screening_matches
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT COUNT(DISTINCT v.id) AS flagged_vendors
                    FROM temp_vendor v
                    JOIN external_screening_results esr ON v.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND sm.match_type IN ('OFAC - sdn', 'PEP')
                      AND v.TenantId = %s
                """, [tenant_id])
                flagged_vendors = cursor.fetchone()[0] or 0
                logger.info(f"Flagged vendors: {flagged_vendors}")
                
                # Step 3: Get breakdown by match type (OFAC vs PEP)
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT 
                        sm.match_type,
                        COUNT(DISTINCT v.id) as vendor_count
                    FROM temp_vendor v
                    JOIN external_screening_results esr ON v.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND sm.match_type IN ('OFAC - sdn', 'PEP')
                      AND v.TenantId = %s
                    GROUP BY sm.match_type
                    ORDER BY sm.match_type
                """, [tenant_id])
                
                match_breakdown = cursor.fetchall()
                logger.info(f"Match breakdown: {match_breakdown}")
                
                # Step 4: Get detailed flagged vendors list for debugging
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT 
                        v.id,
                        v.company_name,
                        sm.match_type,
                        sm.resolution_status,
                        sm.match_score
                    FROM temp_vendor v
                    JOIN external_screening_results esr ON v.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND sm.match_type IN ('OFAC - sdn', 'PEP')
                      AND v.TenantId = %s
                    ORDER BY sm.match_score DESC
                    LIMIT 10
                """, [tenant_id])
                
                flagged_details = cursor.fetchall()
                logger.info(f"Flagged details: {flagged_details}")
                
                # Step 5: Determine status based on flagged count
                target_threshold = 2
                variant = 'success' if flagged_vendors <= target_threshold else 'destructive'
                status_text = 'On Target' if flagged_vendors <= target_threshold else 'Off Target'
                
                # Step 6: Create alert message
                if flagged_vendors == 0:
                    alert_message = "No matches found"
                elif flagged_vendors == 1:
                    alert_message = "1 vendor flagged"
                else:
                    alert_message = f"{flagged_vendors} vendors flagged"
                
                # Step 7: Get historical data for line chart (last 12 months)
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT 
                        DATE_FORMAT(tv.created_at, '%%Y-%%m') AS month,
                        COUNT(DISTINCT tv.id) AS flagged_vendors_count
                    FROM temp_vendor tv
                    JOIN external_screening_results esr ON tv.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND sm.match_type IN ('OFAC - sdn', 'PEP')
                      AND tv.created_at >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
                      AND tv.TenantId = %s
                    GROUP BY DATE_FORMAT(tv.created_at, '%%Y-%%m')
                    ORDER BY month ASC
                """, [tenant_id])
                historical_data = cursor.fetchall()
                logger.info(f"Historical data: {len(historical_data)} months")
                
                # Step 8: Create chart data for line chart
                # Normalize match types for frontend display
                normalized_breakdown = {}
                for match_type, count in match_breakdown:
                    if 'OFAC' in match_type:
                        normalized_breakdown['OFAC - sdn'] = normalized_breakdown.get('OFAC - sdn', 0) + count
                    elif match_type == 'PEP':
                        normalized_breakdown['PEP'] = count
                
                # Ensure both keys exist
                if 'OFAC - sdn' not in normalized_breakdown:
                    normalized_breakdown['OFAC - sdn'] = 0
                if 'PEP' not in normalized_breakdown:
                    normalized_breakdown['PEP'] = 0
                
                # Create historical trend data
                from datetime import datetime, timedelta
                current_date = datetime.now()
                monthly_data = {}
                
                # Populate with actual data
                for month, count in historical_data:
                    monthly_data[month] = count
                
                # Fill in missing months with 0
                trend_values = []
                for i in range(12):
                    month_date = current_date - timedelta(days=30 * i)
                    month_key = month_date.strftime('%Y-%m')
                    trend_values.insert(0, monthly_data.get(month_key, 0))
                
                chart_data = {
                    'flagged': flagged_vendors,
                    'alert': alert_message,
                    'breakdown': normalized_breakdown,
                    'chartType': 'LineChart',
                    'trendData': {
                        'months': 12,
                        'values': trend_values
                    }
                }
                
                response_data = {
                    'value': str(flagged_vendors),
                    'target': "<= 2",
                    'variant': variant,
                    'chartData': chart_data,
                    'flagged': flagged_vendors,
                    'total_vendors': total_vendors,
                    'status': status_text,
                    'debug_info': {
                        'total_vendors': total_vendors,
                        'flagged_vendors': flagged_vendors,
                        'match_breakdown': match_breakdown,
                        'flagged_details': flagged_details,
                        'target_threshold': target_threshold
                    }
                }
                
                logger.info(f"Returning response: {response_data}")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Provide more detailed error information
            logger.error(f"Error in VendorsFlaggedOFACPEPAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to calculate vendors flagged in OFAC/PEP lists',
                'total_vendors': 'unknown',
                'flagged_vendors': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorAcceptanceTimeAPIView(APIView):
    """API view for calculating vendor acceptance/acknowledgment time with RBAC protection"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get vendor acceptance time statistics"""
        logger.info("VendorAcceptanceTimeAPIView.get called")
        
        try:
            with get_db_connection().cursor() as cursor:
                # Step 1: Get overall average acceptance time
                cursor.execute("""
                    SELECT AVG(DATEDIFF(v.created_at, tv.created_at)) AS avg_acceptance_time
                    FROM vendors v
                    JOIN temp_vendor tv ON v.vendor_code = tv.vendor_code
                    WHERE v.status = 'APPROVED'
                """)
                result = cursor.fetchone()
                avg_acceptance_time = result[0] if result[0] is not None else 0
                avg_acceptance_time = round(avg_acceptance_time, 1)
                logger.info(f"Average acceptance time: {avg_acceptance_time} days")
                
                # Step 2: Get vendor-wise breakdown with monthly trend data
                cursor.execute("""
                    SELECT 
                        v.vendor_id,
                        v.company_name,
                        DATEDIFF(v.created_at, tv.created_at) AS acceptance_days,
                        DATE_FORMAT(v.created_at, '%%Y-%%m') AS approval_month
                    FROM vendors v
                    JOIN temp_vendor tv ON v.vendor_code = tv.vendor_code
                    WHERE v.status = 'APPROVED'
                    ORDER BY v.created_at ASC
                """)
                vendor_data = cursor.fetchall()
                logger.info(f"Vendor data: {len(vendor_data)} records")
                
                # Process vendor data to create trendlines
                vendor_trends = {}
                for vendor_id, company_name, acceptance_days, approval_month in vendor_data:
                    if vendor_id not in vendor_trends:
                        vendor_trends[vendor_id] = {
                            'vendor_id': vendor_id,
                            'company_name': company_name,
                            'current_acceptance_days': acceptance_days,
                            'monthly_data': {}
                        }
                    vendor_trends[vendor_id]['monthly_data'][approval_month] = acceptance_days
                
                # Convert to list and limit to top 5 vendors by current performance
                vendor_breakdown = []
                for vendor_id, data in sorted(vendor_trends.items(), key=lambda x: x[1]['current_acceptance_days'])[:5]:
                    vendor_breakdown.append({
                        'vendor_id': data['vendor_id'],
                        'company_name': data['company_name'],
                        'acceptance_days': data['current_acceptance_days'],
                        'monthly_data': data['monthly_data']
                    })
                
                logger.info(f"Processed vendor trends: {len(vendor_breakdown)} vendors")
                
                # Step 3: Get total approved vendors count
                cursor.execute("""
                    SELECT COUNT(*) AS total_approved_vendors
                    FROM vendors v
                    JOIN temp_vendor tv ON v.vendor_code = tv.vendor_code
                    WHERE v.status = 'APPROVED'
                """)
                total_approved_vendors = cursor.fetchone()[0] or 0
                logger.info(f"Total approved vendors: {total_approved_vendors}")
                
                # Step 4: Get monthly trend data (last 12 months)
                cursor.execute("""
                    SELECT 
                        DATE_FORMAT(v.created_at, '%%Y-%%m') AS month,
                        AVG(DATEDIFF(v.created_at, tv.created_at)) AS avg_acceptance_time
                    FROM vendors v
                    JOIN temp_vendor tv ON v.vendor_code = tv.vendor_code
                    WHERE v.status = 'APPROVED'
                      AND v.created_at >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
                    GROUP BY DATE_FORMAT(v.created_at, '%%Y-%%m')
                    ORDER BY month ASC
                """)
                monthly_trend = cursor.fetchall()
                logger.info(f"Monthly trend data: {len(monthly_trend)} months")
                
                # Step 5: Determine status based on average acceptance time
                target_threshold = 2.0
                variant = 'success' if avg_acceptance_time <= target_threshold else 'destructive'
                status_text = 'On Target' if avg_acceptance_time <= target_threshold else 'Off Target'
                
                # Step 6: Format vendor breakdown for frontend (already processed above)
                vendor_list = vendor_breakdown
                
                # Step 7: Format monthly trend data for chart
                trend_data = {
                    'months': [],
                    'values': []
                }
                
                # Create a complete 12-month array
                from datetime import datetime, timedelta
                current_date = datetime.now()
                monthly_data = {}
                
                # Populate with actual data
                for month, avg_time in monthly_trend:
                    monthly_data[month] = round(float(avg_time), 1)
                
                # Fill in missing months with 0
                for i in range(12):
                    month_date = current_date - timedelta(days=30 * i)
                    month_key = month_date.strftime('%Y-%m')
                    trend_data['months'].insert(0, month_key)
                    trend_data['values'].insert(0, monthly_data.get(month_key, 0))
                
                # Step 8: Create chart data for trend line
                chart_data = {
                    'months': 12,
                    'values': trend_data['values']
                }
                
                response_data = {
                    'value': f"{avg_acceptance_time} days",
                    'target': "< 2 days",
                    'variant': variant,
                    'chartData': chart_data,
                    'avg_acceptance_time': avg_acceptance_time,
                    'total_approved_vendors': total_approved_vendors,
                    'vendor_breakdown': vendor_list,
                    'status': status_text,
                    'debug_info': {
                        'avg_acceptance_time': avg_acceptance_time,
                        'total_approved_vendors': total_approved_vendors,
                        'vendor_breakdown_count': len(vendor_breakdown),
                        'monthly_trend_count': len(monthly_trend),
                        'target_threshold': target_threshold
                    }
                }
                
                logger.info(f"Returning response: {response_data}")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Provide more detailed error information
            logger.error(f"Error in VendorAcceptanceTimeAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to calculate vendor acceptance time',
                'avg_acceptance_time': 'unknown',
                'total_approved_vendors': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DashboardMetricsAPIView(APIView):
    """API view for dashboard metrics using temp_vendor table with RBAC protection
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get dashboard metrics and recent activity from temp_vendor table"""
        logger.info("DashboardMetricsAPIView.get called")
        
        try:
            # MULTI-TENANCY: Get tenant ID from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                return Response({'error': 'Tenant context not found'}, status=status.HTTP_403_FORBIDDEN)

            # Get metrics from temp_vendor table
            # MULTI-TENANCY: Filter by tenant
            total_vendors = TempVendor.objects.filter(tenant_id=tenant_id).count()
            high_risk_vendors = TempVendor.objects.filter(tenant_id=tenant_id, risk_level__icontains='HIGH').count()
            critical_vendors = TempVendor.objects.filter(tenant_id=tenant_id, is_critical_vendor=True).count()
            in_review = TempVendor.objects.filter(tenant_id=tenant_id, status__icontains='REVIEW').count()
            approved = TempVendor.objects.filter(tenant_id=tenant_id, status__icontains='APPROVED').count()
            
            # Get recent activity (5 most recently updated vendors)
            # MULTI-TENANCY: Filter by tenant
            recent_vendors = TempVendor.objects.filter(tenant_id=tenant_id).order_by('-updated_at')[:5]
            recent_activity = []
            
            for vendor in recent_vendors:
                # Calculate days active
                days_active = 0
                if vendor.created_at:
                    days_active = (datetime.now().date() - vendor.created_at.date()).days
                
                recent_activity.append({
                    'id': vendor.id,
                    'company_name': vendor.company_name or 'Unknown Company',
                    'status': vendor.status or 'Unknown',
                    'risk_level': vendor.risk_level or 'Unknown',
                    'days_active': days_active,
                    'is_critical': vendor.is_critical_vendor,
                    'has_data_access': vendor.has_data_access,
                    'has_system_access': vendor.has_system_access,
                    'created_at': vendor.created_at.isoformat() if vendor.created_at else None,
                    'updated_at': vendor.updated_at.isoformat() if vendor.updated_at else None
                })
            
            # Calculate questionnaire completion rate
            # MULTI-TENANCY: Filter by tenant through temp_vendor
            total_assignments = QuestionnaireAssignments.objects.filter(temp_vendor__tenant_id=tenant_id).count()
            completed_assignments = QuestionnaireAssignments.objects.filter(temp_vendor__tenant_id=tenant_id, status='RESPONDED').count()
            questionnaire_completion = round((completed_assignments / total_assignments * 100), 2) if total_assignments > 0 else 0
            
            # Calculate average risk score (mock calculation)
            risk_scores = {
                'LOW': 1.0,
                'MEDIUM': 2.0,
                'HIGH': 3.0,
                'CRITICAL': 4.0
            }
            
            total_risk_score = 0
            risk_count = 0
            # MULTI-TENANCY: Filter by tenant
            for vendor in TempVendor.objects.filter(tenant_id=tenant_id).exclude(risk_level__isnull=True).exclude(risk_level=''):
                risk_level = vendor.risk_level.upper()
                if risk_level in risk_scores:
                    total_risk_score += risk_scores[risk_level]
                    risk_count += 1
            
            avg_risk_score = round(total_risk_score / risk_count, 1) if risk_count > 0 else 0
            
            response_data = {
                'metrics': {
                    'total_vendors': total_vendors,
                    'high_risk_vendors': high_risk_vendors,
                    'critical_vendors': critical_vendors,
                    'in_review': in_review,
                    'approved': approved,
                    'questionnaire_completion': questionnaire_completion,
                    'avg_risk_score': avg_risk_score
                },
                'recent_activity': recent_activity,
                'compliance_overview': {
                    'security_assessments': 87,  # Mock data
                    'documentation_complete': 94,  # Mock data
                    'risk_assessments': 76,  # Mock data
                    'fully_compliant': 156,  # Mock data
                    'pending_reviews': 43,  # Mock data
                    'require_attention': 8  # Mock data
                }
            }
            
            logger.info(f"Returning dashboard metrics: {response_data}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error in DashboardMetricsAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to fetch dashboard metrics',
                'metrics': 'unknown',
                'recent_activity': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorAlertsAPIView(APIView):
    """API view for vendor alerts and exceptions with RBAC protection
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get real-time vendor alerts and exceptions"""
        logger.info("VendorAlertsAPIView.get called")
        
        try:
            # MULTI-TENANCY: Get tenant ID from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                return Response({'error': 'Tenant context not found'}, status=status.HTTP_403_FORBIDDEN)

            with get_db_connection().cursor() as cursor:
                alerts = []
                
                # 1. Check for vendors flagged in OFAC list this week
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("""
                    SELECT COUNT(DISTINCT v.id) AS flagged_count
                    FROM temp_vendor v
                    JOIN external_screening_results esr ON v.id = esr.vendor_id
                    JOIN screening_matches sm ON esr.screening_id = sm.screening_id
                    WHERE sm.is_false_positive = 0
                      AND sm.resolution_status IN ('PENDING', 'ESCALATED', 'BLOCKED')
                      AND sm.match_type IN ('OFAC - sdn', 'PEP')
                      AND v.updated_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)
                      AND v.TenantId = %s
                """, [tenant_id])
                ofac_flagged = cursor.fetchone()[0] or 0
                
                if ofac_flagged > 0:
                    alerts.append({
                        'type': 'critical',
                        'icon': 'M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-2.5L13.732 4c-.77-.833-1.964-.833-2.732 0L3.732 16.5c-.77.833.192 2.5 1.732 2.5z',
                        'message': f'{ofac_flagged} Vendor{"s" if ofac_flagged > 1 else ""} flagged in OFAC list this week',
                        'color': 'vendor_text-destructive'
                    })
                
                # 2. Check high-risk vendor percentage
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("SELECT COUNT(*) FROM temp_vendor WHERE TenantId = %s", [tenant_id])
                total_vendors = cursor.fetchone()[0] or 0
                
                # MULTI-TENANCY: Filter by tenant
                cursor.execute("SELECT COUNT(*) FROM temp_vendor WHERE risk_level LIKE '%%HIGH%%' AND TenantId = %s", [tenant_id])
                high_risk_vendors = cursor.fetchone()[0] or 0
                
                if total_vendors > 0:
                    high_risk_percentage = round((high_risk_vendors / total_vendors) * 100, 1)
                    if high_risk_percentage > 15:  # Target threshold
                        alerts.append({
                            'type': 'warning',
                            'icon': 'M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-2.5L13.732 4c-.77-.833-1.964-.833-2.732 0L3.732 16.5c-.77.833.192 2.5 1.732 2.5z',
                            'message': f'High-risk vendor percentage above target at {high_risk_percentage}%',
                            'color': 'vendor_text-warning'
                        })
                
                # 3. Check questionnaire completion rate
                # MULTI-TENANCY: Filter by tenant through temp_vendor
                cursor.execute("""
                    SELECT COUNT(*) FROM questionnaire_assignments qa
                    JOIN temp_vendor tv ON qa.temp_vendor_id = tv.id
                    WHERE tv.TenantId = %s
                """, [tenant_id])
                total_assignments = cursor.fetchone()[0] or 0
                
                # MULTI-TENANCY: Filter by tenant through temp_vendor
                cursor.execute("""
                    SELECT COUNT(*) FROM questionnaire_assignments qa
                    JOIN temp_vendor tv ON qa.temp_vendor_id = tv.id
                    WHERE qa.status = 'RESPONDED' AND tv.TenantId = %s
                """, [tenant_id])
                completed_assignments = cursor.fetchone()[0] or 0
                
                if total_assignments > 0:
                    completion_rate = round((completed_assignments / total_assignments) * 100, 1)
                    if completion_rate < 95:  # Target threshold
                        alerts.append({
                            'type': 'info',
                            'icon': 'M13 16h-1v-4h-1m1-4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z',
                            'message': f'Questionnaire completion rate below 95% target at {completion_rate}%',
                            'color': 'vendor_text-primary'
                        })
                
                # 4. Check for overdue questionnaires
                # MULTI-TENANCY: Filter by tenant through temp_vendor
                cursor.execute("""
                    SELECT COUNT(*) FROM questionnaire_assignments qa
                    JOIN temp_vendor tv ON qa.temp_vendor_id = tv.id
                    WHERE qa.due_date < NOW() AND (qa.submission_date IS NULL OR qa.submission_date > qa.due_date)
                      AND tv.TenantId = %s
                """, [tenant_id])
                overdue_questionnaires = cursor.fetchone()[0] or 0
                
                if overdue_questionnaires > 0:
                    alerts.append({
                        'type': 'warning',
                        'icon': 'M12 8v4m0 4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z',
                        'message': f'{overdue_questionnaires} Questionnaire{"s" if overdue_questionnaires > 1 else ""} overdue',
                        'color': 'vendor_text-warning'
                    })
                
                response_data = {
                    'alerts': alerts,
                    'total_alerts': len(alerts),
                    'critical_alerts': len([a for a in alerts if a['type'] == 'critical']),
                    'warning_alerts': len([a for a in alerts if a['type'] == 'warning']),
                    'info_alerts': len([a for a in alerts if a['type'] == 'info'])
                }
                
                logger.info(f"Returning alerts: {len(alerts)} total alerts")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            logger.error(f"Error in VendorAlertsAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e), 'message': 'Failed to fetch vendor alerts'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorRegistrationCompletionRateAPIView(APIView):
    """API view for calculating vendor registration completion rate with RBAC protection"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get vendor registration completion rate statistics"""
        logger.info("VendorRegistrationCompletionRateAPIView.get called")
        
        try:
            with get_db_connection().cursor() as cursor:
                # Step 1: Get total award notifications
                cursor.execute("SELECT COUNT(*) AS total_notifications FROM rfp_award_notifications")
                total_notifications = cursor.fetchone()[0] or 0
                logger.info(f"Total award notifications: {total_notifications}")
                
                # Step 2: Get vendors who completed registration (response_id exists in temp_vendor)
                cursor.execute("""
                    SELECT COUNT(DISTINCT ran.response_id) AS registered_vendors
                    FROM rfp_award_notifications ran
                    INNER JOIN temp_vendor tv ON ran.response_id = tv.response_id
                    WHERE ran.response_id IS NOT NULL
                """)
                registered_vendors = cursor.fetchone()[0] or 0
                logger.info(f"Registered vendors: {registered_vendors}")
                
                # Step 3: Calculate completion rate percentage
                completion_rate = (registered_vendors / total_notifications * 100) if total_notifications > 0 else 0
                completion_rate = round(completion_rate, 1)
                logger.info(f"Completion rate: {completion_rate}%")
                
                # Step 4: Get pending registrations (notifications sent but not yet registered)
                cursor.execute("""
                    SELECT COUNT(*) AS pending_registrations
                    FROM rfp_award_notifications ran
                    LEFT JOIN temp_vendor tv ON ran.response_id = tv.response_id
                    WHERE ran.response_id IS NOT NULL 
                      AND tv.response_id IS NULL
                """)
                pending_registrations = cursor.fetchone()[0] or 0
                logger.info(f"Pending registrations: {pending_registrations}")
                
                # Step 5: Get breakdown by notification status
                cursor.execute("""
                    SELECT 
                        ran.notification_status,
                        COUNT(*) as count,
                        SUM(CASE WHEN tv.response_id IS NOT NULL THEN 1 ELSE 0 END) as registered_count
                    FROM rfp_award_notifications ran
                    LEFT JOIN temp_vendor tv ON ran.response_id = tv.response_id
                    WHERE ran.response_id IS NOT NULL
                    GROUP BY ran.notification_status
                    ORDER BY ran.notification_status
                """)
                
                status_breakdown = cursor.fetchall()
                logger.info(f"Status breakdown: {status_breakdown}")
                
                # Step 6: Determine status based on completion rate
                target_threshold = 90.0
                variant = 'success' if completion_rate >= target_threshold else 'warning' if completion_rate >= 75 else 'destructive'
                status_text = 'On Target' if completion_rate >= target_threshold else 'At Risk' if completion_rate >= 75 else 'Below Target'
                
                # Step 7: Create chart data for donut chart
                chart_data = {
                    'value': completion_rate,
                    'total': 100,
                    'registered': registered_vendors,
                    'pending': pending_registrations,
                    'total_notifications': total_notifications
                }
                
                response_data = {
                    'value': f"{completion_rate}%",
                    'target': ">= 90%",
                    'variant': variant,
                    'chartData': chart_data,
                    'total_notifications': total_notifications,
                    'registered_vendors': registered_vendors,
                    'pending_registrations': pending_registrations,
                    'completion_rate': completion_rate,
                    'status': status_text,
                    'debug_info': {
                        'total_notifications': total_notifications,
                        'registered_vendors': registered_vendors,
                        'pending_registrations': pending_registrations,
                        'completion_rate': completion_rate,
                        'status_breakdown': status_breakdown
                    }
                }
                
                logger.info(f"Returning response: {response_data}")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Provide more detailed error information
            logger.error(f"Error in VendorRegistrationCompletionRateAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to calculate vendor registration completion rate',
                'total_notifications': 'unknown',
                'registered_vendors': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorRegistrationTimeAPIView(APIView):
    """API view for calculating vendor registration time with RBAC protection"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get vendor registration time statistics"""
        logger.info("VendorRegistrationTimeAPIView.get called")
        
        try:
            with get_db_connection().cursor() as cursor:
                # Step 1: Get average registration time (response_date to created_at)
                cursor.execute("""
                    SELECT AVG(DATEDIFF(tv.created_at, ran.response_date)) AS avg_registration_time
                    FROM temp_vendor tv
                    INNER JOIN rfp_award_notifications ran ON tv.response_id = ran.response_id
                    WHERE ran.response_date IS NOT NULL 
                      AND tv.created_at IS NOT NULL
                      AND ran.response_date <= tv.created_at
                """)
                result = cursor.fetchone()
                avg_registration_time = result[0] if result[0] is not None else 0
                avg_registration_time = round(avg_registration_time, 1)
                logger.info(f"Average registration time: {avg_registration_time} days")
                
                # Step 2: Get breakdown by business type
                cursor.execute("""
                    SELECT 
                        tv.business_type,
                        COUNT(*) as vendor_count,
                        AVG(DATEDIFF(tv.created_at, ran.response_date)) as avg_time,
                        MIN(DATEDIFF(tv.created_at, ran.response_date)) as min_time,
                        MAX(DATEDIFF(tv.created_at, ran.response_date)) as max_time
                    FROM temp_vendor tv
                    INNER JOIN rfp_award_notifications ran ON tv.response_id = ran.response_id
                    WHERE ran.response_date IS NOT NULL 
                      AND tv.created_at IS NOT NULL
                      AND ran.response_date <= tv.created_at
                      AND tv.business_type IS NOT NULL
                      AND tv.business_type != ''
                    GROUP BY tv.business_type
                    ORDER BY avg_time ASC
                """)
                
                business_type_breakdown = cursor.fetchall()
                logger.info(f"Business type breakdown: {business_type_breakdown}")
                
                # Step 3: Process breakdown data for chart
                chart_labels = []
                chart_values = []
                
                # Map business types to standardized labels
                business_type_mapping = {
                    'INDIVIDUAL': 'Individual',
                    'CORPORATION': 'Corporate', 
                    'CORPORATE': 'Corporate',
                    'PARTNERSHIP': 'Partnership',
                    'GOVERNMENT': 'Government',
                    'LLC': 'Corporate',
                    'LLP': 'Partnership',
                    'SOLE_PROPRIETORSHIP': 'Individual'
                }
                
                processed_breakdown = {}
                
                for business_type, vendor_count, avg_time, min_time, max_time in business_type_breakdown:
                    # Standardize business type name
                    standardized_type = business_type_mapping.get(business_type.upper(), business_type.title())
                    processed_breakdown[standardized_type] = {
                        'vendor_count': vendor_count,
                        'avg_time': round(float(avg_time), 1),
                        'min_time': int(min_time),
                        'max_time': int(max_time)
                    }
                
                # Ensure all expected business types are represented
                expected_types = ['Individual', 'Corporate', 'Partnership', 'Government']
                for expected_type in expected_types:
                    if expected_type not in processed_breakdown:
                        processed_breakdown[expected_type] = {
                            'vendor_count': 0,
                            'avg_time': 0,
                            'min_time': 0,
                            'max_time': 0
                        }
                
                # Create chart data in expected order
                for business_type in expected_types:
                    chart_labels.append(business_type)
                    chart_values.append(processed_breakdown[business_type]['avg_time'])
                
                # Step 4: Get total vendors with registration time data
                cursor.execute("""
                    SELECT COUNT(*) FROM temp_vendor tv
                    INNER JOIN rfp_award_notifications ran ON tv.response_id = ran.response_id
                    WHERE ran.response_date IS NOT NULL 
                      AND tv.created_at IS NOT NULL
                      AND ran.response_date <= tv.created_at
                """)
                total_vendors_with_time = cursor.fetchone()[0] or 0
                logger.info(f"Total vendors with registration time: {total_vendors_with_time}")
                
                # Step 5: Get registration time distribution
                cursor.execute("""
                    SELECT 
                        CASE 
                            WHEN DATEDIFF(tv.created_at, ran.response_date) <= 1 THEN '0-1 days'
                            WHEN DATEDIFF(tv.created_at, ran.response_date) <= 3 THEN '1-3 days'
                            WHEN DATEDIFF(tv.created_at, ran.response_date) <= 7 THEN '3-7 days'
                            WHEN DATEDIFF(tv.created_at, ran.response_date) <= 14 THEN '1-2 weeks'
                            ELSE '2+ weeks'
                        END as time_bucket,
                        COUNT(*) as vendor_count
                    FROM temp_vendor tv
                    INNER JOIN rfp_award_notifications ran ON tv.response_id = ran.response_id
                    WHERE ran.response_date IS NOT NULL 
                      AND tv.created_at IS NOT NULL
                      AND ran.response_date <= tv.created_at
                    GROUP BY time_bucket
                    ORDER BY 
                        CASE time_bucket
                            WHEN '0-1 days' THEN 1
                            WHEN '1-3 days' THEN 2
                            WHEN '3-7 days' THEN 3
                            WHEN '1-2 weeks' THEN 4
                            WHEN '2+ weeks' THEN 5
                        END
                """)
                
                time_distribution = cursor.fetchall()
                logger.info(f"Time distribution: {time_distribution}")
                
                # Step 6: Get detailed vendor breakdown for individual performance
                cursor.execute("""
                    SELECT 
                        tv.id,
                        tv.company_name,
                        tv.business_type,
                        DATEDIFF(tv.created_at, ran.response_date) as registration_days,
                        ran.response_date,
                        tv.created_at
                    FROM temp_vendor tv
                    INNER JOIN rfp_award_notifications ran ON tv.response_id = ran.response_id
                    WHERE ran.response_date IS NOT NULL 
                      AND tv.created_at IS NOT NULL
                      AND ran.response_date <= tv.created_at
                    ORDER BY registration_days ASC
                    LIMIT 10
                """)
                
                vendor_details = cursor.fetchall()
                logger.info(f"Vendor details: {len(vendor_details)} records")
                
                # Step 7: Determine status based on average registration time
                target_threshold = 3.0  # Target: < 3 days
                variant = 'success' if avg_registration_time <= target_threshold else 'warning' if avg_registration_time <= 7 else 'destructive'
                status_text = 'On Target' if avg_registration_time <= target_threshold else 'At Risk' if avg_registration_time <= 7 else 'Below Target'
                
                # Step 8: Create chart data for gauge chart
                chart_data = {
                    'value': avg_registration_time,
                    'max': 10,  # Maximum days for gauge scale
                    'target': 3,  # Target threshold
                    'avg_time': avg_registration_time,
                    'total_vendors': total_vendors_with_time,
                    'breakdown': processed_breakdown,
                    'time_distribution': time_distribution,
                    'vendor_details': vendor_details
                }
                
                response_data = {
                    'value': f"{avg_registration_time} days",
                    'target': "< 3 days",
                    'variant': variant,
                    'chartData': chart_data,
                    'avg_registration_time': avg_registration_time,
                    'total_vendors_with_time': total_vendors_with_time,
                    'business_type_breakdown': processed_breakdown,
                    'time_distribution': time_distribution,
                    'vendor_details': vendor_details,
                    'status': status_text,
                    'debug_info': {
                        'avg_registration_time': avg_registration_time,
                        'total_vendors_with_time': total_vendors_with_time,
                        'business_type_breakdown_count': len(business_type_breakdown),
                        'time_distribution_count': len(time_distribution),
                        'vendor_details_count': len(vendor_details),
                        'target_threshold': target_threshold
                    }
                }
                
                logger.info(f"Returning response: {response_data}")
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Provide more detailed error information
            logger.error(f"Error in VendorRegistrationTimeAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            error_details = {
                'error': str(e),
                'traceback': traceback.format_exc(),
                'message': 'Failed to calculate vendor registration time',
                'avg_registration_time': 'unknown',
                'total_vendors_with_time': 'unknown'
            }
            return Response(
                error_details,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorKPICategoriesAPIView(APIView):
    """API view for KPI categories overview with RBAC protection
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def get(self, request):
        """Get real-time KPI categories data"""
        logger.info("VendorKPICategoriesAPIView.get called")
        
        try:
            # MULTI-TENANCY: Get tenant ID from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                return Response({'error': 'Tenant context not found'}, status=status.HTTP_403_FORBIDDEN)

            categories = []
            
            # 1. Initial Validation KPIs - Use Django ORM with TempVendor model
            # MULTI-TENANCY: Filter by tenant
            total_validated = TempVendor.objects.filter(tenant_id=tenant_id, status__in=['APPROVED', 'PENDING_REVIEW']).count()
            approved_vendors = TempVendor.objects.filter(tenant_id=tenant_id, status='APPROVED').count()
            
            validation_score = round((approved_vendors / total_validated * 100), 1) if total_validated > 0 else 0
            
            categories.append({
                'name': 'Initial Validation',
                'count': 3,
                'score': validation_score,
                'bgColor': 'vendor_bg-primary-soft',
                'iconColor': 'vendor_text-primary',
                'scoreColor': 'vendor_text-success' if validation_score >= 90 else 'vendor_text-warning' if validation_score >= 75 else 'vendor_text-destructive',
                'iconPath': 'M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z'
            })
            
            # 2. Due Diligence KPIs - Use Django ORM with QuestionnaireAssignments model
            # MULTI-TENANCY: Filter by tenant through temp_vendor
            total_questionnaires = QuestionnaireAssignments.objects.filter(temp_vendor__tenant_id=tenant_id).count()
            completed_questionnaires = QuestionnaireAssignments.objects.filter(temp_vendor__tenant_id=tenant_id, status='RESPONDED').count()
            
            due_diligence_score = round((completed_questionnaires / total_questionnaires * 100), 1) if total_questionnaires > 0 else 0
            
            categories.append({
                'name': 'Due Diligence',
                'count': 4,
                'score': due_diligence_score,
                'bgColor': 'vendor_bg-success-soft',
                'iconColor': 'vendor_text-success',
                'scoreColor': 'vendor_text-success' if due_diligence_score >= 90 else 'vendor_text-warning' if due_diligence_score >= 75 else 'vendor_text-destructive',
                'iconPath': 'M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z'
            })
            
            # 3. Risk Scoring KPIs - Use Django ORM with TempVendor model
            # MULTI-TENANCY: Filter by tenant
            total_risk_assessed = TempVendor.objects.filter(tenant_id=tenant_id).exclude(risk_level__isnull=True).exclude(risk_level='').count()
            low_medium_risk = TempVendor.objects.filter(tenant_id=tenant_id, risk_level__in=['LOW', 'MEDIUM']).count()
            
            risk_scoring_score = round((low_medium_risk / total_risk_assessed * 100), 1) if total_risk_assessed > 0 else 0
            
            categories.append({
                'name': 'Risk Scoring',
                'count': 2,
                'score': risk_scoring_score,
                'bgColor': 'vendor_bg-warning-soft',
                'iconColor': 'vendor_text-warning',
                'scoreColor': 'vendor_text-success' if risk_scoring_score >= 75 else 'vendor_text-warning' if risk_scoring_score >= 50 else 'vendor_text-destructive',
                'iconPath': 'M9 12l2 2 4-4m5.618-4.016A11.955 11.955 0 0112 2.944a11.955 11.955 0 01-8.618 3.04A12.02 12.02 0 003 9c0 5.591 3.824 10.29 9 11.622 5.176-1.332 9-6.03 9-11.622 0-1.042-.133-2.052-.382-3.016z'
            })
            
            # 4. External Verification KPIs - Use Django ORM (vendor_id is not a ForeignKey, so use filter)
            # MULTI-TENANCY: Filter by tenant through temp_vendor
            tenant_vendor_ids = TempVendor.objects.filter(tenant_id=tenant_id).values_list('id', flat=True)
            vendor_ids_with_screening = ExternalScreeningResult.objects.filter(vendor_id__in=tenant_vendor_ids).values_list('vendor_id', flat=True).distinct()
            total_screened = len(vendor_ids_with_screening)
            
            # Get flagged vendors: those with screening matches that are not false positives and have pending/escalated/blocked status
            flagged_screening_ids = ScreeningMatch.objects.filter(
                is_false_positive=False,
                resolution_status__in=['PENDING', 'ESCALATED', 'BLOCKED']
            ).values_list('screening_id', flat=True).distinct()
            
            # MULTI-TENANCY: Filter by tenant through temp_vendor
            flagged_vendor_ids = ExternalScreeningResult.objects.filter(
                screening_id__in=flagged_screening_ids,
                vendor_id__in=tenant_vendor_ids
            ).values_list('vendor_id', flat=True).distinct()
            
            flagged_vendors = len(flagged_vendor_ids)
            
            external_verification_score = round(((total_screened - flagged_vendors) / total_screened * 100), 1) if total_screened > 0 else 100
            
            categories.append({
                'name': 'External Verification',
                'count': 3,
                'score': external_verification_score,
                'bgColor': 'vendor_bg-destructive-soft',
                'iconColor': 'vendor_text-destructive',
                'scoreColor': 'vendor_text-success' if external_verification_score >= 95 else 'vendor_text-warning' if external_verification_score >= 85 else 'vendor_text-destructive',
                'iconPath': 'M21 21l-6-6m2-5a7 7 0 11-14 0 7 7 0 0114 0z'
            })
            
            response_data = {
                'categories': categories,
                'total_categories': len(categories),
                'average_score': round(sum(cat['score'] for cat in categories) / len(categories), 1) if categories else 0,
                'top_performing_category': max(categories, key=lambda x: x['score'])['name'] if categories else None,
                'needs_attention_category': min(categories, key=lambda x: x['score'])['name'] if categories else None
            }
            
            logger.info(f"Returning KPI categories: {len(categories)} categories")
            return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            logger.error(f"Error in VendorKPICategoriesAPIView: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e), 'message': 'Failed to fetch KPI categories'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VendorDashboardExportPDFAPIView(APIView):
    """
    Export Vendor Dashboard as PDF with RBAC protection
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def post(self, request):
        try:
            data = request.data
            
            # Debug logging
            logger.info(f"PDF export received data: {data}")
            if 'kpis' in data:
                logger.info(f"Number of KPIs: {len(data['kpis'])}")
                for i, kpi in enumerate(data['kpis']):
                    logger.info(f"KPI {i}: {kpi.get('title', 'Unknown')} - Value: {kpi.get('value', 'N/A')} - Target: {kpi.get('target', 'N/A')}")
            
            # Check if reportlab is available
            if not REPORTLAB_AVAILABLE:
                # Fallback: Create simple text export
                return self.create_text_export(data)
            
            # Generate PDF using reportlab with images
            
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
            styles = getSampleStyleSheet()
            
            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#2563eb')
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.HexColor('#374151')
            )
            
            # Build PDF content
            story = []
            story.append(Paragraph("Vendor KPI Dashboard Report", title_style))
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Add KPI data with charts
            if 'kpis' in data:
                story.append(Paragraph("Key Performance Indicators", heading_style))
                
                # Create KPI summary table
                kpi_data = [['KPI Name', 'Value', 'Target', 'Category']]
                for kpi in data['kpis']:
                    kpi_data.append([
                        kpi.get('title', 'N/A'),
                        kpi.get('value', 'N/A'),
                        kpi.get('target', 'N/A'),
                        kpi.get('category', 'N/A')
                    ])
                
                kpi_table = Table(kpi_data)
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(kpi_table)
                story.append(Spacer(1, 20))
                
            # Add summary overview chart
            summary_chart = self.create_summary_chart(data['kpis'])
            if summary_chart:
                story.append(Paragraph("KPI Overview", heading_style))
                story.append(summary_chart)
                story.append(Spacer(1, 20))
            else:
                # Add a test chart if summary chart fails
                test_chart = self.create_test_chart()
                if test_chart:
                    story.append(Paragraph("KPI Overview (Sample Data)", heading_style))
                    story.append(test_chart)
                    story.append(Spacer(1, 20))
                
                # Add individual KPI charts
                story.append(Paragraph("Individual KPI Visualizations", heading_style))
                for i, kpi in enumerate(data['kpis']):
                    if i > 0 and i % 2 == 0:
                        story.append(PageBreak())
                    
                    # Create chart for this KPI
                    chart = self.create_kpi_chart(kpi)
                    if chart:
                        story.append(chart)
                        story.append(Spacer(1, 12))
                    else:
                        # Add a simple text representation if chart fails
                        story.append(Paragraph(f"{kpi.get('title', 'KPI')} - No chart available", styles['Normal']))
                        story.append(Spacer(1, 12))
            
            # Add alerts data
            if 'alerts' in data and data['alerts']:
                story.append(Paragraph("Alerts & Notifications", heading_style))
                alert_data = [['Alert Title', 'Description', 'Severity']]
                for alert in data['alerts']:
                    alert_data.append([
                        alert.get('title', 'N/A'),
                        alert.get('description', 'N/A')[:50] + '...' if len(alert.get('description', '')) > 50 else alert.get('description', 'N/A'),
                        alert.get('severity', 'N/A')
                    ])
                
                alert_table = Table(alert_data)
                alert_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(alert_table)
            
            doc.build(story)
            pdf_buffer.seek(0)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="vendor-dashboard-{datetime.now().strftime("%Y-%m-%d")}.pdf"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error in PDF export: {str(e)}")
            return Response(
                {'error': str(e), 'message': 'Failed to export PDF'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def create_kpi_chart(self, kpi):
        """Create a chart visualization for a KPI"""
        try:
            drawing = Drawing(400, 200)
            
            kpi_title = kpi.get('title', 'KPI')
            kpi_value = kpi.get('value', '0')
            kpi_target = kpi.get('target', '0')
            chart_type = kpi.get('chartType', 'bar')
            
            # Debug logging
            logger.info(f"Creating chart for KPI: {kpi_title}, Value: {kpi_value}, Target: {kpi_target}")
            
            # Extract numeric values with better parsing
            try:
                # Handle percentage values
                value_str = str(kpi_value).replace('%', '').replace(',', '').strip()
                target_str = str(kpi_target).replace('%', '').replace(',', '').strip()
                
                # Handle target format like "<= 5%" or ">= 90%"
                if '<=' in target_str:
                    target_str = target_str.replace('<=', '').strip()
                elif '>=' in target_str:
                    target_str = target_str.replace('>=', '').strip()
                
                value_num = float(value_str) if value_str else 0
                target_num = float(target_str) if target_str else 0
                
                logger.info(f"Parsed values - Value: {value_num}, Target: {target_num}")
                
            except Exception as e:
                logger.error(f"Error parsing KPI values: {e}")
                value_num = 0
                target_num = 0
            
            # Create title
            title = String(200, 180, kpi_title, textAnchor='middle', fontSize=14, fillColor=colors.HexColor('#2563eb'))
            drawing.add(title)
            
            # Ensure we have valid data for charts
            if value_num == 0 and target_num == 0:
                # Use sample data if no real data
                value_num = 50
                target_num = 75
                logger.warning(f"No data for {kpi_title}, using sample data")
            
            # Determine chart type based on KPI title and type
            if 'completion' in kpi_title.lower() or chart_type == 'donut':
                # Create pie chart for completion rates
                pie = Pie()
                pie.x = 150
                pie.y = 80
                pie.width = 100
                pie.height = 100
                
                # Calculate completion percentage
                completion = min(100, max(0, value_num))
                remaining = 100 - completion
                
                pie.data = [completion, remaining]
                pie.labels = [f'Complete\n{completion:.1f}%', f'Pending\n{remaining:.1f}%']
                pie.slices.strokeWidth = 1
                pie.slices.strokeColor = colors.white
                pie.slices[0].fillColor = colors.HexColor('#10b981')  # Green
                pie.slices[1].fillColor = colors.HexColor('#e5e7eb')  # Gray
                
                drawing.add(pie)
                
            elif 'time' in kpi_title.lower() or chart_type == 'gauge':
                # Create horizontal bar chart for time-based KPIs
                chart = VerticalBarChart()
                chart.x = 100
                chart.y = 70
                chart.width = 200
                chart.height = 80
                
                chart.data = [[value_num, target_num]]
                chart.categoryAxis.categoryNames = ['Current', 'Target']
                chart.valueAxis.valueMin = 0
                chart.valueAxis.valueMax = max(value_num, target_num) * 1.2
                
                # Color bars based on performance
                if value_num <= target_num:
                    current_color = colors.HexColor('#10b981')  # Green
                else:
                    current_color = colors.HexColor('#ef4444')  # Red
                
                chart.bars[0].fillColor = current_color
                chart.bars[1].fillColor = colors.HexColor('#3b82f6')  # Blue for target
                
                drawing.add(chart)
                
            elif 'match' in kpi_title.lower() or 'overdue' in kpi_title.lower():
                # Create vertical bar chart for rate-based KPIs
                chart = VerticalBarChart()
                chart.x = 100
                chart.y = 70
                chart.width = 200
                chart.height = 80
                
                chart.data = [[value_num, target_num]]
                chart.categoryAxis.categoryNames = ['Current', 'Target']
                chart.valueAxis.valueMin = 0
                chart.valueAxis.valueMax = max(value_num, target_num) * 1.2
                
                # Color bars based on performance (lower is better for match/overdue rates)
                if value_num <= target_num:
                    current_color = colors.HexColor('#10b981')  # Green
                else:
                    current_color = colors.HexColor('#ef4444')  # Red
                
                chart.bars[0].fillColor = current_color
                chart.bars[1].fillColor = colors.HexColor('#3b82f6')  # Blue for target
                
                drawing.add(chart)
                
            else:
                # Create vertical bar chart for other KPIs
                chart = VerticalBarChart()
                chart.x = 100
                chart.y = 70
                chart.width = 200
                chart.height = 80
                
                chart.data = [[value_num, target_num]]
                chart.categoryAxis.categoryNames = ['Current', 'Target']
                chart.valueAxis.valueMin = 0
                chart.valueAxis.valueMax = max(value_num, target_num) * 1.2
                
                chart.bars[0].fillColor = colors.HexColor('#3b82f6')  # Blue
                chart.bars[1].fillColor = colors.HexColor('#10b981')  # Green for target
                
                drawing.add(chart)
            
            # Add value labels
            value_label = String(200, 30, f"Value: {kpi_value} | Target: {kpi_target}", 
                               textAnchor='middle', fontSize=10, fillColor=colors.HexColor('#6b7280'))
            drawing.add(value_label)
            
            return drawing
            
        except Exception as e:
            logger.error(f"Error creating chart for KPI {kpi.get('title', 'Unknown')}: {str(e)}")
            return None
    
    def create_summary_chart(self, kpis):
        """Create a summary chart showing all KPIs"""
        try:
            drawing = Drawing(400, 250)
            
            # Create title
            title = String(200, 230, "KPI Performance Overview", textAnchor='middle', fontSize=16, fillColor=colors.HexColor('#2563eb'))
            drawing.add(title)
            
            # Create a bar chart showing all KPIs
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 50
            chart.width = 300
            chart.height = 150
            
            # Prepare data
            kpi_names = []
            current_values = []
            target_values = []
            
            for kpi in kpis[:6]:  # Limit to 6 KPIs for readability
                name = kpi.get('title', 'KPI')[:15]  # Truncate long names
                kpi_names.append(name)
                
                try:
                    # Handle percentage values
                    value_str = str(kpi.get('value', '0')).replace('%', '').replace(',', '').strip()
                    target_str = str(kpi.get('target', '0')).replace('%', '').replace(',', '').strip()
                    
                    # Handle target format like "<= 5%" or ">= 90%"
                    if '<=' in target_str:
                        target_str = target_str.replace('<=', '').strip()
                    elif '>=' in target_str:
                        target_str = target_str.replace('>=', '').strip()
                    
                    value_num = float(value_str) if value_str else 0
                    target_num = float(target_str) if target_str else 0
                    
                    # Use sample data if values are 0
                    if value_num == 0 and target_num == 0:
                        value_num = 50 + (i * 10)  # Sample data
                        target_num = 75 + (i * 5)
                        logger.warning(f"No data for {name}, using sample data: {value_num}, {target_num}")
                    
                    current_values.append(value_num)
                    target_values.append(target_num)
                    
                    logger.info(f"Summary chart - {name}: Value={value_num}, Target={target_num}")
                    
                except Exception as e:
                    logger.error(f"Error parsing values for {name}: {e}")
                    # Use sample data as fallback
                    sample_value = 50 + (i * 10)
                    sample_target = 75 + (i * 5)
                    current_values.append(sample_value)
                    target_values.append(sample_target)
            
            chart.data = [current_values, target_values]
            chart.categoryAxis.categoryNames = kpi_names
            chart.valueAxis.valueMin = 0
            
            # Set appropriate max value
            all_values = current_values + target_values
            max_val = max(all_values) if all_values else 100
            chart.valueAxis.valueMax = max_val * 1.2
            
            # Color the bars
            chart.bars[0].fillColor = colors.HexColor('#3b82f6')  # Blue for current
            chart.bars[1].fillColor = colors.HexColor('#10b981')  # Green for target
            
            # Add legend
            legend_current = Rect(50, 20, 15, 10, fillColor=colors.HexColor('#3b82f6'))
            legend_current_text = String(70, 25, "Current Values", fontSize=8)
            legend_target = Rect(150, 20, 15, 10, fillColor=colors.HexColor('#10b981'))
            legend_target_text = String(170, 25, "Target Values", fontSize=8)
            
            drawing.add(chart)
            drawing.add(legend_current)
            drawing.add(legend_current_text)
            drawing.add(legend_target)
            drawing.add(legend_target_text)
            
            return drawing
            
        except Exception as e:
            logger.error(f"Error creating summary chart: {str(e)}")
            return None
    
    def create_test_chart(self):
        """Create a test chart with sample data to verify chart rendering"""
        try:
            drawing = Drawing(400, 250)
            
            # Create title
            title = String(200, 230, "Sample KPI Performance Overview", textAnchor='middle', fontSize=16, fillColor=colors.HexColor('#2563eb'))
            drawing.add(title)
            
            # Create a bar chart with sample data
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 50
            chart.width = 300
            chart.height = 150
            
            # Sample data
            chart.data = [[15.0, 5.0], [66.7, 90.0], [14.8, 5.0], [3.2, 7.0]]
            chart.categoryAxis.categoryNames = ['Screening', 'Registration', 'Questionnaire', 'Acceptance']
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = 100
            
            # Color the bars
            chart.bars[0].fillColor = colors.HexColor('#3b82f6')  # Blue for current
            chart.bars[1].fillColor = colors.HexColor('#10b981')  # Green for target
            
            # Add legend
            legend_current = Rect(50, 20, 15, 10, fillColor=colors.HexColor('#3b82f6'))
            legend_current_text = String(70, 25, "Current Values", fontSize=8)
            legend_target = Rect(150, 20, 15, 10, fillColor=colors.HexColor('#10b981'))
            legend_target_text = String(170, 25, "Target Values", fontSize=8)
            
            drawing.add(chart)
            drawing.add(legend_current)
            drawing.add(legend_current_text)
            drawing.add(legend_target)
            drawing.add(legend_target_text)
            
            return drawing
            
        except Exception as e:
            logger.error(f"Error creating test chart: {str(e)}")
            return None
    


class VendorDashboardExportExcelAPIView(APIView):
    """
    Export Vendor Dashboard as Excel with RBAC protection
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [VendorPermission]
    
    def post(self, request):
        try:
            data = request.data
            
            # Check if openpyxl is available
            if not OPENPYXL_AVAILABLE:
                # Fallback: Create simple CSV if openpyxl is not available
                return self.create_csv_export(data)
            
            workbook = openpyxl.Workbook()
            
            # Create KPI Summary sheet
            kpi_sheet = workbook.active
            kpi_sheet.title = "KPI Summary"
            self.populate_kpi_sheet(kpi_sheet, data.get('kpis', []))
            
            # Create Alerts sheet
            alerts_sheet = workbook.create_sheet("Alerts & Notifications")
            self.populate_alerts_sheet(alerts_sheet, data.get('alerts', []))
            
            # Create Categories sheet
            categories_sheet = workbook.create_sheet("KPI Categories")
            self.populate_categories_sheet(categories_sheet, data.get('categories', []))
            
            # Save to buffer
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="vendor-dashboard-{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error in Excel export: {str(e)}")
            return Response(
                {'error': str(e), 'message': 'Failed to export Excel'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def populate_kpi_sheet(self, sheet, kpis):
        """Populate KPI data in Excel sheet"""
        # Headers
        headers = ['KPI Name', 'Value', 'Target', 'Category', 'Status', 'Responsible', 'Frequency']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row, kpi in enumerate(kpis, 2):
            sheet.cell(row=row, column=1, value=kpi.get('title', ''))
            sheet.cell(row=row, column=2, value=kpi.get('value', ''))
            sheet.cell(row=row, column=3, value=kpi.get('target', ''))
            sheet.cell(row=row, column=4, value=kpi.get('category', ''))
            sheet.cell(row=row, column=5, value=kpi.get('status', ''))
            sheet.cell(row=row, column=6, value=kpi.get('responsible', ''))
            sheet.cell(row=row, column=7, value=kpi.get('frequency', ''))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def populate_alerts_sheet(self, sheet, alerts):
        """Populate alerts data in Excel sheet"""
        headers = ['Alert Title', 'Description', 'Severity', 'Date', 'Status']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, alert in enumerate(alerts, 2):
            sheet.cell(row=row, column=1, value=alert.get('title', ''))
            sheet.cell(row=row, column=2, value=alert.get('description', ''))
            sheet.cell(row=row, column=3, value=alert.get('severity', ''))
            sheet.cell(row=row, column=4, value=alert.get('date', ''))
            sheet.cell(row=row, column=5, value=alert.get('status', ''))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def populate_categories_sheet(self, sheet, categories):
        """Populate categories data in Excel sheet"""
        headers = ['Category Name', 'Score', 'Color', 'Description']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, category in enumerate(categories, 2):
            sheet.cell(row=row, column=1, value=category.get('name', ''))
            sheet.cell(row=row, column=2, value=category.get('score', ''))
            sheet.cell(row=row, column=3, value=category.get('color', ''))
            sheet.cell(row=row, column=4, value=category.get('description', ''))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_csv_export(self, data):
        """Fallback CSV export when openpyxl is not available"""
        
        csv_buffer = io.StringIO()
        
        # Write KPI data
        writer = csv.writer(csv_buffer)
        writer.writerow(['Vendor KPI Dashboard Report'])
        writer.writerow([f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'])
        writer.writerow([])
        writer.writerow(['KPI SUMMARY'])
        writer.writerow(['KPI Name', 'Value', 'Target', 'Category'])
        
        for kpi in data.get('kpis', []):
            writer.writerow([
                kpi.get('title', ''),
                kpi.get('value', ''),
                kpi.get('target', ''),
                kpi.get('category', '')
            ])
        
        writer.writerow([])
        writer.writerow(['ALERTS & NOTIFICATIONS'])
        writer.writerow(['Alert Title', 'Description', 'Severity'])
        
        for alert in data.get('alerts', []):
            writer.writerow([
                alert.get('title', ''),
                alert.get('description', ''),
                alert.get('severity', '')
            ])
        
        csv_content = csv_buffer.getvalue()
        csv_buffer.close()
        
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="vendor-dashboard-{datetime.now().strftime("%Y-%m-%d")}.csv"'
        
        return response
    
    def create_text_export(self, data):
        """Fallback text export when reportlab is not available"""
        text_content = []
        text_content.append("VENDOR KPI DASHBOARD REPORT")
        text_content.append("=" * 50)
        text_content.append(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        text_content.append("")
        
        # Add KPI data
        if 'kpis' in data:
            text_content.append("KEY PERFORMANCE INDICATORS")
            text_content.append("-" * 30)
            for kpi in data['kpis']:
                text_content.append(f"KPI: {kpi.get('title', 'N/A')}")
                text_content.append(f"  Value: {kpi.get('value', 'N/A')}")
                text_content.append(f"  Target: {kpi.get('target', 'N/A')}")
                text_content.append(f"  Category: {kpi.get('category', 'N/A')}")
                text_content.append("")
        
        # Add alerts data
        if 'alerts' in data and data['alerts']:
            text_content.append("ALERTS & NOTIFICATIONS")
            text_content.append("-" * 30)
            for alert in data['alerts']:
                text_content.append(f"Alert: {alert.get('title', 'N/A')}")
                text_content.append(f"  Description: {alert.get('description', 'N/A')}")
                text_content.append(f"  Severity: {alert.get('severity', 'N/A')}")
                text_content.append("")
        
        text_output = "\n".join(text_content)
        
        response = HttpResponse(text_output, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="vendor-dashboard-{datetime.now().strftime("%Y-%m-%d")}.txt"'
        
        return response