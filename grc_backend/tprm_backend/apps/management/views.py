"""
Management views for vendor listing
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Q, OuterRef, Subquery, Exists, Value, CharField
from django.db.models.functions import Coalesce
from tprm_backend.apps.vendor_core.models import Vendors, TempVendor
from tprm_backend.apps.vendor_core.serializers import VendorsSerializer
from .serializers import AllVendorsListSerializer, TempVendorSerializer
import csv

# Try to import tenant utils - use TPRM backend version if available
try:
    from tprm_backend.core.tenant_utils import get_tenant_id_from_request
except ImportError:
    try:
        from grc.utils.tenant_context import get_tenant_id_from_request
    except ImportError:
        def get_tenant_id_from_request(request):
            """Fallback if tenant utils not available"""
            if hasattr(request, 'tenant_id'):
                return request.tenant_id
            elif hasattr(request, 'tenant') and request.tenant:
                if hasattr(request.tenant, 'tenant_id'):
                    return request.tenant.tenant_id
                elif isinstance(request.tenant, int):
                    return request.tenant
            return None


class AllVendorsListView(APIView):
    """
    API endpoint to list all vendors with 4 types:
    1. Vendor onboarded with RFP (in both vendors & temp_vendor with response_id)
    2. Vendor onboarded without RFP (in both vendors & temp_vendor without response_id)
    3. Temporary vendor with RFP (only in temp_vendor with response_id)
    4. Temporary vendor without RFP (only in temp_vendor without response_id)
    """
    permission_classes = [AllowAny]

    def get(self, request):
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[AllVendorsListView] GET request received for /api/v1/management/vendors/all/")
        
        try:
            # Get tenant_id from request - try multiple methods like RFP views
            tenant_id = None
            
            # Method 1: Check if already set on request
            if hasattr(request, 'tenant_id') and request.tenant_id:
                tenant_id = request.tenant_id
                logger.info(f"[AllVendorsListView] Got tenant_id from request.tenant_id: {tenant_id}")
            
            # Method 2: Try tenant utils function
            if not tenant_id:
                tenant_id = get_tenant_id_from_request(request)
                if tenant_id:
                    logger.info(f"[AllVendorsListView] Got tenant_id from get_tenant_id_from_request: {tenant_id}")
            
            # Method 2.5: Try to extract from JWT token payload
            if not tenant_id:
                try:
                    auth_header = request.headers.get('Authorization', '')
                    if auth_header.startswith('Bearer '):
                        token = auth_header.split(' ')[1]
                        # Decode JWT token to get payload
                        import jwt
                        from django.conf import settings
                        try:
                            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'], options={"verify_signature": False})
                            if 'tenant_id' in payload:
                                tenant_id = payload['tenant_id']
                                logger.info(f"[AllVendorsListView] Got tenant_id from JWT token: {tenant_id}")
                        except Exception as jwt_error:
                            logger.debug(f"[AllVendorsListView] Could not decode JWT: {jwt_error}")
                except Exception as e:
                    logger.debug(f"[AllVendorsListView] Error extracting tenant_id from JWT: {e}")
            
            # Method 3: Try to get from authenticated user
            if not tenant_id and hasattr(request, 'user') and request.user and request.user.is_authenticated:
                try:
                    user = request.user
                    # Check if user has tenant_id attribute
                    if hasattr(user, 'tenant_id') and user.tenant_id:
                        tenant_id = user.tenant_id
                        logger.info(f"[AllVendorsListView] Got tenant_id from user.tenant_id: {tenant_id}")
                    elif hasattr(user, 'tenant') and user.tenant:
                        tenant_id = user.tenant.tenant_id
                        logger.info(f"[AllVendorsListView] Got tenant_id from user.tenant.tenant_id: {tenant_id}")
                    else:
                        # Try to get user_id and look up user from database
                        user_id = None
                        if hasattr(user, 'userid'):
                            user_id = user.userid
                        elif hasattr(user, 'id'):
                            user_id = user.id
                        elif hasattr(user, 'UserId'):
                            user_id = user.UserId
                        
                        if user_id:
                            # Try different User models - handle database connection errors gracefully
                            try:
                                from django.db import connection
                                # Check if database connection is available
                                connection.ensure_connection()
                                
                                try:
                                    from grc.models import Users
                                    db_user = Users.objects.get(UserId=user_id)
                                    if hasattr(db_user, 'tenant_id') and db_user.tenant_id:
                                        tenant_id = db_user.tenant_id
                                        logger.info(f"[AllVendorsListView] Got tenant_id from Users model (UserId={user_id}): {tenant_id}")
                                    elif hasattr(db_user, 'TenantId') and db_user.TenantId:
                                        tenant_id = db_user.TenantId
                                        logger.info(f"[AllVendorsListView] Got tenant_id from Users.TenantId (UserId={user_id}): {tenant_id}")
                                except Exception as e1:
                                    # Check if it's a database connection error
                                    error_str = str(e1).lower()
                                    if 'too many connections' in error_str or '1040' in error_str:
                                        logger.warning(f"[AllVendorsListView] Database connection pool exhausted, skipping tenant lookup from Users model")
                                    else:
                                        try:
                                            from bcpdrp.models import Users
                                            db_user = Users.objects.get(user_id=user_id)
                                            if hasattr(db_user, 'tenant_id') and db_user.tenant_id:
                                                tenant_id = db_user.tenant_id
                                                logger.info(f"[AllVendorsListView] Got tenant_id from bcpdrp.Users (user_id={user_id}): {tenant_id}")
                                            elif hasattr(db_user, 'TenantId') and db_user.TenantId:
                                                tenant_id = db_user.TenantId
                                                logger.info(f"[AllVendorsListView] Got tenant_id from bcpdrp.Users.TenantId (user_id={user_id}): {tenant_id}")
                                        except Exception as e2:
                                            error_str2 = str(e2).lower()
                                            if 'too many connections' in error_str2 or '1040' in error_str2:
                                                logger.warning(f"[AllVendorsListView] Database connection pool exhausted, skipping tenant lookup from bcpdrp.Users model")
                                            else:
                                                logger.warning(f"[AllVendorsListView] Could not get tenant_id from user lookup: {e1}, {e2}")
                            except Exception as db_error:
                                error_str = str(db_error).lower()
                                if 'too many connections' in error_str or '1040' in error_str:
                                    logger.warning(f"[AllVendorsListView] Database connection pool exhausted, cannot lookup tenant_id. Will show all vendors.")
                                else:
                                    logger.warning(f"[AllVendorsListView] Database error during tenant lookup: {db_error}")
                except Exception as e:
                    logger.warning(f"[AllVendorsListView] Error extracting tenant_id from user: {e}")
            
            # If still no tenant_id, use default tenant (1) for development
            if not tenant_id:
                tenant_id = 1
                logger.info(f"[AllVendorsListView] Using default tenant_id: {tenant_id}")
            
            logger.info(f"[AllVendorsListView] Final tenant_id: {tenant_id}")
            
            # Build querysets with tenant filtering - handle database errors gracefully
            try:
                from django.db import connection
                connection.ensure_connection()
                
                vendors_qs = Vendors.objects.all()
                temp_vendors_qs = TempVendor.objects.all()
                
                # Apply tenant filter if tenant_id is available
                if tenant_id:
                    vendors_qs = vendors_qs.filter(tenant_id=tenant_id)
                    temp_vendors_qs = temp_vendors_qs.filter(tenant_id=tenant_id)
                    logger.info(f"[AllVendorsListView] Filtering by tenant_id: {tenant_id}")
                else:
                    logger.warning(f"[AllVendorsListView] No tenant_id found - showing ALL vendors (no tenant filtering)")
            except Exception as db_error:
                error_str = str(db_error).lower()
                if 'too many connections' in error_str or '1040' in error_str:
                    logger.error(f"[AllVendorsListView] Database connection pool exhausted! Cannot query vendors.")
                    return Response({
                        'success': False,
                        'error': 'Database connection pool exhausted. Please try again later or contact administrator.',
                        'error_code': 'DB_CONNECTION_ERROR'
                    }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
                else:
                    logger.error(f"[AllVendorsListView] Database error: {db_error}")
                    return Response({
                        'success': False,
                        'error': f'Database error: {str(db_error)}'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Get all vendor codes from both tables - handle database errors
            try:
                vendor_codes_vendors = set(vendors_qs.values_list('vendor_code', flat=True))
                vendor_codes_temp = set(temp_vendors_qs.values_list('vendor_code', flat=True))
            except Exception as db_error:
                error_str = str(db_error).lower()
                if 'too many connections' in error_str or '1040' in error_str:
                    logger.error(f"[AllVendorsListView] Database connection pool exhausted while fetching vendor codes!")
                    return Response({
                        'success': False,
                        'error': 'Database connection pool exhausted. Please try again later or contact administrator.',
                        'error_code': 'DB_CONNECTION_ERROR'
                    }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
                else:
                    raise  # Re-raise other database errors
            
            all_vendors = []
            
            # Type 1 & 2: Vendors in both tables (onboarded)
            onboarded_codes = vendor_codes_vendors.intersection(vendor_codes_temp)
            for vendor_code in onboarded_codes:
                vendor_filter = {'vendor_code': vendor_code}
                temp_filter = {'vendor_code': vendor_code}
                
                if tenant_id:
                    vendor_filter['tenant_id'] = tenant_id
                    temp_filter['tenant_id'] = tenant_id
                
                vendor = Vendors.objects.filter(**vendor_filter).first()
                temp_vendor = TempVendor.objects.filter(**temp_filter).first()
                
                if vendor:
                    vendor_data = VendorsSerializer(vendor).data
                    
                    # Check if has response_id
                    if temp_vendor and temp_vendor.response_id:
                        vendor_data['vendor_type'] = 'ONBOARDED_WITH_RFP'
                        vendor_data['vendor_type_label'] = 'Vendor Onboarded with RFP'
                    else:
                        vendor_data['vendor_type'] = 'ONBOARDED_WITHOUT_RFP'
                        vendor_data['vendor_type_label'] = 'Vendor Onboarded without RFP'
                    
                    vendor_data['response_id'] = temp_vendor.response_id if temp_vendor else None
                    vendor_data['is_temporary'] = False
                    all_vendors.append(vendor_data)
            
            # Type 3 & 4: Temporary vendors (only in temp_vendor)
            temp_only_codes = vendor_codes_temp - vendor_codes_vendors
            for vendor_code in temp_only_codes:
                temp_filter = {'vendor_code': vendor_code}
                if tenant_id:
                    temp_filter['tenant_id'] = tenant_id
                
                temp_vendor = TempVendor.objects.filter(**temp_filter).first()
                
                if temp_vendor:
                    vendor_data = TempVendorSerializer(temp_vendor).data
                    
                    # Check if has response_id
                    if temp_vendor.response_id:
                        vendor_data['vendor_type'] = 'TEMPORARY_WITH_RFP'
                        vendor_data['vendor_type_label'] = 'Temporary Vendor with RFP'
                    else:
                        vendor_data['vendor_type'] = 'TEMPORARY_WITHOUT_RFP'
                        vendor_data['vendor_type_label'] = 'Temporary Vendor without RFP'
                    
                    vendor_data['is_temporary'] = True
                    all_vendors.append(vendor_data)
            
            # Sort by created_at (most recent first)
            all_vendors.sort(key=lambda x: x.get('created_at', ''), reverse=True)
            
            return Response({
                'success': True,
                'data': all_vendors,
                'total': len(all_vendors),
                'counts': {
                    'onboarded_with_rfp': len([v for v in all_vendors if v['vendor_type'] == 'ONBOARDED_WITH_RFP']),
                    'onboarded_without_rfp': len([v for v in all_vendors if v['vendor_type'] == 'ONBOARDED_WITHOUT_RFP']),
                    'temporary_with_rfp': len([v for v in all_vendors if v['vendor_type'] == 'TEMPORARY_WITH_RFP']),
                    'temporary_without_rfp': len([v for v in all_vendors if v['vendor_type'] == 'TEMPORARY_WITHOUT_RFP']),
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VendorDetailView(APIView):
    """
    API endpoint to get detailed vendor information
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, vendor_code):
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            # Get tenant_id from request - try multiple methods like RFP views
            tenant_id = None
            
            # Method 1: Check if already set on request
            if hasattr(request, 'tenant_id') and request.tenant_id:
                tenant_id = request.tenant_id
            
            # Method 2: Try tenant utils function
            if not tenant_id:
                tenant_id = get_tenant_id_from_request(request)
            
            # Method 2.5: Try to extract from JWT token payload
            if not tenant_id:
                try:
                    auth_header = request.headers.get('Authorization', '')
                    if auth_header.startswith('Bearer '):
                        token = auth_header.split(' ')[1]
                        # Decode JWT token to get payload
                        import jwt
                        from django.conf import settings
                        try:
                            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'], options={"verify_signature": False})
                            if 'tenant_id' in payload:
                                tenant_id = payload['tenant_id']
                                logger.info(f"[VendorDetailView] Got tenant_id from JWT token: {tenant_id}")
                        except Exception as jwt_error:
                            logger.debug(f"[VendorDetailView] Could not decode JWT: {jwt_error}")
                except Exception as e:
                    logger.debug(f"[VendorDetailView] Error extracting tenant_id from JWT: {e}")
            
            # Method 3: Try to get from authenticated user
            if not tenant_id and hasattr(request, 'user') and request.user and request.user.is_authenticated:
                try:
                    user = request.user
                    if hasattr(user, 'tenant_id') and user.tenant_id:
                        tenant_id = user.tenant_id
                        logger.info(f"[VendorDetailView] Got tenant_id from user.tenant_id: {tenant_id}")
                    elif hasattr(user, 'tenant') and user.tenant:
                        tenant_id = user.tenant.tenant_id
                        logger.info(f"[VendorDetailView] Got tenant_id from user.tenant.tenant_id: {tenant_id}")
                    else:
                        # Try to get user_id and look up user from database
                        user_id = None
                        if hasattr(user, 'userid'):
                            user_id = user.userid
                        elif hasattr(user, 'id'):
                            user_id = user.id
                        elif hasattr(user, 'UserId'):
                            user_id = user.UserId
                        
                        if user_id:
                            # Try different User models - handle database connection errors gracefully
                            try:
                                from django.db import connection
                                connection.ensure_connection()
                                
                                try:
                                    from grc.models import Users
                                    db_user = Users.objects.get(UserId=user_id)
                                    if hasattr(db_user, 'tenant_id') and db_user.tenant_id:
                                        tenant_id = db_user.tenant_id
                                        logger.info(f"[VendorDetailView] Got tenant_id from Users model (UserId={user_id}): {tenant_id}")
                                    elif hasattr(db_user, 'TenantId') and db_user.TenantId:
                                        tenant_id = db_user.TenantId
                                        logger.info(f"[VendorDetailView] Got tenant_id from Users.TenantId (UserId={user_id}): {tenant_id}")
                                except Exception as e1:
                                    error_str = str(e1).lower()
                                    if 'too many connections' in error_str or '1040' in error_str:
                                        logger.warning(f"[VendorDetailView] Database connection pool exhausted, skipping tenant lookup")
                                    else:
                                        try:
                                            from bcpdrp.models import Users
                                            db_user = Users.objects.get(user_id=user_id)
                                            if hasattr(db_user, 'tenant_id') and db_user.tenant_id:
                                                tenant_id = db_user.tenant_id
                                                logger.info(f"[VendorDetailView] Got tenant_id from bcpdrp.Users (user_id={user_id}): {tenant_id}")
                                            elif hasattr(db_user, 'TenantId') and db_user.TenantId:
                                                tenant_id = db_user.TenantId
                                                logger.info(f"[VendorDetailView] Got tenant_id from bcpdrp.Users.TenantId (user_id={user_id}): {tenant_id}")
                                        except Exception as e2:
                                            error_str2 = str(e2).lower()
                                            if 'too many connections' in error_str2 or '1040' in error_str2:
                                                logger.warning(f"[VendorDetailView] Database connection pool exhausted, skipping tenant lookup")
                                            else:
                                                logger.warning(f"[VendorDetailView] Could not get tenant_id from user lookup: {e1}, {e2}")
                            except Exception as db_error:
                                error_str = str(db_error).lower()
                                if 'too many connections' in error_str or '1040' in error_str:
                                    logger.warning(f"[VendorDetailView] Database connection pool exhausted, cannot lookup tenant_id")
                                else:
                                    logger.warning(f"[VendorDetailView] Database error during tenant lookup: {db_error}")
                except Exception as e:
                    logger.warning(f"[VendorDetailView] Error extracting tenant_id from user: {e}")
            
            logger.info(f"[VendorDetailView] Final tenant_id: {tenant_id}")
            
            # Build filters
            vendor_filter = {'vendor_code': vendor_code}
            temp_filter = {'vendor_code': vendor_code}
            
            # Apply tenant filter if tenant_id is available
            if tenant_id:
                vendor_filter['tenant_id'] = tenant_id
                temp_filter['tenant_id'] = tenant_id
                logger.info(f"[VendorDetailView] Filtering by tenant_id: {tenant_id}")
            else:
                logger.warning(f"[VendorDetailView] No tenant_id found - searching all tenants")
            
            # Get vendor with optional tenant filtering
            vendor = Vendors.objects.filter(**vendor_filter).first()
            temp_vendor = TempVendor.objects.filter(**temp_filter).first()
            
            if vendor:
                # Return data from vendors table
                vendor_data = VendorsSerializer(vendor).data
                
                if temp_vendor and temp_vendor.response_id:
                    vendor_data['vendor_type'] = 'ONBOARDED_WITH_RFP'
                    vendor_data['vendor_type_label'] = 'Vendor Onboarded with RFP'
                else:
                    vendor_data['vendor_type'] = 'ONBOARDED_WITHOUT_RFP'
                    vendor_data['vendor_type_label'] = 'Vendor Onboarded without RFP'
                
                vendor_data['response_id'] = temp_vendor.response_id if temp_vendor else None
                vendor_data['is_temporary'] = False
                
            elif temp_vendor:
                # Return data from temp_vendor table
                vendor_data = TempVendorSerializer(temp_vendor).data
                
                if temp_vendor.response_id:
                    vendor_data['vendor_type'] = 'TEMPORARY_WITH_RFP'
                    vendor_data['vendor_type_label'] = 'Temporary Vendor with RFP'
                else:
                    vendor_data['vendor_type'] = 'TEMPORARY_WITHOUT_RFP'
                    vendor_data['vendor_type_label'] = 'Temporary Vendor without RFP'
                
                vendor_data['is_temporary'] = True
            else:
                return Response({
                    'success': False,
                    'error': 'Vendor not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get vendor_id for fetching related data
            vendor_id = None
            if vendor:
                vendor_id = vendor.vendor_id
            elif temp_vendor:
                # For temporary vendors, we need to find the vendor_id from vendors table if it exists
                try:
                    matching_vendor = Vendors.objects.filter(vendor_code=vendor_code).first()
                    if matching_vendor:
                        vendor_id = matching_vendor.vendor_id
                except Exception:
                    pass
            
            # Fetch related data if vendor_id exists
            related_data = {}
            if vendor_id:
                try:
                    # Import models with correct paths
                    from tprm_backend.contracts.models import VendorContract, ContractTerm, ContractClause
                    from tprm_backend.slas.models import VendorSLA, SLAMetric
                    from tprm_backend.bcpdrp.models import Plan
                    
                    # Build tenant filter for related data - use tenant_id for ForeignKey filtering
                    related_filter = {}
                    if tenant_id:
                        related_filter['tenant_id'] = tenant_id
                    
                    # Fetch contracts - VendorContract uses vendor as ForeignKey
                    # Try vendor_id first (Django auto-created field), then vendor__vendor_id (through relationship)
                    try:
                        contracts_qs = VendorContract.objects.filter(vendor_id=vendor_id)
                    except Exception as e1:
                        logger.warning(f"[VendorDetailView] vendor_id filter failed, trying vendor__vendor_id: {e1}")
                        try:
                            contracts_qs = VendorContract.objects.filter(vendor__vendor_id=vendor_id)
                        except Exception as e2:
                            logger.error(f"[VendorDetailView] Both vendor_id and vendor__vendor_id filters failed: {e2}")
                            contracts_qs = VendorContract.objects.none()
                    
                    if tenant_id:
                        contracts_qs = contracts_qs.filter(tenant_id=tenant_id)
                    contracts = list(contracts_qs.values())
                    related_data['contracts'] = contracts
                    logger.info(f"[VendorDetailView] Found {len(contracts)} contracts for vendor_id={vendor_id}")
                    
                    # Fetch contract terms and clauses for each contract
                    contract_ids = [c['contract_id'] for c in contracts]
                    if contract_ids:
                        terms_qs = ContractTerm.objects.filter(contract_id__in=contract_ids)
                        clauses_qs = ContractClause.objects.filter(contract_id__in=contract_ids)
                        if tenant_id:
                            terms_qs = terms_qs.filter(tenant_id=tenant_id)
                            clauses_qs = clauses_qs.filter(tenant_id=tenant_id)
                        related_data['contract_terms'] = list(terms_qs.values())
                        related_data['contract_clauses'] = list(clauses_qs.values())
                    else:
                        related_data['contract_terms'] = []
                        related_data['contract_clauses'] = []
                    
                    # Fetch SLAs - VendorSLA uses vendor as ForeignKey with db_column='vendor_id'
                    # So we can use vendor_id directly
                    slas_qs = VendorSLA.objects.filter(vendor_id=vendor_id)
                    if tenant_id:
                        slas_qs = slas_qs.filter(tenant_id=tenant_id)
                    slas = list(slas_qs.values())
                    related_data['slas'] = slas
                    logger.info(f"[VendorDetailView] Found {len(slas)} SLAs for vendor_id={vendor_id}")
                    
                    # Fetch SLA metrics for each SLA
                    sla_ids = [s['sla_id'] for s in slas]
                    if sla_ids:
                        metrics_qs = SLAMetric.objects.filter(sla_id__in=sla_ids)
                        if tenant_id:
                            metrics_qs = metrics_qs.filter(tenant_id=tenant_id)
                        related_data['sla_metrics'] = list(metrics_qs.values())
                    else:
                        related_data['sla_metrics'] = []
                    
                    # Fetch BCP/DRP plans - Plan uses vendor_id as IntegerField
                    plans_qs = Plan.objects.filter(vendor_id=vendor_id)
                    if tenant_id:
                        plans_qs = plans_qs.filter(tenant_id=tenant_id)
                    plans = list(plans_qs.values())
                    related_data['bcp_drp_plans'] = plans
                    logger.info(f"[VendorDetailView] Found {len(plans)} BCP/DRP plans for vendor_id={vendor_id}")
                    
                    # Fetch Contract Audits - based on contracts linked to this vendor
                    if contract_ids:
                        try:
                            from tprm_backend.audits_contract.models import ContractAudit, ContractAuditFinding
                            
                            contract_audits_qs = ContractAudit.objects.filter(contract_id__in=contract_ids)
                            if tenant_id:
                                contract_audits_qs = contract_audits_qs.filter(tenant_id=tenant_id)
                            contract_audits = list(contract_audits_qs.values())
                            related_data['contract_audits'] = contract_audits
                            logger.info(f"[VendorDetailView] Found {len(contract_audits)} contract audits for contracts {contract_ids}")
                            
                            # Fetch contract audit findings
                            contract_audit_ids = [ca['audit_id'] for ca in contract_audits]
                            if contract_audit_ids:
                                contract_findings_qs = ContractAuditFinding.objects.filter(audit_id__in=contract_audit_ids)
                                if tenant_id:
                                    contract_findings_qs = contract_findings_qs.filter(tenant_id=tenant_id)
                                related_data['contract_audit_findings'] = list(contract_findings_qs.values())
                                logger.info(f"[VendorDetailView] Found {len(related_data['contract_audit_findings'])} contract audit findings")
                            else:
                                related_data['contract_audit_findings'] = []
                        except ImportError as ie:
                            logger.warning(f"[VendorDetailView] Could not import contract audit models: {ie}")
                            related_data['contract_audits'] = []
                            related_data['contract_audit_findings'] = []
                        except Exception as ae:
                            logger.warning(f"[VendorDetailView] Error fetching contract audits: {ae}")
                            related_data['contract_audits'] = []
                            related_data['contract_audit_findings'] = []
                    else:
                        related_data['contract_audits'] = []
                        related_data['contract_audit_findings'] = []
                    
                    # Fetch SLA Audits - based on SLAs linked to this vendor
                    if sla_ids:
                        try:
                            from tprm_backend.audits.models import Audit, AuditFinding
                            
                            sla_audits_qs = Audit.objects.filter(sla_id__in=sla_ids)
                            if tenant_id:
                                sla_audits_qs = sla_audits_qs.filter(tenant_id=tenant_id)
                            sla_audits = list(sla_audits_qs.values())
                            related_data['sla_audits'] = sla_audits
                            logger.info(f"[VendorDetailView] Found {len(sla_audits)} SLA audits for SLAs {sla_ids}")
                            
                            # Fetch SLA audit findings
                            sla_audit_ids = [sa['audit_id'] for sa in sla_audits]
                            if sla_audit_ids:
                                sla_findings_qs = AuditFinding.objects.filter(audit_id__in=sla_audit_ids)
                                if tenant_id:
                                    sla_findings_qs = sla_findings_qs.filter(tenant_id=tenant_id)
                                related_data['sla_audit_findings'] = list(sla_findings_qs.values())
                                logger.info(f"[VendorDetailView] Found {len(related_data['sla_audit_findings'])} SLA audit findings")
                            else:
                                related_data['sla_audit_findings'] = []
                        except ImportError as ie:
                            logger.warning(f"[VendorDetailView] Could not import SLA audit models: {ie}")
                            related_data['sla_audits'] = []
                            related_data['sla_audit_findings'] = []
                        except Exception as ae:
                            logger.warning(f"[VendorDetailView] Error fetching SLA audits: {ae}")
                            related_data['sla_audits'] = []
                            related_data['sla_audit_findings'] = []
                    else:
                        related_data['sla_audits'] = []
                        related_data['sla_audit_findings'] = []
                    
                except Exception as e:
                    logger.exception(f"[VendorDetailView] Error fetching related data: {e}")
                    import traceback
                    logger.error(f"[VendorDetailView] Traceback: {traceback.format_exc()}")
                    related_data = {
                        'contracts': [],
                        'contract_terms': [],
                        'contract_clauses': [],
                        'slas': [],
                        'sla_metrics': [],
                        'bcp_drp_plans': [],
                        'contract_audits': [],
                        'contract_audit_findings': [],
                        'sla_audits': [],
                        'sla_audit_findings': []
                    }
            else:
                related_data = {
                    'contracts': [],
                    'contract_terms': [],
                    'contract_clauses': [],
                    'slas': [],
                    'sla_metrics': [],
                    'bcp_drp_plans': [],
                    'contract_audits': [],
                    'contract_audit_findings': [],
                    'sla_audits': [],
                    'sla_audit_findings': []
                }
            
            # Add related data to response
            vendor_data['related_data'] = related_data
            
            return Response({
                'success': True,
                'data': vendor_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception(f"[VendorDetailView] Unexpected error: {e}")
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
"""
Views for Management app - TempVendor operations
"""
 
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Q
from django.utils import timezone
import logging
 
from tprm_backend.apps.vendor_core.models import TempVendor, Users, ExternalScreeningResult, ScreeningMatch
from tprm_backend.apps.vendor_core.services import OFACService
from tprm_backend.apps.vendor_core.vendor_authentication import (
    JWTAuthentication,
    VendorAuthenticationMixin
)
from tprm_backend.core.tenant_utils import get_tenant_id_from_request
from tprm_backend.rbac.models import RBACTPRM
from .serializers import TempVendorManagementSerializer
from django.contrib.auth.hashers import make_password
from django.core.mail import EmailMessage, send_mail
from django.conf import settings
from django.db import connections
import secrets
import string
import threading
import json
 
# Initialize logger
logger = logging.getLogger('management')
 
 
class TempVendorManagementViewSet(VendorAuthenticationMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing TempVendor records
    Provides CRUD operations for temporary vendor registrations
    MULTI-TENANCY: Filters by tenant to ensure tenant isolation
    """
    queryset = TempVendor.objects.all()
    serializer_class = TempVendorManagementSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
   
    def get_queryset(self):
        """Get temp vendors with tenant filtering"""
        tenant_id = get_tenant_id_from_request(self.request)
       
        # Use 'tprm' database connection which points to tprm_integrations schema
        if tenant_id:
            queryset = TempVendor.objects.using('tprm').filter(tenant_id=tenant_id)
            logger.info(f"Filtering temp vendors by tenant_id: {tenant_id}")
        else:
            queryset = TempVendor.objects.using('tprm').all()
            logger.warning("No tenant_id found - returning all temp vendors")
       
        queryset = queryset.order_by('-created_at')
       
        # Apply search filter
        search_term = self.request.query_params.get('search', None)
        if search_term:
            search_term = search_term.strip()
            if search_term:
                queryset = queryset.filter(
                    Q(company_name__icontains=search_term) |
                    Q(legal_name__icontains=search_term) |
                    Q(vendor_code__icontains=search_term)
                )
                logger.info(f"Filtered temp vendors by search term: '{search_term}'")
       
        return queryset
   
    def create(self, request, *args, **kwargs):
        """Create a new temp vendor"""
        try:
            tenant_id = get_tenant_id_from_request(request)
            user_id = request.data.get('UserId') or request.data.get('userid')
           
            # Prepare data with tenant information
            data = request.data.copy()
            if tenant_id:
                data['TenantId'] = tenant_id
            if user_id:
                data['userid'] = user_id
                data['UserId'] = user_id
           
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
           
            # Save with tenant and user info
            save_kwargs = {}
            if tenant_id:
                save_kwargs['tenant_id'] = tenant_id
            if user_id:
                save_kwargs['userid'] = user_id
           
            # Save using 'tprm' database connection which points to tprm_integrations schema
            temp_vendor = serializer.save(**save_kwargs)
            # Ensure we're using the correct database
            if hasattr(temp_vendor, '_state'):
                temp_vendor._state.db = 'tprm'
           
            logger.info(f"Created temp vendor {temp_vendor.id} for user {user_id}, tenant {tenant_id}")
           
            # Automatically create vendor credentials (user account and RBAC)
            created_user_id = None
            try:
                created_user_id = self._create_vendor_credentials(temp_vendor, tenant_id)
                if created_user_id:
                    # Update temp_vendor with the created user_id
                    temp_vendor.userid = created_user_id
                    temp_vendor.save(using='tprm')
                    logger.info(f"Updated temp vendor {temp_vendor.id} with user_id {created_user_id}")
            except Exception as cred_error:
                logger.error(f"Error creating vendor credentials: {str(cred_error)}")
                # Don't fail the vendor creation if credentials creation fails
                # Just log the error
           
            # Trigger external screening in parallel (non-blocking)
            try:
                screening_thread = threading.Thread(
                    target=self._trigger_external_screening,
                    args=(temp_vendor.id,),
                    daemon=True  # Thread will terminate when main thread exits
                )
                screening_thread.start()
                logger.info(f"Started external screening thread for vendor {temp_vendor.id}")
            except Exception as screening_error:
                logger.error(f"Failed to start external screening for vendor {temp_vendor.id}: {str(screening_error)}")
                # Don't fail vendor creation if screening thread fails to start
           
            return Response({
                'status': 'success',
                'message': 'Vendor created successfully',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
           
        except Exception as e:
            logger.error(f"Error creating temp vendor: {str(e)}")
            return Response({
                'status': 'error',
                'message': f'Failed to create vendor: {str(e)}',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
   
    def update(self, request, *args, **kwargs):
        """Update a temp vendor"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
           
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            serializer.save()
           
            logger.info(f"Updated temp vendor {instance.id}")
           
            return Response({
                'status': 'success',
                'message': 'Vendor updated successfully',
                'data': serializer.data
            }, status=status.HTTP_200_OK)
           
        except Exception as e:
            logger.error(f"Error updating temp vendor: {str(e)}")
            return Response({
                'status': 'error',
                'message': f'Failed to update vendor: {str(e)}',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
   
    def list(self, request, *args, **kwargs):
        """List temp vendors"""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
           
            return Response({
                'status': 'success',
                'count': len(serializer.data),
                'results': serializer.data
            }, status=status.HTTP_200_OK)
           
        except Exception as e:
            logger.error(f"Error listing temp vendors: {str(e)}")
            return Response({
                'status': 'error',
                'message': f'Failed to list vendors: {str(e)}',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
    def retrieve(self, request, *args, **kwargs):
        """Retrieve a single temp vendor"""
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
           
            return Response({
                'status': 'success',
                'data': serializer.data
            }, status=status.HTTP_200_OK)
           
        except Exception as e:
            logger.error(f"Error retrieving temp vendor: {str(e)}")
            return Response({
                'status': 'error',
                'message': f'Failed to retrieve vendor: {str(e)}',
                'error': str(e)
            }, status=status.HTTP_404_NOT_FOUND)
   
    def destroy(self, request, *args, **kwargs):
        """Delete a temp vendor"""
        try:
            instance = self.get_object()
            instance_id = instance.id
            instance.delete()
           
            logger.info(f"Deleted temp vendor {instance_id}")
           
            return Response({
                'status': 'success',
                'message': 'Vendor deleted successfully'
            }, status=status.HTTP_204_NO_CONTENT)
           
        except Exception as e:
            logger.error(f"Error deleting temp vendor: {str(e)}")
            return Response({
                'status': 'error',
                'message': f'Failed to delete vendor: {str(e)}',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
   
    def _create_vendor_credentials(self, temp_vendor, tenant_id):
        """
        Create vendor credentials: user account and RBAC entry
        Returns the created user_id
        """
        from django.db import connections
       
        # Generate username from vendor code or company name
        username_base = temp_vendor.vendor_code or temp_vendor.company_name or f"vendor_{temp_vendor.id}"
        # Clean username: remove special characters, spaces, lowercase
        username_base = ''.join(c.lower() if c.isalnum() else '_' for c in username_base)
       
        # Generate unique username by checking if it exists using raw SQL
        db_connection = connections['tprm']
        username = username_base
        counter = 1
        with db_connection.cursor() as cursor:
            while True:
                cursor.execute("SELECT COUNT(*) FROM users WHERE UserName = %s", [username])
                if cursor.fetchone()[0] == 0:
                    break
                username = f"{username_base}_{counter}"
                counter += 1
       
        # Generate a random password
        password_length = 12
        alphabet = string.ascii_letters + string.digits
        random_password = ''.join(secrets.choice(alphabet) for i in range(password_length))
       
        # Hash the password
        hashed_password = make_password(random_password)
       
        # Get all primary contacts (contacts where isPrimary or primary_contact is true)
        primary_contacts = []
       
        if temp_vendor.contacts and isinstance(temp_vendor.contacts, list):
            # Find all contacts marked as primary
            primary_contacts = [
                c for c in temp_vendor.contacts
                if (c.get('isPrimary') == True or c.get('primary_contact') == True)
                and c.get('email')  # Must have an email
            ]
           
            # If no primary contacts found, use first contact with email as fallback
            if not primary_contacts:
                first_contact_with_email = next((c for c in temp_vendor.contacts if c.get('email')), None)
                if first_contact_with_email:
                    primary_contacts = [first_contact_with_email]
       
        # Get email for user account (use first primary contact email or generate one)
        email = primary_contacts[0].get('email') if primary_contacts else None
        if not email:
            # If no email found, generate one from company name
            email_base = temp_vendor.company_name or username_base
            email_base = ''.join(c.lower() if c.isalnum() else '' for c in email_base)
            email = f"{email_base}@vendor.tprm.local"
       
        # Split company name for first/last name
        company_name = temp_vendor.company_name or 'Vendor'
        name_parts = company_name.split(' ', 1)
        first_name = name_parts[0][:45] if name_parts else 'Vendor'
        last_name = name_parts[1][:45] if len(name_parts) > 1 else ''
       
        # Create user in users table using raw SQL to handle all fields
        try:
            with db_connection.cursor() as cursor:
                # Insert user with all required fields
                cursor.execute("""
                    INSERT INTO users
                    (UserName, Password, Email, FirstName, LastName, IsActive, TenantId, CreatedAt, UpdatedAt)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    username,
                    hashed_password,
                    email[:100],  # Ensure email doesn't exceed max length
                    first_name[:45],
                    last_name[:45],
                    'Y',  # IsActive
                    tenant_id,
                    timezone.now(),
                    timezone.now()
                ])
                created_user_id = cursor.lastrowid
                logger.info(f"Created user account for vendor: UserId={created_user_id}, UserName={username}, Email={email}")
           
            # Create RBAC entry with role "Vendor" and only "SubmitQuestionnaireResponses" permission using raw SQL
            with db_connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO rbac_tprm
                    (UserId, UserName, Role, SubmitQuestionnaireResponses, TenantId, CreatedAt, UpdatedAt, IsActive)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    created_user_id,
                    username,
                    'Vendor',
                    1,  # SubmitQuestionnaireResponses = True
                    tenant_id,
                    timezone.now(),
                    timezone.now(),
                    'Y'  # IsActive
                ])
                logger.info(f"Created RBAC entry for vendor: Role=Vendor, UserId={created_user_id}, SubmitQuestionnaireResponses=True")
           
            # Send credentials email to all primary contacts
            if primary_contacts:
                email_count = 0
                email_errors = []
               
                for primary_contact in primary_contacts:
                    contact_email = primary_contact.get('email')
                    contact_name = primary_contact.get('name') or temp_vendor.company_name or 'Vendor'
                   
                    if contact_email:
                        try:
                            self._send_vendor_credentials_email(
                                vendor_name=contact_name,
                                vendor_email=contact_email,
                                company_name=temp_vendor.company_name or 'Your Company',
                                username=username,
                                password=random_password  # Send plain password in email
                            )
                            email_count += 1
                            logger.info(f"Sent credentials email to primary contact: {contact_email} ({contact_name})")
                        except Exception as email_error:
                            # Log email error but continue sending to other primary contacts
                            error_msg = f"Failed to send credentials email to {contact_email}: {str(email_error)}"
                            logger.error(error_msg)
                            email_errors.append(error_msg)
               
                if email_count > 0:
                    logger.info(f"Successfully sent credentials emails to {email_count} primary contact(s) for vendor {temp_vendor.id}")
                else:
                    logger.warning(f"Failed to send credentials emails to any primary contacts for vendor {temp_vendor.id}")
                    if email_errors:
                        logger.error(f"Email errors: {'; '.join(email_errors)}")
            else:
                logger.warning(f"No primary contact email found for vendor {temp_vendor.id}. Credentials email not sent.")
           
            return created_user_id
           
        except Exception as e:
            logger.error(f"Error creating vendor credentials: {str(e)}")
            raise
   
    def _send_vendor_credentials_email(self, vendor_name, vendor_email, company_name, username, password):
        """
        Send vendor credentials via email to welcome them to the system.
       
        Args:
            vendor_name: Name of the primary contact or vendor
            vendor_email: Email address to send to (primary contact email)
            company_name: Company name
            username: Generated username
            password: Generated password (plain text for email)
        """
        try:
            subject = 'Welcome to TPRM Vendor Portal - Your Access Credentials'
 
            portal_base_url = getattr(settings, 'FRONTEND_URL', None) or getattr(settings, 'SITE_URL', None) or 'http://localhost:3000'
            portal_login_url = f"{portal_base_url.rstrip('/')}/login"
 
            # Create a professional HTML email
            html_message = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0; }}
                    .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                    .header {{ background: linear-gradient(135deg, #3b82f6 0%, #1e40af 100%); color: white; padding: 30px; border-radius: 8px 8px 0 0; text-align: center; }}
                    .content {{ background: #f9fafb; padding: 30px; border-radius: 0 0 8px 8px; }}
                    .credentials-box {{ background: white; padding: 25px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #3b82f6; }}
                    .credential-item {{ margin: 15px 0; padding: 10px; background: #f3f4f6; border-radius: 4px; }}
                    .label {{ font-weight: 600; color: #1f2937; display: inline-block; min-width: 120px; }}
                    .value {{ color: #374151; font-family: monospace; }}
                    .warning {{ background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 4px; }}
                    .footer {{ text-align: center; color: #6b7280; font-size: 12px; margin-top: 30px; padding-top: 20px; border-top: 1px solid #e5e7eb; }}
                    .button {{ display: inline-block; background: #3b82f6; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; margin: 20px 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1 style="margin: 0;">Welcome to TPRM Vendor Portal</h1>
                    </div>
                    <div class="content">
                        <p>Dear <strong>{vendor_name}</strong>,</p>
                        <p>We are pleased to welcome <strong>{company_name}</strong> to our Third-Party Risk Management (TPRM) system. Your vendor account has been successfully created, and you now have access to the vendor portal.</p>
                       
                        <div class="credentials-box">
                            <h3 style="margin-top: 0; color: #1f2937;">Your Access Credentials</h3>
                            <div class="credential-item">
                                <span class="label">Portal URL:</span>
                                <span class="value"><a href="{portal_login_url}" style="color: #3b82f6;">{portal_login_url}</a></span>
                            </div>
                            <div class="credential-item">
                                <span class="label">Username:</span>
                                <span class="value">{username}</span>
                            </div>
                            <div class="credential-item">
                                <span class="label">Email:</span>
                                <span class="value">{vendor_email}</span>
                            </div>
                            <div class="credential-item">
                                <span class="label">Password:</span>
                                <span class="value">{password}</span>
                            </div>
                        </div>
                       
                        <div class="warning">
                            <strong>⚠️ Important Security Notice:</strong> For security reasons, we strongly recommend that you change your password upon first login.
                        </div>
                       
                        <p>Through the vendor portal, you will be able to:</p>
                        <ul style="line-height: 2;">
                            <li>Submit questionnaire responses</li>
                            <li>View and manage your vendor profile</li>
                            <li>Upload required documents</li>
                            <li>Track your vendor lifecycle status</li>
                            <li>View performance metrics and assessments</li>
                        </ul>
                       
                        <p style="text-align: center;">
                            <a href="{portal_login_url}" class="button">Access Vendor Portal</a>
                        </p>
                       
                        <p>If you have any questions or need assistance, please don't hesitate to contact our support team.</p>
                       
                        <p>Best regards,<br><strong>TPRM Team</strong></p>
                    </div>
                    <div class="footer">
                        <p>This is an automated message. Please do not reply to this email.<br>
                        For support, contact us through the vendor portal or your designated account manager.</p>
                    </div>
                </div>
            </body>
            </html>
            """
           
            # Plain text version for email clients that don't support HTML
            plain_message = f"""
Dear {vendor_name},
 
We are pleased to welcome {company_name} to our Third-Party Risk Management (TPRM) system. Your vendor account has been successfully created, and you now have access to the vendor portal.
 
Below are your credentials to access the Vendor Portal:
 
Portal URL: {portal_login_url}
Username: {username}
Email: {vendor_email}
Password: {password}
 
IMPORTANT: For security reasons, we strongly recommend that you change your password upon first login.
 
Through the vendor portal, you will be able to:
- Submit questionnaire responses
- View and manage your vendor profile
- Upload required documents
- Track your vendor lifecycle status
- View performance metrics and assessments
 
If you have any questions or need assistance, please don't hesitate to contact our support team.
 
Best regards,
TPRM Team
 
---
This is an automated message. Please do not reply to this email.
For support, contact us through the vendor portal or your designated account manager.
            """
           
            # Use EmailMessage for HTML support
            from_email = settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@tprm.com'
           
            logger.info(f"Preparing to send credentials email to {vendor_email} from {from_email}")
           
            # Create email message
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=from_email,
                to=[vendor_email],
            )
            email.content_subtype = "html"  # Set content type to HTML
           
            # Send email with explicit error handling
            try:
                logger.info(f"Sending credentials email via EmailMessage to {vendor_email}...")
                result = email.send(fail_silently=False)
                logger.info(f"Credentials email sent successfully to {vendor_email}")
            except Exception as send_error:
                logger.error(f"EmailMessage.send() failed: {str(send_error)}")
               
                # Try fallback with send_mail
                try:
                    logger.info(f"Attempting fallback email send using send_mail to {vendor_email}...")
                    send_mail(
                        subject=subject,
                        message=plain_message,
                        from_email=from_email,
                        recipient_list=[vendor_email],
                        fail_silently=False,
                    )
                    logger.info(f"Fallback email sent successfully to {vendor_email}")
                except Exception as fallback_error:
                    error_msg = f"Both EmailMessage and send_mail failed. EmailMessage error: {str(send_error)}, send_mail error: {str(fallback_error)}"
                    logger.error(error_msg, exc_info=True)
                    raise Exception(error_msg) from send_error
           
        except Exception as e:
            logger.error(f"Error sending vendor credentials email: {str(e)}", exc_info=True)
            raise
   
    def _trigger_external_screening(self, vendor_id):
        """
        Trigger external screening for a vendor in parallel (non-blocking).
        This method runs in a separate thread after vendor creation.
        """
        try:
            logger.info(f"Starting external screening for vendor {vendor_id}")
           
            # Get the temp vendor instance
            try:
                vendor = TempVendor.objects.using('tprm').get(id=vendor_id)
            except TempVendor.DoesNotExist:
                logger.error(f"Vendor {vendor_id} not found for external screening")
                return
           
            # Perform comprehensive automatic screening
            screening_results = self._perform_automatic_screening(vendor)
           
            logger.info(f"Completed external screening for vendor {vendor_id} with {len(screening_results) if screening_results else 0} results")
           
        except Exception as e:
            logger.error(f"Error in external screening thread for vendor {vendor_id}: {str(e)}", exc_info=True)
   
    def _perform_automatic_screening(self, vendor):
        """
        Perform comprehensive automatic screening for a vendor across multiple sources.
        Adapted from vendor_core/views.py
        """
        logger.info(f"Starting comprehensive screening for vendor {vendor.id}: {vendor.company_name}")
       
        screening_results = []
       
        # 1. OFAC Screening
        try:
            ofac_result = self._perform_ofac_screening(vendor)
            if ofac_result:
                screening_results.append(ofac_result)
        except Exception as e:
            logger.error(f"OFAC screening failed for vendor {vendor.id}: {str(e)}")
       
        # 2. PEP (Politically Exposed Person) Screening
        try:
            pep_result = self._perform_pep_screening(vendor)
            if pep_result:
                screening_results.append(pep_result)
        except Exception as e:
            logger.error(f"PEP screening failed for vendor {vendor.id}: {str(e)}")
       
        # 3. Sanctions Screening
        try:
            sanctions_result = self._perform_sanctions_screening(vendor)
            if sanctions_result:
                screening_results.append(sanctions_result)
        except Exception as e:
            logger.error(f"Sanctions screening failed for vendor {vendor.id}: {str(e)}")
       
        # 4. Adverse Media Screening
        try:
            adverse_media_result = self._perform_adverse_media_screening(vendor)
            if adverse_media_result:
                screening_results.append(adverse_media_result)
        except Exception as e:
            logger.error(f"Adverse media screening failed for vendor {vendor.id}: {str(e)}")
       
        logger.info(f"Completed comprehensive screening for vendor {vendor.id} with {len(screening_results)} results")
        return screening_results
   
    def _perform_ofac_screening(self, vendor):
        """Perform OFAC screening for a vendor"""
        try:
            logger.info(f"Starting OFAC screening for vendor {vendor.id}: {vendor.company_name}")
           
            ofac_service = OFACService()
           
            # Create screening record using raw SQL for tprm database
            db_connection = connections['tprm']
            # Get tenant_id from vendor
            tenant_id = None
            if hasattr(vendor, 'tenant_id') and vendor.tenant_id:
                tenant_id = vendor.tenant_id
            elif hasattr(vendor, 'TenantId') and vendor.TenantId:
                tenant_id = vendor.TenantId
            elif hasattr(vendor, 'tenant') and vendor.tenant:
                if hasattr(vendor.tenant, 'tenant_id'):
                    tenant_id = vendor.tenant.tenant_id
                elif isinstance(vendor.tenant, int):
                    tenant_id = vendor.tenant
            
            with db_connection.cursor() as cursor:
                if tenant_id:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated, TenantId)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'OFAC',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name,
                            'tax_id': vendor.tax_id
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now(),
                        tenant_id
                    ])
                else:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'OFAC',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name,
                            'tax_id': vendor.tax_id
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now()
                    ])
                screening_id = cursor.lastrowid
           
            # Get the screening instance
            screening = ExternalScreeningResult.objects.using('tprm').get(screening_id=screening_id)
           
            # Search OFAC database
            search_name = vendor.company_name or vendor.legal_name
            logger.info(f"Searching OFAC database for: {search_name}")
           
            search_results = ofac_service.search_entity(search_name)
            logger.info(f"OFAC search results: {search_results}")
           
            if 'error' in search_results:
                logger.warning(f"OFAC API error for vendor {vendor.id}: {search_results.get('error')}")
                screening.status = 'CLEAR'
                screening.save(using='tprm')
                return None
           
            matches = search_results.get('matches', [])
            logger.info(f"Found {len(matches)} matches for vendor {vendor.id}")
           
            high_risk_count = 0
           
            # Process matches
            for match in matches:
                match_score = ofac_service.calculate_risk_score(match)
                risk_level = ofac_service.determine_risk_level(match_score)
               
                if risk_level == 'HIGH':
                    high_risk_count += 1
               
                # Create match record
                match_details = {
                    **ofac_service.extract_match_details(match),
                    'risk_level': risk_level,
                    'screening_date': timezone.now().isoformat()
                }
               
                with db_connection.cursor() as cursor:
                    if tenant_id:
                        cursor.execute("""
                            INSERT INTO screening_matches
                            (screening_id, match_type, match_score, match_details, TenantId)
                            VALUES (%s, %s, %s, %s, %s)
                        """, [
                            screening_id,
                            f"OFAC - {match.get('source', 'Unknown')}",
                            match_score,
                            json.dumps(match_details),
                            tenant_id
                        ])
                    else:
                        cursor.execute("""
                            INSERT INTO screening_matches
                            (screening_id, match_type, match_score, match_details)
                            VALUES (%s, %s, %s, %s)
                        """, [
                            screening_id,
                            f"OFAC - {match.get('source', 'Unknown')}",
                            match_score,
                            json.dumps(match_details)
                        ])
           
            # Update screening status
            with db_connection.cursor() as cursor:
                status = 'POTENTIAL_MATCH' if high_risk_count > 0 else ('UNDER_REVIEW' if len(matches) > 0 else 'CLEAR')
                cursor.execute("""
                    UPDATE external_screening_results
                    SET total_matches = %s, high_risk_matches = %s, status = %s, last_updated = %s
                    WHERE screening_id = %s
                """, [len(matches), high_risk_count, status, timezone.now(), screening_id])
           
            logger.info(f"OFAC screening completed for vendor {vendor.id} with status: {status}")
            return {'screening_type': 'OFAC', 'status': status, 'total_matches': len(matches)}
           
        except Exception as e:
            logger.error(f"OFAC screening failed for vendor {vendor.id}: {str(e)}", exc_info=True)
            return None
   
    def _perform_pep_screening(self, vendor):
        """Perform PEP (Politically Exposed Person) screening"""
        try:
            db_connection = connections['tprm']
            # Get tenant_id from vendor
            tenant_id = None
            if hasattr(vendor, 'tenant_id') and vendor.tenant_id:
                tenant_id = vendor.tenant_id
            elif hasattr(vendor, 'TenantId') and vendor.TenantId:
                tenant_id = vendor.TenantId
            elif hasattr(vendor, 'tenant') and vendor.tenant:
                if hasattr(vendor.tenant, 'tenant_id'):
                    tenant_id = vendor.tenant.tenant_id
                elif isinstance(vendor.tenant, int):
                    tenant_id = vendor.tenant
            
            with db_connection.cursor() as cursor:
                if tenant_id:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated, TenantId)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'PEP',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now(),
                        tenant_id
                    ])
                else:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'PEP',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now()
                    ])
                screening_id = cursor.lastrowid
           
            # Simulate PEP search (replace with actual PEP API in production)
            matches = []
            high_risk_count = 0
           
            status = 'CLEAR' if not matches else ('POTENTIAL_MATCH' if high_risk_count > 0 else 'UNDER_REVIEW')
           
            with db_connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE external_screening_results
                    SET total_matches = %s, high_risk_matches = %s, status = %s, last_updated = %s
                    WHERE screening_id = %s
                """, [len(matches), high_risk_count, status, timezone.now(), screening_id])
           
            return {'screening_type': 'PEP', 'status': status, 'total_matches': len(matches)}
           
        except Exception as e:
            logger.error(f"PEP screening failed for vendor {vendor.id}: {str(e)}", exc_info=True)
            return None
   
    def _perform_sanctions_screening(self, vendor):
        """Perform sanctions screening"""
        try:
            db_connection = connections['tprm']
            # Get tenant_id from vendor
            tenant_id = None
            if hasattr(vendor, 'tenant_id') and vendor.tenant_id:
                tenant_id = vendor.tenant_id
            elif hasattr(vendor, 'TenantId') and vendor.TenantId:
                tenant_id = vendor.TenantId
            elif hasattr(vendor, 'tenant') and vendor.tenant:
                if hasattr(vendor.tenant, 'tenant_id'):
                    tenant_id = vendor.tenant.tenant_id
                elif isinstance(vendor.tenant, int):
                    tenant_id = vendor.tenant
            
            with db_connection.cursor() as cursor:
                if tenant_id:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated, TenantId)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'SANCTIONS',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now(),
                        tenant_id
                    ])
                else:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'SANCTIONS',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now()
                    ])
                screening_id = cursor.lastrowid
           
            # Simulate sanctions search (replace with actual sanctions API in production)
            matches = []
            high_risk_count = 0
           
            status = 'CLEAR' if not matches else ('POTENTIAL_MATCH' if high_risk_count > 0 else 'UNDER_REVIEW')
           
            with db_connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE external_screening_results
                    SET total_matches = %s, high_risk_matches = %s, status = %s, last_updated = %s
                    WHERE screening_id = %s
                """, [len(matches), high_risk_count, status, timezone.now(), screening_id])
           
            return {'screening_type': 'SANCTIONS', 'status': status, 'total_matches': len(matches)}
           
        except Exception as e:
            logger.error(f"Sanctions screening failed for vendor {vendor.id}: {str(e)}", exc_info=True)
            return None
   
    def _perform_adverse_media_screening(self, vendor):
        """Perform adverse media screening"""
        try:
            db_connection = connections['tprm']
            # Get tenant_id from vendor
            tenant_id = None
            if hasattr(vendor, 'tenant_id') and vendor.tenant_id:
                tenant_id = vendor.tenant_id
            elif hasattr(vendor, 'TenantId') and vendor.TenantId:
                tenant_id = vendor.TenantId
            elif hasattr(vendor, 'tenant') and vendor.tenant:
                if hasattr(vendor.tenant, 'tenant_id'):
                    tenant_id = vendor.tenant.tenant_id
                elif isinstance(vendor.tenant, int):
                    tenant_id = vendor.tenant
            
            with db_connection.cursor() as cursor:
                if tenant_id:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated, TenantId)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'ADVERSE_MEDIA',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now(),
                        tenant_id
                    ])
                else:
                    cursor.execute("""
                        INSERT INTO external_screening_results
                        (vendor_id, screening_type, search_terms, status, screening_date, last_updated)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [
                        vendor.id,
                        'ADVERSE_MEDIA',
                        json.dumps({
                            'company_name': vendor.company_name,
                            'legal_name': vendor.legal_name
                        }),
                        'UNDER_REVIEW',
                        timezone.now(),
                        timezone.now()
                    ])
                screening_id = cursor.lastrowid
           
            # Simulate adverse media search (replace with actual adverse media API in production)
            matches = []
            high_risk_count = 0
           
            status = 'CLEAR' if not matches else ('POTENTIAL_MATCH' if high_risk_count > 0 else 'UNDER_REVIEW')
           
            with db_connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE external_screening_results
                    SET total_matches = %s, high_risk_matches = %s, status = %s, last_updated = %s
                    WHERE screening_id = %s
                """, [len(matches), high_risk_count, status, timezone.now(), screening_id])
           
            return {'screening_type': 'ADVERSE_MEDIA', 'status': status, 'total_matches': len(matches)}
           
        except Exception as e:
            logger.error(f"Adverse media screening failed for vendor {vendor.id}: {str(e)}", exc_info=True)
            return None


class ExternalScreeningView(APIView):
    """
    API endpoint to trigger external screening for a vendor by vendor_code.
    If vendor is onboarded (exists in vendors table), gets vendor_id from temp_vendor
    and performs comprehensive screening, saving results to external_screening_results
    and screening_matches tables.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request, vendor_code):
        """
        Trigger external screening for a vendor
        
        Args:
            vendor_code: Vendor code to screen
            
        Returns:
            Response with screening results
        """
        logger.info(f"[ExternalScreeningView] Starting external screening for vendor_code: {vendor_code}")
        
        try:
            # Get tenant_id from request
            tenant_id = get_tenant_id_from_request(request)
            if not tenant_id:
                logger.warning(f"[ExternalScreeningView] No tenant_id found in request")
            
            # Step 1: Check if vendor is onboarded (exists in vendors table)
            onboarded_vendor = None
            try:
                onboarded_vendor = Vendors.objects.filter(vendor_code=vendor_code).first()
                if onboarded_vendor:
                    logger.info(f"[ExternalScreeningView] Vendor {vendor_code} is onboarded (vendor_id: {onboarded_vendor.vendor_id})")
            except Exception as e:
                logger.error(f"[ExternalScreeningView] Error checking vendors table: {str(e)}")
            
            # Step 2: Get vendor_id from temp_vendor table using vendor_code
            temp_vendor = None
            try:
                # Try to find in temp_vendor table
                temp_vendor = TempVendor.objects.using('tprm').filter(vendor_code=vendor_code).first()
                if not temp_vendor:
                    # If not found by vendor_code, try to find by matching with onboarded vendor
                    if onboarded_vendor:
                        # Try to find temp_vendor that matches the onboarded vendor's details
                        temp_vendor = TempVendor.objects.using('tprm').filter(
                            company_name=onboarded_vendor.company_name
                        ).first()
                
                if temp_vendor:
                    logger.info(f"[ExternalScreeningView] Found temp_vendor (id: {temp_vendor.id}) for vendor_code: {vendor_code}")
                else:
                    logger.warning(f"[ExternalScreeningView] No temp_vendor found for vendor_code: {vendor_code}")
            except Exception as e:
                logger.error(f"[ExternalScreeningView] Error finding temp_vendor: {str(e)}")
            
            # Step 3: If no temp_vendor found, return error
            if not temp_vendor:
                return Response({
                    'success': False,
                    'error': f'Vendor with code {vendor_code} not found in temp_vendor table. Cannot perform screening.',
                    'vendor_code': vendor_code
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Step 4: Perform comprehensive automatic screening
            # Use the existing screening methods from TempVendorManagementViewSet
            viewset = TempVendorManagementViewSet()
            screening_results = viewset._perform_automatic_screening(temp_vendor)
            
            # Step 5: Format response
            response_data = {
                'success': True,
                'message': f'External screening completed for vendor {vendor_code}',
                'vendor_code': vendor_code,
                'vendor_id': temp_vendor.id,
                'is_onboarded': onboarded_vendor is not None,
                'screening_results': screening_results or [],
                'total_screening_types': len(screening_results) if screening_results else 0
            }
            
            logger.info(f"[ExternalScreeningView] Screening completed for vendor_code: {vendor_code}, results: {len(screening_results) if screening_results else 0}")
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[ExternalScreeningView] Error performing external screening for vendor_code {vendor_code}: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'error': f'Failed to perform external screening: {str(e)}',
                'vendor_code': vendor_code
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VendorScreeningResultsView(APIView):
    """
    API endpoint to get external screening results for a vendor by vendor_code.
    Returns screening results and matches grouped by screening_date (version).
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, vendor_code):
        """
        Get screening results for a vendor
        
        Args:
            vendor_code: Vendor code to get screening results for
            start_date (optional): Filter results from this date (YYYY-MM-DD)
            end_date (optional): Filter results to this date (YYYY-MM-DD)
            
        Returns:
            Response with screening results grouped by version (screening_date)
        """
        logger.info(f"[VendorScreeningResultsView] Getting screening results for vendor_code: {vendor_code}")
        
        try:
            # Get date filters from query params
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Get tenant_id from request
            tenant_id = get_tenant_id_from_request(request)
            
            # Step 1: Find temp_vendor by vendor_code
            temp_vendor = None
            try:
                temp_vendor = TempVendor.objects.using('tprm').filter(vendor_code=vendor_code).first()
                if not temp_vendor:
                    # Try to find by matching with onboarded vendor
                    onboarded_vendor = Vendors.objects.filter(vendor_code=vendor_code).first()
                    if onboarded_vendor:
                        temp_vendor = TempVendor.objects.using('tprm').filter(
                            company_name=onboarded_vendor.company_name
                        ).first()
            except Exception as e:
                logger.error(f"[VendorScreeningResultsView] Error finding temp_vendor: {str(e)}")
            
            if not temp_vendor:
                return Response({
                    'success': False,
                    'error': f'Vendor with code {vendor_code} not found',
                    'vendor_code': vendor_code
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Step 2: Get screening results for this vendor
            screening_results_query = ExternalScreeningResult.objects.using('tprm').filter(
                vendor_id=temp_vendor.id
            )
            
            # Apply date filters
            if start_date:
                try:
                    from datetime import datetime
                    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                    screening_results_query = screening_results_query.filter(screening_date__gte=start_datetime)
                except ValueError:
                    logger.warning(f"[VendorScreeningResultsView] Invalid start_date format: {start_date}")
            
            if end_date:
                try:
                    from datetime import datetime
                    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                    # Add one day to include the entire end date
                    from datetime import timedelta
                    end_datetime = end_datetime + timedelta(days=1)
                    screening_results_query = screening_results_query.filter(screening_date__lt=end_datetime)
                except ValueError:
                    logger.warning(f"[VendorScreeningResultsView] Invalid end_date format: {end_date}")
            
            # Order by screening_date descending (newest first)
            screening_results = screening_results_query.order_by('-screening_date')
            
            # Step 3: Group results by screening_date (version) and get matches
            grouped_results = {}
            for result in screening_results:
                # Use screening_date as version key (format: YYYY-MM-DD)
                version_key = result.screening_date.strftime('%Y-%m-%d') if result.screening_date else 'unknown'
                
                if version_key not in grouped_results:
                    grouped_results[version_key] = {
                        'version': version_key,
                        'screening_date': result.screening_date.isoformat() if result.screening_date else None,
                        'results': []
                    }
                
                # Get matches for this screening result
                matches = ScreeningMatch.objects.using('tprm').filter(
                    screening_id=result.screening_id
                ).order_by('-match_score')
                
                result_data = {
                    'screening_id': result.screening_id,
                    'screening_type': result.screening_type,
                    'screening_date': result.screening_date.isoformat() if result.screening_date else None,
                    'search_terms': result.search_terms,
                    'total_matches': result.total_matches,
                    'high_risk_matches': result.high_risk_matches,
                    'status': result.status,
                    'last_updated': result.last_updated.isoformat() if result.last_updated else None,
                    'reviewed_by': result.reviewed_by,
                    'review_date': result.review_date.isoformat() if result.review_date else None,
                    'review_comments': result.review_comments,
                    'matches': [
                        {
                            'match_id': match.match_id,
                            'match_type': match.match_type,
                            'match_score': float(match.match_score) if match.match_score else 0,
                            'match_details': match.match_details,
                            'is_false_positive': match.is_false_positive,
                            'resolution_status': match.resolution_status,
                            'resolution_notes': match.resolution_notes,
                            'resolved_by': match.resolved_by,
                            'resolved_date': match.resolved_date.isoformat() if match.resolved_date else None
                        }
                        for match in matches
                    ]
                }
                
                grouped_results[version_key]['results'].append(result_data)
            
            # Convert to list and sort by version (date) descending
            versions_list = list(grouped_results.values())
            versions_list.sort(key=lambda x: x['screening_date'] or '', reverse=True)
            
            response_data = {
                'success': True,
                'vendor_code': vendor_code,
                'vendor_id': temp_vendor.id,
                'vendor_name': temp_vendor.company_name,
                'versions': versions_list,
                'total_versions': len(versions_list),
                'total_results': sum(len(v['results']) for v in versions_list)
            }
            
            logger.info(f"[VendorScreeningResultsView] Found {len(versions_list)} versions with {response_data['total_results']} total results")
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[VendorScreeningResultsView] Error getting screening results for vendor_code {vendor_code}: {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'error': f'Failed to get screening results: {str(e)}',
                'vendor_code': vendor_code
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VendorsListForDropdownView(APIView):
    """
    API endpoint to list only vendors from vendors table (for dropdown)
    Returns only onboarded vendors, not temporary vendors
    """
    permission_classes = [IsAuthenticated]
 
    def get(self, request):
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[VendorsListForDropdownView] GET request received")
        
        try:
            # Get tenant_id from request
            tenant_id = None
            if hasattr(request, 'tenant_id') and request.tenant_id:
                tenant_id = request.tenant_id
                logger.info(f"[VendorsListForDropdownView] Got tenant_id from request.tenant_id: {tenant_id}")
            elif not tenant_id:
                tenant_id = get_tenant_id_from_request(request)
                if tenant_id:
                    logger.info(f"[VendorsListForDropdownView] Got tenant_id from get_tenant_id_from_request: {tenant_id}")
                else:
                    logger.warning(f"[VendorsListForDropdownView] No tenant_id found - will fetch all vendors")
           
            # Get vendors only from vendors table (onboarded vendors)
            from tprm_backend.apps.vendor_core.models import Vendors
            from .serializers import AllVendorsListSerializer
           
            vendors_qs = Vendors.objects.all()
            initial_count = vendors_qs.count()
            logger.info(f"[VendorsListForDropdownView] Initial vendors queryset count (before tenant filter): {initial_count}")
            
            if tenant_id:
                # Try tenant_id first (Django auto-created field for ForeignKey)
                try:
                    vendors_qs = vendors_qs.filter(tenant_id=tenant_id)
                    filtered_count = vendors_qs.count()
                    logger.info(f"[VendorsListForDropdownView] Filtered by tenant_id={tenant_id}, count: {filtered_count}")
                except Exception as filter_error:
                    # Fallback: try using tenant__id if tenant_id doesn't work
                    logger.warning(f"[VendorsListForDropdownView] tenant_id filter failed, trying tenant__id: {filter_error}")
                    try:
                        vendors_qs = Vendors.objects.filter(tenant__id=tenant_id)
                        filtered_count = vendors_qs.count()
                        logger.info(f"[VendorsListForDropdownView] Filtered by tenant__id={tenant_id}, count: {filtered_count}")
                    except Exception as filter_error2:
                        logger.error(f"[VendorsListForDropdownView] Both tenant_id and tenant__id filters failed: {filter_error2}")
                        # Continue without filter - return all vendors
                        vendors_qs = Vendors.objects.all()
            else:
                logger.info(f"[VendorsListForDropdownView] No tenant_id filter applied, fetching all vendors")
           
            vendors_qs = vendors_qs.order_by('company_name')
            vendors = list(vendors_qs)
            logger.info(f"[VendorsListForDropdownView] Retrieved {len(vendors)} vendor objects from database")
           
            # Serialize vendors
            serializer = AllVendorsListSerializer(vendors, many=True)
            vendor_list = serializer.data
            logger.info(f"[VendorsListForDropdownView] Serialized {len(vendor_list)} vendors")
           
            # Ensure vendor_id is set and set vendor_type fields for onboarded vendors
            for vendor in vendor_list:
                # Ensure vendor_id is present (should already be there from serializer)
                if 'vendor_id' not in vendor:
                    vendor['vendor_id'] = vendor.get('id')
                    logger.debug(f"[VendorsListForDropdownView] Added vendor_id from id field for vendor: {vendor.get('company_name')}")
                # Set vendor type fields for onboarded vendors
                vendor['vendor_type'] = 'ONBOARDED'
                vendor['vendor_type_label'] = 'Onboarded Vendor'
                vendor['is_temporary'] = False
                vendor['response_id'] = None  # Will be None for vendors-only endpoint
           
            logger.info(f"[VendorsListForDropdownView] Returning {len(vendor_list)} vendors from vendors table")
            
            # Log sample vendor data for debugging
            if vendor_list:
                sample_vendor = vendor_list[0]
                logger.info(f"[VendorsListForDropdownView] Sample vendor data: vendor_id={sample_vendor.get('vendor_id')}, company_name={sample_vendor.get('company_name')}, vendor_code={sample_vendor.get('vendor_code')}")
           
            return Response({
                'success': True,
                'data': vendor_list,
                'total': len(vendor_list)
            }, status=status.HTTP_200_OK)
           
        except Exception as e:
            logger.exception(f"[VendorsListForDropdownView] Error: {e}")
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
 
class VendorRisksListView(APIView):
    """
    API endpoint to list all vendor risks from risk_tprm table
    """
    permission_classes = [IsAuthenticated]
 
    def get(self, request):
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[VendorRisksListView] GET request received for /api/v1/management/vendor-risks/")
       
        try:
            # Get tenant_id from request
            tenant_id = None
           
            # Method 1: Check if already set on request
            if hasattr(request, 'tenant_id') and request.tenant_id:
                tenant_id = request.tenant_id
                logger.info(f"[VendorRisksListView] Got tenant_id from request.tenant_id: {tenant_id}")
           
            # Method 2: Try tenant utils function
            if not tenant_id:
                tenant_id = get_tenant_id_from_request(request)
                if tenant_id:
                    logger.info(f"[VendorRisksListView] Got tenant_id from get_tenant_id_from_request: {tenant_id}")
           
            # Method 3: Try to extract from JWT token payload
            if not tenant_id:
                try:
                    auth_header = request.headers.get('Authorization', '')
                    if auth_header.startswith('Bearer '):
                        token = auth_header.split(' ')[1]
                        import jwt
                        from django.conf import settings
                        try:
                            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'], options={"verify_signature": False})
                            if 'tenant_id' in payload:
                                tenant_id = payload['tenant_id']
                                logger.info(f"[VendorRisksListView] Got tenant_id from JWT token: {tenant_id}")
                        except Exception as jwt_error:
                            logger.debug(f"[VendorRisksListView] Could not decode JWT: {jwt_error}")
                except Exception as e:
                    logger.debug(f"[VendorRisksListView] Error extracting tenant_id from JWT: {e}")
           
            logger.info(f"[VendorRisksListView] Final tenant_id: {tenant_id}")
           
            # Import RiskTPRM model
            from tprm_backend.apps.vendor_risk.models import RiskTPRM
            from .serializers import VendorRiskSerializer
            from django.db import connections
           
            # Use 'tprm' database connection which points to tprm_integration schema
            db_connection = 'tprm'
           
            # Build query with tenant filtering
            queryset = RiskTPRM.objects.using(db_connection).all()
           
            # Filter by entity only if explicitly provided (default: show all risks)
            entity_filter = request.query_params.get('entity', None)
            if entity_filter:
                queryset = queryset.filter(entity=entity_filter)
                logger.info(f"[VendorRisksListView] Filtering by entity: {entity_filter}")
            else:
                logger.info(f"[VendorRisksListView] Showing all risks (no entity filter)")
           
            # Apply tenant filter if tenant_id is available
            # RiskTPRM has a ForeignKey 'tenant' with db_column='TenantId', so use tenant_id
            if tenant_id:
                from django.db.models import Q
                # Include records with matching tenant_id OR NULL tenant_id (for backward compatibility)
                queryset = queryset.filter(Q(tenant_id=tenant_id) | Q(tenant_id__isnull=True))
                logger.info(f"[VendorRisksListView] Filtering by tenant_id: {tenant_id}")
            else:
                logger.warning(f"[VendorRisksListView] No tenant_id found - showing ALL risks (no tenant filtering)")
           
            # Apply additional filters
            status_filter = request.query_params.get('status', None)
            if status_filter:
                queryset = queryset.filter(status=status_filter)
           
            priority_filter = request.query_params.get('priority', None)
            if priority_filter:
                queryset = queryset.filter(priority=priority_filter)
           
            risk_type_filter = request.query_params.get('risk_type', None)
            if risk_type_filter:
                queryset = queryset.filter(risk_type=risk_type_filter)
           
            # Vendor ID filter - filters risks across all sources (vendor, BCP/DRP, RFP, Contracts)
            # IMPORTANT: Direct vendor risks use temp_vendor.id, while RFP/BCP/DRP/Contracts use vendors.vendor_id
            vendor_id_filter = request.query_params.get('vendor_id', None)
            if vendor_id_filter:
                try:
                    vendor_id = int(vendor_id_filter)
                    from django.db.models import Q
                    from tprm_backend.apps.vendor_core.models import Vendors, TempVendor
                   
                    # Determine which ID type we have and find the corresponding IDs
                    temp_vendor_id = None
                    vendors_vendor_id = None
                    vendor_code = None
                   
                    # Check if it's a vendors.vendor_id
                    try:
                        vendor_filter = {'vendor_id': vendor_id}
                        if tenant_id:
                            vendor_filter['tenant_id'] = tenant_id
                        vendor = Vendors.objects.using(db_connection).filter(**vendor_filter).first()
                        if vendor:
                            vendors_vendor_id = vendor_id
                            vendor_code = vendor.vendor_code
                            # Find corresponding temp_vendor by vendor_code
                            if vendor_code:
                                temp_filter = {'vendor_code': vendor_code}
                                if tenant_id:
                                    temp_filter['tenant_id'] = tenant_id
                                temp_vendor = TempVendor.objects.using(db_connection).filter(**temp_filter).first()
                                if temp_vendor:
                                    temp_vendor_id = temp_vendor.id
                                    logger.info(f"[VendorRisksListView] Found vendors.vendor_id={vendor_id}, mapped to temp_vendor.id={temp_vendor_id}")
                    except Exception as e:
                        logger.debug(f"[VendorRisksListView] vendor_id {vendor_id} not found in vendors table: {e}")
                   
                    # Check if it's a temp_vendor.id
                    if not temp_vendor_id:
                        try:
                            temp_filter = {'id': vendor_id}
                            if tenant_id:
                                temp_filter['tenant_id'] = tenant_id
                            temp_vendor = TempVendor.objects.using(db_connection).filter(**temp_filter).first()
                            if temp_vendor:
                                temp_vendor_id = vendor_id
                                vendor_code = temp_vendor.vendor_code
                                # Find corresponding vendors.vendor_id by vendor_code
                                if vendor_code:
                                    vendor_filter = {'vendor_code': vendor_code}
                                    if tenant_id:
                                        vendor_filter['tenant_id'] = tenant_id
                                    vendor = Vendors.objects.using(db_connection).filter(**vendor_filter).first()
                                    if vendor:
                                        vendors_vendor_id = vendor.vendor_id
                                        logger.info(f"[VendorRisksListView] Found temp_vendor.id={vendor_id}, mapped to vendors.vendor_id={vendors_vendor_id}")
                        except Exception as e:
                            logger.debug(f"[VendorRisksListView] vendor_id {vendor_id} not found in temp_vendor table: {e}")
                   
                    # Build conditions for different risk sources
                    vendor_conditions = []
                   
                    # 1. Direct vendor risks - use temp_vendor.id (linked to temp_vendor)
                    if temp_vendor_id:
                        vendor_risk_condition = Q(
                            entity='vendor_management',
                            row=str(temp_vendor_id)
                        )
                        # Handle JSONField data - check if it contains 'temp_vendor' or equals it
                        vendor_risk_condition &= (
                            Q(data__icontains='temp_vendor') |
                            Q(data='temp_vendor')
                        )
                        vendor_conditions.append(vendor_risk_condition)
                        logger.info(f"[VendorRisksListView] Filtering direct vendor risks by temp_vendor.id={temp_vendor_id}")
                   
                    # 2. BCP/DRP Plan risks - use vendors.vendor_id (linked to vendors table)
                    if vendors_vendor_id:
                        try:
                            from tprm_backend.bcpdrp.models import Plan
                            bcp_plan_ids = list(Plan.objects.using(db_connection).filter(vendor_id=vendors_vendor_id).values_list('plan_id', flat=True))
                            bcp_plan_ids_str = [str(pid) for pid in bcp_plan_ids]
                           
                            if bcp_plan_ids_str:
                                bcp_risk_condition = Q(
                                    entity__in=['bcp_drp_module', 'BCP_DRP', 'bcp_drp'],
                                    row__in=bcp_plan_ids_str
                                )
                                # Handle JSONField data - check for BCP/DRP related data values
                                # Comprehensive risks use 'comprehensive_plan_data', regular risks use 'bcp_drp_plans' or 'bcp_drp_evaluations'
                                bcp_risk_condition &= (
                                    Q(data__icontains='bcp_drp_plans') |
                                    Q(data='bcp_drp_plans') |
                                    Q(data__icontains='bcp_drp_evaluations') |
                                    Q(data='bcp_drp_evaluations') |
                                    Q(data__icontains='comprehensive_plan_data') |
                                    Q(data='comprehensive_plan_data')
                                )
                                vendor_conditions.append(bcp_risk_condition)
                                logger.info(f"[VendorRisksListView] Found {len(bcp_plan_ids)} BCP/DRP plans for vendors.vendor_id={vendors_vendor_id}")
                        except Exception as bcp_error:
                            logger.warning(f"[VendorRisksListView] Error fetching BCP/DRP plans: {bcp_error}")
                   
                    # 3. RFP Response risks - use vendors.vendor_id (linked to vendors table)
                    if vendors_vendor_id:
                        try:
                            from tprm_backend.rfp.models import RFPResponse
                            rfp_response_ids = list(RFPResponse.objects.using(db_connection).filter(vendor_id=vendors_vendor_id).values_list('response_id', flat=True))
                            rfp_response_ids_str = [str(rid) for rid in rfp_response_ids]
                           
                            if rfp_response_ids_str:
                                rfp_risk_condition = Q(
                                    entity__in=['rfp_module', 'RFP', 'rfp'],
                                    row__in=rfp_response_ids_str
                                )
                                # Handle JSONField data - check for RFP related data values
                                # Regular risks use 'rfp_responses' or 'rfps', comprehensive risks may use 'comprehensive_rfp_data'
                                rfp_risk_condition &= (
                                    Q(data__icontains='rfp_responses') |
                                    Q(data='rfp_responses') |
                                    Q(data__icontains='rfps') |
                                    Q(data='rfps') |
                                    Q(data__icontains='comprehensive_rfp_data') |
                                    Q(data='comprehensive_rfp_data')
                                )
                                vendor_conditions.append(rfp_risk_condition)
                                logger.info(f"[VendorRisksListView] Found {len(rfp_response_ids)} RFP responses for vendors.vendor_id={vendors_vendor_id}")
                        except Exception as rfp_error:
                            logger.warning(f"[VendorRisksListView] Error fetching RFP responses: {rfp_error}")
                   
                    # 4. Contract risks - use vendors.vendor_id (linked to vendors table via ForeignKey)
                    if vendors_vendor_id:
                        try:
                            from tprm_backend.apps.vendor_lifecycle.models import VendorContracts
                            contract_ids = list(VendorContracts.objects.using(db_connection).filter(vendor_id=vendors_vendor_id).values_list('contract_id', flat=True))
                            contract_ids_str = [str(cid) for cid in contract_ids]
                           
                            if contract_ids_str:
                                contract_risk_condition = Q(
                                    entity__in=['contract_module', 'Contract', 'contract'],
                                    row__in=contract_ids_str
                                )
                                # Handle JSONField data - check for Contract related data values
                                # Regular risks use 'vendor_contracts', comprehensive risks may use 'comprehensive_contract_data'
                                contract_risk_condition &= (
                                    Q(data__icontains='vendor_contracts') |
                                    Q(data='vendor_contracts') |
                                    Q(data__icontains='contracts') |
                                    Q(data='contracts') |
                                    Q(data__icontains='contract_terms') |
                                    Q(data='contract_terms') |
                                    Q(data__icontains='contract_clauses') |
                                    Q(data='contract_clauses') |
                                    Q(data__icontains='comprehensive_contract_data') |
                                    Q(data='comprehensive_contract_data')
                                )
                                vendor_conditions.append(contract_risk_condition)
                                logger.info(f"[VendorRisksListView] Found {len(contract_ids)} contracts for vendors.vendor_id={vendors_vendor_id}")
                        except Exception as contract_error:
                            logger.warning(f"[VendorRisksListView] Error fetching contracts: {contract_error}")
                   
                    # Combine all conditions with OR
                    if vendor_conditions:
                        combined_condition = vendor_conditions[0]
                        for condition in vendor_conditions[1:]:
                            combined_condition |= condition
                        queryset = queryset.filter(combined_condition)
                        logger.info(f"[VendorRisksListView] Filtering by vendor_id: {vendor_id} (temp_vendor.id={temp_vendor_id}, vendors.vendor_id={vendors_vendor_id}) with {len(vendor_conditions)} condition(s)")
                    else:
                        # No matching conditions found - return empty queryset
                        queryset = queryset.none()
                        logger.info(f"[VendorRisksListView] No risks found for vendor_id: {vendor_id} (temp_vendor.id={temp_vendor_id}, vendors.vendor_id={vendors_vendor_id})")
                       
                except (ValueError, TypeError) as e:
                    logger.warning(f"[VendorRisksListView] Invalid vendor_id filter: {vendor_id_filter}, error: {e}")
                    # Don't filter if vendor_id is invalid
           
            # Search filter
            search_term = request.query_params.get('search', None)
            if search_term:
                queryset = queryset.filter(
                    Q(title__icontains=search_term) |
                    Q(description__icontains=search_term)
                )
           
            # Order by created_at descending
            queryset = queryset.order_by('-created_at')
           
            # Pagination
            page_size = int(request.query_params.get('page_size', 50))
            page = int(request.query_params.get('page', 1))
            start = (page - 1) * page_size
            end = start + page_size
           
            total_count = queryset.count()
            risks = queryset[start:end]
           
            # Get all unique entity values from the database (for dropdown)
            unique_entities = RiskTPRM.objects.using(db_connection).exclude(
                entity__isnull=True
            ).exclude(
                entity=''
            ).values_list('entity', flat=True).distinct().order_by('entity')
            unique_entities_list = list(unique_entities)
           
            # Serialize the data
            serializer = VendorRiskSerializer(risks, many=True)
           
            return Response({
                'success': True,
                'data': serializer.data,
                'total': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size if page_size > 0 else 1,
                'unique_entities': unique_entities_list
            }, status=status.HTTP_200_OK)
           
        except Exception as e:
            logger.exception(f"[VendorRisksListView] Unexpected error: {e}")
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VendorRisksExportExcelView(APIView):
    """
    API endpoint to export vendor risks to Excel file
    Exports all filtered risks with all columns from risk_tprm table
    """
    permission_classes = [IsAuthenticated]
 
    def get(self, request):
        import logging
        import io
        from datetime import datetime
        from django.http import HttpResponse
       
        logger = logging.getLogger(__name__)
        logger.info(f"[VendorRisksExportExcelView] GET request received for Excel export")
       
        try:
            # Try to import openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter
                OPENPYXL_AVAILABLE = True
            except ImportError:
                OPENPYXL_AVAILABLE = False
                logger.warning("[VendorRisksExportExcelView] openpyxl not available, using CSV fallback")
           
            # Get tenant_id from request (same logic as VendorRisksListView)
            tenant_id = None
           
            if hasattr(request, 'tenant_id') and request.tenant_id:
                tenant_id = request.tenant_id
                logger.info(f"[VendorRisksExportExcelView] Got tenant_id from request.tenant_id: {tenant_id}")
           
            if not tenant_id:
                tenant_id = get_tenant_id_from_request(request)
                if tenant_id:
                    logger.info(f"[VendorRisksExportExcelView] Got tenant_id from get_tenant_id_from_request: {tenant_id}")
           
            if not tenant_id:
                try:
                    auth_header = request.headers.get('Authorization', '')
                    if auth_header.startswith('Bearer '):
                        token = auth_header.split(' ')[1]
                        import jwt
                        from django.conf import settings
                        try:
                            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'], options={"verify_signature": False})
                            if 'tenant_id' in payload:
                                tenant_id = payload['tenant_id']
                                logger.info(f"[VendorRisksExportExcelView] Got tenant_id from JWT token: {tenant_id}")
                        except Exception as jwt_error:
                            logger.debug(f"[VendorRisksExportExcelView] Could not decode JWT: {jwt_error}")
                except Exception as e:
                    logger.debug(f"[VendorRisksExportExcelView] Error extracting tenant_id from JWT: {e}")
           
            logger.info(f"[VendorRisksExportExcelView] Final tenant_id: {tenant_id}")
           
            # Import RiskTPRM model
            from tprm_backend.apps.vendor_risk.models import RiskTPRM
            from django.db import connections
            from django.db.models import Q
           
            # Use 'tprm' database connection
            db_connection = 'tprm'
           
            # Build query with same filters as VendorRisksListView
            queryset = RiskTPRM.objects.using(db_connection).all()
           
            # Filter by entity if provided
            entity_filter = request.query_params.get('entity', None)
            if entity_filter:
                queryset = queryset.filter(entity=entity_filter)
                logger.info(f"[VendorRisksExportExcelView] Filtering by entity: {entity_filter}")
           
            # Apply tenant filter
            if tenant_id:
                queryset = queryset.filter(Q(tenant_id=tenant_id) | Q(tenant_id__isnull=True))
                logger.info(f"[VendorRisksExportExcelView] Filtering by tenant_id: {tenant_id}")
            else:
                logger.warning(f"[VendorRisksExportExcelView] No tenant_id found - showing ALL risks")
           
            # Apply additional filters
            status_filter = request.query_params.get('status', None)
            if status_filter:
                queryset = queryset.filter(status=status_filter)
           
            priority_filter = request.query_params.get('priority', None)
            if priority_filter:
                queryset = queryset.filter(priority=priority_filter)
           
            risk_type_filter = request.query_params.get('risk_type', None)
            if risk_type_filter:
                queryset = queryset.filter(risk_type=risk_type_filter)
           
            # Vendor ID filter (same complex logic as VendorRisksListView)
            vendor_id_filter = request.query_params.get('vendor_id', None)
            if vendor_id_filter:
                try:
                    vendor_id = int(vendor_id_filter)
                    from tprm_backend.apps.vendor_core.models import Vendors, TempVendor
                   
                    temp_vendor_id = None
                    vendors_vendor_id = None
                    vendor_code = None
                   
                    # Check if it's a vendors.vendor_id
                    try:
                        vendor_filter = {'vendor_id': vendor_id}
                        if tenant_id:
                            vendor_filter['tenant_id'] = tenant_id
                        vendor = Vendors.objects.using(db_connection).filter(**vendor_filter).first()
                        if vendor:
                            vendors_vendor_id = vendor_id
                            vendor_code = vendor.vendor_code
                            if vendor_code:
                                temp_filter = {'vendor_code': vendor_code}
                                if tenant_id:
                                    temp_filter['tenant_id'] = tenant_id
                                temp_vendor = TempVendor.objects.using(db_connection).filter(**temp_filter).first()
                                if temp_vendor:
                                    temp_vendor_id = temp_vendor.id
                    except Exception as e:
                        logger.debug(f"[VendorRisksExportExcelView] vendor_id {vendor_id} not found in vendors table: {e}")
                   
                    # Check if it's a temp_vendor.id
                    if not temp_vendor_id:
                        try:
                            temp_filter = {'id': vendor_id}
                            if tenant_id:
                                temp_filter['tenant_id'] = tenant_id
                            temp_vendor = TempVendor.objects.using(db_connection).filter(**temp_filter).first()
                            if temp_vendor:
                                temp_vendor_id = vendor_id
                                vendor_code = temp_vendor.vendor_code
                                if vendor_code:
                                    vendor_filter = {'vendor_code': vendor_code}
                                    if tenant_id:
                                        vendor_filter['tenant_id'] = tenant_id
                                    vendor = Vendors.objects.using(db_connection).filter(**vendor_filter).first()
                                    if vendor:
                                        vendors_vendor_id = vendor.vendor_id
                        except Exception as e:
                            logger.debug(f"[VendorRisksExportExcelView] vendor_id {vendor_id} not found in temp_vendor table: {e}")
                   
                    # Build conditions for different risk sources
                    vendor_conditions = []
                   
                    if temp_vendor_id:
                        vendor_risk_condition = Q(
                            entity='vendor_management',
                            row=str(temp_vendor_id)
                        )
                        vendor_risk_condition &= (
                            Q(data__icontains='temp_vendor') |
                            Q(data='temp_vendor')
                        )
                        vendor_conditions.append(vendor_risk_condition)
                   
                    if vendors_vendor_id:
                        try:
                            from tprm_backend.bcpdrp.models import Plan
                            bcp_plan_ids = list(Plan.objects.using(db_connection).filter(vendor_id=vendors_vendor_id).values_list('plan_id', flat=True))
                            bcp_plan_ids_str = [str(pid) for pid in bcp_plan_ids]
                           
                            if bcp_plan_ids_str:
                                bcp_risk_condition = Q(
                                    entity__in=['bcp_drp_module', 'BCP_DRP', 'bcp_drp'],
                                    row__in=bcp_plan_ids_str
                                )
                                bcp_risk_condition &= (
                                    Q(data__icontains='bcp_drp_plans') |
                                    Q(data='bcp_drp_plans') |
                                    Q(data__icontains='bcp_drp_evaluations') |
                                    Q(data='bcp_drp_evaluations') |
                                    Q(data__icontains='comprehensive_plan_data') |
                                    Q(data='comprehensive_plan_data')
                                )
                                vendor_conditions.append(bcp_risk_condition)
                        except Exception as bcp_error:
                            logger.warning(f"[VendorRisksExportExcelView] Error fetching BCP/DRP plans: {bcp_error}")
                       
                        try:
                            from tprm_backend.rfp.models import RFPResponse
                            rfp_response_ids = list(RFPResponse.objects.using(db_connection).filter(vendor_id=vendors_vendor_id).values_list('response_id', flat=True))
                            rfp_response_ids_str = [str(rid) for rid in rfp_response_ids]
                           
                            if rfp_response_ids_str:
                                rfp_risk_condition = Q(
                                    entity__in=['rfp_module', 'RFP', 'rfp'],
                                    row__in=rfp_response_ids_str
                                )
                                rfp_risk_condition &= (
                                    Q(data__icontains='rfp_responses') |
                                    Q(data='rfp_responses') |
                                    Q(data__icontains='rfps') |
                                    Q(data='rfps') |
                                    Q(data__icontains='comprehensive_rfp_data') |
                                    Q(data='comprehensive_rfp_data')
                                )
                                vendor_conditions.append(rfp_risk_condition)
                        except Exception as rfp_error:
                            logger.warning(f"[VendorRisksExportExcelView] Error fetching RFP responses: {rfp_error}")
                       
                        try:
                            from tprm_backend.apps.vendor_lifecycle.models import VendorContracts
                            contract_ids = list(VendorContracts.objects.using(db_connection).filter(vendor_id=vendors_vendor_id).values_list('contract_id', flat=True))
                            contract_ids_str = [str(cid) for cid in contract_ids]
                           
                            if contract_ids_str:
                                contract_risk_condition = Q(
                                    entity__in=['contract_module', 'Contract', 'contract'],
                                    row__in=contract_ids_str
                                )
                                contract_risk_condition &= (
                                    Q(data__icontains='vendor_contracts') |
                                    Q(data='vendor_contracts') |
                                    Q(data__icontains='contracts') |
                                    Q(data='contracts') |
                                    Q(data__icontains='contract_terms') |
                                    Q(data='contract_terms') |
                                    Q(data__icontains='contract_clauses') |
                                    Q(data='contract_clauses') |
                                    Q(data__icontains='comprehensive_contract_data') |
                                    Q(data='comprehensive_contract_data')
                                )
                                vendor_conditions.append(contract_risk_condition)
                        except Exception as contract_error:
                            logger.warning(f"[VendorRisksExportExcelView] Error fetching contracts: {contract_error}")
                   
                    if vendor_conditions:
                        combined_condition = vendor_conditions[0]
                        for condition in vendor_conditions[1:]:
                            combined_condition |= condition
                        queryset = queryset.filter(combined_condition)
                except (ValueError, TypeError) as e:
                    logger.warning(f"[VendorRisksExportExcelView] Invalid vendor_id filter: {vendor_id_filter}, error: {e}")
           
            # Search filter
            search_term = request.query_params.get('search', None)
            if search_term:
                queryset = queryset.filter(
                    Q(title__icontains=search_term) |
                    Q(description__icontains=search_term)
                )
           
            # Order by created_at descending
            queryset = queryset.order_by('-created_at')
           
            # Get all risks (no pagination for export)
            risks = list(queryset.values())
            total_count = len(risks)
           
            logger.info(f"[VendorRisksExportExcelView] Exporting {total_count} risks to Excel")
           
            if OPENPYXL_AVAILABLE:
                # Create Excel workbook
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Vendor Risks"
               
                # Define all columns from risk_tprm table
                headers = [
                    'ID', 'Title', 'Description', 'Likelihood', 'Impact', 'Score',
                    'Priority', 'Status', 'Risk Type', 'Entity', 'Exposure Rating',
                    'AI Explanation', 'Suggested Mitigations', 'Data', 'Row',
                    'Created At', 'Updated At', 'Tenant ID'
                ]
               
                # Write headers with styling
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
               
                # Write data rows
                for row_idx, risk in enumerate(risks, 2):
                    # Handle JSONField data - convert to string if it's a dict/list
                    data_value = risk.get('data')
                    if isinstance(data_value, (dict, list)):
                        import json
                        data_value = json.dumps(data_value)
                   
                    # Handle datetime fields
                    created_at = risk.get('created_at')
                    if created_at:
                        if isinstance(created_at, str):
                            try:
                                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                            except:
                                pass
                        if isinstance(created_at, datetime):
                            created_at = created_at.strftime('%Y-%m-%d %H:%M:%S')
                   
                    updated_at = risk.get('updated_at')
                    if updated_at:
                        if isinstance(updated_at, str):
                            try:
                                updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
                            except:
                                pass
                        if isinstance(updated_at, datetime):
                            updated_at = updated_at.strftime('%Y-%m-%d %H:%M:%S')
                   
                    sheet.cell(row=row_idx, column=1, value=risk.get('id', ''))
                    sheet.cell(row=row_idx, column=2, value=risk.get('title', ''))
                    sheet.cell(row=row_idx, column=3, value=risk.get('description', ''))
                    sheet.cell(row=row_idx, column=4, value=risk.get('likelihood', ''))
                    sheet.cell(row=row_idx, column=5, value=risk.get('impact', ''))
                    sheet.cell(row=row_idx, column=6, value=risk.get('score', ''))
                    sheet.cell(row=row_idx, column=7, value=risk.get('priority', ''))
                    sheet.cell(row=row_idx, column=8, value=risk.get('status', ''))
                    sheet.cell(row=row_idx, column=9, value=risk.get('risk_type', ''))
                    sheet.cell(row=row_idx, column=10, value=risk.get('entity', ''))
                    sheet.cell(row=row_idx, column=11, value=risk.get('exposure_rating', ''))
                    sheet.cell(row=row_idx, column=12, value=risk.get('ai_explanation', ''))
                    sheet.cell(row=row_idx, column=13, value=risk.get('suggested_mitigations', ''))
                    sheet.cell(row=row_idx, column=14, value=data_value)
                    sheet.cell(row=row_idx, column=15, value=risk.get('row', ''))
                    sheet.cell(row=row_idx, column=16, value=created_at)
                    sheet.cell(row=row_idx, column=17, value=updated_at)
                    sheet.cell(row=row_idx, column=18, value=risk.get('tenant_id', ''))
               
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
               
                # Save to buffer
                excel_buffer = io.BytesIO()
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
               
                # Create response
                response = HttpResponse(
                    excel_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"vendor-risks-export-{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
               
                logger.info(f"[VendorRisksExportExcelView] Excel export completed: {total_count} risks exported")
                return response
            else:
                # Fallback to CSV if openpyxl is not available
                response = HttpResponse(content_type='text/csv')
                filename = f"vendor-risks-export-{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.csv"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
               
                writer = csv.writer(response)
               
                # Write headers
                headers = [
                    'ID', 'Title', 'Description', 'Likelihood', 'Impact', 'Score',
                    'Priority', 'Status', 'Risk Type', 'Entity', 'Exposure Rating',
                    'AI Explanation', 'Suggested Mitigations', 'Data', 'Row',
                    'Created At', 'Updated At', 'Tenant ID'
                ]
                writer.writerow(headers)
               
                # Write data rows
                for risk in risks:
                    data_value = risk.get('data')
                    if isinstance(data_value, (dict, list)):
                        import json
                        data_value = json.dumps(data_value)
                   
                    created_at = risk.get('created_at')
                    if created_at and isinstance(created_at, datetime):
                        created_at = created_at.strftime('%Y-%m-%d %H:%M:%S')
                   
                    updated_at = risk.get('updated_at')
                    if updated_at and isinstance(updated_at, datetime):
                        updated_at = updated_at.strftime('%Y-%m-%d %H:%M:%S')
                   
                    writer.writerow([
                        risk.get('id', ''),
                        risk.get('title', ''),
                        risk.get('description', ''),
                        risk.get('likelihood', ''),
                        risk.get('impact', ''),
                        risk.get('score', ''),
                        risk.get('priority', ''),
                        risk.get('status', ''),
                        risk.get('risk_type', ''),
                        risk.get('entity', ''),
                        risk.get('exposure_rating', ''),
                        risk.get('ai_explanation', ''),
                        risk.get('suggested_mitigations', ''),
                        data_value,
                        risk.get('row', ''),
                        created_at,
                        updated_at,
                        risk.get('tenant_id', '')
                    ])
               
                logger.info(f"[VendorRisksExportExcelView] CSV export completed: {total_count} risks exported")
                return response
           
        except Exception as e:
            logger.exception(f"[VendorRisksExportExcelView] Unexpected error: {e}")
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)